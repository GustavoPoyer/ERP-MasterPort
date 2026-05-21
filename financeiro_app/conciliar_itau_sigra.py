"""
Script de Conciliação de Extratos Itaú com Comprovantes SIGRA

Este script concilia lançamentos de extratos bancários do Itaú com comprovantes
do arquivo "pgtos sigra", identificando quais comprovantes compõem cada lançamento do extrato.

Autor: Automação Financeira
Data: 2026-02-02
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from itertools import combinations
from collections import Counter
import os
import re
import glob
import sys
import warnings
warnings.filterwarnings('ignore')

if sys.platform == 'win32':
    # Evita substituir sys.stdout/sys.stderr por novos wrappers.
    # A substituição pode fechar o stream antigo no GC e causar:
    # ValueError('I/O operation on closed file.') / lost sys.stderr.
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False

# ============================================================================
# CONFIGURAÇÕES
# ============================================================================

# Conciliação exige o MESMO dia de calendário (extrato e comprovante). 0 = sem folga em dias.
TOLERANCIA_DATA = 0
# Valores: conciliação exige igualdade exata em centavos (extrato vs comprovante ou soma de comprovantes)

# O extrato costuma ser acumulativo (vários meses); o PGTO costuma ser só o mês corrente.
# Se True, restringe quais datas do extrato entram na conciliação SIGRA (ver opção abaixo).
CONCILIAR_APENAS_EXTRATO_DENTRO_PERIODO_PGTO = True
# Se True (e a opção acima for True): só tenta SIGRA em extratos cuja data é exatamente um dia em que
# existe pelo menos um comprovante no PGTO (coluna Criação). Evita incluir dias “vazios” entre min e max.
# Se False: usa faixa contínua da menor à maior data do PGTO.
CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO = True

# Empate: vários comprovantes com o mesmo valor no mesmo dia (ou vários processos RF com a mesma soma),
# sem como desempatar pelo texto/referência do extrato.
# True  = não concilia nesses casos (mais seguro; taxa cai muito se o PGTO repete valores no dia).
# False = escolhe de forma determinística (primeiro candidato / menor ID / Ref. Sigra ordenada) — data e valor continuam exatos.
EMPATES_SEM_REF_EXIGIR_DESISTENCIA = False

# Caminhos dos arquivos (serão buscados automaticamente na pasta do dia)
CAMINHO_EXTRATO = None  # Será definido automaticamente
CAMINHO_COMPROVANTES = None  # Será definido automaticamente

# Nomes das colunas esperadas - Formato Itaú
# Formato antigo (aba Lançamentos): A=Data, B=Lançamento, C=Razão Social, E=Valor (R$)
# Formato novo (aba Extrato_*): B=Data, C=Lançamento, F=Ag./Origem, G=Valor (R$), H=Saldo (R$) - SEM Razão Social
COLUNAS_EXTRATO = {
    'data': None,
    'valor': 'Valor (R$)',
    'favorecido': 'Razão Social',  # Formato antigo; novo formato usa Lançamento como fallback
    'lancamento': 'Lançamento',
    'descricao': None,
    'valor_alt': ['VALOR', 'Valor', 'Valor (R$)'],
    'favorecido_alt': ['DESTINO', 'Razão Social', 'Razao Social', 'Favorecido', 'Lançamento'],  # Lançamento como fallback no formato novo
    'lancamento_alt': ['Lançamento', 'Descrição', 'Histórico', 'Descricao']
}

# Comprovantes SIGRA (pgtos sigra):
# Coluna D = Ref. Sigra (referência do processo), Coluna H = Valor, Coluna J = Criação (data),
# Coluna O = Fornecedor (filtrar "Receita Federal" para casar cada linha com PUCOMEX no extrato).
COLUNAS_COMPROVANTES = {
    'data': 'Criação',  # Coluna J: Criação (DD/MM/YY HH:MM)
    'valor': 'Valor',  # Coluna H: Valor
    'favorecido': 'Fornecedor',  # Coluna O: Fornecedor
    'descricao': 'Categoria',  # Coluna M: Categoria
    'rps': 'RPS'  # Coluna L: RPS (opcional, para busca por código)
}

from numerario_itau import COLUNAS_NUMERARIO, conciliar_extrato_numerario, ler_numerario

# Caminho de saída
# Pode ser absoluto (recomendado) ou relativo ao diretório deste script.
CAMINHO_SAIDA = r'G:\Drives compartilhados\automação\Conciliações\conciliacao_itau_sigra.xlsx'


# ============================================================================
# FUNÇÕES DE LEITURA E NORMALIZAÇÃO
# ============================================================================

def ler_extrato(caminho):
    """
    Lê o arquivo Excel do extrato bancário Itaú.
    Suporta formatos: aba "Lançamentos" (antigo) ou aba "Extrato_*" (novo - Data, Lançamento, Ag./Origem, Valor (R$)).
    
    Args:
        caminho: Caminho do arquivo Excel
        
    Returns:
        DataFrame com os dados do extrato
    """
    try:
        xl = pd.ExcelFile(caminho, engine='openpyxl')

        def pontuar_colunas(df_teste):
            cols = [str(c).lower() for c in df_teste.columns]
            score = 0
            if any('data' in c for c in cols):
                score += 2
            if any('valor' in c for c in cols):
                score += 2
            if any('destino' in c or 'lançamento' in c or 'lancamento' in c or 'razão social' in c or 'razao social' in c for c in cols):
                score += 1
            if any('unnamed' in c for c in cols):
                score -= 1
            return score

        dfs_extrato = []
        abas_lidas = 0

        # Lê TODAS as abas e consolida (layout mensal: JAN/FEV/MAR em abas separadas)
        for sheet_name in xl.sheet_names:
            try:
                candidatos = []
                # Candidato 1: formato com cabeçalho na linha 6
                try:
                    c1 = pd.read_excel(caminho, sheet_name=sheet_name, header=5, engine='openpyxl')
                    if len(c1) > 0:
                        candidatos.append(c1)
                except Exception:
                    pass

                # Candidato 2: leitura padrão
                try:
                    c2 = pd.read_excel(caminho, sheet_name=sheet_name, engine='openpyxl')
                    if len(c2) > 0:
                        candidatos.append(c2)
                except Exception:
                    pass

                if not candidatos:
                    continue

                # Escolhe o melhor candidato para a aba
                df_aba = max(candidatos, key=pontuar_colunas).dropna(how='all')
                if len(df_aba) == 0:
                    continue

                # Filtra abas que não parecem extrato
                if pontuar_colunas(df_aba) < 3:
                    continue

                df_aba = df_aba.copy()
                df_aba['aba_origem_extrato'] = str(sheet_name)
                dfs_extrato.append(df_aba)
                abas_lidas += 1
                print(f"[INFO] Aba de extrato carregada: '{sheet_name}' ({len(df_aba)} linhas)")
            except Exception:
                continue

        if not dfs_extrato:
            raise ValueError("Nenhuma aba de extrato válida foi encontrada no arquivo.")

        df = pd.concat(dfs_extrato, ignore_index=True).dropna(how='all')
        print(f"[OK] Extrato consolidado: {len(df)} lançamentos ({abas_lidas} aba(s))")
        return df
    except Exception as e:
        print(f"[ERRO] Erro ao ler extrato: {e}")
        raise


def ler_comprovantes(caminho):
    """
    Lê o arquivo Excel dos comprovantes SIGRA (pgtos sigra).
    
    Args:
        caminho: Caminho do arquivo Excel
        
    Returns:
        DataFrame com os dados dos comprovantes
    """
    try:
        df = pd.read_excel(caminho, engine='openpyxl')
        print(f"[INFO] Arquivo lido inicialmente: {len(df)} linhas")
        
        # Se tem colunas "Unnamed", tenta encontrar o cabeçalho correto
        if any('Unnamed' in str(col) for col in df.columns):
            for skip_rows in range(0, 10):
                try:
                    df_teste = pd.read_excel(caminho, skiprows=skip_rows, engine='openpyxl')
                    if not any('Unnamed' in str(col) for col in df_teste.columns[:3]):
                        df = df_teste
                        print(f"[INFO] Cabeçalho encontrado na linha {skip_rows + 1}")
                        break
                except:
                    continue
        
        # Remove linhas completamente vazias
        df = df.dropna(how='all')
        
        print(f"[OK] Comprovantes carregados: {len(df)} registros")
        return df
    except Exception as e:
        print(f"[ERRO] Erro ao ler comprovantes: {e}")
        raise


def normalizar_data(serie_data):
    """
    Normaliza uma série de datas para datetime.
    
    Args:
        serie_data: Série com datas em vários formatos
        
    Returns:
        Série com datas normalizadas (datetime)
    """
    if serie_data.empty:
        return pd.Series(dtype='datetime64[ns]')
    
    # Converte para string primeiro
    serie_limpa = serie_data.astype(str).str.strip()
    
    # Remove hora se houver (formato DD/MM/YY HH:MM)
    def extrair_data(valor):
        if pd.isna(valor) or valor == 'nan':
            return None
        valor_str = str(valor).strip()
        if ' ' in valor_str:
            # Tem hora, pega só a parte da data
            return valor_str.split()[0]
        return valor_str
    
    serie_processada = serie_limpa.apply(extrair_data)
    
    # Tenta diferentes formatos e combina resultados (não retorna no primeiro match parcial)
    formatos = [
        '%d/%m/%Y',      # 02/01/2026
        '%d/%m/%y',      # 02/01/26
        '%Y-%m-%d',      # 2026-01-02
    ]
    resultado_final = pd.Series(pd.NaT, index=serie_processada.index, dtype='datetime64[ns]')
    for formato in formatos:
        try:
            parsed = pd.to_datetime(serie_processada, format=formato, errors='coerce')
            resultado_final = resultado_final.fillna(parsed.dt.normalize())
        except:
            continue
    
    # Formato DD/MM sem ano (novo extrato Itaú) - usa ano atual (só preenche os que ainda estão NaT)
    def parse_dd_mm(val):
        if pd.isna(val) or str(val).strip() in ('', 'nan'):
            return pd.NaT
        s = str(val).strip()
        if re.match(r'^\d{1,2}/\d{1,2}$', s):
            try:
                dia, mes = map(int, s.split('/'))
                ano = datetime.now().year
                return pd.Timestamp(year=ano, month=mes, day=dia)
            except:
                return pd.NaT
        return pd.NaT
    parsed_ddmm = serie_processada.apply(parse_dd_mm)
    resultado_final = resultado_final.fillna(parsed_ddmm)
    
    # Conversão genérica para o que ainda restar
    parsed_generico = pd.to_datetime(serie_processada, errors='coerce')
    resultado_final = resultado_final.fillna(parsed_generico.dt.normalize())
    
    return resultado_final


def normalizar_valor(valor):
    """
    Normaliza valores monetários para float.
    
    Args:
        valor: Valor a ser normalizado
        
    Returns:
        Float com o valor normalizado (sempre positivo)
    """
    if pd.isna(valor):
        return 0.0
    
    # Converte para string
    valor_str = str(valor).strip()
    
    # Remove caracteres especiais
    valor_str = valor_str.replace('R$', '').replace('$', '').replace(' ', '').replace('\xa0', '')
    
    # Se ficou vazio, retorna 0
    if not valor_str:
        return 0.0
    
    # Formato brasileiro: ponto como milhar, vírgula como decimal
    if '.' in valor_str and ',' in valor_str:
        # Remove pontos (milhar) e substitui vírgula por ponto (decimal)
        valor_str = valor_str.replace('.', '').replace(',', '.')
    elif ',' in valor_str:
        # Só tem vírgula - pode ser decimal
        valor_str = valor_str.replace(',', '.')
    
    try:
        resultado = float(valor_str)
        return abs(round(resultado, 2))
    except:
        return 0.0


def normalizar_texto(texto):
    """
    Normaliza texto para comparação.
    
    Args:
        texto: String a ser normalizada
        
    Returns:
        String normalizada (maiúsculas, sem espaços extras)
    """
    if pd.isna(texto):
        return ''
    
    texto = str(texto).upper().strip()
    texto = ' '.join(texto.split())  # Remove espaços múltiplos
    
    return texto


def _texto_extrato_favorecido_busca(extrato_row):
    """
    Texto do extrato para cruzar com Fornecedor no PGTO.
    No Itaú, o DESTINO/Lançamento costuma descrever o pagamento; a Razão Social às vezes é genérica.
    """
    obs_extra = ''
    for col in extrato_row.index:
        cl = str(col).lower()
        if 'obs' in cl or 'observa' in cl:
            v = extrato_row.get(col, '')
            if pd.notna(v) and str(v).strip():
                obs_extra = str(v).strip()
                break
    parts = [
        extrato_row.get('lancamento_original', '') or '',
        extrato_row.get('favorecido_original', '') or '',
        obs_extra,
    ]
    return normalizar_texto(' '.join(str(p) for p in parts if str(p).strip()))


def _fornecedor_compativel_extrato(fav_comp_norm, texto_extrato_norm):
    """Casa fornecedor do comprovante com lançamento e/ou razão do extrato (sem relaxar valor/data)."""
    if not fav_comp_norm or not texto_extrato_norm:
        return False
    if fav_comp_norm == texto_extrato_norm:
        return True
    if fav_comp_norm in texto_extrato_norm or texto_extrato_norm in fav_comp_norm:
        return True
    for w in fav_comp_norm.split():
        if len(w) >= 4 and w in texto_extrato_norm:
            return True
    return False


def _tokens_referencia_no_extrato(extrato_row, min_len=6):
    """Trechos alfanuméricos longos do extrato (DI, processo, etc.) para cruzar com Ref./RPS/Cliente."""
    texto = _texto_extrato_favorecido_busca(extrato_row).replace(' ', '')
    if not texto:
        return []
    found = re.findall(r'[A-Z0-9]{%d,}' % min_len, texto)
    return list(dict.fromkeys(found))


def _score_referencia_cruzada(extrato_row, comp_row):
    """Quantas referências do extrato aparecem nos campos-chave do comprovante (desempate exato)."""
    tokens = _tokens_referencia_no_extrato(extrato_row, min_len=6)
    if not tokens:
        return 0
    parts = []
    for c in comp_row.index:
        cl = str(c).lower()
        if any(k in cl for k in ('sigra', 'rps', 'cliente', 'fornecedor', 'categoria', 'cnpj')):
            v = comp_row.get(c)
            if pd.notna(v):
                parts.append(str(v).upper().replace(' ', ''))
    blob = ''.join(parts)
    return sum(1 for t in tokens if t in blob)


def formatar_valor_br(valor):
    """
    Formata valor monetário no padrão brasileiro.
    
    Args:
        valor: Valor numérico
        
    Returns:
        String formatada (ex: "4.601,70")
    """
    if pd.isna(valor) or valor == 0:
        return "0,00"
    
    valor_float = round(float(valor), 2)
    parte_inteira = int(abs(valor_float))
    parte_decimal = abs(valor_float) - parte_inteira
    centavos = int(round(parte_decimal * 100))
    
    parte_inteira_str = f"{parte_inteira:,}".replace(",", ".")
    parte_decimal_str = f"{centavos:02d}"
    
    return f"{parte_inteira_str},{parte_decimal_str}"


def preparar_dados(df, tipo='extrato'):
    """
    Prepara e normaliza os dados do DataFrame.
    
    Args:
        df: DataFrame a ser preparado
        tipo: 'extrato' ou 'comprovantes'
        
    Returns:
        DataFrame preparado com colunas normalizadas
    """
    df_prep = df.copy()
    
    # Seleciona colunas baseado no tipo
    if tipo == 'extrato':
        colunas = COLUNAS_EXTRATO
        prefixo = 'extrato'
    elif tipo == 'numerario':
        colunas = COLUNAS_NUMERARIO
        prefixo = 'comprovante'  # Usa mesmo prefixo para compatibilidade com funções de conciliação
    else:
        colunas = COLUNAS_COMPROVANTES
        prefixo = 'comprovante'
    
    print(f"\nColunas disponíveis em {tipo}: {list(df_prep.columns)}")
    
    # Busca colunas por nome ou posição
    col_data = None
    col_valor = None
    col_favorecido = None
    col_lancamento = None  # Coluna "Lançamento" do extrato Itaú (ex: "PUCOMEX...", "RECEBIMENTOS LUXDATA COMERCIO")
    
    # Tenta encontrar por nome exato
    col_config_data = colunas.get('data')
    col_config_valor = colunas.get('valor')
    col_config_favorecido = colunas.get('favorecido')
    col_config_lancamento = colunas.get('lancamento') if tipo == 'extrato' else None
    
    if col_config_data and col_config_data in df_prep.columns:
        col_data = col_config_data
    if col_config_valor and col_config_valor in df_prep.columns:
        col_valor = col_config_valor
    if col_config_favorecido and col_config_favorecido in df_prep.columns:
        col_favorecido = col_config_favorecido
    if col_config_lancamento and col_config_lancamento in df_prep.columns:
        col_lancamento = col_config_lancamento
    
    # Para extrato: tenta nomes alternativos (formato Itaú: Valor (R$), Razão Social, Lançamento)
    if tipo == 'extrato':
        for alt in colunas.get('valor_alt', []):
            if alt and alt in df_prep.columns and col_valor is None:
                col_valor = alt
                break
        for alt in colunas.get('favorecido_alt', []):
            if alt and alt in df_prep.columns and col_favorecido is None:
                col_favorecido = alt
                break
        for alt in colunas.get('lancamento_alt', []):
            if alt and alt in df_prep.columns and col_lancamento is None:
                col_lancamento = alt
                break
    
    # Se não encontrou, busca por palavras-chave
    for col in df_prep.columns:
        col_lower = str(col).lower()
        if ('data' in col_lower or 'cria' in col_lower or 'data lan' in col_lower) and col_data is None:
            col_data = col
        if ('valor' in col_lower or 'value' in col_lower or '(r$)' in col_lower) and col_valor is None:
            col_valor = col
        if ('destino' in col_lower or 'favorecido' in col_lower or 'fornecedor' in col_lower or 
            'cliente' in col_lower or 'razão social' in col_lower or 'razao social' in col_lower) and col_favorecido is None:
            col_favorecido = col
        if tipo == 'extrato' and ('lançamento' in col_lower or 'lancamento' in col_lower or 'histórico' in col_lower or 'descrição' in col_lower or 'descricao' in col_lower) and col_lancamento is None:
            col_lancamento = col
    
    # Se ainda não encontrou, usa posição (formato Itaú: A=Data, B=Lançamento, C=Razão Social, D=CPF/CNPJ, E=Valor)
    if not col_data:
        col_data = df_prep.columns[0] if len(df_prep.columns) > 0 else None
    if not col_valor:
        # Coluna E = índice 4 no Itaú (0-based)
        col_valor = df_prep.columns[4] if len(df_prep.columns) > 4 else (df_prep.columns[3] if len(df_prep.columns) > 3 else df_prep.columns[0])
    if not col_favorecido:
        # Coluna C = índice 2 no Itaú (Razão Social)
        col_favorecido = df_prep.columns[2] if len(df_prep.columns) > 2 else col_valor
    if tipo == 'extrato' and not col_lancamento and len(df_prep.columns) > 1:
        # Coluna B = índice 1 no Itaú (Lançamento)
        col_lancamento = df_prep.columns[1]
    
    # Comprovantes: garante que a coluna de valor é numérica (evita pegar coluna de texto por engano)
    if tipo == 'comprovantes' and col_valor in df_prep.columns:
        serie_valor = pd.to_numeric(df_prep[col_valor].apply(normalizar_valor), errors='coerce')
        qtd_numericos = serie_valor.notna().sum()
        if qtd_numericos < max(1, len(df_prep) // 2):
            # Coluna atual não parece ser de valores; procura coluna com mais números
            for c in df_prep.columns:
                if c == col_valor:
                    continue
                test = pd.to_numeric(df_prep[c].apply(normalizar_valor), errors='coerce')
                if test.notna().sum() > qtd_numericos:
                    col_valor = c
                    qtd_numericos = test.notna().sum()
    
    # Verifica se as colunas existem (data pode ser opcional no extrato em alguns formatos)
    if col_data is not None and col_data not in df_prep.columns:
        col_data = df_prep.columns[0] if len(df_prep.columns) > 0 else None  # fallback
    if col_valor not in df_prep.columns:
        raise ValueError(f"Coluna de valor '{col_valor}' não encontrada. Colunas disponíveis: {list(df_prep.columns)}")
    if col_favorecido not in df_prep.columns:
        raise ValueError(f"Coluna de favorecido '{col_favorecido}' não encontrada. Colunas disponíveis: {list(df_prep.columns)}")
    if col_lancamento is not None and col_lancamento not in df_prep.columns:
        col_lancamento = None
    
    print(f"Usando colunas: Data='{col_data}', Valor='{col_valor}', Favorecido='{col_favorecido}'" + (f", Lançamento='{col_lancamento}'" if col_lancamento else ""))
    
    # Cria DataFrame normalizado
    df_normalizado = pd.DataFrame()
    if col_data is not None:
        df_normalizado['data_original'] = df_prep[col_data]
        df_normalizado['data'] = normalizar_data(df_prep[col_data])
    else:
        # Extrato sem coluna de data (ex.: só Lançamentos com Razão Social e Valor)
        n = len(df_prep)
        df_normalizado['data_original'] = [pd.NaT] * n
        df_normalizado['data'] = pd.Series([pd.NaT] * n)
    df_normalizado['valor_original'] = df_prep[col_valor]
    df_normalizado['valor'] = df_prep[col_valor].apply(normalizar_valor)
    df_normalizado['favorecido_original'] = df_prep[col_favorecido].fillna('')
    df_normalizado['favorecido'] = df_prep[col_favorecido].apply(normalizar_texto)
    if tipo == 'extrato' and col_lancamento:
        df_normalizado['lancamento_original'] = df_prep[col_lancamento].fillna('').astype(str).str.strip()
    elif tipo == 'extrato':
        df_normalizado['lancamento_original'] = ''
    
    # Adiciona todas as colunas originais (incluindo ID do numerário, Assunto (Email), etc.)
    for col in df_prep.columns:
        if col not in df_normalizado.columns:
            df_normalizado[col] = df_prep[col]
    
    # Para numerário, preserva coluna ID original se existir
    if tipo == 'numerario':
        col_id_numerario = COLUNAS_NUMERARIO.get('id')
        if col_id_numerario and col_id_numerario in df_prep.columns:
            df_normalizado['ID'] = df_prep[col_id_numerario]
    
    if tipo == 'comprovantes':
        # Garante que campos de referência/códigos permaneçam como texto
        colunas_ref_like = []
        for col in df_normalizado.columns:
            col_lower = str(col).lower()
            if any(palavra in col_lower for palavra in ['ref', 'rps', 'protocolo', 'codigo']):
                colunas_ref_like.append(col)

        if colunas_ref_like:
            for col in colunas_ref_like:
                df_normalizado[col] = df_normalizado[col].apply(
                    lambda x: '' if pd.isna(x) else str(x).strip()
                )

    # Remove linhas inválidas
    if tipo == 'extrato':
        # Para extratos, mantém todas as linhas (mesmo com data inválida ou valor zero)
        df_normalizado = df_normalizado[
            (df_normalizado['data'].notna()) | (df_normalizado['valor'] != 0)
        ]
    else:
        # Para comprovantes e numerário: remove só valor inválido; mantém linhas sem data para match por valor
        df_normalizado = df_normalizado[df_normalizado['valor'] != 0]
        if tipo == 'numerario':
            df_normalizado = df_normalizado[df_normalizado['data'].notna()]
    
    # Adiciona ID único
    df_normalizado[f'ID_{prefixo}'] = range(1, len(df_normalizado) + 1)
    
    print(f"[OK] Dados {tipo} preparados: {len(df_normalizado)} registros válidos")
    
    return df_normalizado


# ============================================================================
# FUNÇÕES DE CONCILIAÇÃO
# ============================================================================

def _to_centavos(valor):
    """
    Converte valor (float, string ou qualquer formato) para centavos (int).
    Usa normalizar_valor para garantir consistência na comparação.
    """
    v = normalizar_valor(valor) if valor is not None else 0.0
    if pd.isna(v) or v == 0:
        return None  # Indica inválido para pular na busca
    return int(round(float(v) * 100, 0))


def _valores_equivalentes(valor_extrato_int, valor_comp_int):
    """True somente se extrato e comprovante (ou soma) coincidem no mesmo valor em centavos."""
    if valor_comp_int is None:
        return False
    return valor_comp_int == valor_extrato_int


def _extrato_e_pucomex_ou_receita_federal(extrato_row):
    """
    Lançamentos do extrato que correspondem ao fluxo Receita Federal / PUCOMEX:
    no comprovante filtra-se Fornecedor "Receita Federal", agrupa por processo
    (Ref. Sigra / coluna D) e soma os valores (coluna H) para casar no extrato.
    """
    lanc = normalizar_texto(extrato_row.get('lancamento_original', '') or '')
    fav = normalizar_texto(extrato_row.get('favorecido_original', '') or '')
    return 'PUCOMEX' in lanc or 'PUCOMEX' in fav or (
        'RECEITA' in fav and 'FEDERAL' in fav
    ) or ('RECEITA' in lanc and 'FEDERAL' in lanc)


def _comprovante_fornecedor_receita_federal(row):
    """True se a coluna Fornecedor (orig. favorecido_original) indica Receita Federal."""
    fav = normalizar_texto(row.get('favorecido_original', '') or '')
    return 'RECEITA' in fav and 'FEDERAL' in fav


def _extrato_e_recebimento(row):
    """
    True quando o lançamento do extrato representa RECEBIMENTO(S).
    Esses itens são tratados em outra planilha de comprovantes e não devem
    ser conciliados neste fluxo (SIGRA/Numerário).
    """
    lanc = str(row.get('lancamento_original', '') or '').upper()
    fav = str(row.get('favorecido_original', '') or '').upper()
    texto = f"{lanc} {fav}"
    return 'RECEBIMENTO' in texto or 'RECEBIMENTOS' in texto


def encontrar_grupo_receita_federal_por_processo(extrato_row, comprovantes_disponiveis):
    """
    Para extrato PUCOMEX / Receita Federal:
    1) filtra comprovantes com Fornecedor = Receita Federal (col. O);
    2) agrupa por processo (Ref. Sigra / col. D);
    3) soma coluna H por processo;
    4) casa a soma do processo com o valor do extrato.

    Só entram comprovantes cuja data é a mesma do extrato (evita casar valor
    repetido em outro mês). Dentro desse dia, o critério é a soma por processo.

    Returns:
        Lista com índices do processo encontrado, ou None
    """
    if not _extrato_e_pucomex_ou_receita_federal(extrato_row):
        return None

    valor_extrato = abs(float(extrato_row['valor']))
    if valor_extrato == 0:
        return None

    valor_extrato_int = _to_centavos(valor_extrato) or int(round(valor_extrato * 100, 0))

    # Filtra somente Receita Federal no fornecedor (coluna O no arquivo original)
    rf = comprovantes_disponiveis[
        comprovantes_disponiveis.apply(_comprovante_fornecedor_receita_federal, axis=1)
    ]
    if len(rf) == 0:
        return None

    data_extrato_raw = extrato_row.get('data')
    if pd.notna(data_extrato_raw):
        data_extrato_norm = pd.Timestamp(data_extrato_raw).normalize()

        def _rf_mesmo_dia_extrato(row):
            d = row.get('data')
            if pd.isna(d):
                return False
            return pd.Timestamp(d).normalize() == data_extrato_norm

        rf = rf[rf.apply(_rf_mesmo_dia_extrato, axis=1)]
        if len(rf) == 0:
            return None

    # Descobre coluna de processo (Ref. Sigra / D)
    col_ref = None
    for c in rf.columns:
        cl = str(c).lower()
        if 'sigra' in cl and ('ref' in cl or 'referencia' in cl or 'processo' in cl):
            col_ref = c
            break
    if col_ref is None:
        for c in rf.columns:
            if 'sigra' in str(c).lower():
                col_ref = c
                break
    if col_ref is None:
        print("    [AVISO] RF/PUCOMEX: coluna de processo (Ref. Sigra) não encontrada.")
        return None

    # Agrupa por processo e soma col. H (valor)
    grupos = {}
    for idx, row in rf.iterrows():
        proc_raw = row.get(col_ref, '')
        proc = '' if pd.isna(proc_raw) else str(proc_raw).strip()
        if not proc:
            continue
        grupos.setdefault(proc, []).append((idx, abs(float(row.get('valor', 0) or 0))))

    candidatos = []
    for proc, itens in grupos.items():
        soma_int = int(round(sum(v for _, v in itens) * 100, 0))
        if _valores_equivalentes(valor_extrato_int, soma_int):
            candidatos.append((proc, [i for i, _ in itens]))

    if not candidatos:
        return None

    # Mais de um processo com a mesma soma no mesmo dia: desempate por Ref. Sigra no texto do extrato (sem relaxar valor).
    if len(candidatos) > 1:
        texto_busca = _texto_extrato_favorecido_busca(extrato_row).replace(' ', '')
        desempate = []
        for proc, indices in candidatos:
            proc_u = str(proc).upper().replace(' ', '')
            score_txt = 0
            if texto_busca and proc_u and proc_u in texto_busca:
                score_txt = 2
            desempate.append((score_txt, proc, indices))
        desempate.sort(key=lambda x: -x[0])
        if desempate[0][0] > desempate[1][0]:
            _, proc_sel, indices_sel = desempate[0]
        elif EMPATES_SEM_REF_EXIGIR_DESISTENCIA:
            print(f"    [INFO] RF/PUCOMEX: {len(candidatos)} processos com mesma soma e sem Ref. no texto do extrato; não associar.")
            return None
        else:
            desempate.sort(key=lambda x: str(x[1]))
            _, proc_sel, indices_sel = desempate[0]
            print(f"    [INFO] RF/PUCOMEX: empate entre processos; usando Ref. Sigra '{proc_sel}' (determinístico).")
    else:
        proc_sel, indices_sel = candidatos[0]

    print(
        f"    [OK] RF/PUCOMEX por processo (Ref. Sigra {proc_sel}): "
        f"{len(indices_sel)} comprovante(s) = R$ {formatar_valor_br(valor_extrato)}"
    )
    return indices_sel


def buscar_por_siscomex_prot(extrato_row, comprovantes_disponiveis):
    """
    Busca específica para SISCOMEX: extrai o número após "SISCOMEX PROT" no extrato,
    adiciona hífen antes do último dígito (ex.: 3542022200 → 354202220-0) e busca
    nos comprovantes por RPS. Soma os valores dos comprovantes encontrados.
    
    Returns:
        Lista de índices dos comprovantes, ou None
    """
    # Verifica se o extrato contém SISCOMEX PROT
    lancamento = str(extrato_row.get('lancamento_original', ''))
    favorecido = str(extrato_row.get('favorecido_original', ''))
    texto_busca = lancamento + ' ' + favorecido
    
    if 'SISCOMEX' not in texto_busca.upper() or 'PROT' not in texto_busca.upper():
        return None
    
    # Extrai o número após "SISCOMEX PROT" (ex.: "SISCOMEX PROT 3542022200" → "3542022200")
    match = re.search(r'SISCOMEX\s+PROT\s+(\d+)', texto_busca, re.IGNORECASE)
    if not match:
        return None
    
    numero_original = match.group(1)
    if len(numero_original) < 2:
        return None
    
    # Adiciona hífen antes do último dígito (ex.: 3542022200 → 354202220-0)
    numero_formatado = numero_original[:-1] + '-' + numero_original[-1]
    variantes_rps = [numero_formatado, numero_original]
    # Evita duplicar se protocolo já tinha hífen no mesmo lugar
    variantes_rps = list(dict.fromkeys(variantes_rps))
    
    print(f"    [INFO] SISCOMEX detectado: '{numero_original}' -> buscando RPS em {variantes_rps}")
    
    # Busca coluna RPS nos comprovantes
    col_rps = None
    for c in comprovantes_disponiveis.columns:
        if 'rps' in str(c).lower():
            col_rps = c
            break
    
    if col_rps is None:
        print(f"    [AVISO] Coluna RPS não encontrada nos comprovantes")
        return None
    
    # Busca comprovantes cujo RPS contém protocolo com ou sem hífen (valor e data continuam exatos depois)
    def rps_match(row):
        rps = str(row.get(col_rps, '')).strip()
        if not rps:
            return False
        rps_limpo = rps.replace(' ', '').replace('_', '').replace('.', '')
        for variante in variantes_rps:
            v = str(variante).replace(' ', '').replace('_', '').replace('.', '')
            if not v:
                continue
            if v in rps_limpo or rps_limpo == v:
                return True
        return False
    
    comprovantes_siscomex = comprovantes_disponiveis[
        comprovantes_disponiveis.apply(rps_match, axis=1)
    ]
    
    if len(comprovantes_siscomex) == 0:
        print(f"    [INFO] Nenhum comprovante com RPS compatível com {variantes_rps}")
        return None

    data_extrato_sis = extrato_row.get('data')
    if pd.notna(data_extrato_sis):
        de_norm = pd.Timestamp(data_extrato_sis).normalize()

        def _siscomex_mesmo_dia(row):
            d = row.get('data')
            if pd.isna(d):
                return False
            return pd.Timestamp(d).normalize() == de_norm

        comprovantes_siscomex = comprovantes_siscomex[
            comprovantes_siscomex.apply(_siscomex_mesmo_dia, axis=1)
        ]
        if len(comprovantes_siscomex) == 0:
            print(f"    [INFO] SISCOMEX: RPS '{numero_formatado}' sem comprovante na mesma data do extrato ({de_norm.strftime('%d/%m/%Y')})")
            return None

    print(f"    [INFO] Encontrados {len(comprovantes_siscomex)} comprovante(s) com RPS '{numero_formatado}'")
    
    # Soma os valores dos comprovantes
    valor_extrato = abs(extrato_row['valor'])
    soma_comprovantes = sum(abs(float(v)) for v in comprovantes_siscomex['valor'])
    
    # Verifica se a soma bate (em centavos)
    valor_extrato_int = int(round(valor_extrato * 100, 0))
    soma_int = int(round(soma_comprovantes * 100, 0))
    
    if _valores_equivalentes(valor_extrato_int, soma_int):
        indices = comprovantes_siscomex.index.tolist()
        print(f"    [OK] SISCOMEX conciliado por RPS: {len(indices)} comprovante(s) = R$ {formatar_valor_br(valor_extrato)}")
        return indices
    else:
        print(f"    [AVISO] SISCOMEX: soma dos comprovantes (R$ {formatar_valor_br(soma_comprovantes)}) != valor extrato (R$ {formatar_valor_br(valor_extrato)})")
        return None


def encontrar_match_exato(extrato_row, comprovantes_disponiveis, exigir_mesmo_favorecido=False, evitar_ambiguo=True):
    """
    Encontra comprovante com o mesmo valor (exato em centavos) na MESMA data de calendário do extrato.
    Não usa comprovante sem data. Se o extrato não tiver data, retorna None (não concilia sem data).
    
    Args:
        extrato_row: Linha do extrato
        comprovantes_disponiveis: DataFrame com comprovantes disponíveis
        exigir_mesmo_favorecido: Se True, só aceita quando fornecedor casa com lançamento+razão do extrato
        evitar_ambiguo: Se True e há vários candidatos (mesmo dia + valor), tenta desempate por Ref./RPS
                       no texto do extrato; se continuar ambíguo, retorna None (não chuta o primeiro).
        
    Returns:
        Lista com índice do comprovante encontrado, ou None
    """
    valor_extrato = abs(extrato_row['valor'])
    data_extrato = extrato_row['data']
    
    if valor_extrato == 0:
        return None
    
    if pd.isna(data_extrato):
        return None
    
    valor_extrato_int = _to_centavos(valor_extrato) or int(round(valor_extrato * 100, 0))
    data_extrato_normalizada = pd.Timestamp(data_extrato).normalize()
    texto_extrato = _texto_extrato_favorecido_busca(extrato_row)
    fav_extrato_norm = normalizar_texto(extrato_row.get('favorecido_original', '') or '') if exigir_mesmo_favorecido else ''
    
    candidatos = []  # Lista de (idx, diff_dias) que batem em valor e data
    for idx, row in comprovantes_disponiveis.iterrows():
        valor_comp_int = _to_centavos(row.get('valor'))
        if valor_comp_int is None:
            continue
        
        if not _valores_equivalentes(valor_extrato_int, valor_comp_int):
            continue
        
        data_comp = row['data']
        if pd.isna(data_comp):
            continue
        data_comp_normalizada = pd.Timestamp(data_comp).normalize()
        diff_dias = abs((data_comp_normalizada - data_extrato_normalizada).days)
        if diff_dias <= TOLERANCIA_DATA:
            if exigir_mesmo_favorecido:
                fav_comp_norm = normalizar_texto(row.get('favorecido_original', '') or '')
                if not fav_comp_norm:
                    continue
                # Razão social OU descrição no lançamento (ex.: SISPAG / GRU / nome no DESTINO)
                ok = False
                if fav_extrato_norm:
                    if (
                        fav_extrato_norm == fav_comp_norm
                        or fav_extrato_norm in fav_comp_norm
                        or fav_comp_norm in fav_extrato_norm
                    ):
                        ok = True
                if not ok and texto_extrato:
                    ok = _fornecedor_compativel_extrato(fav_comp_norm, texto_extrato)
                if not ok:
                    continue
            candidatos.append((idx, diff_dias))
    
    if not candidatos:
        return None
    
    candidatos.sort(key=lambda x: x[1])
    if len(candidatos) == 1:
        idx_match, diff_dias = candidatos[0]
        print(f"    [OK] Match exato (mesmo dia): Comprovante ID {comprovantes_disponiveis.loc[idx_match].get('ID_comprovante', idx_match)} | R$ {formatar_valor_br(valor_extrato)} | diff: {diff_dias} dia(s)")
        return [idx_match]

    # Vários candidatos: desempate só por referência cruzada (valor e data já idênticos)
    scores = []
    for idx, diff_dias in candidatos:
        comp_row = comprovantes_disponiveis.loc[idx]
        sc = _score_referencia_cruzada(extrato_row, comp_row)
        scores.append((sc, -diff_dias, idx, diff_dias))
    scores.sort(key=lambda x: (-x[0], x[1]))
    best_sc, _, idx_best, diff_best = scores[0]
    segundo = scores[1][0] if len(scores) > 1 else -1
    if best_sc > 0 and best_sc > segundo:
        print(
            f"    [OK] Match exato (mesmo dia, desempate Ref/RPS no texto extrato): "
            f"Comprovante ID {comprovantes_disponiveis.loc[idx_best].get('ID_comprovante', idx_best)} | "
            f"R$ {formatar_valor_br(valor_extrato)} | diff: {diff_best} dia(s)"
        )
        return [idx_best]
    if evitar_ambiguo:
        print(f"    [INFO] Match exato ambíguo: {len(candidatos)} comprovantes no mesmo dia e valor; sem referência única no extrato.")
        return None

    # Determinístico: menor ID_comprovante entre os empatados (mesmo dia e valor já garantidos)
    candidatos_ord = sorted(
        candidatos,
        key=lambda t: int(comprovantes_disponiveis.loc[t[0]].get('ID_comprovante', 10**9) or 10**9),
    )
    idx_match, diff_dias = candidatos_ord[0]
    print(f"    [OK] Match exato (mesmo dia): Comprovante ID {comprovantes_disponiveis.loc[idx_match].get('ID_comprovante', idx_match)} | R$ {formatar_valor_br(valor_extrato)} | diff: {diff_dias} dia(s)")
    return [idx_match]


def encontrar_combinacoes_comprovantes(extrato_row, comprovantes_disponiveis, max_itens=10):
    """
    Tenta encontrar uma combinação de comprovantes do mesmo dia do extrato
    cuja soma seja igual ao valor do extrato.
    
    Args:
        extrato_row: Linha do extrato
        comprovantes_disponiveis: DataFrame com comprovantes disponíveis
        max_itens: Número máximo de comprovantes na combinação
        
    Returns:
        Lista de índices dos comprovantes que somam o valor, ou None
    """
    valor_extrato = abs(extrato_row['valor'])
    data_extrato = extrato_row['data']
    
    if valor_extrato == 0:
        return None
    
    valor_extrato_int = int(round(valor_extrato * 100, 0))
    tem_data_extrato = pd.notna(data_extrato)
    data_extrato_norm = pd.Timestamp(data_extrato).normalize() if tem_data_extrato else None
    
    # Só tenta combinações quando o extrato tem data (comprovantes do mesmo dia)
    if not tem_data_extrato or data_extrato_norm is None:
        return None
    
    def mesma_data(row):
        d = row['data']
        if pd.isna(d):
            return False
        d_norm = pd.Timestamp(d).normalize()
        diff = abs((d_norm - data_extrato_norm).days)
        return diff <= TOLERANCIA_DATA
    
    comprovantes_do_dia = comprovantes_disponiveis[
        comprovantes_disponiveis.apply(mesma_data, axis=1)
    ]
    
    if len(comprovantes_do_dia) < 2:
        return None
    
    # Limita candidatos para não travar (C(50,5) já são milhões de combinações)
    MAX_CANDIDATOS_DIA = 25
    MAX_COMBINACOES_TOTAL = 20000  # Abandona após testar esse número de combinações (evita travar)
    if len(comprovantes_do_dia) > MAX_CANDIDATOS_DIA:
        comprovantes_do_dia = comprovantes_do_dia.copy()
        comprovantes_do_dia['_valor_abs'] = comprovantes_do_dia['valor'].apply(lambda x: abs(float(x)))
        alvo_medio = valor_extrato / 5
        comprovantes_do_dia['_dist'] = (comprovantes_do_dia['_valor_abs'] - alvo_medio).abs()
        comprovantes_do_dia = comprovantes_do_dia.nsmallest(MAX_CANDIDATOS_DIA, '_dist')
        comprovantes_do_dia = comprovantes_do_dia.drop(columns=['_valor_abs', '_dist'], errors='ignore')
    
    # Valores dos comprovantes (em centavos) com índice
    valores_com_indice = []
    for idx, row in comprovantes_do_dia.iterrows():
        v = abs(float(row['valor']))
        v_int = int(round(v * 100, 0))
        valores_com_indice.append((idx, v_int))
    
    n = len(valores_com_indice)
    # Para valores muito altos, limita tamanho máximo da combinação (evita travar)
    if valor_extrato > 100000:
        max_itens = min(max_itens, 4)
    elif valor_extrato > 50000:
        max_itens = min(max_itens, 5)
    
    combinacoes_testadas = 0
    # Tenta combinações de 2, 3, ... até max_itens (soma deve bater exatamente em centavos com o extrato)
    for tamanho in range(2, min(max_itens + 1, n + 1)):
        for combo in combinations(range(n), tamanho):
            combinacoes_testadas += 1
            if combinacoes_testadas > MAX_COMBINACOES_TOTAL:
                return None
            soma_int = sum(valores_com_indice[i][1] for i in combo)
            if _valores_equivalentes(valor_extrato_int, soma_int):
                indices = [valores_com_indice[i][0] for i in combo]
                print(f"    [OK] Combinação de {tamanho} comprovantes: R$ {formatar_valor_br(valor_extrato)} | Data: {data_extrato_norm.strftime('%d/%m/%Y')}")
                return indices
    
    return None


def encontrar_grupo_por_processo(extrato_row, comprovantes_disponiveis):
    """
    Concilia pelo "processo" (Ref. Sigra): agrupa comprovantes do mesmo dia por Ref. Sigra,
    soma os valores de cada grupo e casa com o valor do extrato.

    Para PUCOMEX / Receita Federal no extrato, não usar este modo: use
    encontrar_grupo_receita_federal_por_processo (Fornecedor RF + soma por processo).
    
    Returns:
        Lista de índices dos comprovantes do grupo que bateu com o valor do extrato, ou None
    """
    valor_extrato = abs(extrato_row['valor'])
    data_extrato = extrato_row['data']
    
    if valor_extrato == 0:
        return None
    if pd.isna(data_extrato):
        return None
    
    valor_extrato_int = int(round(valor_extrato * 100, 0))
    data_extrato_norm = pd.Timestamp(data_extrato).normalize()
    
    def mesma_data(row):
        d = row['data']
        if pd.isna(d):
            return False
        d_norm = pd.Timestamp(d).normalize()
        return abs((d_norm - data_extrato_norm).days) <= TOLERANCIA_DATA
    
    comprovantes_do_dia = comprovantes_disponiveis[
        comprovantes_disponiveis.apply(mesma_data, axis=1)
    ]
    if len(comprovantes_do_dia) < 1:
        return None
    
    # Nome da coluna Ref. Sigra (pode variar: "Ref. Sigra", "Ref Sigra", "Sigra", etc.)
    col_ref = None
    for c in comprovantes_do_dia.columns:
        cl = str(c).lower()
        if 'sigra' in cl and ('ref' in cl or 'referencia' in cl or 'processo' in cl):
            col_ref = c
            break
    if col_ref is None:
        for c in comprovantes_do_dia.columns:
            if 'sigra' in str(c).lower():
                col_ref = c
                break
    # Se ainda não achou coluna Ref. Sigra, agrupa só por cliente (favorecido) para não misturar clientes
    usa_ref_sigra = col_ref is not None
    
    # Coluna Ref. Cliente (ou favorecido/fornecedor): não pode somar comprovantes de clientes diferentes
    col_ref_cliente = None
    for c in comprovantes_do_dia.columns:
        if 'ref' in str(c).lower() and 'cliente' in str(c).lower():
            col_ref_cliente = c
            break
    if col_ref_cliente is None and 'favorecido_original' in comprovantes_do_dia.columns:
        col_ref_cliente = 'favorecido_original'  # fallback: usa favorecido (ex.: Fornecedor)
    elif col_ref_cliente is None and 'Fornecedor' in comprovantes_do_dia.columns:
        col_ref_cliente = 'Fornecedor'
    
    # Agrupa por (Ref. Sigra, Ref. Cliente): só soma comprovantes do mesmo processo E do mesmo cliente
    grupos = {}
    for idx, row in comprovantes_do_dia.iterrows():
        ref = '-'
        if usa_ref_sigra and col_ref:
            r = row.get(col_ref, '-')
            ref = '-' if pd.isna(r) or not str(r).strip() else str(r).strip()
        ref_cliente = '-'
        if col_ref_cliente:
            v = row.get(col_ref_cliente, '-')
            ref_cliente = '-' if pd.isna(v) or not str(v).strip() else str(v).strip()
        chave = (ref, ref_cliente)
        if chave not in grupos:
            grupos[chave] = []
        grupos[chave].append((idx, abs(float(row['valor']))))
    
    # Para cada grupo, soma e verifica se bate com o extrato (exige Ref. Sigra preenchida se a coluna existir)
    for (ref, ref_cliente), itens in grupos.items():
        if usa_ref_sigra and ref == '-':
            continue  # Quando tem coluna Ref. Sigra, ignora grupo sem processo
        if not usa_ref_sigra and ref_cliente == '-':
            continue  # Sem coluna Ref. Sigra: só aceita grupos com mesmo cliente (não mistura)
        soma = sum(v for _, v in itens)
        soma_int = int(round(soma * 100, 0))
        if _valores_equivalentes(valor_extrato_int, soma_int):
            indices = [idx for idx, _ in itens]
            msg_cliente = f" | Cliente: {ref_cliente}" if ref_cliente != '-' else ""
            print(f"    [OK] Grupo por processo (Ref. Sigra {ref}){msg_cliente}: {len(indices)} comprovante(s) = R$ {formatar_valor_br(valor_extrato)} | Data: {data_extrato_norm.strftime('%d/%m/%Y')}")
            return indices
    
    return None


def _periodo_calendario_comprovantes(df_comprovantes):
    """Menor e maior data (normalizada) nos comprovantes, ou (None, None)."""
    if df_comprovantes is None or len(df_comprovantes) == 0 or 'data' not in df_comprovantes.columns:
        return None, None
    s = df_comprovantes['data'].dropna()
    if len(s) == 0:
        return None, None
    return pd.Timestamp(s.min()).normalize(), pd.Timestamp(s.max()).normalize()


def _datas_distintas_comprovantes(df_comprovantes):
    """Conjunto de datas de calendário (normalizadas) que existem na coluna data dos comprovantes."""
    if df_comprovantes is None or len(df_comprovantes) == 0 or 'data' not in df_comprovantes.columns:
        return None
    s = df_comprovantes['data'].dropna()
    if len(s) == 0:
        return None
    return {pd.Timestamp(x).normalize() for x in s}


def conciliar_extrato_comprovantes(df_extrato, df_comprovantes):
    """
    Concilia extratos com comprovantes.
    
    Args:
        df_extrato: DataFrame com extratos
        df_comprovantes: DataFrame com comprovantes
        
    Returns:
        Tupla com (df_conciliacao, df_extratos_pendentes, df_comprovantes_pendentes,
                   periodo_pgto_min, periodo_pgto_max, datas_distintas_pgto)
        periodo_* = limites de data do arquivo PGTO (None se indeterminado).
        datas_distintas_pgto = set de Timestamps normalizados com dias que existem no PGTO, ou None.
    """
    print("\n" + "="*80)
    print("INICIANDO CONCILIAÇÃO")
    print("="*80)
    
    periodo_pgto_min, periodo_pgto_max = _periodo_calendario_comprovantes(df_comprovantes)
    datas_distintas_pgto = _datas_distintas_comprovantes(df_comprovantes)
    if periodo_pgto_min is not None and periodo_pgto_max is not None:
        print(
            f"[INFO] Período coberto pelo PGTO (datas dos comprovantes): "
            f"{periodo_pgto_min.strftime('%d/%m/%Y')} a {periodo_pgto_max.strftime('%d/%m/%Y')}"
        )
        if datas_distintas_pgto:
            print(f"[INFO] Dias com pelo menos um comprovante no PGTO: {len(datas_distintas_pgto)} dia(s).")
        if CONCILIAR_APENAS_EXTRATO_DENTRO_PERIODO_PGTO:
            if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO and datas_distintas_pgto:
                print(
                    "[INFO] Só entram na conciliação SIGRA extratos cuja data coincide com um desses dias "
                    "(não todos os dias entre min e max)."
                )
            else:
                print(
                    "[INFO] Extratos com data fora do intervalo min–max do PGTO não entram na conciliação SIGRA nesta rodagem."
                )
    else:
        print("[AVISO] Não foi possível definir período pelo PGTO (sem datas válidas); todos os extratos serão tentados.")
    
    print(
        f"[INFO] Empates (mesmo dia + valor, sem ref. única no extrato): "
        f"{'não conciliar' if EMPATES_SEM_REF_EXIGIR_DESISTENCIA else 'desempate por menor ID / Ref. Sigra (determinístico)'}"
    )
    
    conciliacoes = []
    comprovantes_usados = set()
    diagnostico_nao_conciliado = {
        'recebimento': 0,
        'extrato_fora_periodo_pgto': 0,
        'sem_comprovantes_disponiveis': 0,
        'sem_data_extrato': 0,
        'sem_comprovantes_mesmo_dia': 0,
        'sem_valor_no_mesmo_dia': 0,
        'com_valor_no_mesmo_dia_mas_favorecido_diferente': 0,
        'com_valor_no_mesmo_dia_mas_ambiguidade': 0,
        'siscomex_sem_rps': 0,
        'grupo_processo_nao_encontrado': 0,
    }
    exemplos_nao_conciliado = {k: [] for k in diagnostico_nao_conciliado}

    def _registrar_motivo(motivo, extrato_row, detalhe=''):
        diagnostico_nao_conciliado[motivo] = diagnostico_nao_conciliado.get(motivo, 0) + 1
        exemplos = exemplos_nao_conciliado.setdefault(motivo, [])
        if len(exemplos) < 5:
            exemplos.append(
                f"ID {extrato_row.get('ID_extrato')} | "
                f"R$ {formatar_valor_br(abs(extrato_row.get('valor', 0) or 0))} | "
                f"{str(extrato_row.get('lancamento_original', '') or '')[:60]}"
                + (f" | {detalhe}" if detalhe else "")
            )
    
    # Ordena extratos por valor (maiores primeiro)
    df_extrato_ordenado = df_extrato.sort_values('valor', key=abs, ascending=False).copy()
    
    total_extratos = len(df_extrato_ordenado)
    print(f"\nTotal de extratos: {total_extratos}")
    print(f"Total de comprovantes: {len(df_comprovantes)}")
    
    for idx, extrato_row in df_extrato_ordenado.iterrows():
        if _extrato_e_recebimento(extrato_row):
            print(
                f"\nProcessando extrato {extrato_row['ID_extrato']}/{total_extratos} "
                f"(R$ {formatar_valor_br(abs(extrato_row['valor']))})"
            )
            print("  [INFO] Extrato de RECEBIMENTO detectado: mantido pendente para conciliação em planilha específica.")
            _registrar_motivo('recebimento', extrato_row)
            continue

        data_extrato = extrato_row.get('data')
        if CONCILIAR_APENAS_EXTRATO_DENTRO_PERIODO_PGTO and pd.notna(data_extrato):
            d_e = pd.Timestamp(data_extrato).normalize()
            fora_base = False
            detalhe = ''
            if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO and datas_distintas_pgto:
                if d_e not in datas_distintas_pgto:
                    fora_base = True
                    detalhe = f"extrato {d_e.strftime('%d/%m/%Y')} sem comprovante nesse dia no PGTO"
            elif periodo_pgto_min is not None and periodo_pgto_max is not None:
                if d_e < periodo_pgto_min or d_e > periodo_pgto_max:
                    fora_base = True
                    detalhe = (
                        f"extrato {d_e.strftime('%d/%m/%Y')} vs PGTO "
                        f"{periodo_pgto_min.strftime('%d/%m/%Y')}-{periodo_pgto_max.strftime('%d/%m/%Y')}"
                    )
            if fora_base:
                _registrar_motivo('extrato_fora_periodo_pgto', extrato_row, detalhe)
                continue

        valor_extrato_abs = abs(extrato_row['valor'])
        print(f"\nProcessando extrato {extrato_row['ID_extrato']}/{total_extratos} (R$ {formatar_valor_br(valor_extrato_abs)})")
        
        # Filtra comprovantes ainda não usados
        comprovantes_disponiveis = df_comprovantes[
            ~df_comprovantes.index.isin(comprovantes_usados)
        ]
        
        if len(comprovantes_disponiveis) == 0:
            print(f"  [INFO] Nenhum comprovante disponível")
            _registrar_motivo('sem_comprovantes_disponiveis', extrato_row)
            continue
        
        # Diagnóstico pré-match para entender por que o extrato não conciliou.
        valor_extrato = abs(extrato_row.get('valor', 0) or 0)
        tem_data = pd.notna(data_extrato)
        comprovantes_mesmo_dia = comprovantes_disponiveis.iloc[0:0]
        comprovantes_mesmo_dia_valor = comprovantes_disponiveis.iloc[0:0]
        comprovantes_mesmo_dia_valor_fav = comprovantes_disponiveis.iloc[0:0]

        if tem_data:
            data_extrato_norm = pd.Timestamp(data_extrato).normalize()

            def _mesmo_dia_diag(row):
                d = row.get('data')
                if pd.isna(d):
                    return False
                return pd.Timestamp(d).normalize() == data_extrato_norm

            comprovantes_mesmo_dia = comprovantes_disponiveis[
                comprovantes_disponiveis.apply(_mesmo_dia_diag, axis=1)
            ]
            valor_extrato_int = _to_centavos(valor_extrato)
            if valor_extrato_int is not None and len(comprovantes_mesmo_dia) > 0:
                comprovantes_mesmo_dia_valor = comprovantes_mesmo_dia[
                    comprovantes_mesmo_dia['valor'].apply(lambda v: _valores_equivalentes(valor_extrato_int, _to_centavos(v)))
                ]
                fav_ext = normalizar_texto(extrato_row.get('favorecido_original', '') or '')
                if fav_ext and len(comprovantes_mesmo_dia_valor) > 0:
                    comprovantes_mesmo_dia_valor_fav = comprovantes_mesmo_dia_valor[
                        comprovantes_mesmo_dia_valor['favorecido_original'].apply(
                            lambda v: (
                                fav_ext == normalizar_texto(v or '')
                                or fav_ext in normalizar_texto(v or '')
                                or normalizar_texto(v or '') in fav_ext
                            )
                        )
                    ]

        # 0) SISCOMEX: busca específica por RPS (transforma número do protocolo e busca nos comprovantes)
        combinacao = buscar_por_siscomex_prot(extrato_row, comprovantes_disponiveis)

        # 0b) PUCOMEX / Receita Federal: soma por processo no mesmo dia (exato; evita depender só da passada final)
        if not combinacao and _extrato_e_pucomex_ou_receita_federal(extrato_row):
            combinacao = encontrar_grupo_receita_federal_por_processo(extrato_row, comprovantes_disponiveis)

        # 1) Prioridade: soma por processo no mesmo dia (Ref. Sigra + Ref. Cliente),
        #    para não casar um valor "solto" quando existe grupo fechado por processo.
        if not combinacao:
            combinacao = encontrar_grupo_por_processo(extrato_row, comprovantes_disponiveis)

        # 2) Se não fechou por processo, tenta match exato com favorecido
        #    (um comprovante = valor + data + favorecido similar)
        if not combinacao:
            combinacao = encontrar_match_exato(
                extrato_row,
                comprovantes_disponiveis,
                exigir_mesmo_favorecido=True,
                evitar_ambiguo=EMPATES_SEM_REF_EXIGIR_DESISTENCIA,
            )

        # 3) Fallback: match só por dia + valor (ex.: SISPAG TRIBUTOS -> GRU AIRP)
        if not combinacao:
            combinacao = encontrar_match_exato(
                extrato_row,
                comprovantes_disponiveis,
                exigir_mesmo_favorecido=False,
                evitar_ambiguo=EMPATES_SEM_REF_EXIGIR_DESISTENCIA,
            )
        
        # Não tenta combinações arbitrárias de comprovantes do mesmo dia: isso somaria valores de
        # processos diferentes para bater o extrato, o que não é permitido.
        
        if combinacao:
            # Marca comprovantes como usados
            comprovantes_usados.update(combinacao)
            
            # Cria registro de conciliação
            for comprovante_idx in combinacao:
                comprovante_row = df_comprovantes.loc[comprovante_idx]

                ref_sigra_val = comprovante_row.get('Ref. Sigra', '-')
                if pd.isna(ref_sigra_val) or not str(ref_sigra_val).strip():
                    ref_sigra_val = '-'

                categoria_val = comprovante_row.get('Categoria', '-')
                if pd.isna(categoria_val) or not str(categoria_val).strip():
                    categoria_val = '-'

                # Cliente do comprovante (coluna F no PGTO MASTER; NÃO usar "Ref Cliente" que é coluna E)
                cliente_val = '-'
                for col in comprovante_row.index:
                    cn = str(col).lower()
                    if 'cliente' in cn and 'ref' not in cn:
                        v = comprovante_row.get(col)
                        if pd.notna(v) and str(v).strip():
                            cliente_val = str(v).strip()
                            break

                conciliacoes.append({
                    'ID_extrato': extrato_row['ID_extrato'],
                    'Data_extrato': extrato_row['data_original'],
                    'Valor_extrato': extrato_row['valor'],
                    'Favorecido_extrato': extrato_row['favorecido_original'],
                    'ID_comprovante': comprovante_row['ID_comprovante'],
                    'Data_comprovante': comprovante_row['data_original'],
                    'Valor_comprovante': comprovante_row['valor'],
                    'Favorecido_comprovante': comprovante_row['favorecido_original'],
                    'Ref. Sigra': ref_sigra_val,
                    'Categoria': categoria_val,
                    'Cliente': cliente_val,
                    'Origem': 'SIGRA'  # Indica que foi conciliado via SIGRA
                })

            print(f"  [✓] Conciliação confirmada: {len(combinacao)} comprovante(s)")
        else:
            txt = f"{extrato_row.get('lancamento_original', '')} {extrato_row.get('favorecido_original', '')}".upper()
            if 'SISCOMEX' in txt and 'PROT' in txt:
                _registrar_motivo('siscomex_sem_rps', extrato_row)
            if not tem_data:
                _registrar_motivo('sem_data_extrato', extrato_row)
            elif len(comprovantes_mesmo_dia) == 0:
                _registrar_motivo('sem_comprovantes_mesmo_dia', extrato_row)
            elif len(comprovantes_mesmo_dia_valor) == 0:
                _registrar_motivo('sem_valor_no_mesmo_dia', extrato_row)
            elif len(comprovantes_mesmo_dia_valor_fav) == 0:
                _registrar_motivo('com_valor_no_mesmo_dia_mas_favorecido_diferente', extrato_row)
            elif len(comprovantes_mesmo_dia_valor) > 1:
                _registrar_motivo('com_valor_no_mesmo_dia_mas_ambiguidade', extrato_row, f"{len(comprovantes_mesmo_dia_valor)} candidatos")
            else:
                _registrar_motivo('grupo_processo_nao_encontrado', extrato_row)
    
    # Passada final (após conciliação normal):
    # Receita Federal por processo (col. D): filtra fornecedor RF, agrupa por processo,
    # soma col. H e busca extrato pendente por valor igual.
    extratos_conciliados_ids = set(c['ID_extrato'] for c in conciliacoes)
    extratos_pendentes_rf = df_extrato[
        (~df_extrato['ID_extrato'].isin(extratos_conciliados_ids)) &
        (~df_extrato.apply(_extrato_e_recebimento, axis=1))
    ]
    if CONCILIAR_APENAS_EXTRATO_DENTRO_PERIODO_PGTO and len(extratos_pendentes_rf) > 0:
        def _rf_dentro_base_pgto(row):
            d = row.get('data')
            if pd.isna(d):
                return False
            dn = pd.Timestamp(d).normalize()
            if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO and datas_distintas_pgto:
                return dn in datas_distintas_pgto
            if periodo_pgto_min is not None and periodo_pgto_max is not None:
                return periodo_pgto_min <= dn <= periodo_pgto_max
            return True

        extratos_pendentes_rf = extratos_pendentes_rf[
            extratos_pendentes_rf.apply(_rf_dentro_base_pgto, axis=1)
        ]
    comprovantes_pendentes_rf = df_comprovantes[~df_comprovantes.index.isin(comprovantes_usados)]

    rf = comprovantes_pendentes_rf[
        comprovantes_pendentes_rf.apply(_comprovante_fornecedor_receita_federal, axis=1)
    ]
    if len(rf) > 0 and len(extratos_pendentes_rf) > 0:
        col_ref_rf = None
        for c in rf.columns:
            cl = str(c).lower()
            if 'sigra' in cl and ('ref' in cl or 'referencia' in cl or 'processo' in cl):
                col_ref_rf = c
                break
        if col_ref_rf is None:
            for c in rf.columns:
                if 'sigra' in str(c).lower():
                    col_ref_rf = c
                    break

        if col_ref_rf is not None:
            print("\n[RF] PASSADA FINAL POR PROCESSO (Fornecedor = Receita Federal)")
            grupos_rf = {}
            for idx_rf, row_rf in rf.iterrows():
                proc_raw = row_rf.get(col_ref_rf, '')
                proc = '' if pd.isna(proc_raw) else str(proc_raw).strip()
                if not proc:
                    continue
                grupos_rf.setdefault(proc, []).append(idx_rf)

            extratos_usados_rf = set()
            for proc, idxs in grupos_rf.items():
                soma_proc = sum(abs(float(rf.loc[i].get('valor', 0) or 0)) for i in idxs)
                soma_proc_int = int(round(soma_proc * 100, 0))

                datas_grupo_norm = set()
                for i in idxs:
                    d = rf.loc[i].get('data')
                    if pd.notna(d):
                        datas_grupo_norm.add(pd.Timestamp(d).normalize())

                candidatos = []
                for idx_ext, row_ext in extratos_pendentes_rf.iterrows():
                    id_ext = row_ext.get('ID_extrato')
                    if id_ext in extratos_usados_rf:
                        continue
                    valor_ext_int = _to_centavos(abs(float(row_ext.get('valor', 0) or 0)))
                    if valor_ext_int is None or valor_ext_int != soma_proc_int:
                        continue
                    # Mesmo dia de calendário entre extrato e o grupo de comprovantes (sem "data mais próxima")
                    if datas_grupo_norm:
                        d_ext = row_ext.get('data')
                        if pd.isna(d_ext):
                            continue
                        d_ext_n = pd.Timestamp(d_ext).normalize()
                        if d_ext_n not in datas_grupo_norm:
                            continue
                    # Desempate: Ref. processo / tokens do extrato (valor já é igual)
                    score_ref = 0
                    texto_e = _texto_extrato_favorecido_busca(row_ext).replace(' ', '')
                    proc_u = str(proc).upper().replace(' ', '')
                    if proc_u and proc_u in texto_e:
                        score_ref += 3
                    for tok in _tokens_referencia_no_extrato(row_ext, min_len=6):
                        if tok and tok in proc_u:
                            score_ref += 1
                    candidatos.append((score_ref, int(id_ext) if pd.notna(id_ext) else 0, idx_ext))

                if not candidatos:
                    continue

                candidatos.sort(key=lambda x: (-x[0], x[1]))
                if (
                    len(candidatos) > 1
                    and candidatos[0][0] == candidatos[1][0]
                    and candidatos[0][0] == 0
                    and EMPATES_SEM_REF_EXIGIR_DESISTENCIA
                ):
                    continue

                _, _, idx_ext_sel = candidatos[0]
                row_ext_sel = extratos_pendentes_rf.loc[idx_ext_sel]
                id_ext_sel = row_ext_sel.get('ID_extrato')
                extratos_usados_rf.add(id_ext_sel)
                comprovantes_usados.update(idxs)

                print(
                    f"    [OK] RF processo {proc}: {len(idxs)} comprovante(s) "
                    f"= R$ {formatar_valor_br(soma_proc)} -> Extrato ID {id_ext_sel}"
                )

                for comprovante_idx in idxs:
                    comprovante_row = df_comprovantes.loc[comprovante_idx]
                    ref_sigra_val = comprovante_row.get('Ref. Sigra', '-')
                    if pd.isna(ref_sigra_val) or not str(ref_sigra_val).strip():
                        ref_sigra_val = '-'

                    categoria_val = comprovante_row.get('Categoria', '-')
                    if pd.isna(categoria_val) or not str(categoria_val).strip():
                        categoria_val = '-'

                    cliente_val = '-'
                    for col in comprovante_row.index:
                        cn = str(col).lower()
                        if 'cliente' in cn and 'ref' not in cn:
                            v = comprovante_row.get(col)
                            if pd.notna(v) and str(v).strip():
                                cliente_val = str(v).strip()
                                break

                    conciliacoes.append({
                        'ID_extrato': row_ext_sel['ID_extrato'],
                        'Data_extrato': row_ext_sel['data_original'],
                        'Valor_extrato': row_ext_sel['valor'],
                        'Favorecido_extrato': row_ext_sel['favorecido_original'],
                        'ID_comprovante': comprovante_row['ID_comprovante'],
                        'Data_comprovante': comprovante_row['data_original'],
                        'Valor_comprovante': comprovante_row['valor'],
                        'Favorecido_comprovante': comprovante_row['favorecido_original'],
                        'Ref. Sigra': ref_sigra_val,
                        'Categoria': categoria_val,
                        'Cliente': cliente_val,
                        'Origem': 'SIGRA_RF_PROCESSO'
                    })
        else:
            print("[RF] PASSADA FINAL: coluna Ref. Sigra/processo não encontrada; etapa RF ignorada.")

    # Cria DataFrames de resultado
    df_conciliacao = pd.DataFrame(conciliacoes)
    
    # Extratos pendentes (não conciliados). Mantém todos para a passada de Numerário e para o Excel;
    # o Status Extrato marca “Fora do período PGTO” quando não há base SIGRA para aquela data.
    extratos_conciliados = set(df_conciliacao['ID_extrato'].unique()) if len(df_conciliacao) > 0 else set()
    df_extratos_pendentes = df_extrato[~df_extrato['ID_extrato'].isin(extratos_conciliados)].copy()
    
    # Comprovantes pendentes (não usados)
    df_comprovantes_pendentes = df_comprovantes[~df_comprovantes.index.isin(comprovantes_usados)].copy()
    
    print("\n" + "="*80)
    print("RESUMO DA CONCILIAÇÃO")
    print("="*80)
    print(f"Conciliações encontradas: {len(df_conciliacao)}")
    print(f"Extratos pendentes: {len(df_extratos_pendentes)}")
    print(f"Comprovantes pendentes: {len(df_comprovantes_pendentes)}")
    if diagnostico_nao_conciliado.get('extrato_fora_periodo_pgto', 0) > 0:
        print(
            f"[INFO] {diagnostico_nao_conciliado['extrato_fora_periodo_pgto']} extrato(s) fora da base de datas do PGTO "
            "(intervalo ou dias sem linha no arquivo, conforme configuração)."
        )
    print("-"*80)
    print("DIAGNÓSTICO (não conciliados no fluxo principal):")
    total_diag = sum(diagnostico_nao_conciliado.values())
    if total_diag == 0:
        print("Sem ocorrências para diagnóstico.")
    else:
        ordenado = sorted(diagnostico_nao_conciliado.items(), key=lambda x: x[1], reverse=True)
        for motivo, qtd in ordenado:
            if qtd <= 0:
                continue
            perc = (qtd / total_diag) * 100
            print(f"- {motivo}: {qtd} ({perc:.1f}%)")
            for ex in exemplos_nao_conciliado.get(motivo, []):
                print(f"    ex.: {ex}")
    
    return (
        df_conciliacao,
        df_extratos_pendentes,
        df_comprovantes_pendentes,
        periodo_pgto_min,
        periodo_pgto_max,
        datas_distintas_pgto,
    )


def criar_aba_status_extrato(
    df_extrato,
    df_conciliacao,
    periodo_pgto_min=None,
    periodo_pgto_max=None,
    datas_distintas_pgto=None,
):
    """
    Cria a aba 'Status Extrato' com o resumo de cada lançamento.
    Se datas_distintas_pgto for informado, 'fora do PGTO' = data do extrato não está nesse conjunto.
    """
    conciliacoes_por_extrato = {}

    if len(df_conciliacao) > 0:
        for _, row in df_conciliacao.iterrows():
            id_extrato = row.get('ID_extrato')
            if pd.isna(id_extrato):
                continue

            if id_extrato not in conciliacoes_por_extrato:
                conciliacoes_por_extrato[id_extrato] = {
                    'ids_comprovantes': [],
                    'valores': [],
                    'ref_sigra': [],
                    'categorias': [],
                    'clientes': [],  # Cliente (coluna F do PGTO MASTER)
                    'origens': []  # SIGRA ou Numerário
                }

            conciliacoes_por_extrato[id_extrato]['ids_comprovantes'].append(row.get('ID_comprovante'))

            valor_comp = row.get('Valor_comprovante', 0)
            try:
                valor_comp_float = abs(float(valor_comp))
            except (TypeError, ValueError):
                valor_comp_float = 0.0
            conciliacoes_por_extrato[id_extrato]['valores'].append(valor_comp_float)

            ref_sigra_val_raw = row.get('Ref. Sigra', '-')
            if pd.isna(ref_sigra_val_raw):
                ref_sigra_val = '-'
            else:
                ref_sigra_val = str(ref_sigra_val_raw).strip()
                if not ref_sigra_val:
                    ref_sigra_val = '-'
            conciliacoes_por_extrato[id_extrato]['ref_sigra'].append(ref_sigra_val)

            categoria_val_raw = row.get('Categoria', '-')
            if pd.isna(categoria_val_raw):
                categoria_val = '-'
            else:
                categoria_val = str(categoria_val_raw).strip()
                if not categoria_val:
                    categoria_val = '-'
            conciliacoes_por_extrato[id_extrato]['categorias'].append(categoria_val)

            cliente_val = row.get('Cliente', '-')
            if pd.isna(cliente_val):
                cliente_val = '-'
            else:
                cliente_val = str(cliente_val).strip() or '-'
            conciliacoes_por_extrato[id_extrato]['clientes'].append(cliente_val)
            
            # Origem da conciliação (SIGRA ou Numerário)
            origem_val = row.get('Origem', 'SIGRA')
            if pd.isna(origem_val):
                origem_val = 'SIGRA'
            conciliacoes_por_extrato[id_extrato]['origens'].append(str(origem_val).strip())

    status_lista = []

    for _, extrato_row in df_extrato.iterrows():
        id_extrato = extrato_row.get('ID_extrato')
        valor_extrato = extrato_row.get('valor', 0)
        # Preferir nome do lançamento (ex: "PUCOMEX...", "RECEBIMENTOS LUXDATA COMERCIO") para Favorecido/Descrição
        favorecido = extrato_row.get('lancamento_original', '')
        if not favorecido or (isinstance(favorecido, str) and not favorecido.strip()):
            col_fav_extrato = COLUNAS_EXTRATO.get('favorecido')
            if col_fav_extrato and col_fav_extrato in extrato_row.index:
                favorecido = extrato_row.get(col_fav_extrato, '')
            if not favorecido:
                favorecido = extrato_row.get('favorecido_original', '')
        # Usar coluna 'data' já normalizada (DD/MM) — evita pd.to_datetime em 'data_original'
        # sem dayfirst, que interpreta 12/02/2026 como 2/dez (MM/DD) e exibe 02/12/2026.
        data_extrato_norm = extrato_row.get('data')
        data_extrato_raw = extrato_row.get('data_original', '')

        try:
            valor_extrato_float = abs(float(valor_extrato))
        except (TypeError, ValueError):
            valor_extrato_float = 0.0

        data_formatada = ''
        if pd.notna(data_extrato_norm):
            data_formatada = pd.Timestamp(data_extrato_norm).strftime('%d/%m/%Y')
        elif pd.notna(data_extrato_raw) and data_extrato_raw != '':
            data_dt = pd.to_datetime(data_extrato_raw, errors='coerce', dayfirst=True)
            if pd.notna(data_dt):
                data_formatada = data_dt.strftime('%d/%m/%Y')
            else:
                data_formatada = str(data_extrato_raw)

        categoria_extrato = '-'
        for coluna in extrato_row.index:
            nome_coluna = str(coluna).lower()
            if 'categoria' in nome_coluna:
                valor_coluna = extrato_row.get(coluna)
                if pd.notna(valor_coluna) and str(valor_coluna).strip():
                    categoria_extrato = str(valor_coluna).strip()
                    break

        if id_extrato in conciliacoes_por_extrato:
            info = conciliacoes_por_extrato[id_extrato]
            qtd_comprovantes = len(info['ids_comprovantes'])
            valor_total = round(sum(info['valores']), 2)
            extrato_cents = int(round(valor_extrato_float * 100))
            total_cents = sum(int(round(abs(v) * 100)) for v in info['valores'])
            diferenca = round((extrato_cents - total_cents) / 100.0, 2)
            ids_comprovantes = ', '.join(str(x) for x in info['ids_comprovantes'] if pd.notna(x))

            refs_unicas = []
            for r in info['ref_sigra']:
                if r == '-' or pd.isna(r):
                    continue
                r_str = str(r).strip()
                if not r_str:
                    continue
                if r_str not in refs_unicas:
                    refs_unicas.append(r_str)
            ref_sigra_str = ', '.join(refs_unicas) if refs_unicas else '-'

            categorias_validas = []
            for c in info['categorias']:
                if c == '-' or pd.isna(c):
                    continue
                c_str = str(c).strip()
                if not c_str:
                    continue
                if c_str not in categorias_validas:
                    categorias_validas.append(c_str)
            categoria_comprovantes_str = ', '.join(categorias_validas) if categorias_validas else '-'

            clientes_validos = []
            for c in info.get('clientes', []):
                if c == '-' or pd.isna(c):
                    continue
                c_str = str(c).strip()
                if not c_str:
                    continue
                if c_str not in clientes_validos:
                    clientes_validos.append(c_str)
            cliente_str = ', '.join(clientes_validos) if clientes_validos else '-'

            if categoria_extrato != '-' and categoria_comprovantes_str == '-':
                categoria_final = categoria_extrato
            elif categoria_extrato != '-' and categoria_comprovantes_str != '-':
                categoria_final = f"{categoria_extrato} ({categoria_comprovantes_str})"
            else:
                categoria_final = categoria_comprovantes_str if categoria_comprovantes_str != '-' else categoria_extrato

            # Identifica origem da conciliação (SIGRA ou Numerário)
            origens_unicas = list(set(info['origens']))
            origem_str = ', '.join(origens_unicas) if origens_unicas else 'SIGRA'
            
            if extrato_cents != total_cents:
                sinal = "faltam" if diferenca > 0 else "sobram"
                observacao = (
                    f"⚠️ Valor não bate (R$ {formatar_valor_br(abs(diferenca))} — {sinal} nos comprovantes) | Origem: {origem_str}"
                )
            elif qtd_comprovantes > 1:
                observacao = f"Conciliado com {qtd_comprovantes} comprovantes | Origem: {origem_str}"
            else:
                observacao = f"Perfeitamente conciliado | Origem: {origem_str}"

            if extrato_cents != total_cents:
                status = '⚠️ Divergência de valor'
            else:
                status = '✅ Conciliado'
        else:
            qtd_comprovantes = 0
            valor_total = 0.0
            diferenca = valor_extrato_float
            ids_comprovantes = '-'
            ref_sigra_str = '-'
            categoria_final = categoria_extrato
            cliente_str = '-'
            fora_pgto = False
            if pd.notna(data_extrato_norm):
                de = pd.Timestamp(data_extrato_norm).normalize()
                if datas_distintas_pgto is not None and len(datas_distintas_pgto) > 0:
                    if de not in datas_distintas_pgto:
                        fora_pgto = True
                elif periodo_pgto_min is not None and periodo_pgto_max is not None:
                    if de < periodo_pgto_min or de > periodo_pgto_max:
                        fora_pgto = True
            if fora_pgto:
                if datas_distintas_pgto is not None and len(datas_distintas_pgto) > 0:
                    observacao = (
                        "Data do extrato sem nenhum comprovante nesse dia no arquivo PGTO. "
                        "Nesta rodagem a base SIGRA só cobre dias que existem no PGTO."
                    )
                else:
                    observacao = (
                        f"Data fora do período do arquivo PGTO "
                        f"({periodo_pgto_min.strftime('%d/%m/%Y')} a {periodo_pgto_max.strftime('%d/%m/%Y')}). "
                        f"Nesta rodagem não há comprovantes para essa data."
                    )
                status = 'ℹ️ Fora do período PGTO'
            elif _extrato_e_recebimento(extrato_row):
                observacao = 'Recebimento: não conciliado neste fluxo SIGRA/Numerário.'
                status = 'ℹ️ Recebimento'
            else:
                observacao = 'Não encontrou comprovantes correspondentes'
                status = '❌ Pendente'

        status_lista.append({
            'ID Extrato': id_extrato,
            'Data': data_formatada,
            'Valor Extrato': valor_extrato_float,
            'Favorecido/Descrição': str(favorecido) if favorecido else '',
            'Status': status,
            'Qtd Comprovantes': qtd_comprovantes,
            'Valor Total Conciliado': valor_total,
            'Diferença': diferenca,
            'Ref. Sigra': ref_sigra_str,
            'Categoria': categoria_final,
            'Cliente': cliente_str,
            'IDs Comprovantes': ids_comprovantes,
            'Observação': observacao
        })

    df_status = pd.DataFrame(status_lista)
    return df_status


def dividir_status_extrato_por_mes(df_status):
    """
    Divide o DataFrame de status em múltiplas abas mensais.
    Ex.: 'Status Extrato - Jan26', 'Status Extrato - Fev26', etc.
    """
    if df_status is None or len(df_status) == 0:
        return {'Status Extrato': df_status}

    if 'Data' not in df_status.columns:
        return {'Status Extrato': df_status}

    df_tmp = df_status.copy()
    df_tmp['_data_mes'] = pd.to_datetime(df_tmp['Data'], errors='coerce', dayfirst=True)

    # Se não conseguiu interpretar datas, mantém aba única
    if df_tmp['_data_mes'].notna().sum() == 0:
        return {'Status Extrato': df_status}

    meses_abrev = {
        1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
        7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
    }

    resultado = {}
    df_validas = df_tmp[df_tmp['_data_mes'].notna()].copy()
    df_invalidas = df_tmp[df_tmp['_data_mes'].isna()].copy()

    periodos = sorted(df_validas['_data_mes'].dt.to_period('M').unique())
    for periodo in periodos:
        mask = df_validas['_data_mes'].dt.to_period('M') == periodo
        df_mes = df_validas.loc[mask].drop(columns=['_data_mes']).copy()
        mes = int(periodo.month)
        ano2 = str(periodo.year)[-2:]
        nome_aba = f"Status Extrato - {meses_abrev.get(mes, str(mes))}{ano2}"
        resultado[nome_aba] = df_mes

    # Registros sem data válida ficam em aba separada para não perder informação
    if len(df_invalidas) > 0:
        resultado['Status Extrato - SemData'] = df_invalidas.drop(columns=['_data_mes'])

    return resultado


def formatar_aba_status_extrato(workbook):
    """
    Aplica formatação profissional às abas de Status Extrato.
    Cabeçalho em destaque, bordas, cores por status, colunas monetárias formatadas.
    """
    if not OPENPYXL_DISPONIVEL:
        return
    abas_status = [s for s in workbook.sheetnames if str(s).startswith('Status Extrato')]
    if not abas_status:
        return

    # Cores (paleta profissional)
    COR_CABECALHO = "2F5496"       # Azul escuro
    COR_CONCILIADO = "D4EDDA"     # Verde claro
    COR_PENDENTE = "FFF3CD"        # Âmbar claro
    COR_DIFERENCA = "F8D7DA"       # Vermelho claro (alerta)
    COR_INFO = "E7F3FF"            # Azul bem claro (fora do período PGTO / recebimento)
    COR_ALTERNADA = "F8F9FA"       # Cinza muito claro
    COR_BORDA = "CCCCCC"

    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    preenchimento_cabecalho = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    borda_fina = Border(
        left=Side(style='thin', color=COR_BORDA),
        right=Side(style='thin', color=COR_BORDA),
        top=Side(style='thin', color=COR_BORDA),
        bottom=Side(style='thin', color=COR_BORDA)
    )
    alinhamento_centro = Alignment(horizontal='center', vertical='center', wrap_text=False)
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for nome_aba in abas_status:
        sheet = workbook[nome_aba]
        # Identifica índices das colunas pelo cabeçalho
        col_valor_extrato = None
        col_valor_conciliado = None
        col_diferenca = None
        col_status = None
        col_data = None
        headers_row = list(sheet[1])
        for col_idx, cell in enumerate(headers_row, 1):
            if not cell.value:
                continue
            h = str(cell.value).lower()
            if 'valor extrato' in h:
                col_valor_extrato = col_idx
            elif 'valor total conciliado' in h:
                col_valor_conciliado = col_idx
            elif 'diferença' in h or 'diferenca' in h:
                col_diferenca = col_idx
            elif h == 'status':
                col_status = col_idx
            elif 'data' in h:
                col_data = col_idx

        # Cabeçalho
        for cell in sheet[1]:
            cell.font = fonte_cabecalho
            cell.fill = preenchimento_cabecalho
            cell.border = borda_fina
            cell.alignment = alinhamento_centro

        # Congelar painel (cabeçalho fixo)
        sheet.freeze_panes = 'A2'

        # Larguras das colunas (nome do cabeçalho -> largura)
        larguras = {
            'ID Extrato': 12,
            'Data': 12,
            'Valor Extrato': 18,
            'Favorecido/Descrição': 38,
            'Status': 16,
            'Qtd Comprovantes': 14,
            'Valor Total Conciliado': 20,
            'Diferença': 14,
            'Ref. Sigra': 18,
            'Categoria': 28,
            'Cliente': 32,
            'IDs Comprovantes': 28,
            'Observação': 42
        }
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value and str(cell.value).strip() in larguras:
                sheet.column_dimensions[get_column_letter(col_idx)].width = larguras[str(cell.value).strip()]

        # Dados: bordas, alinhamentos, cores por status e valores
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            for col_idx, cell in enumerate(row, 1):
                cell.border = borda_fina
                celula_com_cor = False

                # Valores monetários
                if col_valor_extrato and col_idx == col_valor_extrato:
                    if cell.value is not None and cell.value != '':
                        try:
                            cell.number_format = '"R$" #,##0.00'
                            cell.alignment = alinhamento_direita
                        except Exception:
                            pass
                elif col_valor_conciliado and col_idx == col_valor_conciliado:
                    if cell.value is not None and cell.value != '':
                        try:
                            cell.number_format = '"R$" #,##0.00'
                            cell.alignment = alinhamento_direita
                        except Exception:
                            pass
                elif col_diferenca and col_idx == col_diferenca:
                    if cell.value is not None and cell.value != '':
                        try:
                            cell.number_format = '"R$" #,##0.00'
                            cell.alignment = alinhamento_direita
                            if abs(float(cell.value)) > 0.005:
                                cell.fill = PatternFill(start_color=COR_DIFERENCA, end_color=COR_DIFERENCA, fill_type="solid")
                                celula_com_cor = True
                        except Exception:
                            pass
                # Coluna Status: cor por valor
                elif col_status and col_idx == col_status and cell.value:
                    cell.alignment = alinhamento_centro
                    txt = str(cell.value)
                    if '✅' in txt or 'Conciliado' in txt:
                        cell.fill = PatternFill(start_color=COR_CONCILIADO, end_color=COR_CONCILIADO, fill_type="solid")
                        celula_com_cor = True
                    elif '❌' in txt or 'Pendente' in txt:
                        cell.fill = PatternFill(start_color=COR_PENDENTE, end_color=COR_PENDENTE, fill_type="solid")
                        celula_com_cor = True
                    elif '⚠️' in txt:
                        cell.fill = PatternFill(start_color=COR_DIFERENCA, end_color=COR_DIFERENCA, fill_type="solid")
                        celula_com_cor = True
                    elif 'ℹ️' in txt:
                        cell.fill = PatternFill(start_color=COR_INFO, end_color=COR_INFO, fill_type="solid")
                        celula_com_cor = True
                # Data: centralizado
                elif col_data and col_idx == col_data:
                    cell.alignment = alinhamento_centro
                # Favorecido/Descrição e Observação: wrap e esquerda
                elif col_idx <= len(headers_row) and headers_row[col_idx - 1].value:
                    h = str(headers_row[col_idx - 1].value)
                    if 'Favorecido' in h or 'Descrição' in h or 'Observação' in h:
                        cell.alignment = alinhamento_esquerda
                # Linhas alternadas (sem sobrescrever status/diferença)
                if row_idx % 2 == 0 and not celula_com_cor:
                    cell.fill = PatternFill(start_color=COR_ALTERNADA, end_color=COR_ALTERNADA, fill_type="solid")

        print(f"  [OK] Aba '{nome_aba}' formatada")


# ============================================================================
# FUNÇÕES DE BUSCA DE ARQUIVOS
# ============================================================================

def buscar_arquivos_itau_sigra():
    """
    Busca automaticamente os arquivos do Itaú, SIGRA e Numerário na pasta do dia de hoje.
    
    Returns:
        Tupla com (caminho_extrato, caminho_comprovantes, caminho_numerario)
        caminho_numerario pode ser None se o arquivo não for encontrado
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_hoje = datetime.now().strftime('%Y-%m-%d')
    pasta_downloads = os.path.join(script_dir, 'downloads', data_hoje)
    
    print(f"\n[INFO] Buscando arquivos na pasta: {pasta_downloads}")
    
    # Verifica se a pasta existe - se não, busca a mais recente
    if not os.path.exists(pasta_downloads):
        pasta_base = os.path.join(script_dir, 'downloads')
        if os.path.exists(pasta_base):
            pastas = []
            for item in os.listdir(pasta_base):
                caminho_item = os.path.join(pasta_base, item)
                if os.path.isdir(caminho_item) and len(item) == 10 and item.count('-') == 2:
                    try:
                        datetime.strptime(item, '%Y-%m-%d')
                        pastas.append(item)
                    except:
                        pass
            
            if pastas:
                pastas.sort()
                pasta_mais_recente = pastas[-1]
                pasta_downloads = os.path.join(pasta_base, pasta_mais_recente)
                print(f"[AVISO] Pasta de hoje ({data_hoje}) não encontrada.")
                print(f"[INFO] Usando pasta mais recente disponível: {pasta_mais_recente}")
            else:
                raise FileNotFoundError(f"Pasta não encontrada: {pasta_downloads}")
        else:
            raise FileNotFoundError(f"Pasta base 'downloads' não existe: {pasta_base}")
    
    # Busca arquivo do extrato Itaú (vários padrões - formato pode variar com nova estrutura do banco)
    padroes_extrato = [
        '*itau*.xlsx', '*ITAU*.xlsx',
        '*extrato*itau*.xlsx', '*EXTRATO*ITAU*.xlsx',
        '*extrato*master*.xlsx', '*EXTRATO*MASTER*.xlsx',
        '*master*itau*.xlsx', '*MASTER*ITAU*.xlsx',
    ]
    arquivos_extrato = []
    for padrao in padroes_extrato:
        arquivos_extrato.extend(glob.glob(os.path.join(pasta_downloads, padrao)))
    
    # Busca arquivo de comprovantes SIGRA / PGTO MASTER (nomes variam no Drive)
    arquivos_comprovantes = glob.glob(os.path.join(pasta_downloads, '*pgto*sigra*.xlsx'))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*PGTO*SIGRA*.xlsx')))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*pgto*master*.xlsx')))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*PGTO*MASTER*.xlsx')))
    
    # Busca arquivo de numerário (opcional)
    arquivos_numerario = glob.glob(os.path.join(pasta_downloads, '*numerario*.xlsx'))
    arquivos_numerario.extend(glob.glob(os.path.join(pasta_downloads, '*NUMERARIO*.xlsx')))
    arquivos_numerario.extend(glob.glob(os.path.join(pasta_downloads, '*numerário*.xlsx')))
    arquivos_numerario.extend(glob.glob(os.path.join(pasta_downloads, '*NUMERÁRIO*.xlsx')))
    
    # Remove duplicatas e arquivos temporários do Excel (~$) - exclui EXTRATO BB (Banco do Brasil)
    arquivos_extrato = list(set(arquivos_extrato))
    arquivos_extrato = [f for f in arquivos_extrato if ('bb' not in os.path.basename(f).lower()) or ('itau' in os.path.basename(f).lower())]
    arquivos_comprovantes = list(set(arquivos_comprovantes))
    arquivos_numerario = list(set(arquivos_numerario))
    arquivos_extrato = [
        f for f in arquivos_extrato if not os.path.basename(f).startswith('~$')
    ]
    arquivos_comprovantes = [
        f for f in arquivos_comprovantes if not os.path.basename(f).startswith('~$')
    ]
    arquivos_numerario = [
        f for f in arquivos_numerario if not os.path.basename(f).startswith('~$')
    ]

    def ordenar_por_relevancia(caminhos):
        """
        Ordena por data de modificação (mais novo primeiro) e tamanho (maior primeiro).
        Evita escolher arquivo antigo/errado quando há múltiplos candidatos no diretório.
        """
        validos = [c for c in caminhos if os.path.isfile(c)]
        return sorted(
            validos,
            key=lambda p: (os.path.getmtime(p), os.path.getsize(p)),
            reverse=True
        )

    arquivos_extrato = ordenar_por_relevancia(arquivos_extrato)
    arquivos_comprovantes = ordenar_por_relevancia(arquivos_comprovantes)
    arquivos_numerario = ordenar_por_relevancia(arquivos_numerario)
    
    # Verifica se encontrou os arquivos
    if not arquivos_extrato:
        raise FileNotFoundError(f"Nenhum arquivo de extrato Itaú encontrado na pasta {pasta_downloads}")
    
    if not arquivos_comprovantes:
        raise FileNotFoundError(
            f"Nenhum arquivo de comprovantes (PGTO/SIGRA) encontrado na pasta {pasta_downloads}. "
            "Esperado nome contendo pgto+sigra ou pgto+master (.xlsx)."
        )
    
    caminho_extrato = arquivos_extrato[0]
    caminho_comprovantes = arquivos_comprovantes[0]
    caminho_numerario = arquivos_numerario[0] if arquivos_numerario else None
    
    print(f"[OK] Extrato encontrado: {os.path.basename(caminho_extrato)}")
    print(f"[OK] Comprovantes encontrados: {os.path.basename(caminho_comprovantes)}")
    if caminho_numerario:
        print(f"[OK] Numerário encontrado: {os.path.basename(caminho_numerario)}")
    else:
        print(f"[INFO] Arquivo de numerário não encontrado (opcional)")
    
    return caminho_extrato, caminho_comprovantes, caminho_numerario


# ============================================================================
# COMPLETAR PLANILHA (ACUMULAR EM VEZ DE SUBSTITUIR)
# ============================================================================

def carregar_dados_existentes_itau(caminho_saida):
    """
    Carrega as abas do arquivo de conciliação existente (se existir).
    Usado para completar a planilha com novos dados em vez de substituir.
    Returns:
        Dict com DataFrames por aba (extrato, comprovantes, conciliacao, etc.) ou None se arquivo não existir/vazio.
    """
    if not os.path.exists(caminho_saida):
        return None
    try:
        xl = pd.ExcelFile(caminho_saida, engine='openpyxl')
        result = {}
        for sheet in ['extrato', 'comprovantes', 'conciliacao', 'pendencias_extratos', 'pendencias_comprovantes', 'Status Extrato']:
            if sheet in xl.sheet_names:
                df = pd.read_excel(caminho_saida, sheet_name=sheet, engine='openpyxl')
                if len(df) > 0 or sheet in ('conciliacao', 'pendencias_extratos', 'pendencias_comprovantes'):
                    result[sheet] = df
        if 'numerario' in xl.sheet_names:
            result['numerario'] = pd.read_excel(caminho_saida, sheet_name='numerario', engine='openpyxl')
        if 'extrato' not in result or len(result.get('extrato', [])) == 0:
            return None
        return result
    except Exception as e:
        print(f"[AVISO] Não foi possível carregar planilha existente: {e}")
        return None


def mesclar_dados_itau(existentes, df_extrato, df_comprovantes, df_conciliacao,
                       df_extratos_pendentes, df_comprovantes_pendentes, df_numerario):
    """
    Mescla dados da execução atual com os já existentes na planilha.
    Renumera ID_extrato e ID_comprovante da execução atual para não colidir com os existentes.
    Retorna (df_extrato_total, df_comprovantes_total, df_conciliacao_total,
             df_extratos_pendentes_total, df_comprovantes_pendentes_total, df_numerario_total).
    """
    df_ext = existentes.get('extrato')
    df_comp = existentes.get('comprovantes')
    df_conc = existentes.get('conciliacao')
    df_pend_ext = existentes.get('pendencias_extratos')
    df_pend_comp = existentes.get('pendencias_comprovantes')
    df_num = existentes.get('numerario')

    def _normalizar_data_chave(v):
        if pd.isna(v):
            return ''
        try:
            return pd.Timestamp(v).normalize().strftime('%Y-%m-%d')
        except Exception:
            try:
                dt = pd.to_datetime(v, errors='coerce', dayfirst=True)
                if pd.notna(dt):
                    return pd.Timestamp(dt).normalize().strftime('%Y-%m-%d')
            except Exception:
                pass
        return str(v).strip()

    def _normalizar_texto_chave(v):
        if pd.isna(v):
            return ''
        return normalizar_texto(v)

    def _chave_extrato_linha(row):
        data_ref = row.get('data', pd.NA)
        if pd.isna(data_ref):
            data_ref = row.get('data_original', pd.NA)
        valor_centavos = int(round(normalizar_valor(row.get('valor', row.get('valor_original', 0))) * 100, 0))
        lanc = _normalizar_texto_chave(row.get('lancamento_original', ''))
        fav = _normalizar_texto_chave(row.get('favorecido_original', row.get('Razão Social', '')))
        aba = _normalizar_texto_chave(row.get('aba_origem_extrato', ''))
        return (
            _normalizar_data_chave(data_ref),
            valor_centavos,
            lanc,
            fav,
            aba,
        )

    # Evita duplicar extrato quando a base já contém linhas antigas e a equipe apenas acrescenta novas.
    # Estratégia com contagem por chave: preserva duplicatas legítimas e remove apenas reimportações.
    if df_ext is not None and len(df_ext) > 0 and len(df_extrato) > 0 and 'ID_extrato' in df_extrato.columns:
        chaves_existentes = [_chave_extrato_linha(row) for _, row in df_ext.iterrows()]
        contagem_existentes = Counter(chaves_existentes)
        contagem_novo = Counter()
        manter_flags = []
        duplicadas_removidas = 0

        for _, row in df_extrato.iterrows():
            chave = _chave_extrato_linha(row)
            contagem_novo[chave] += 1
            if contagem_novo[chave] <= contagem_existentes.get(chave, 0):
                manter_flags.append(False)
                duplicadas_removidas += 1
            else:
                manter_flags.append(True)

        if duplicadas_removidas > 0:
            df_extrato = df_extrato.loc[manter_flags].copy()
            ids_extrato_mantidos = set(df_extrato['ID_extrato'].tolist()) if len(df_extrato) > 0 else set()
            if len(df_conciliacao) > 0 and 'ID_extrato' in df_conciliacao.columns:
                df_conciliacao = df_conciliacao[df_conciliacao['ID_extrato'].isin(ids_extrato_mantidos)].copy()
            if len(df_extratos_pendentes) > 0 and 'ID_extrato' in df_extratos_pendentes.columns:
                df_extratos_pendentes = df_extratos_pendentes[
                    df_extratos_pendentes['ID_extrato'].isin(ids_extrato_mantidos)
                ].copy()
            print(f"[INFO] Extrato: {duplicadas_removidas} linha(s) já existente(s) ignorada(s) para evitar duplicação.")

    max_id_extrato = int(df_ext['ID_extrato'].max()) if df_ext is not None and len(df_ext) > 0 and 'ID_extrato' in df_ext.columns else 0
    max_id_comprovante = int(df_comp['ID_comprovante'].max()) if df_comp is not None and len(df_comp) > 0 and 'ID_comprovante' in df_comp.columns else 0
    max_id_numerario = int(df_num['ID_comprovante'].max()) if df_num is not None and len(df_num) > 0 and 'ID_comprovante' in df_num.columns else 0

    # Mapeamento: ID antigo (da execução atual) -> ID novo (após append)
    mapa_extrato = {}
    if len(df_extrato) > 0 and 'ID_extrato' in df_extrato.columns:
        for i, old_id in enumerate(df_extrato['ID_extrato'].values, start=1):
            mapa_extrato[old_id] = max_id_extrato + i
    mapa_comprovante = {}
    if len(df_comprovantes) > 0 and 'ID_comprovante' in df_comprovantes.columns:
        for i, old_id in enumerate(df_comprovantes['ID_comprovante'].values, start=1):
            mapa_comprovante[old_id] = max_id_comprovante + i
    mapa_numerario = {}
    if df_numerario is not None and len(df_numerario) > 0 and 'ID_comprovante' in df_numerario.columns:
        for i, old_id in enumerate(df_numerario['ID_comprovante'].values, start=1):
            mapa_numerario[old_id] = max_id_numerario + i

    # Renumera extrato e comprovantes da execução atual
    df_extrato_novo = df_extrato.copy()
    df_comprovantes_novo = df_comprovantes.copy()
    if 'ID_extrato' in df_extrato_novo.columns:
        df_extrato_novo['ID_extrato'] = df_extrato_novo['ID_extrato'].map(mapa_extrato).fillna(df_extrato_novo['ID_extrato']).astype(int)
    if 'ID_comprovante' in df_comprovantes_novo.columns:
        df_comprovantes_novo['ID_comprovante'] = df_comprovantes_novo['ID_comprovante'].map(mapa_comprovante).fillna(df_comprovantes_novo['ID_comprovante']).astype(int)

    # Conciliação: remapeia ID_extrato e ID_comprovante (ID_comprovante conforme Origem)
    df_conc_novo = df_conciliacao.copy()
    if len(df_conc_novo) > 0:
        if 'ID_extrato' in df_conc_novo.columns:
            df_conc_novo['ID_extrato'] = df_conc_novo['ID_extrato'].map(mapa_extrato).fillna(df_conc_novo['ID_extrato'])
        if 'ID_comprovante' in df_conc_novo.columns and 'Origem' in df_conc_novo.columns:
            def remap_id_comp(row):
                o = row.get('Origem', 'SIGRA')
                old = row.get('ID_comprovante')
                if pd.isna(old):
                    return old
                if str(o).strip().upper() == 'NUMERÁRIO' or 'Numerário' in str(o):
                    return mapa_numerario.get(old, old)
                return mapa_comprovante.get(old, old)
            df_conc_novo['ID_comprovante'] = df_conc_novo.apply(remap_id_comp, axis=1)
        elif 'ID_comprovante' in df_conc_novo.columns:
            df_conc_novo['ID_comprovante'] = df_conc_novo['ID_comprovante'].map(mapa_comprovante).fillna(df_conc_novo['ID_comprovante'])

    # Pendências: remapeia IDs
    df_pend_ext_novo = df_extratos_pendentes.copy()
    df_pend_comp_novo = df_comprovantes_pendentes.copy()
    if len(df_pend_ext_novo) > 0 and 'ID_extrato' in df_pend_ext_novo.columns:
        df_pend_ext_novo['ID_extrato'] = df_pend_ext_novo['ID_extrato'].map(mapa_extrato).fillna(df_pend_ext_novo['ID_extrato']).astype(int)
    if len(df_pend_comp_novo) > 0 and 'ID_comprovante' in df_pend_comp_novo.columns:
        df_pend_comp_novo['ID_comprovante'] = df_pend_comp_novo['ID_comprovante'].map(mapa_comprovante).fillna(df_pend_comp_novo['ID_comprovante']).astype(int)

    # Numerário novo: renumera
    df_numerario_novo = None
    if df_numerario is not None and len(df_numerario) > 0:
        df_numerario_novo = df_numerario.copy()
        if 'ID_comprovante' in df_numerario_novo.columns:
            df_numerario_novo['ID_comprovante'] = df_numerario_novo['ID_comprovante'].map(mapa_numerario).fillna(df_numerario_novo['ID_comprovante']).astype(int)

    # Concatena existente + novo (alinhando colunas)
    def concat_alinhado(df_antigo, df_novo):
        if df_antigo is None or len(df_antigo) == 0:
            return df_novo.copy() if df_novo is not None else df_novo
        if df_novo is None or len(df_novo) == 0:
            return df_antigo.copy()
        cols = list(df_antigo.columns) + [c for c in df_novo.columns if c not in df_antigo.columns]
        a = df_antigo.copy()
        b = df_novo.copy()
        for c in cols:
            if c not in a.columns:
                a[c] = pd.NA
            if c not in b.columns:
                b[c] = pd.NA
        return pd.concat([a[cols], b[cols]], ignore_index=True)

    df_extrato_total = concat_alinhado(df_ext, df_extrato_novo)
    df_comprovantes_total = concat_alinhado(df_comp, df_comprovantes_novo)
    df_conciliacao_total = concat_alinhado(df_conc, df_conc_novo)
    df_pend_ext_total = concat_alinhado(df_pend_ext, df_pend_ext_novo)
    df_pend_comp_total = concat_alinhado(df_pend_comp, df_pend_comp_novo)
    if df_num is not None and (df_numerario_novo is not None and len(df_numerario_novo) > 0):
        df_numerario_total = concat_alinhado(df_num, df_numerario_novo)
    elif df_num is not None and len(df_num) > 0:
        df_numerario_total = df_num
    elif df_numerario_novo is not None and len(df_numerario_novo) > 0:
        df_numerario_total = df_numerario_novo
    else:
        df_numerario_total = df_numerario  # pode ser None

    return (df_extrato_total, df_comprovantes_total, df_conciliacao_total,
            df_pend_ext_total, df_pend_comp_total, df_numerario_total)


# ============================================================================
# FUNÇÃO PRINCIPAL
# ============================================================================

def main():
    """
    Função principal que executa toda a conciliação.
    """
    print("="*80)
    print("CONCILIAÇÃO ITAÚ COM COMPROVANTES SIGRA")
    print("="*80)
    print(f"Data/Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80)
    
    try:
        # 1. Busca arquivos
        print("\n[1/7] Buscando arquivos...")
        caminho_extrato, caminho_comprovantes, caminho_numerario = buscar_arquivos_itau_sigra()
        
        # 2. Lê arquivos
        print("\n[2/7] Lendo arquivos...")
        df_extrato_raw = ler_extrato(caminho_extrato)
        df_comprovantes_raw = ler_comprovantes(caminho_comprovantes)
        
        # 3. Prepara dados
        print("\n[3/7] Preparando dados...")
        df_extrato = preparar_dados(df_extrato_raw, tipo='extrato')
        df_comprovantes = preparar_dados(df_comprovantes_raw, tipo='comprovantes')
        
        # 4. Concilia com SIGRA
        print("\n[4/7] Executando conciliação com SIGRA...")
        (
            df_conciliacao,
            df_extratos_pendentes,
            df_comprovantes_pendentes,
            periodo_pgto_min,
            periodo_pgto_max,
            datas_distintas_pgto,
        ) = conciliar_extrato_comprovantes(df_extrato, df_comprovantes)
        
        # 5. Concilia com Numerário (se arquivo existir)
        if caminho_numerario and os.path.exists(caminho_numerario):
            print("\n[5/7] Executando conciliação com Numerário...")
            df_numerario_raw = ler_numerario(caminho_numerario)
            df_numerario = preparar_dados(df_numerario_raw, tipo='numerario')
            
            # Concilia extratos pendentes com numerário
            novas_conciliacoes = conciliar_extrato_numerario(
                df_extratos_pendentes, 
                df_numerario, 
                df_conciliacao.to_dict('records') if len(df_conciliacao) > 0 else []
            )
            
            # Adiciona novas conciliações às existentes
            if novas_conciliacoes:
                df_novas_conciliacoes = pd.DataFrame(novas_conciliacoes)
                df_conciliacao = pd.concat([df_conciliacao, df_novas_conciliacoes], ignore_index=True)
                
                # Atualiza lista de extratos pendentes (remove os que foram conciliados com numerário)
                ids_conciliados_numerario = set(c['ID_extrato'] for c in novas_conciliacoes)
                df_extratos_pendentes = df_extratos_pendentes[
                    ~df_extratos_pendentes['ID_extrato'].isin(ids_conciliados_numerario)
                ].copy()
        else:
            print("\n[5/7] Pulando conciliação com Numerário (arquivo não encontrado)")
            df_numerario = None

        # Métricas da execução atual (antes de mesclar com histórico)
        total_extratos_execucao = len(df_extrato)
        total_comprovantes_execucao = len(df_comprovantes)
        total_conciliacoes_execucao = len(df_conciliacao)
        total_extratos_pendentes_execucao = len(df_extratos_pendentes)
        total_comprovantes_pendentes_execucao = len(df_comprovantes_pendentes)
        extratos_elegiveis_sigra_execucao = 0
        if CONCILIAR_APENAS_EXTRATO_DENTRO_PERIODO_PGTO:

            def _extrato_na_base_pgto_row(r):
                if pd.isna(r.get('data')):
                    return False
                dn = pd.Timestamp(r['data']).normalize()
                if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO and datas_distintas_pgto:
                    return dn in datas_distintas_pgto
                if periodo_pgto_min is not None and periodo_pgto_max is not None:
                    return periodo_pgto_min <= dn <= periodo_pgto_max
                return False

            extratos_elegiveis_sigra_execucao = int(df_extrato.apply(_extrato_na_base_pgto_row, axis=1).sum())
        else:
            extratos_elegiveis_sigra_execucao = len(df_extrato)
        conciliados_unicos_periodo_pgto = 0
        if len(df_conciliacao) > 0 and extratos_elegiveis_sigra_execucao > 0:
            id_para_data = df_extrato.set_index('ID_extrato')['data']
            ok = set()
            for eid in df_conciliacao['ID_extrato'].dropna().unique():
                d = id_para_data.get(eid)
                if d is None or pd.isna(d):
                    continue
                dn = pd.Timestamp(d).normalize()
                if CONCILIAR_APENAS_EXTRATO_DENTRO_PERIODO_PGTO:
                    if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO and datas_distintas_pgto:
                        if dn not in datas_distintas_pgto:
                            continue
                    elif periodo_pgto_min is None or periodo_pgto_max is None:
                        continue
                    elif not (periodo_pgto_min <= dn <= periodo_pgto_max):
                        continue
                ok.add(eid)
            conciliados_unicos_periodo_pgto = len(ok)

        # 6. Monta aba de status (será recalculada após mesclar se houver dados existentes)
        print("\n[6/7] Montando aba de status do extrato...")
        datas_para_status = (
            datas_distintas_pgto if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO and datas_distintas_pgto else None
        )
        df_status_extrato = criar_aba_status_extrato(
            df_extrato,
            df_conciliacao,
            periodo_pgto_min,
            periodo_pgto_max,
            datas_distintas_pgto=datas_para_status,
        )
        
        # 7. Gera Excel (completa planilha existente em vez de substituir)
        print("\n[7/7] Gerando arquivo Excel...")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        caminho_saida = CAMINHO_SAIDA
        if not os.path.isabs(caminho_saida):
            caminho_saida = os.path.join(script_dir, caminho_saida)
        os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
        
        dados_existentes = carregar_dados_existentes_itau(caminho_saida)
        if dados_existentes is not None:
            print("[INFO] Planilha existente encontrada. Completando com dados desta execução (não substituindo).")
            (df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes,
             df_comprovantes_pendentes, df_numerario) = mesclar_dados_itau(
                dados_existentes, df_extrato, df_comprovantes, df_conciliacao,
                df_extratos_pendentes, df_comprovantes_pendentes, df_numerario
            )
            periodo_status_min, periodo_status_max = _periodo_calendario_comprovantes(df_comprovantes)
            datas_status = _datas_distintas_comprovantes(df_comprovantes)
            df_status_extrato = criar_aba_status_extrato(
                df_extrato,
                df_conciliacao,
                periodo_status_min,
                periodo_status_max,
                datas_distintas_pgto=datas_status if CONCILIAR_EXTRATO_SO_DATAS_COM_PGTO else None,
            )
        
        def _gravar_excel(destino_saida):
            with pd.ExcelWriter(destino_saida, engine='openpyxl') as writer:
                df_extrato.to_excel(writer, sheet_name='extrato', index=False)
                df_comprovantes.to_excel(writer, sheet_name='comprovantes', index=False)
                if df_numerario is not None and len(df_numerario) > 0:
                    df_numerario.to_excel(writer, sheet_name='numerario', index=False)
                abas_status = dividir_status_extrato_por_mes(df_status_extrato)
                for nome_aba_status, df_status_mes in abas_status.items():
                    df_status_mes.to_excel(writer, sheet_name=nome_aba_status, index=False)
                if len(df_conciliacao) > 0:
                    df_conciliacao.to_excel(writer, sheet_name='conciliacao', index=False)
                # Formatação profissional das abas de Status Extrato
                formatar_aba_status_extrato(writer.book)

        caminho_saida_final = caminho_saida
        try:
            _gravar_excel(caminho_saida_final)
        except PermissionError:
            base, ext = os.path.splitext(caminho_saida)
            caminho_saida_final = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            print(
                f"[AVISO] Arquivo de saída estava aberto/sem permissão: {caminho_saida}. "
                f"Salvando em novo arquivo: {caminho_saida_final}"
            )
            _gravar_excel(caminho_saida_final)
        
        print(f"\n[OK] Arquivo Excel gerado: {caminho_saida_final}")
        
        # Resumo final
        print("\n" + "="*80)
        print("RESUMO FINAL")
        print("="*80)
        print("Execução atual (sem histórico):")
        print(f"Total de extratos: {total_extratos_execucao}")
        if extratos_elegiveis_sigra_execucao > 0:
            print(
                f"Extratos na base de datas do PGTO (elegíveis SIGRA nesta rodagem): "
                f"{extratos_elegiveis_sigra_execucao}"
            )
        print(f"Total de comprovantes: {total_comprovantes_execucao}")
        print(f"Conciliações encontradas: {total_conciliacoes_execucao}")
        print(f"Extratos pendentes: {total_extratos_pendentes_execucao}")
        print(f"Comprovantes pendentes: {total_comprovantes_pendentes_execucao}")
        if total_extratos_execucao > 0:
            taxa_execucao = (total_conciliacoes_execucao / total_extratos_execucao) * 100
            print(f"Taxa de conciliação (execução atual): {taxa_execucao:.1f}%")
        if extratos_elegiveis_sigra_execucao > 0:
            taxa_sobre_pgto = (conciliados_unicos_periodo_pgto / extratos_elegiveis_sigra_execucao) * 100
            print(
                f"Taxa sobre extratos do período PGTO (conciliados / elegíveis SIGRA): "
                f"{taxa_sobre_pgto:.1f}% ({conciliados_unicos_periodo_pgto}/{extratos_elegiveis_sigra_execucao})"
            )

        print("-"*80)
        print("Acumulado na planilha (após mescla):")
        print(f"Total de extratos: {len(df_extrato)}")
        print(f"Total de comprovantes: {len(df_comprovantes)}")
        print(f"Conciliações encontradas: {len(df_conciliacao)}")
        print(f"Extratos pendentes: {len(df_extratos_pendentes)}")
        print(f"Comprovantes pendentes: {len(df_comprovantes_pendentes)}")
        if len(df_extrato) > 0:
            taxa_acumulada = (len(df_conciliacao) / len(df_extrato)) * 100
            print(f"Taxa de conciliação (acumulada): {taxa_acumulada:.1f}%")
        print("="*80)
        
    except Exception as e:
        print(f"\n[ERRO] Erro durante execução: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == '__main__':
    main()
