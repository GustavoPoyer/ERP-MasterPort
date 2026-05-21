"""
Script de Conciliação de Extratos Bancários - Banco do Brasil

Este script concilia lançamentos de extratos bancários do Banco do Brasil com comprovantes individuais,
identificando quais comprovantes compõem cada lançamento do extrato.

Autor: Automação Financeira
Data: 2026-01-19
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from itertools import combinations
import os
import glob
import re
import sys
import warnings
warnings.filterwarnings('ignore')

if sys.platform == 'win32':
    # Evita falhas de UnicodeEncodeError ao imprimir símbolos no terminal do Windows.
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# ============================================================================
# CONFIGURAÇÕES
# ============================================================================

# Tolerância de data (dias) - permite diferença de ±N dias
TOLERANCIA_DATA = 2
# Tolerância de valor (centavos) - extrato pode vir arredondado; aceita até N centavos de diferença
TOLERANCIA_CENTAVOS = 10

# Caminhos dos arquivos (serão buscados automaticamente na pasta do dia)
CAMINHO_EXTRATO = None  # Será definido automaticamente
CAMINHO_COMPROVANTES = None  # Será definido automaticamente

# Nomes das colunas esperadas para BB
# EXTRATO BB tem: Data, Historico, Documento, valor, Inf., Detalhamento Hist., SIGRA
# PGTOS tem: Ref. Cliente, Cliente, CNPJ, Valor, Pago Despachante, Criação, Vencimento, RPS, Categoria, Reembolso, Fornecedor
COLUNAS_EXTRATO = {
    'data': 'Data',  # Coluna A: Data (DD/MM/YYYY)
    'valor': 'valor',  # Coluna D: valor (minúscula)
    'favorecido': 'Historico',  # Coluna B: Historico (descrição/favorecido)
    'descricao': 'Detalhamento Hist.',  # Coluna F: Detalhamento Hist. (opcional)
    'documento': 'Documento'  # Coluna C: Documento (código que corresponde ao RPS dos comprovantes)
}

COLUNAS_COMPROVANTES = {
    'data': 'Criação',  # Coluna J: Criação (DD/MM/YY HH:MM)
    'valor': 'Valor',  # Coluna H: Valor
    'favorecido': 'Fornecedor',  # Coluna O: Fornecedor (ou pode ser Cliente ou Categoria)
    'descricao': 'Categoria',  # Coluna M: Categoria (tem "MARINHA MERCANTE" aqui!)
    'rps': 'RPS'  # Coluna L: RPS (código que corresponde ao Documento do extrato)
}

# Caminho de saída
CAMINHO_SAIDA = r'G:\Drives compartilhados\automação\Conciliações\conciliacao_bb_final.xlsx'


# ============================================================================
# FUNÇÕES DE LEITURA E NORMALIZAÇÃO
# ============================================================================

def ler_extrato(caminho):
    """
    Lê o arquivo Excel do extrato bancário.
    Para BB: detecta automaticamente a aba do mês atual (ex.: FEVEREIRO26).
    
    Args:
        caminho: Caminho do arquivo Excel
        
    Returns:
        DataFrame com os dados do extrato
    """
    try:
        # Abre o arquivo para detectar abas
        xl = pd.ExcelFile(caminho, engine='openpyxl')
        
        # Detecta aba do mês atual (ex.: fevereiro → "FEVEREIRO26" ou "FEV26")
        mes_atual = datetime.now().month
        meses = {1: ['JANEIRO', 'JAN'], 2: ['FEVEREIRO', 'FEV'], 3: ['MARÇO', 'MARCO', 'MAR'], 
                 4: ['ABRIL', 'ABR'], 5: ['MAIO', 'MAI'], 6: ['JUNHO', 'JUN'],
                 7: ['JULHO', 'JUL'], 8: ['AGOSTO', 'AGO'], 9: ['SETEMBRO', 'SET'],
                 10: ['OUTUBRO', 'OUT'], 11: ['NOVEMBRO', 'NOV'], 12: ['DEZEMBRO', 'DEZ']}
        
        sheet_name = None
        nomes_mes_atual = meses.get(mes_atual, [])
        for nome_mes in nomes_mes_atual:
            for sheet in xl.sheet_names:
                if nome_mes.upper() in sheet.upper():
                    sheet_name = sheet
                    break
            if sheet_name:
                break
        
        if sheet_name:
            print(f"[INFO] Usando aba do mês atual: '{sheet_name}'")
        else:
            # Fallback: usa a última aba (normalmente o mês mais recente)
            sheet_name = xl.sheet_names[-1] if xl.sheet_names else 0
            print(f"[INFO] Aba do mês atual não encontrada, usando aba mais recente: '{sheet_name}'")
        
        df = pd.read_excel(caminho, sheet_name=sheet_name, engine='openpyxl')
        
        # Se tem colunas "Unnamed", tenta encontrar o cabeçalho correto
        if any('Unnamed' in str(col) for col in df.columns):
            for skip_rows in range(0, 10):
                try:
                    df_teste = pd.read_excel(caminho, sheet_name=sheet_name, skiprows=skip_rows, engine='openpyxl')
                    if not any('Unnamed' in str(col) for col in df_teste.columns[:3]):
                        df = df_teste
                        print(f"[INFO] Cabeçalho encontrado na linha {skip_rows + 1}")
                        break
                except:
                    continue
        
        # Remove linhas completamente vazias
        df = df.dropna(how='all')
        
        # Remove "Saldo Anterior" e outras linhas que não são transações
        if 'Historico' in df.columns:
            df = df[~df['Historico'].str.contains('Saldo Anterior', case=False, na=False)]
        elif 'historico' in df.columns:
            df = df[~df['historico'].str.contains('Saldo Anterior', case=False, na=False)]
        
        print(f"[OK] Extrato carregado: {len(df)} lançamentos (aba: {sheet_name})")
        return df
    except Exception as e:
        print(f"[ERRO] Erro ao ler extrato: {e}")
        raise


def ler_comprovantes(caminho):
    """
    Lê o arquivo Excel dos comprovantes de pagamento.
    
    Args:
        caminho: Caminho do arquivo Excel
        
    Returns:
        DataFrame com os dados dos comprovantes
    """
    try:
        # Tenta ler normalmente primeiro
        df = pd.read_excel(caminho)
        
        # Se tem colunas "Unnamed", tenta encontrar o cabeçalho correto
        if any('Unnamed' in str(col) for col in df.columns):
            # Tenta ler pulando algumas linhas até encontrar cabeçalho válido
            for skip_rows in range(0, 10):
                try:
                    df_teste = pd.read_excel(caminho, skiprows=skip_rows)
                    # Verifica se encontrou colunas válidas
                    if not any('Unnamed' in str(col) for col in df_teste.columns[:3]):
                        df = df_teste
                        print(f"[INFO] Cabeçalho encontrado na linha {skip_rows + 1}")
                        break
                except:
                    continue
        
        # Remove linhas completamente vazias
        df = df.dropna(how='all')
        
        print(f"[OK] Comprovantes carregados: {len(df)} registros")
        return df
    except Exception as e:
        print(f"[ERRO] Erro ao ler comprovantes: {e}")
        raise


def normalizar_data(serie_data):
    """
    Normaliza uma série de datas para datetime.
    
    Args:
        serie_data: Série pandas com datas
        
    Returns:
        Série com datas normalizadas
    """
    # Remove caracteres especiais (\xa0, espaços, etc)
    if serie_data.dtype == 'object':
        serie_limpa = serie_data.astype(str).str.replace('\xa0', '', regex=False).str.strip()
    else:
        serie_limpa = serie_data
    
    # Tenta converter para datetime
    if serie_limpa.dtype == 'object':
        # Tenta diferentes formatos (incluindo formato com hora para BB: DD/MM/YY HH:MM)
        formatos = [
            '%Y-%m-%d %H:%M:%S',  # datetime já serializado como texto
            '%d/%m/%Y %H:%M',  # Formato com hora (BB comprovantes: "19/01/26 21:16")
            '%d/%m/%y %H:%M',  # Formato com hora (ano 2 dígitos)
            '%d/%m/%Y',        # Formato padrão brasileiro
            '%Y-%m-%d',        # Formato ISO
            '%d-%m-%Y',        # Formato com traço
            '%Y/%m/%d'         # Formato alternativo
        ]
        
        for formato in formatos:
            try:
                resultado = pd.to_datetime(serie_limpa, format=formato, errors='coerce')
                # Verifica se conseguiu converter pelo menos alguns valores
                if resultado.notna().sum() > 0:
                    return resultado
            except:
                continue
        
        # Se nenhum formato funcionar, tenta conversão genérica assumindo dia primeiro (pt-BR)
        return pd.to_datetime(serie_limpa, errors='coerce', dayfirst=True)
    else:
        return pd.to_datetime(serie_limpa, errors='coerce', dayfirst=True)


def normalizar_texto(texto):
    """
    Normaliza texto para comparação (remove acentos, espaços extras, etc).
    
    Args:
        texto: String a ser normalizada
        
    Returns:
        String normalizada
    """
    if pd.isna(texto):
        return ''
    
    texto = str(texto).upper().strip()
    
    # Remove espaços múltiplos
    texto = ' '.join(texto.split())
    
    # Remove caracteres especiais comuns
    texto = texto.replace('.', '').replace(',', '').replace('-', '').replace('/', '')
    
    return texto


def normalizar_codigo_texto(valor):
    """
    Normaliza códigos numéricos/textuais (Documento/RPS) sem quebrar por notação científica.
    """
    if pd.isna(valor):
        return ''

    # Tipos numéricos diretos
    if isinstance(valor, (int, np.integer)):
        return str(int(valor))
    if isinstance(valor, (float, np.floating)):
        if np.isnan(valor):
            return ''
        if float(valor).is_integer():
            return str(int(valor))
        return f"{valor:.15f}".rstrip('0').rstrip('.')

    texto = str(valor).replace('\xa0', '').strip()
    if not texto or texto.lower() in ('nan', 'none'):
        return ''

    # Ex.: "2,1026E+12" / "2.1026E+12"
    padrao_cientifico = r'^[+-]?\d+(?:[.,]\d+)?[eE][+-]?\d+$'
    if re.match(padrao_cientifico, texto):
        try:
            num = float(texto.replace(',', '.'))
            if float(num).is_integer():
                return str(int(num))
            return f"{num:.15f}".rstrip('0').rstrip('.')
        except Exception:
            pass

    if texto.endswith('.0'):
        texto = texto[:-2]
    return texto


def normalizar_valor(valor):
    """
    Normaliza valores monetários.
    
    Args:
        valor: Valor a ser normalizado
        
    Returns:
        Float com o valor normalizado
    """
    if pd.isna(valor):
        return 0.0
    
    # Se for string, remove formatação
    if isinstance(valor, str):
        # Remove caracteres especiais (\xa0, R$, espaços, etc)
        valor = valor.replace('\xa0', '').replace('R$', '').replace('$', '').strip()
        valor = valor.replace(' ', '')
        
        # Formato brasileiro comum: 50.000,00 ou 50,000,00 (ambos significam 50000.00)
        # Detecta padrão: se tem vírgula seguida de 2 dígitos no final, vírgula é decimal
        # Se tem ponto seguido de 2 dígitos no final, ponto pode ser decimal
        
        # Remove pontos de milhar primeiro (formato 50.000,00)
        if valor.count('.') > 0 and valor.count(',') > 0:
            # Tem ambos: verifica qual é o separador decimal
            ultimo_ponto = valor.rfind('.')
            ultima_virgula = valor.rfind(',')
            
            if ultimo_ponto > ultima_virgula:
                # Ponto vem depois: formato 50,000.00 (internacional) ou 50000.00
                if len(valor.split('.')[-1]) <= 2:
                    # Ponto é decimal
                    valor = valor.replace(',', '')
                else:
                    # Vírgula é decimal: formato 50.000,00
                    valor = valor.replace('.', '').replace(',', '.')
            else:
                # Vírgula vem depois: formato brasileiro 50.000,00
                valor = valor.replace('.', '').replace(',', '.')
        elif valor.count(',') > 0:
            # Só tem vírgula
            partes = valor.split(',')
            # Se tem múltiplas vírgulas (ex: 50,000,00), a última é decimal
            if len(partes) > 2:
                # Formato estranho: 50,000,00 - trata como se vírgulas fossem separadores de milhar
                # e a última parte (2 dígitos) é decimal
                if len(partes[-1]) == 2:
                    # Última parte tem 2 dígitos: é decimal
                    valor = ''.join(partes[:-1]) + '.' + partes[-1]
                else:
                    # Remove todas as vírgulas
                    valor = valor.replace(',', '')
            elif len(partes) == 2 and len(partes[1]) == 2:
                # Formato: 50000,00 (vírgula é decimal)
                valor = valor.replace(',', '.')
            else:
                # Formato: 50,000 (vírgula é separador de milhar - raro mas possível)
                valor = valor.replace(',', '')
        elif valor.count('.') > 0:
            # Só tem ponto
            partes = valor.split('.')
            if len(partes) == 2 and len(partes[1]) <= 2:
                # Formato: 50000.00 (ponto é decimal) - já está OK
                pass
            else:
                # Formato: 50.000 (ponto é separador de milhar)
                valor = valor.replace('.', '')
    
    try:
        resultado = float(valor)
        return abs(resultado)  # Sempre positivo
    except Exception as e:
        return 0.0


def preparar_dados(df, tipo='extrato'):
    """
    Prepara e normaliza os dados do DataFrame.
    
    Args:
        df: DataFrame a ser preparado
        tipo: 'extrato' ou 'comprovantes'
        
    Returns:
        DataFrame preparado com colunas normalizadas
    """
    df_prep = df.copy()
    
    # Seleciona colunas baseado no tipo
    if tipo == 'extrato':
        colunas = COLUNAS_EXTRATO
        prefixo = 'extrato'
    else:
        colunas = COLUNAS_COMPROVANTES
        prefixo = 'comprovante'
    
    # Mostra colunas disponíveis
    print(f"\nColunas disponíveis em {tipo}: {list(df_prep.columns)}")
    
    # Normaliza colunas (tenta encontrar por nome similar se não existir exato)
    col_data = None
    col_valor = None
    col_favorecido = None
    col_documento = None  # Para extrato BB
    col_rps = None  # Para comprovantes BB
    
    # Primeiro, tenta encontrar por nome exato nas configurações
    col_config_data = colunas.get('data')
    col_config_valor = colunas.get('valor')
    col_config_favorecido = colunas.get('favorecido')
    col_config_documento = colunas.get('documento')
    col_config_rps = colunas.get('rps')
    
    # Match exato ou case-insensitive (Excel pode ter "Valor" ou "valor")
    def _achar_coluna(nome, cols):
        if not nome:
            return None
        if nome in cols:
            return nome
        nome_lower = str(nome).lower()
        for c in cols:
            if str(c).lower() == nome_lower:
                return c
        return None
    col_data = _achar_coluna(col_config_data, df_prep.columns)
    col_valor = _achar_coluna(col_config_valor, df_prep.columns)
    col_favorecido = _achar_coluna(col_config_favorecido, df_prep.columns)
    col_documento = _achar_coluna(col_config_documento, df_prep.columns)
    col_rps = _achar_coluna(col_config_rps, df_prep.columns)
    
    # Se não encontrou por nome exato/config, busca por palavras-chave
    for col in df_prep.columns:
        col_lower = str(col).lower()
        # Para BB: busca "criação" também (coluna de data dos comprovantes)
        if ('data' in col_lower or 'date' in col_lower or 'criação' in col_lower or 'criacao' in col_lower) and col_data is None:
            col_data = col
        if ('valor' in col_lower or 'value' in col_lower or 'amount' in col_lower or '(r$)' in col_lower) and col_valor is None:
            col_valor = col
        # Para BB: busca "historico" (extrato) e "fornecedor" ou "categoria" (comprovantes)
        if ('favorecido' in col_lower or 'benefici' in col_lower or 'descri' in col_lower or 'nome' in col_lower or 
            'historico' in col_lower or 'complementar' in col_lower or 'fornecedor' in col_lower or 'categoria' in col_lower) and col_favorecido is None:
            col_favorecido = col
        # Para BB: busca "documento" (extrato) e "rps" (comprovantes)
        if 'documento' in col_lower and col_documento is None:
            col_documento = col
        if 'rps' in col_lower and col_rps is None:
            col_rps = col
    
    # Se ainda não encontrou, usa posição (mas ignora "Unnamed")
    if not col_data:
        for col in df_prep.columns:
            if 'Unnamed' not in str(col):
                col_data = col
                break
        if not col_data:
            col_data = df_prep.columns[0] if len(df_prep.columns) > 0 else None
    
    if not col_valor:
        for col in df_prep.columns:
            if 'Unnamed' not in str(col) and col != col_data:
                col_valor = col
                break
        if not col_valor:
            col_valor = df_prep.columns[1] if len(df_prep.columns) > 1 else df_prep.columns[0]
    
    if not col_favorecido:
        for col in df_prep.columns:
            if 'Unnamed' not in str(col) and col != col_data and col != col_valor:
                col_favorecido = col
                break
        if not col_favorecido:
            col_favorecido = df_prep.columns[2] if len(df_prep.columns) > 2 else col_valor
    
    # Verifica se as colunas existem
    if col_data not in df_prep.columns:
        raise ValueError(f"Coluna de data '{col_data}' não encontrada. Colunas disponíveis: {list(df_prep.columns)}")
    if col_valor not in df_prep.columns:
        raise ValueError(f"Coluna de valor '{col_valor}' não encontrada. Colunas disponíveis: {list(df_prep.columns)}")
    if col_favorecido not in df_prep.columns:
        raise ValueError(f"Coluna de favorecido '{col_favorecido}' não encontrada. Colunas disponíveis: {list(df_prep.columns)}")
    
    print(f"Usando colunas: Data='{col_data}', Valor='{col_valor}', Favorecido='{col_favorecido}'")
    
    # Cria DataFrame normalizado
    df_normalizado = pd.DataFrame()
    df_normalizado['data_original'] = df_prep[col_data]
    df_normalizado['data'] = normalizar_data(df_prep[col_data])
    df_normalizado['valor'] = df_prep[col_valor].apply(normalizar_valor)
    
    # Para BB: se for comprovantes, pode usar "Categoria" além de "Fornecedor" para favorecido
    if tipo == 'comprovantes':
        # Tenta usar "Fornecedor" primeiro, mas também adiciona "Categoria" como info adicional
        df_normalizado['favorecido_original'] = df_prep[col_favorecido].fillna('')
        df_normalizado['favorecido'] = df_prep[col_favorecido].apply(normalizar_texto)
        
        # Adiciona categoria como informação adicional (para matching com "MARINHA MERCANTE")
        col_cat = _achar_coluna('Categoria', df_prep.columns) or _achar_coluna('categoria', df_prep.columns)
        if col_cat:
            df_normalizado['categoria_original'] = df_prep[col_cat].fillna('')
            df_normalizado['categoria'] = df_prep[col_cat].apply(normalizar_texto)
    else:
        df_normalizado['favorecido_original'] = df_prep[col_favorecido].fillna('')
        df_normalizado['favorecido'] = df_prep[col_favorecido].apply(normalizar_texto)
        
        # Adiciona coluna Documento se existir (para matching com RPS dos comprovantes)
        if col_documento and col_documento in df_prep.columns:
            df_normalizado['documento_original'] = df_prep[col_documento].fillna('')
            df_normalizado['documento'] = df_prep[col_documento].apply(normalizar_codigo_texto)
    
    # Para comprovantes: adiciona RPS se existir
    if tipo == 'comprovantes':
        col_rps_final = col_rps or _achar_coluna('RPS', df_prep.columns) or _achar_coluna('rps', df_prep.columns)
        if col_rps_final:
            df_normalizado['rps_original'] = df_prep[col_rps_final].fillna('')
            df_normalizado['rps'] = df_prep[col_rps_final].apply(normalizar_codigo_texto)
    
    # Adiciona todas as colunas originais
    for col in df_prep.columns:
        if col not in df_normalizado.columns:
            df_normalizado[col] = df_prep[col]
    
    # Debug: mostra estatísticas antes de filtrar
    total_antes = len(df_normalizado)
    datas_validas = df_normalizado['data'].notna().sum()
    valores_validos = (df_normalizado['valor'] != 0).sum()
    print(f"  Total de linhas: {total_antes}")
    print(f"  Datas válidas: {datas_validas}")
    print(f"  Valores válidos (diferentes de zero): {valores_validos}")
    
    # Mostra alguns exemplos de valores se todos forem zero
    if valores_validos == 0:
        print(f"  [DEBUG] Primeiros 5 valores originais da coluna '{col_valor}':")
        for i, val in enumerate(df_prep[col_valor].head(5)):
            print(f"    [{i+1}] {repr(val)} (tipo: {type(val).__name__})")
    
    # Remove linhas com valores inválidos
    df_normalizado = df_normalizado[df_normalizado['data'].notna()]
    df_normalizado = df_normalizado[df_normalizado['valor'] != 0]
    
    # Debug: mostra quantos foram removidos
    removidos = total_antes - len(df_normalizado)
    if removidos > 0:
        print(f"  [INFO] {removidos} linha(s) removida(s) por dados inválidos")
    
    # Adiciona ID único
    df_normalizado[f'ID_{prefixo}'] = range(1, len(df_normalizado) + 1)
    
    print(f"[OK] Dados {tipo} preparados: {len(df_normalizado)} registros válidos")
    
    return df_normalizado


# ============================================================================
# FUNÇÕES DE CONCILIAÇÃO
# ============================================================================

def datas_proximas(data1, data2, tolerancia=TOLERANCIA_DATA):
    """
    Verifica se duas datas estão dentro da tolerância especificada.
    
    Args:
        data1: Primeira data
        data2: Segunda data
        tolerancia: Tolerância em dias
        
    Returns:
        True se as datas estão próximas, False caso contrário
    """
    if pd.isna(data1) or pd.isna(data2):
        return False
    
    diferenca = abs((data1 - data2).days)
    return diferenca <= tolerancia


def textos_similares(texto1, texto2, min_similaridade=0.6):
    """
    Verifica se dois textos são similares (usando comparação simples).
    
    Args:
        texto1: Primeiro texto
        texto2: Segundo texto
        min_similaridade: Similaridade mínima (0-1)
        
    Returns:
        True se os textos são similares, False caso contrário
    """
    texto1 = normalizar_texto(texto1)
    texto2 = normalizar_texto(texto2)
    
    if not texto1 or not texto2:
        return True  # Se algum estiver vazio, considera compatível
    
    # Calcula similaridade simples (palavras em comum)
    palavras1 = set(texto1.split())
    palavras2 = set(texto2.split())
    
    if not palavras1 or not palavras2:
        return True
    
    intersecao = palavras1.intersection(palavras2)
    uniao = palavras1.union(palavras2)
    
    similaridade = len(intersecao) / len(uniao) if uniao else 0
    
    return similaridade >= min_similaridade


def normalizar_codigo_documento(codigo):
    """
    Normaliza código do documento do extrato para buscar no RPS dos comprovantes.
    
    Para SISCOMEX: extrai o código a partir do "35"
    Exemplo: "2103524154320" → "3524154320" (que corresponde a "352415432-0" nos comprovantes)
    
    Args:
        codigo: Código do extrato (ex: "2103524154320" ou "00002103524154320")
        
    Returns:
        Código normalizado a partir do "35" (ex: "3524154320")
    """
    if pd.isna(codigo):
        return ''
    
    codigo_str = normalizar_codigo_texto(codigo)
    
    # Remove traços, espaços e caracteres especiais
    codigo_limpo = codigo_str.replace('-', '').replace(' ', '').replace('_', '').replace('.', '')
    
    # Para SISCOMEX: extrai o código a partir do "35"
    # Procura a primeira ocorrência de "35" no código
    pos_35 = codigo_limpo.find('35')
    if pos_35 >= 0:
        # Extrai a partir do "35"
        codigo_normalizado = codigo_limpo[pos_35:]
    else:
        # Se não encontrar "35", remove zeros à esquerda (fallback)
        codigo_normalizado = codigo_limpo.lstrip('0')
    
    return codigo_normalizado


def buscar_por_rps(extrato_row, comprovantes_disponiveis):
    """
    Busca comprovantes pelo RPS quando o extrato tem código de documento.
    Usado principalmente para SISCOMEX que está dividido em várias linhas nos comprovantes.
    
    Args:
        extrato_row: Linha do extrato (Series)
        comprovantes_disponiveis: DataFrame com comprovantes disponíveis
        
    Returns:
        Lista de índices dos comprovantes que correspondem ao RPS, ou None se não encontrar
    """
    # Verifica se tem código de documento
    codigo_documento = extrato_row.get('documento_original', '')
    if pd.isna(codigo_documento) or not str(codigo_documento).strip():
        return None
    
    # Normaliza o código (remove traço)
    codigo_normalizado = normalizar_codigo_documento(codigo_documento)
    
    if not codigo_normalizado:
        return None
    
    print(f"    [INFO] Buscando comprovantes por RPS: código extrato '{codigo_documento}' → normalizado '{codigo_normalizado}'")
    
    # Verifica se os comprovantes têm coluna RPS
    if 'rps' not in comprovantes_disponiveis.columns and 'rps_original' not in comprovantes_disponiveis.columns:
        print(f"    [AVISO] Comprovantes não têm coluna RPS disponível")
        return None
    
    # Busca comprovantes cujo RPS contenha o código normalizado
    # No extrato: "00002103531852835" → normalizado: "3531852835"
    # Nos comprovantes: "353185283-5" → normalizado: "3531852835"
    col_rps = 'rps' if 'rps' in comprovantes_disponiveis.columns else 'rps_original'
    if col_rps not in comprovantes_disponiveis.columns:
        col_rps = 'RPS' if 'RPS' in comprovantes_disponiveis.columns else None
    
    if col_rps is None:
        print(f"    [AVISO] Coluna RPS não encontrada nos comprovantes")
        return None
    
    def rps_contem_codigo(row):
        """
        Verifica se o RPS do comprovante corresponde EXATAMENTE ao código do extrato.
        NÃO aceita buscas flexíveis - apenas match exato do RPS normalizado.
        """
        rps_comp = str(row.get(col_rps, '')).strip()
        if not rps_comp:
            return False
        
        # Remove traço e espaços do RPS dos comprovantes (ex: "352415432-0" → "3524154320")
        rps_limpo = rps_comp.replace('-', '').replace(' ', '').replace('_', '')
        # Remove zeros à esquerda
        rps_normalizado = rps_limpo.lstrip('0')
        
        # COMPARAÇÃO EXATA APENAS - não mistura RPS diferentes
        # O código do extrato já foi normalizado (ex: "3531852835")
        # O RPS do comprovante normalizado deve ser EXATAMENTE igual
        if codigo_normalizado == rps_normalizado:
            return True
        
        # Se não bater exatamente, retorna False
        # NÃO tenta busca flexível para evitar misturar documentos diferentes
        return False
    
    # Debug: mostra algumas informações sobre RPS
    print(f"    [DEBUG] Verificando RPS nos comprovantes...")
    print(f"    [DEBUG] Coluna RPS usada: '{col_rps}'")
    print(f"    [DEBUG] Colunas disponíveis nos comprovantes: {list(comprovantes_disponiveis.columns)}")
    
    # Verifica quantos comprovantes têm RPS preenchido
    comprovantes_com_rps = comprovantes_disponiveis[comprovantes_disponiveis[col_rps].notna() & 
                                                      (comprovantes_disponiveis[col_rps].astype(str).str.strip() != '')]
    print(f"    [DEBUG] Comprovantes com RPS preenchido: {len(comprovantes_com_rps)} de {len(comprovantes_disponiveis)}")
    
    if len(comprovantes_disponiveis) > 0:
        print(f"    [DEBUG] Primeiros 5 RPS encontrados nos comprovantes:")
        for idx, (_, comp_row) in enumerate(comprovantes_disponiveis.head(5).iterrows()):
            rps_comp_raw = str(comp_row.get(col_rps, '')).strip()
            rps_comp_limpo = rps_comp_raw.replace('-', '').replace(' ', '').replace('_', '')
            rps_comp_norm = rps_comp_limpo.lstrip('0')
            print(f"      [{idx+1}] RPS original: '{rps_comp_raw}' → limpo: '{rps_comp_limpo}' → normalizado: '{rps_comp_norm}'")
        
        # Mostra um exemplo de comprovante que TEM RPS preenchido
        if len(comprovantes_com_rps) > 0:
            print(f"    [DEBUG] Exemplo de comprovante COM RPS preenchido:")
            exemplo = comprovantes_com_rps.iloc[0]
            rps_exemplo = str(exemplo.get(col_rps, '')).strip()
            print(f"      RPS: '{rps_exemplo}'")
    
    # Aplica a função de comparação para encontrar comprovantes com RPS correspondente
    comprovantes_rps = comprovantes_disponiveis[
        comprovantes_disponiveis.apply(rps_contem_codigo, axis=1)
    ].copy()
    
    if len(comprovantes_rps) == 0:
        print(f"    [INFO] Nenhum comprovante encontrado com RPS contendo '{codigo_normalizado}'")
        print(f"    [DEBUG] Código extrato normalizado: '{codigo_normalizado}' (tamanho: {len(codigo_normalizado)})")
        return None
    
    print(f"    [INFO] Encontrados {len(comprovantes_rps)} comprovante(s) com RPS correspondente")
    
    # Mostra TODOS os comprovantes encontrados (não apenas os primeiros 5)
    for idx, (_, comp_row) in enumerate(comprovantes_rps.iterrows()):
        valor_comp = abs(comp_row.get('valor', 0))
        rps_comp = str(comp_row.get(col_rps, ''))
        categoria_comp = str(comp_row.get('categoria_original', ''))[:40] if 'categoria_original' in comp_row else 'N/A'
        data_comp = comp_row.get('data_original', comp_row.get('data', ''))
        print(f"      [{idx+1}] RPS: {rps_comp} | Valor: R$ {valor_comp:,.2f} | Categoria: {categoria_comp} | Data: {data_comp}")
    
    # Soma os valores dos comprovantes encontrados
    valor_extrato = abs(extrato_row['valor'])
    soma_comprovantes = sum(abs(float(v)) for v in comprovantes_rps['valor'].tolist())
    
    print(f"    [INFO] Soma dos comprovantes encontrados: R$ {soma_comprovantes:,.2f} | Valor do extrato: R$ {valor_extrato:,.2f}")
    
    # Converte para centavos para comparação exata
    valor_extrato_centavos = int(round(valor_extrato * 100))
    soma_comprovantes_centavos = int(round(soma_comprovantes * 100))
    
    # Verifica se a soma bate (exato ou dentro da tolerância)
    diff_centavos = abs(soma_comprovantes_centavos - valor_extrato_centavos)
    if diff_centavos <= TOLERANCIA_CENTAVOS:
        print(f"    [OK] Match por RPS! Soma encontrada (diferença: {diff_centavos} centavos).")
        indices_retornados = comprovantes_rps.index.tolist()
        print(f"    [DEBUG] Retornando {len(indices_retornados)} índices: {indices_retornados[:5]}...")
        return indices_retornados
    else:
        print(f"    [AVISO] Soma não bate. Diferença: R$ {abs(soma_comprovantes - valor_extrato):,.2f} ({diff_centavos} centavos)")
        print(f"    [DEBUG] Comprovantes encontrados: {len(comprovantes_rps)} | Esperado: 4 comprovantes somando R$ {valor_extrato:,.2f}")
        
        # NÃO tenta busca ampla ou flexível
        # Se a soma não bater exatamente, retorna None
        # RPS deve ser EXATO e TODOS os comprovantes com esse RPS devem somar o valor do extrato
        return None


def encontrar_combinacoes_comprovantes(extrato_row, comprovantes_disponiveis, max_combinacoes=10):
    """
    Encontra combinações de comprovantes que somam o valor do extrato.
    Algoritmo inteligente que usa múltiplas estratégias para maximizar conciliações.
    
    Args:
        extrato_row: Linha do extrato (Series)
        comprovantes_disponiveis: DataFrame com comprovantes disponíveis
        max_combinacoes: Número máximo de itens na combinação
        
    Returns:
        Lista de índices dos comprovantes que formam a combinação
    """
    valor_extrato = abs(extrato_row['valor'])
    data_extrato = extrato_row['data']
    favorecido_extrato = str(extrato_row.get('favorecido_original', '')).upper()
    
    # Debug: mostra informações do extrato sendo processado
    print(f"    [DEBUG] Processando extrato: R$ {valor_extrato:,.2f} | Data: {data_extrato} | Favorecido: {favorecido_extrato[:50]}")
    print(f"    [DEBUG] Comprovantes disponíveis: {len(comprovantes_disponiveis)}")
    
    # Caso especial: SISCOMEX com código de documento
    # Se for SISCOMEX e tiver código de documento, busca APENAS por RPS
    # NÃO mistura com outros comprovantes - ou encontra todos pelo RPS ou retorna None
    if 'SISCOMEX' in favorecido_extrato:
        codigo_documento = extrato_row.get('documento_original', '')
        if pd.notna(codigo_documento) and str(codigo_documento).strip():
            print(f"    [INFO] Extrato contém 'SISCOMEX' com código de documento '{codigo_documento}' - buscando APENAS por RPS")
            resultado_rps = buscar_por_rps(extrato_row, comprovantes_disponiveis)
            if resultado_rps:
                print(f"    [OK] Conciliação por RPS confirmada - usando apenas comprovantes com o mesmo RPS")
                return resultado_rps
            else:
                print(f"    [AVISO] Não encontrou comprovantes por RPS com valor exato - NÃO tenta busca normal (restrição por RPS)")
                return None  # NÃO tenta busca normal - RPS é obrigatório para SISCOMEX
        else:
            print(f"    [INFO] Extrato contém 'SISCOMEX' sem código de documento - não busca comprovantes (taxa/imposto)")
            return None
    
    # Informação especial para BB: 
    # - "AFRMM-Adicional Frete Ren" no extrato corresponde a comprovantes com "MARINHA MERCANTE" na categoria
    # - AFRMM é match DIRETO 1:1 por valor, NÃO combina múltiplos comprovantes
    tem_afrmm = 'AFRMM' in favorecido_extrato or 'MARINHA MERCANTE' in favorecido_extrato
    
    if tem_afrmm:
        print(f"    [INFO] Extrato contém 'AFRMM' ou 'MARINHA MERCANTE' - buscando match DIRETO (1:1) por valor")
        # Filtra comprovantes que tenham "MARINHA MERCANTE" na categoria
        if 'categoria' in comprovantes_disponiveis.columns:
            comprovantes_afrmm = comprovantes_disponiveis[
                comprovantes_disponiveis['categoria'].str.contains('MARINHA MERCANTE', case=False, na=False)
            ].copy()
            print(f"    [INFO] Filtrados {len(comprovantes_afrmm)} comprovantes com categoria 'MARINHA MERCANTE'")
            
            # Busca match DIRETO 1:1 por valor exato (não combina múltiplos)
            valor_extrato_centavos = int(round(valor_extrato * 100))
            
            for idx, comp_row in comprovantes_afrmm.iterrows():
                valor_comp = abs(comp_row.get('valor', 0))
                valor_comp_centavos = int(round(valor_comp * 100))
                
                # Match exato ou dentro da tolerância (arredondamento)
                if abs(valor_comp_centavos - valor_extrato_centavos) <= TOLERANCIA_CENTAVOS:
                    print(f"    [OK] Match direto encontrado: R$ {valor_comp:,.2f} (categoria MARINHA MERCANTE)")
                    return [idx]  # Retorna apenas este comprovante (match 1:1)
            
            # Se não encontrou match, retorna None (não tenta combinar)
            print(f"    [INFO] Nenhum comprovante com valor R$ {valor_extrato:,.2f} (±{TOLERANCIA_CENTAVOS} centavos) encontrado")
            return None
        else:
            print(f"    [INFO] AFRMM mas sem coluna Categoria nos comprovantes - tenta busca normal")
    
    # Verifica se há comprovantes disponíveis
    if len(comprovantes_disponiveis) == 0:
        print(f"    [DEBUG] Nenhum comprovante disponível para busca")
        return None
    
    # Debug: mostra alguns comprovantes disponíveis
    if len(comprovantes_disponiveis) > 0:
        print(f"    [DEBUG] Primeiros 3 comprovantes disponíveis:")
        for idx, (_, comp_row) in enumerate(comprovantes_disponiveis.head(3).iterrows()):
            valor_comp = abs(comp_row.get('valor', 0))
            data_comp = comp_row.get('data', '')
            favorecido_comp = str(comp_row.get('favorecido_original', ''))[:50]
            categoria_comp = str(comp_row.get('categoria_original', ''))[:30] if 'categoria_original' in comp_row else 'N/A'
            print(f"      [{idx+1}] R$ {valor_comp:,.2f} | Data: {data_comp} | Fornecedor: {favorecido_comp} | Categoria: {categoria_comp}")
    
    # Estratégia: Busca progressiva só por data (favorecido não precisa ser parecido)
    # Prioriza mesma data, depois ±1, ±2, ±3 dias. Máximo 10 comprovantes por combinação.
    estrategias_busca = [
        (0, False, "mesma data"),
        (1, False, "±1 dia"),
        (2, False, "±2 dias"),
        (3, False, "±3 dias"),
    ]
    
    melhor_resultado = None
    melhor_tamanho = float('inf')
    
    for tolerancia, filtrar_favorecido, descricao in estrategias_busca:
        # Filtra por data conforme tolerância
        comprovantes_candidatos = comprovantes_disponiveis[
            comprovantes_disponiveis.apply(
                lambda row: abs((row['data'] - data_extrato).days) <= tolerancia,
                axis=1
            )
        ].copy()
        
        if len(comprovantes_candidatos) == 0:
            continue
        
        # Favorecido não precisa ser parecido: não filtra por similaridade de texto.
        
        # Calcula métricas de relevância (data e valor apenas)
        comprovantes_candidatos = comprovantes_candidatos.copy()
        comprovantes_candidatos['dias_diff'] = (comprovantes_candidatos['data'] - data_extrato).abs().dt.days
        comprovantes_candidatos['valor_abs'] = comprovantes_candidatos['valor'].abs()
        
        # Relevância só por data e valor (favorecido não precisa ser parecido)
        peso_data = 1.0 if tolerancia == 0 else 1.0 / (1 + tolerancia * 2)
        comprovantes_candidatos['relevancia'] = (
            peso_data * 0.5 +  # Proximidade de data (50%)
            (1.0 / (1 + abs(comprovantes_candidatos['valor_abs'] - valor_extrato) / max(valor_extrato, 1))) * 0.5  # Proximidade de valor (50%)
        )
        
        # Ordena por relevância (mais relevantes primeiro)
        comprovantes_candidatos = comprovantes_candidatos.sort_values(
            by='relevancia',
            ascending=False
        ).copy()
        
        # Limita candidatos mas aumenta o pool para valores maiores
        # Para valores muito grandes, precisa de mais candidatos
        if valor_extrato > 100000:
            limite_candidatos = 250  # Valores muito grandes precisam de muitas opções
        elif valor_extrato > 50000:
            limite_candidatos = 220
        elif valor_extrato > 10000:
            limite_candidatos = 200
        else:
            limite_candidatos = 180
        
        comprovantes_candidatos = comprovantes_candidatos.head(limite_candidatos)
        
        # Remove colunas auxiliares
        colunas_aux = ['dias_diff', 'valor_abs', 'relevancia']
        for col in colunas_aux:
            if col in comprovantes_candidatos.columns:
                comprovantes_candidatos = comprovantes_candidatos.drop(columns=[col])
        
        # Prepara dados
        indices = comprovantes_candidatos.index.tolist()
        valores = [abs(float(v)) for v in comprovantes_candidatos['valor'].tolist()]
        
        # Converte para centavos
        valor_extrato_centavos = int(round(valor_extrato * 100))
        valores_centavos = [int(round(v * 100)) for v in valores]
        
        # Verifica soma total
        soma_total_disponivel = sum(valores_centavos)
        if soma_total_disponivel < valor_extrato_centavos:
            print(f"    [DEBUG] Estratégia '{descricao}': Soma disponível (R$ {soma_total_disponivel/100:,.2f}) < Valor extrato (R$ {valor_extrato_centavos/100:,.2f})")
            continue
        
        print(f"    [DEBUG] Estratégia '{descricao}': {len(comprovantes_candidatos)} candidatos, soma total: R$ {soma_total_disponivel/100:,.2f}, valor alvo: R$ {valor_extrato_centavos/100:,.2f}")
        
        # Estratégia 2: Tenta algoritmo dinâmico primeiro (mais rápido)
        resultado = subset_sum_dinamico(valores_centavos, valor_extrato_centavos, max_combinacoes, TOLERANCIA_CENTAVOS)
        
        if resultado:
            indices_encontrados = [indices[i] for i in resultado]
            # Se encontrou uma solução melhor (menor número de itens), guarda
            if len(indices_encontrados) < melhor_tamanho:
                melhor_resultado = indices_encontrados
                melhor_tamanho = len(indices_encontrados)
                # Se encontrou solução na mesma data, retorna imediatamente (prioridade máxima)
                if tolerancia == 0:
                    soma_encontrada = sum(valores[i] for i in resultado)
                    print(f"    [OK] Encontrada combinação ótima na mesma data: {len(indices_encontrados)} comprovante(s) somando R$ {soma_encontrada:,.2f}")
                    return melhor_resultado
                # Se encontrou solução perfeita com 1 item, também retorna
                elif len(indices_encontrados) == 1:
                    soma_encontrada = sum(valores[i] for i in resultado)
                    print(f"    [OK] Encontrada combinação perfeita (1 item): {len(indices_encontrados)} comprovante(s) em {descricao} somando R$ {soma_encontrada:,.2f}")
                    return melhor_resultado
        
        # Estratégia 3: Para valores grandes, tenta heurística gulosa primeiro
        # Ordena por valor (maior primeiro) e tenta encontrar combinação grande primeiro
        if valor_extrato > 20000:
            indices_valores = sorted(enumerate(valores_centavos), key=lambda x: x[1], reverse=True)
            # Tenta heurística: pega os maiores valores que cabem
            soma_atual = 0
            indices_heuristica = []
            for idx, valor in indices_valores:
                if soma_atual + valor <= valor_extrato_centavos:  # Sem tolerância
                    soma_atual += valor
                    indices_heuristica.append(indices[idx])
                    if soma_atual == valor_extrato_centavos:  # Valor deve ser EXATO
                        if not melhor_resultado or len(indices_heuristica) < melhor_tamanho:
                            melhor_resultado = indices_heuristica
                            melhor_tamanho = len(indices_heuristica)
                            print(f"    [OK] Heurística gulosa encontrou: {len(indices_heuristica)} comprovante(s) em {descricao}")
                            break
        
        # Estratégia 4: Busca exaustiva limitada
        # SEMPRE tenta busca exaustiva se não encontrou ainda
        # LIMITA a 10 comprovantes máximo para evitar misturar muitos documentos sem relação
        # Mas garante que sempre tenta quando há candidatos suficientes
        if not melhor_resultado and len(comprovantes_candidatos) > 0:
            # Cria lista de (posicao, valor) e ordena por valor
            indices_valores = list(zip(range(len(indices)), valores_centavos))
            indices_valores.sort(key=lambda x: x[1], reverse=True)
            # posicoes_ordenadas: posições na lista indices ordenadas por valor
            posicoes_ordenadas = [i for i, v in indices_valores]
            valores_ordenados = [v for i, v in indices_valores]
            
            # Testa combinações de tamanho crescente, mas com limites RESTRITIVOS
            # MÁXIMO de 10 comprovantes por combinação para evitar misturar documentos sem relação
            for tamanho in range(1, min(11, len(posicoes_ordenadas) + 1)):  # Até 10 itens (reduzido de 25)
                # Ajusta limites dinamicamente baseado no valor
                # LIMITES REDUZIDOS para evitar combinações muito grandes sem relação
                if valor_extrato > 100000:
                    # Valores muito grandes: limites mais restritivos
                    if tamanho <= 3:
                        max_combinacoes_teste = 2000
                        max_candidatos = min(30, len(posicoes_ordenadas))
                    elif tamanho <= 6:
                        max_combinacoes_teste = 1000
                        max_candidatos = min(25, len(posicoes_ordenadas))
                    else:  # tamanho <= 10
                        max_combinacoes_teste = 300
                        max_candidatos = min(20, len(posicoes_ordenadas))
                elif valor_extrato > 50000:
                    if tamanho <= 3:
                        max_combinacoes_teste = 1500
                        max_candidatos = min(25, len(posicoes_ordenadas))
                    elif tamanho <= 6:
                        max_combinacoes_teste = 800
                        max_candidatos = min(20, len(posicoes_ordenadas))
                    else:  # tamanho <= 10
                        max_combinacoes_teste = 200
                        max_candidatos = min(18, len(posicoes_ordenadas))
                else:
                    # Para valores menores, aumenta limites para garantir que encontra
                    # Especialmente importante para casos como AFRMM com poucos candidatos
                    if tamanho <= 4:
                        max_combinacoes_teste = 2000  # Aumentado para garantir busca completa
                        max_candidatos = min(30, len(posicoes_ordenadas))  # Aumentado
                    elif tamanho <= 7:
                        max_combinacoes_teste = 1000  # Aumentado
                        max_candidatos = min(25, len(posicoes_ordenadas))  # Aumentado
                    else:  # tamanho <= 10
                        max_combinacoes_teste = 500  # Aumentado
                        max_candidatos = min(20, len(posicoes_ordenadas))  # Aumentado
                
                combinacoes_testadas = 0
                for combo in combinations(range(max_candidatos), tamanho):
                    combinacoes_testadas += 1
                    if combinacoes_testadas > max_combinacoes_teste:
                        break
                    
                    soma_centavos = sum(valores_ordenados[i] for i in combo)
                    diff_centavos = abs(soma_centavos - valor_extrato_centavos)
                    if diff_centavos <= TOLERANCIA_CENTAVOS:  # Exato ou dentro da tolerância
                        # Mapeia posições de volta para os índices originais do DataFrame
                        # combo contém posições em posicoes_ordenadas (0, 1, 2, ...)
                        # posicoes_ordenadas[i] contém a posição na lista indices
                        # indices[pos] contém o índice original do DataFrame
                        indices_encontrados = [indices[posicoes_ordenadas[i]] for i in combo]
                        
                        # Se a combinação tem muitos comprovantes (>5), exige que pelo menos 40% sejam da mesma data
                        # (favorecido não precisa ser parecido)
                        if len(indices_encontrados) > 5:
                            try:
                                comprovantes_encontrados = comprovantes_candidatos.loc[indices_encontrados]
                            except KeyError:
                                comprovantes_encontrados = comprovantes_disponiveis.loc[indices_encontrados]
                            mesma_data = sum(1 for _, row in comprovantes_encontrados.iterrows() 
                                           if abs((row['data'] - data_extrato).days) == 0)
                            percentual_mesma_data = mesma_data / len(indices_encontrados)
                            if percentual_mesma_data < 0.4:
                                print(f"    [AVISO] Combinação rejeitada: {len(indices_encontrados)} comprovantes, apenas {percentual_mesma_data*100:.0f}% na mesma data")
                                continue
                        
                        if not melhor_resultado or len(indices_encontrados) < melhor_tamanho:
                            melhor_resultado = indices_encontrados
                            melhor_tamanho = len(indices_encontrados)
                            soma_encontrada = soma_centavos / 100
                            
                            # Mostra detalhes da combinação encontrada
                            print(f"    [MATCH] Busca exaustiva encontrou {len(indices_encontrados)} comprovante(s):")
                            for idx_found in indices_encontrados:
                                # Usa comprovantes_candidatos porque indices_encontrados vem dele
                                if idx_found in comprovantes_candidatos.index:
                                    found_row = comprovantes_candidatos.loc[idx_found]
                                elif idx_found in comprovantes_disponiveis.index:
                                    found_row = comprovantes_disponiveis.loc[idx_found]
                                else:
                                    print(f"      [ERRO] Índice {idx_found} não encontrado em nenhum DataFrame")
                                    continue
                                
                                data_found = found_row.get('data_original', found_row.get('data', ''))
                                favorecido_found = str(found_row.get('favorecido_original', ''))[:50]
                                valor_found = abs(found_row['valor'])
                                print(f"      ✓ ID {found_row.get('ID_comprovante', idx_found)}: R$ {valor_found:,.2f} | {data_found} | {favorecido_found}")
                            print(f"    [SOMA] Total: R$ {soma_encontrada:,.2f} (desejado: R$ {valor_extrato:,.2f})")
                            
                            # Se encontrou solução pequena, retorna
                            if len(indices_encontrados) <= 3:
                                return melhor_resultado
    
    # Retorna o melhor resultado encontrado, se houver
    if melhor_resultado:
        # Calcula soma final tentando usar comprovantes_disponiveis primeiro
        soma_final = 0
        for idx in melhor_resultado:
            try:
                if idx in comprovantes_disponiveis.index:
                    soma_final += abs(comprovantes_disponiveis.loc[idx]['valor'])
            except (KeyError, IndexError):
                # Se não encontrar, tenta buscar usando outro método ou pula
                pass
        
        print(f"    [RESULTADO FINAL] Melhor combinação encontrada: {len(melhor_resultado)} comprovante(s) somando R$ {soma_final:,.2f}")
        print(f"    [DETALHE] Comprovantes relacionados:")
        for idx_final in melhor_resultado:
            try:
                if idx_final in comprovantes_disponiveis.index:
                    final_row = comprovantes_disponiveis.loc[idx_final]
                else:
                    print(f"    [AVISO] Índice {idx_final} não encontrado no DataFrame")
                    continue
            except (KeyError, IndexError):
                print(f"    [AVISO] Erro ao acessar índice {idx_final}")
                continue
            data_final = final_row.get('data_original', final_row.get('data', ''))
            favorecido_final = str(final_row.get('favorecido_original', ''))[:50]
            valor_final = abs(final_row['valor'])
            print(f"      → ID {final_row.get('ID_comprovante', idx_final)}: R$ {valor_final:,.2f} | {data_final} | {favorecido_final}")
        return melhor_resultado
    
    print(f"    [INFO] Nenhuma combinação encontrada para R$ {valor_extrato:,.2f} do extrato '{favorecido_extrato[:50]}'")
    
    # Debug adicional: mostra estatísticas dos comprovantes disponíveis
    if len(comprovantes_disponiveis) > 0:
        valores_disponiveis = [abs(float(v)) for v in comprovantes_disponiveis['valor'].tolist()]
        valores_disponiveis_centavos = [int(round(v * 100)) for v in valores_disponiveis]
        soma_total = sum(valores_disponiveis_centavos)
        valor_extrato_centavos = int(round(valor_extrato * 100))
        
        print(f"    [DEBUG] Estatísticas dos comprovantes disponíveis:")
        print(f"      Total de comprovantes: {len(comprovantes_disponiveis)}")
        print(f"      Soma total disponível: R$ {soma_total/100:,.2f}")
        print(f"      Valor necessário: R$ {valor_extrato_centavos/100:,.2f}")
        print(f"      Diferença: R$ {(soma_total - valor_extrato_centavos)/100:,.2f}")
        if len(valores_disponiveis) > 0:
            print(f"      Valor mínimo: R$ {min(valores_disponiveis):,.2f}")
            print(f"      Valor máximo: R$ {max(valores_disponiveis):,.2f}")
            print(f"      Valor médio: R$ {sum(valores_disponiveis)/len(valores_disponiveis):,.2f}")
    
    return None


def subset_sum_dinamico(valores_centavos, valor_alvo_centavos, max_itens, tolerancia_centavos=None):
    """
    Usa programação dinâmica otimizada para encontrar subconjunto que soma exatamente o valor alvo.
    Versão melhorada para lidar com valores grandes e muitas combinações.
    
    Args:
        valores_centavos: Lista de valores em centavos
        valor_alvo_centavos: Valor alvo em centavos
        max_itens: Número máximo de itens na combinação
        
    Returns:
        Lista de índices que formam a combinação, ou None se não encontrar
    """
    n = len(valores_centavos)
    
    # Se o valor alvo é muito grande ou pequeno, retorna None
    if valor_alvo_centavos <= 0 or valor_alvo_centavos > sum(valores_centavos):
        return None
    
    # Aumenta o limite para valores maiores, permitindo conciliar valores até R$ 1.000.000,00
    # Para valores acima de 100 milhões de centavos (R$ 1.000.000,00), usa abordagem diferente
    if valor_alvo_centavos > 100000000:
        return None
    
    # Ordena valores em ordem decrescente para tentar combinações maiores primeiro
    valores_com_indices = sorted(enumerate(valores_centavos), key=lambda x: x[1], reverse=True)
    indices_ordenados = [i for i, v in valores_com_indices]
    valores_ordenados = [v for i, v in valores_com_indices]
    
    # Usa abordagem de programação dinâmica mais eficiente
    # Dicionário: soma_centavos -> lista de tuplas (numero_itens, indices_usados)
    # Mantém apenas as melhores combinações (menor número de itens)
    soma_possivel = {0: [(0, [])]}  # Soma 0 pode ser obtida com 0 itens
    
    # Aumenta o limite de estados para permitir encontrar mais combinações
    # Mas ainda mantém limite para evitar explosão de memória
    # Para poucos candidatos (como AFRMM com 2-7 candidatos), aumenta muito o limite
    if n <= 20:
        max_estados = 500000  # Para poucos candidatos, permite muito mais estados
    elif n <= 50:
        max_estados = 200000  # Para quantidade média
    else:
        max_estados = 100000  # Para muitos candidatos, mantém limite
    
    for idx_global, valor in enumerate(valores_ordenados):
        # Se já passou do limite de itens, para
        if idx_global >= max_itens:
            break
        
        novas_somas = {}
        estados_criados = 0
        
        # Adiciona todas as somas anteriores
        for soma_atual, combinacoes in soma_possivel.items():
            # Mantém apenas a melhor combinação para cada soma (menos itens)
            if combinacoes:
                melhor_combinacao = min(combinacoes, key=lambda x: x[0])
                if soma_atual not in novas_somas:
                    novas_somas[soma_atual] = [melhor_combinacao]
                    estados_criados += 1
            else:
                # Se não há combinações, cria uma vazia
                melhor_combinacao = (0, [])
            
            # Tenta adicionar o valor atual
            nova_soma = soma_atual + valor
            novo_num_itens = melhor_combinacao[0] + 1
            novos_indices = melhor_combinacao[1] + [indices_ordenados[idx_global]]
            
            # Só adiciona se não ultrapassar limites
            if novo_num_itens <= max_itens and nova_soma <= valor_alvo_centavos + 100:  # Pequena margem para arredondamentos
                if estados_criados < max_estados:
                    if nova_soma not in novas_somas:
                        novas_somas[nova_soma] = [(novo_num_itens, novos_indices)]
                        estados_criados += 1
                    else:
                        # Verifica se esta combinação é melhor (menos itens)
                        combinacao_existente = novas_somas[nova_soma][0]
                        if novo_num_itens < combinacao_existente[0]:
                            novas_somas[nova_soma] = [(novo_num_itens, novos_indices)]
                    
                    # Se encontrou valor exato, retorna imediatamente
                    if nova_soma == valor_alvo_centavos:
                        return novos_indices
        
        # Se criou muitos estados, filtra apenas os mais relevantes (mais próximos do alvo)
        if len(novas_somas) > max_estados:
            soma_possivel = dict(sorted(novas_somas.items(), 
                                      key=lambda x: abs(x[0] - valor_alvo_centavos))[:max_estados])
        else:
            soma_possivel = novas_somas
    
    # Melhor match: exato primeiro; senão, dentro da tolerância (menor diferença, depois menos itens)
    tol = tolerancia_centavos if tolerancia_centavos is not None else 0
    melhor_match = None
    melhor_diff = float('inf')
    for soma, combinacoes in soma_possivel.items():
        if not combinacoes:
            continue
        diff = abs(soma - valor_alvo_centavos)
        if diff == 0:
            melhor_match = combinacoes[0][1]
            break
        if tol > 0 and diff <= tol and diff < melhor_diff:
            melhor_diff = diff
            melhor_match = combinacoes[0][1]
    return melhor_match


def conciliar_extrato_comprovantes(df_extrato, df_comprovantes):
    """
    Concilia extratos com comprovantes usando estratégia inteligente de múltiplas passadas.
    
    Args:
        df_extrato: DataFrame do extrato preparado
        df_comprovantes: DataFrame dos comprovantes preparado
        
    Returns:
        DataFrame com a conciliação
    """
    print("\n" + "="*80)
    print("INICIANDO CONCILIAÇÃO INTELIGENTE")
    print("="*80)

    # Quando o incremental filtra para um período sem linhas, evita apply em DF vazio
    # (pandas pode retornar DataFrame, quebrando a atribuição de coluna escalar).
    if df_extrato is None or len(df_extrato) == 0:
        print("[INFO] Nenhum extrato para conciliar nesta rodada.")
        return pd.DataFrame(), set()
    
    conciliacoes = []
    comprovantes_usados = set()
    
    # Estratégia 1: Processa extratos mais simples primeiro (1-3 comprovantes prováveis)
    # Calcula "complexidade estimada" de cada extrato baseado em:
    # - Tamanho do valor (valores menores são mais simples)
    # - Número de comprovantes disponíveis na mesma data
    def calcular_complexidade(row):
        valor = abs(row['valor'])
        data_extrato = row['data']
        
        # Conta comprovantes disponíveis na mesma data (±2 dias)
        comprovantes_proximos = df_comprovantes[
            df_comprovantes.apply(
                lambda r: abs((r['data'] - data_extrato).days) <= 2,
                axis=1
            )
        ]
        qtd_proximos = len(comprovantes_proximos)
        
        # Complexidade = valor / número de opções
        # Quanto menor o valor e mais opções, mais simples
        if qtd_proximos > 0:
            complexidade = valor / qtd_proximos
        else:
            complexidade = valor  # Sem opções = complexo
        
        return complexidade
    
    df_extrato['complexidade'] = df_extrato.apply(calcular_complexidade, axis=1)
    
    # Ordena por complexidade (mais simples primeiro) - isso permite encontrar matches simples primeiro
    # e deixar mais comprovantes disponíveis para matches complexos
    df_extrato_ordenado = df_extrato.sort_values('complexidade', ascending=True).copy()
    
    total_extratos = len(df_extrato_ordenado)
    
    # Primeira passada: tenta conciliar todos
    print("\n[PASSADA 1] Processando todos os extratos (ordem: mais simples primeiro)...")
    extratos_nao_conciliados = []
    
    for idx, extrato_row in df_extrato_ordenado.iterrows():
        valor_extrato_abs = abs(extrato_row['valor'])
        print(f"\nProcessando extrato {extrato_row['ID_extrato']}/{total_extratos} "
              f"(R$ {valor_extrato_abs:,.2f}) [complexidade: {extrato_row['complexidade']:,.0f}]")
        
        # Filtra comprovantes ainda não usados
        comprovantes_disponiveis = df_comprovantes[
            ~df_comprovantes.index.isin(comprovantes_usados)
        ]
        
        if len(comprovantes_disponiveis) == 0:
            print(f"  [INFO] Nenhum comprovante disponível")
            extratos_nao_conciliados.append(extrato_row)
            continue
        
        # Encontra combinação de comprovantes
        combinacao = encontrar_combinacoes_comprovantes(
            extrato_row,
            comprovantes_disponiveis
        )
        
        if combinacao:
            print(f"  [DEBUG] Combinação recebida: {len(combinacao)} comprovante(s), índices: {combinacao[:5]}...")
            
            # Verifica se todos os índices existem no DataFrame original
            indices_validos = []
            indices_invalidos = []
            for c in combinacao:
                if c in df_comprovantes.index:
                    indices_validos.append(c)
                else:
                    indices_invalidos.append(c)
                    print(f"  [AVISO] Índice {c} não encontrado no DataFrame original de comprovantes")
            
            if len(indices_validos) != len(combinacao):
                print(f"  [ERRO] Alguns índices não são válidos. Válidos: {len(indices_validos)}/{len(combinacao)}")
                print(f"  [DEBUG] Índices inválidos: {indices_invalidos}")
                print(f"  [DEBUG] Primeiros 10 índices disponíveis no DataFrame: {list(df_comprovantes.index[:10])}")
                extratos_nao_conciliados.append(extrato_row)
                continue
            
            # Verifica se a combinação está correta (soma dos valores)
            try:
                valores_encontrados = []
                for c in combinacao:
                    valor_comp = abs(df_comprovantes.loc[c]['valor'])
                    valores_encontrados.append(valor_comp)
                
                soma_comprovantes = sum(valores_encontrados)
                diferenca = abs(soma_comprovantes - valor_extrato_abs)
                
                # Converte para centavos para comparação exata (evita problemas de ponto flutuante)
                soma_comprovantes_centavos = int(round(soma_comprovantes * 100))
                valor_extrato_centavos = int(round(valor_extrato_abs * 100))
                diferenca_centavos = abs(soma_comprovantes_centavos - valor_extrato_centavos)
                
                print(f"  [DEBUG] Valores dos comprovantes: {[f'R$ {v:,.2f}' for v in valores_encontrados]}")
                print(f"  [DEBUG] Soma: R$ {soma_comprovantes:,.2f} | Extrato: R$ {valor_extrato_abs:,.2f} | Diferença: R$ {diferenca:,.2f}")
                print(f"  [DEBUG] Comparação em centavos: {soma_comprovantes_centavos} vs {valor_extrato_centavos} | Diferença: {diferenca_centavos} centavos")
            except (KeyError, IndexError) as e:
                print(f"  [ERRO] Erro ao acessar comprovantes: {e}")
                print(f"  [DEBUG] Índices na combinação: {combinacao}")
                print(f"  [DEBUG] Tipo dos índices: {[type(c) for c in combinacao[:3]]}")
                print(f"  [DEBUG] Primeiros 10 índices disponíveis no DataFrame: {list(df_comprovantes.index[:10])}")
                extratos_nao_conciliados.append(extrato_row)
                continue
            
            if diferenca_centavos <= TOLERANCIA_CENTAVOS:  # Exato ou dentro da tolerância (arredondamento)
                # Marca comprovantes como usados
                comprovantes_usados.update(combinacao)
                
                # Cria registros de conciliação
                for comprovante_idx in combinacao:
                    try:
                        comprovante_row = df_comprovantes.loc[comprovante_idx]
                        
                        # Busca Ref. Sigra do comprovante
                        ref_sigra = '-'
                        for col_name in ['Ref. Sigra', 'Ref Sigra', 'REF. SIGRA', 'ref_sigra']:
                            if col_name in comprovante_row.index:
                                ref_val = comprovante_row.get(col_name)
                                if pd.notna(ref_val) and str(ref_val).strip() and str(ref_val).strip() != '':
                                    ref_sigra = str(ref_val).strip()
                                    break
                        
                        # Debug: mostra se encontrou Ref. Sigra
                        if ref_sigra != '-':
                            print(f"      [DEBUG] Ref. Sigra encontrada: {ref_sigra} (comprovante ID {comprovante_row['ID_comprovante']})")
                        else:
                            # Verifica se a coluna existe mas está vazia
                            if 'Ref. Sigra' in comprovante_row.index:
                                print(f"      [DEBUG] Coluna 'Ref. Sigra' existe mas está vazia para comprovante ID {comprovante_row['ID_comprovante']}")
                        
                        # Busca Categoria do comprovante
                        categoria = '-'
                        for col_name in ['Categoria', 'categoria', 'categoria_original']:
                            if col_name in comprovante_row.index:
                                cat_val = comprovante_row.get(col_name)
                                if pd.notna(cat_val) and str(cat_val).strip() and str(cat_val).strip() != '':
                                    categoria = str(cat_val).strip()
                                    break
                        
                        # Busca Cliente do comprovante (coluna F no PGTO MASTER; NÃO usar "Ref Cliente" que é coluna E)
                        cliente = '-'
                        for col_name in comprovante_row.index:
                            cn = str(col_name).lower()
                            if 'cliente' in cn and 'ref' not in cn:
                                v = comprovante_row.get(col_name)
                                if pd.notna(v) and str(v).strip():
                                    cliente = str(v).strip()
                                    break
                        
                        conciliacoes.append({
                            'ID_extrato': extrato_row['ID_extrato'],
                            'Data_extrato': extrato_row['data_original'],
                            'Valor_extrato': extrato_row['valor'],
                            'Favorecido_extrato': extrato_row['favorecido_original'],
                            'ID_comprovante': comprovante_row['ID_comprovante'],
                            'Data_comprovante': comprovante_row['data_original'],
                            'Valor_comprovante': comprovante_row['valor'],
                            'Favorecido_comprovante': comprovante_row['favorecido_original'],
                            'Ref. Sigra': ref_sigra,
                            'Categoria': categoria,
                            'Cliente': cliente
                        })
                    except (KeyError, IndexError) as e:
                        print(f"  [ERRO] Erro ao processar comprovante {comprovante_idx}: {e}")
                        continue
                
                print(f"\n  [✓ CONCILIAÇÃO CONFIRMADA]")
                print(f"  Extrato ID {extrato_row['ID_extrato']}: R$ {valor_extrato_abs:,.2f}")
                print(f"  Relacionado com {len(combinacao)} comprovante(s) totalizando R$ {soma_comprovantes:,.2f}")
                print(f"  Diferença: R$ {diferenca:,.2f}")
            else:
                print(f"  [ERRO] Diferença encontrada: R$ {diferenca:,.2f}")
                extratos_nao_conciliados.append(extrato_row)
        else:
            extratos_nao_conciliados.append(extrato_row)
            # Mostra informações para debug
            comprovantes_mesma_data = comprovantes_disponiveis[
                comprovantes_disponiveis.apply(
                    lambda row: datas_proximas(extrato_row['data'], row['data']),
                    axis=1
                )
            ]
            if len(comprovantes_mesma_data) > 0:
                soma_disponivel = abs(comprovantes_mesma_data['valor']).sum()
                print(f"  [INFO] {len(comprovantes_mesma_data)} comprovante(s) na mesma data, soma: R$ {soma_disponivel:,.2f}")
    
    # Segunda passada: reprocessa extratos não conciliados com estratégia diferente
    if len(extratos_nao_conciliados) > 0:
        print(f"\n[PASSADA 2] Reprocessando {len(extratos_nao_conciliados)} extratos não conciliados...")
        print("           (Ordenando por valor: maiores primeiro para encontrar agrupamentos)")
        
        # Na segunda passada, ordena por valor (maior primeiro) para priorizar agrupamentos grandes
        df_nao_conciliados = pd.DataFrame(extratos_nao_conciliados)
        df_nao_conciliados = df_nao_conciliados.sort_values('valor', ascending=False).copy()
        
        for idx, extrato_row in df_nao_conciliados.iterrows():
            valor_extrato_abs = abs(extrato_row['valor'])
            print(f"\nReprocessando extrato {extrato_row['ID_extrato']} (R$ {valor_extrato_abs:,.2f})")
            
            # Filtra comprovantes ainda não usados
            comprovantes_disponiveis = df_comprovantes[
                ~df_comprovantes.index.isin(comprovantes_usados)
            ]
            
            if len(comprovantes_disponiveis) == 0:
                continue
            
            # Tenta novamente com mais agressividade (mais candidatos, mais tolerância)
            combinacao = encontrar_combinacoes_comprovantes(
                extrato_row,
                comprovantes_disponiveis
            )
            
            if combinacao:
                # Verifica se todos os índices existem no DataFrame original
                indices_validos = []
                for c in combinacao:
                    if c in df_comprovantes.index:
                        indices_validos.append(c)
                    else:
                        print(f"  [AVISO] Índice {c} não encontrado no DataFrame original de comprovantes")
                
                if len(indices_validos) != len(combinacao):
                    print(f"  [ERRO] Alguns índices não são válidos. Válidos: {len(indices_validos)}/{len(combinacao)}")
                    continue
                
                try:
                    valores_encontrados = []
                    for c in combinacao:
                        valor_comp = abs(df_comprovantes.loc[c]['valor'])
                        valores_encontrados.append(valor_comp)
                    
                    soma_comprovantes = sum(valores_encontrados)
                    diferenca = abs(soma_comprovantes - valor_extrato_abs)
                    
                    # Converte para centavos para comparação exata (evita problemas de ponto flutuante)
                    soma_comprovantes_centavos = int(round(soma_comprovantes * 100))
                    valor_extrato_centavos = int(round(valor_extrato_abs * 100))
                    diferenca_centavos = abs(soma_comprovantes_centavos - valor_extrato_centavos)
                except (KeyError, IndexError) as e:
                    print(f"  [ERRO] Erro ao acessar comprovantes: {e}")
                    print(f"  [DEBUG] Índices na combinação: {combinacao}")
                    continue
                
                if diferenca_centavos <= TOLERANCIA_CENTAVOS:  # Exato ou dentro da tolerância
                    comprovantes_usados.update(combinacao)
                    
                    for comprovante_idx in combinacao:
                        try:
                            comprovante_row = df_comprovantes.loc[comprovante_idx]
                            
                            # Busca Ref. Sigra do comprovante
                            ref_sigra = '-'
                            for col_name in ['Ref. Sigra', 'Ref Sigra', 'REF. SIGRA', 'ref_sigra']:
                                if col_name in comprovante_row.index:
                                    ref_val = comprovante_row.get(col_name)
                                    if pd.notna(ref_val) and str(ref_val).strip() and str(ref_val).strip() != '':
                                        ref_sigra = str(ref_val).strip()
                                        break
                            
                            # Busca Categoria do comprovante
                            categoria = '-'
                            for col_name in ['Categoria', 'categoria', 'categoria_original']:
                                if col_name in comprovante_row.index:
                                    cat_val = comprovante_row.get(col_name)
                                    if pd.notna(cat_val) and str(cat_val).strip() and str(cat_val).strip() != '':
                                        categoria = str(cat_val).strip()
                                        break
                            
                            # Busca Cliente do comprovante (coluna F no PGTO MASTER; NÃO usar "Ref Cliente" que é coluna E)
                            cliente = '-'
                            for col_name in comprovante_row.index:
                                cn = str(col_name).lower()
                                if 'cliente' in cn and 'ref' not in cn:
                                    v = comprovante_row.get(col_name)
                                    if pd.notna(v) and str(v).strip():
                                        cliente = str(v).strip()
                                        break
                            
                            conciliacoes.append({
                                'ID_extrato': extrato_row['ID_extrato'],
                                'Data_extrato': extrato_row['data_original'],
                                'Valor_extrato': extrato_row['valor'],
                                'Favorecido_extrato': extrato_row['favorecido_original'],
                                'ID_comprovante': comprovante_row['ID_comprovante'],
                                'Data_comprovante': comprovante_row['data_original'],
                                'Valor_comprovante': comprovante_row['valor'],
                                'Favorecido_comprovante': comprovante_row['favorecido_original'],
                                'Ref. Sigra': ref_sigra,
                                'Categoria': categoria,
                                'Cliente': cliente
                            })
                        except (KeyError, IndexError) as e:
                            print(f"  [ERRO] Erro ao processar comprovante {comprovante_idx}: {e}")
                            continue
                    
                    print(f"  [OK] Conciliação encontrada na 2ª passada: {len(combinacao)} comprovante(s)")
    
    df_conciliacao = pd.DataFrame(conciliacoes)
    
    # Debug: verifica se as colunas foram criadas no DataFrame de conciliação
    if len(conciliacoes) > 0:
        print(f"\n[DEBUG] ===== CRIAÇÃO DF CONCILIAÇÃO =====")
        print(f"[DEBUG] Total de registros de conciliação: {len(conciliacoes)}")
        if len(conciliacoes) > 0:
            primeiro_registro = conciliacoes[0]
            print(f"[DEBUG] Chaves do primeiro registro: {list(primeiro_registro.keys())}")
            if 'Ref. Sigra' in primeiro_registro:
                print(f"[DEBUG] ✓ 'Ref. Sigra' encontrada no primeiro registro: '{primeiro_registro['Ref. Sigra']}'")
            else:
                print(f"[DEBUG] ✗ 'Ref. Sigra' NÃO encontrada no primeiro registro!")
            if 'Categoria' in primeiro_registro:
                print(f"[DEBUG] ✓ 'Categoria' encontrada no primeiro registro: '{primeiro_registro['Categoria']}'")
            else:
                print(f"[DEBUG] ✗ 'Categoria' NÃO encontrada no primeiro registro!")
        print(f"[DEBUG] Colunas no DataFrame de conciliação: {list(df_conciliacao.columns)}")
        print(f"[DEBUG] ======================================\n")
    
    print(f"\n[OK] Conciliação concluída: {len(df_conciliacao)} relacionamentos encontrados")
    print(f"     Extratos conciliados: {len(df_conciliacao['ID_extrato'].unique()) if len(df_conciliacao) > 0 else 0}/{total_extratos}")
    
    return df_conciliacao, comprovantes_usados


# ============================================================================
# FUNÇÕES DE SAÍDA
# ============================================================================

def identificar_pendencias(df_extrato, df_comprovantes, df_conciliacao, comprovantes_usados):
    """
    Identifica pendências (extratos e comprovantes não conciliados).
    
    Args:
        df_extrato: DataFrame do extrato
        df_comprovantes: DataFrame dos comprovantes
        df_conciliacao: DataFrame da conciliação
        comprovantes_usados: Set com índices dos comprovantes usados
        
    Returns:
        Tupla com (DataFrame de extratos pendentes, DataFrame de comprovantes pendentes)
    """
    # Extratos sem conciliação
    extratos_conciliados = set(df_conciliacao['ID_extrato'].unique()) if len(df_conciliacao) > 0 else set()
    df_extratos_pendentes = df_extrato[
        ~df_extrato['ID_extrato'].isin(extratos_conciliados)
    ].copy()
    
    # Comprovantes não utilizados
    df_comprovantes_pendentes = df_comprovantes[
        ~df_comprovantes.index.isin(comprovantes_usados)
    ].copy()
    
    return df_extratos_pendentes, df_comprovantes_pendentes




def criar_aba_conciliacao_bancaria(df_extrato, df_comprovantes, df_conciliacao, 
                                    df_extratos_pendentes, df_comprovantes_pendentes):
    """
    Cria uma aba de conciliação bancária no formato tradicional.
    
    Args:
        df_extrato: DataFrame do extrato completo
        df_comprovantes: DataFrame dos comprovantes
        df_conciliacao: DataFrame da conciliação
        df_extratos_pendentes: DataFrame de extratos pendentes
        df_comprovantes_pendentes: DataFrame de comprovantes pendentes
    
    Returns:
        DataFrame com a conciliação bancária formatada
    """
    # Separa créditos e débitos
    def identificar_tipo(linha):
        favorecido = str(linha.get('favorecido_original', '')).upper()
        valor = linha.get('valor', 0)
        
        if any(palavra in favorecido for palavra in ['PIX RECEBIDO', 'RECEBIMENTO', 'CREDITO', 'CRÉDITO']):
            return 'credito'
        elif any(palavra in favorecido for palavra in ['PIX ENVIADO', 'PAGAMENTO', 'DEBITO', 'DÉBITO', 'SISPAG', 'SISCOMEX']):
            return 'debito'
        elif valor < 0:
            return 'debito'
        else:
            return 'credito'
    
    # Calcula totais
    total_creditos = 0
    total_debitos = 0
    creditos_conciliados = 0
    debitos_conciliados = 0
    creditos_pendentes = 0
    debitos_pendentes = 0
    
    extratos_conciliados_ids = set(df_conciliacao['ID_extrato'].unique()) if len(df_conciliacao) > 0 else set()
    
    for _, row in df_extrato.iterrows():
        valor = abs(row.get('valor', 0))
        tipo = identificar_tipo(row)
        
        if tipo == 'credito':
            total_creditos += valor
            if row['ID_extrato'] in extratos_conciliados_ids:
                creditos_conciliados += valor
            else:
                creditos_pendentes += valor
        else:
            total_debitos += valor
            if row['ID_extrato'] in extratos_conciliados_ids:
                debitos_conciliados += valor
            else:
                debitos_pendentes += valor
    
    # Saldo
    saldo_anterior = 0  # Pode ser calculado se houver histórico
    saldo_atual = saldo_anterior + total_creditos - total_debitos
    saldo_conciliado = saldo_anterior + creditos_conciliados - debitos_conciliados
    
    # Cria estrutura da conciliação
    conciliacao_data = []
    
    # Cabeçalho
    conciliacao_data.append({
        'Categoria': 'CONCILIAÇÃO BANCÁRIA',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    conciliacao_data.append({
        'Categoria': '',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    # Saldo anterior
    conciliacao_data.append({
        'Categoria': 'SALDO ANTERIOR',
        'Descrição': 'Saldo do período anterior',
        'Valor': saldo_anterior,
        'Status': '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    # CRÉDITOS (ENTRADAS)
    conciliacao_data.append({
        'Categoria': 'CRÉDITOS (ENTRADAS)',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    conciliacao_data.append({
        'Categoria': '  Total de Créditos',
        'Descrição': 'Total de recebimentos no período',
        'Valor': total_creditos,
        'Status': '✅' if creditos_pendentes == 0 else '⚠️'
    })
    
    creditos_conc = len([x for x in df_extrato.iterrows() if x[1]['ID_extrato'] in extratos_conciliados_ids and identificar_tipo(x[1]) == 'credito'])
    creditos_pend = len([x for x in df_extratos_pendentes.iterrows() if identificar_tipo(x[1]) == 'credito'])
    
    conciliacao_data.append({
        'Categoria': '    - Conciliados',
        'Descrição': f'{creditos_conc} transação(ões)',
        'Valor': creditos_conciliados,
        'Status': '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '    - Pendentes',
        'Descrição': f'{creditos_pend} transação(ões)',
        'Valor': creditos_pendentes,
        'Status': '❌' if creditos_pendentes > 0 else '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    # DÉBITOS (SAÍDAS)
    conciliacao_data.append({
        'Categoria': 'DÉBITOS (SAÍDAS)',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    conciliacao_data.append({
        'Categoria': '  Total de Débitos',
        'Descrição': 'Total de pagamentos no período',
        'Valor': total_debitos,
        'Status': '✅' if debitos_pendentes == 0 else '⚠️'
    })
    
    debitos_conc = len([x for x in df_extrato.iterrows() if x[1]['ID_extrato'] in extratos_conciliados_ids and identificar_tipo(x[1]) == 'debito'])
    debitos_pend = len([x for x in df_extratos_pendentes.iterrows() if identificar_tipo(x[1]) == 'debito'])
    
    conciliacao_data.append({
        'Categoria': '    - Conciliados',
        'Descrição': f'{debitos_conc} transação(ões)',
        'Valor': debitos_conciliados,
        'Status': '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '    - Pendentes',
        'Descrição': f'{debitos_pend} transação(ões)',
        'Valor': debitos_pendentes,
        'Status': '❌' if debitos_pendentes > 0 else '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    # SALDO
    conciliacao_data.append({
        'Categoria': 'SALDO ATUAL',
        'Descrição': 'Saldo após créditos e débitos',
        'Valor': saldo_atual,
        'Status': '✅' if (creditos_pendentes == 0 and debitos_pendentes == 0) else '⚠️'
    })
    
    conciliacao_data.append({
        'Categoria': 'SALDO CONCILIADO',
        'Descrição': 'Saldo considerando apenas transações conciliadas',
        'Valor': saldo_conciliado,
        'Status': '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    # RESUMO
    total_transacoes = len(df_extrato)
    transacoes_conciliadas = len(extratos_conciliados_ids)
    transacoes_pendentes = len(df_extratos_pendentes)
    taxa_conciliacao = (transacoes_conciliadas / total_transacoes * 100) if total_transacoes > 0 else 0
    
    conciliacao_data.append({
        'Categoria': 'RESUMO',
        'Descrição': '',
        'Valor': None,
        'Status': ''
    })
    
    conciliacao_data.append({
        'Categoria': '  Total de Transações',
        'Descrição': f'Total de lançamentos no extrato',
        'Valor': total_transacoes,
        'Status': 'ℹ️'
    })
    
    conciliacao_data.append({
        'Categoria': '  Transações Conciliadas',
        'Descrição': f'{taxa_conciliacao:.1f}% do total',
        'Valor': transacoes_conciliadas,
        'Status': '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '  Transações Pendentes',
        'Descrição': f'{100 - taxa_conciliacao:.1f}% do total',
        'Valor': transacoes_pendentes,
        'Status': '❌' if transacoes_pendentes > 0 else '✅'
    })
    
    conciliacao_data.append({
        'Categoria': '  Comprovantes Não Utilizados',
        'Descrição': f'Comprovantes sem extrato correspondente',
        'Valor': len(df_comprovantes_pendentes),
        'Status': '⚠️' if len(df_comprovantes_pendentes) > 0 else '✅'
    })
    
    return pd.DataFrame(conciliacao_data)


def criar_aba_status_extrato(df_extrato, df_conciliacao, df_extratos_pendentes, df_comprovantes=None):
    """
    Cria uma aba que mostra o status de cada valor do extrato.
    
    Args:
        df_extrato: DataFrame do extrato completo
        df_conciliacao: DataFrame da conciliação
        df_extratos_pendentes: DataFrame de extratos pendentes
        df_comprovantes: DataFrame dos comprovantes (opcional, para buscar Ref. Sigra)
    
    Returns:
        DataFrame com status de cada extrato
    """
    # Normaliza IDs para garantir match mesmo com arquivos antigos (int/float/string)
    df_extrato = df_extrato.copy()
    df_conciliacao = df_conciliacao.copy()
    if 'ID_extrato' in df_extrato.columns:
        df_extrato['ID_extrato'] = pd.to_numeric(df_extrato['ID_extrato'], errors='coerce')
    if 'ID_extrato' in df_conciliacao.columns:
        df_conciliacao['ID_extrato'] = pd.to_numeric(df_conciliacao['ID_extrato'], errors='coerce')
    if 'ID_comprovante' in df_conciliacao.columns:
        df_conciliacao['ID_comprovante'] = pd.to_numeric(df_conciliacao['ID_comprovante'], errors='coerce')

    # Prepara dados de conciliação agrupados por ID_extrato
    # Debug: verifica se as colunas estão no DataFrame de conciliação
    print(f"\n[DEBUG] ===== VERIFICAÇÃO CONCILIAÇÃO =====")
    print(f"[DEBUG] Total de registros na conciliação: {len(df_conciliacao)}")
    if len(df_conciliacao) > 0:
        print(f"[DEBUG] Colunas no DataFrame de conciliação: {list(df_conciliacao.columns)}")
        if 'Ref. Sigra' in df_conciliacao.columns:
            refs_com_valor = (df_conciliacao['Ref. Sigra'] != '-').sum()
            print(f"[DEBUG] ✓ Coluna 'Ref. Sigra' encontrada: {refs_com_valor}/{len(df_conciliacao)} registros com valores")
            if refs_com_valor > 0:
                print(f"[DEBUG]   Exemplos: {df_conciliacao[df_conciliacao['Ref. Sigra'] != '-']['Ref. Sigra'].head(3).tolist()}")
        else:
            print(f"[DEBUG] ✗ ERRO: Coluna 'Ref. Sigra' NÃO encontrada no DataFrame de conciliação!")
        if 'Categoria' in df_conciliacao.columns:
            cats_com_valor = (df_conciliacao['Categoria'] != '-').sum()
            print(f"[DEBUG] ✓ Coluna 'Categoria' encontrada: {cats_com_valor}/{len(df_conciliacao)} registros com valores")
        else:
            print(f"[DEBUG] ✗ ERRO: Coluna 'Categoria' NÃO encontrada no DataFrame de conciliação!")
    print(f"[DEBUG] ====================================\n")
    
    conciliacoes_por_extrato = {}
    if len(df_conciliacao) > 0:
        for _, row in df_conciliacao.iterrows():
            id_extrato = row['ID_extrato']
            if pd.isna(id_extrato):
                continue
            id_extrato = int(id_extrato)
            if id_extrato not in conciliacoes_por_extrato:
                conciliacoes_por_extrato[id_extrato] = {
                    'comprovantes': [],
                    'valores': [],
                    'total_conciliado': 0,
                    'ids_comprovantes': [],
                    'refs_sigra': [],  # Nova: armazena Ref. Sigra
                    'categorias': [],  # Nova: armazena Categorias
                    'clientes': []  # Cliente (coluna F do PGTO MASTER)
                }
            conciliacoes_por_extrato[id_extrato]['comprovantes'].append(row)
            id_comprovante = row.get('ID_comprovante', pd.NA)
            if pd.notna(id_comprovante):
                try:
                    id_comprovante = int(id_comprovante)
                except Exception:
                    pass
            conciliacoes_por_extrato[id_extrato]['ids_comprovantes'].append(id_comprovante)
            
            # Ref. Sigra, Categoria e Cliente já vêm salvas na conciliação (foram capturadas quando criamos o registro)!
            ref_sigra = row.get('Ref. Sigra', '-')
            categoria = row.get('Categoria', '-')
            cliente = row.get('Cliente', '-')
            
            conciliacoes_por_extrato[id_extrato]['refs_sigra'].append(ref_sigra if pd.notna(ref_sigra) else '-')
            conciliacoes_por_extrato[id_extrato]['categorias'].append(categoria if pd.notna(categoria) else '-')
            conciliacoes_por_extrato[id_extrato]['clientes'].append(cliente if pd.notna(cliente) else '-')
            
            try:
                valor = float(row.get('Valor_comprovante', 0))
                conciliacoes_por_extrato[id_extrato]['valores'].append(valor)
                conciliacoes_por_extrato[id_extrato]['total_conciliado'] += valor
            except:
                pass
    
    # Cria lista de status
    status_lista = []
    
    # Debug: verifica se Ref. Sigra e Categoria estão na conciliação
    if len(df_conciliacao) > 0:
        if 'Ref. Sigra' in df_conciliacao.columns:
            refs_preenchidas = (df_conciliacao['Ref. Sigra'] != '-').sum()
            print(f"[DEBUG] Ref. Sigra encontrada na conciliação: {refs_preenchidas}/{len(df_conciliacao)} valores preenchidos")
        else:
            print(f"[DEBUG] ✗ Coluna 'Ref. Sigra' NÃO encontrada na conciliação!")
        
        if 'Categoria' in df_conciliacao.columns:
            cats_preenchidas = (df_conciliacao['Categoria'] != '-').sum()
            print(f"[DEBUG] Categoria encontrada na conciliação: {cats_preenchidas}/{len(df_conciliacao)} valores preenchidos")
        else:
            print(f"[DEBUG] ✗ Coluna 'Categoria' NÃO encontrada na conciliação!")
    
    for _, extrato_row in df_extrato.iterrows():
        id_extrato = extrato_row['ID_extrato']
        if pd.isna(id_extrato):
            continue
        id_extrato = int(id_extrato)
        valor_extrato = extrato_row.get('valor', 0)
        data_extrato = extrato_row.get('data_original', '')
        favorecido = extrato_row.get('favorecido_original', '')
        
        # Busca categoria do extrato (se existir)
        categoria_extrato = '-'
        for col_cat in extrato_row.index:
            col_cat_lower = str(col_cat).lower()
            if 'categoria' in col_cat_lower or col_cat_lower == 'cate':
                cat_val = extrato_row.get(col_cat)
                if pd.notna(cat_val) and str(cat_val).strip():
                    categoria_extrato = str(cat_val).strip()
                    break
        
        # Verifica se foi conciliado
        if id_extrato in conciliacoes_por_extrato:
            info_conciliacao = conciliacoes_por_extrato[id_extrato]
            status = '✅ Conciliado'
            qtd_comprovantes = len(info_conciliacao['comprovantes'])
            valor_total_conciliado = info_conciliacao['total_conciliado']
            ids_comprovantes = ', '.join(map(str, info_conciliacao['ids_comprovantes']))
            
            # Ref. Sigra: junta todas as refs (separadas por vírgula se múltiplas)
            refs_sigra = info_conciliacao.get('refs_sigra', [])
            if refs_sigra:
                # Remove duplicatas e normaliza para string (evita erro com int no join)
                refs_unicas = list(dict.fromkeys([str(r).strip() for r in refs_sigra if str(r).strip() and str(r).strip() != '-']))
                ref_sigra_str = ', '.join(refs_unicas) if refs_unicas else '-'
            else:
                ref_sigra_str = '-'
            
            # Debug: mostra Ref. Sigra encontrada
            if ref_sigra_str != '-':
                print(f"  [DEBUG] Extrato ID {id_extrato}: Ref. Sigra = {ref_sigra_str}")
            
            # Categoria: junta todas as categorias dos comprovantes (separadas por vírgula se múltiplas)
            categorias_comprovantes = info_conciliacao.get('categorias', [])
            if categorias_comprovantes:
                # Remove duplicatas e normaliza para string
                categorias_unicas = list(dict.fromkeys([str(c).strip() for c in categorias_comprovantes if str(c).strip() and str(c).strip() != '-']))
                categoria_comprovantes_str = ', '.join(categorias_unicas) if categorias_unicas else '-'
            else:
                categoria_comprovantes_str = '-'
            
            # Cliente: junta todos os clientes dos comprovantes (coluna F do PGTO MASTER)
            clientes_comprovantes = info_conciliacao.get('clientes', [])
            if clientes_comprovantes:
                clientes_unicos = list(dict.fromkeys([str(c).strip() for c in clientes_comprovantes if str(c).strip() and str(c).strip() != '-']))
                cliente_str = ', '.join(clientes_unicos) if clientes_unicos else '-'
            else:
                cliente_str = '-'
            
            # Usa categoria do extrato se não tiver categoria dos comprovantes
            if categoria_extrato != '-' and categoria_comprovantes_str == '-':
                categoria_str = categoria_extrato
            elif categoria_comprovantes_str != '-' and categoria_extrato != '-':
                # Se tiver ambas, mostra ambas
                categoria_str = f"{categoria_extrato} ({categoria_comprovantes_str})"
            else:
                categoria_str = categoria_comprovantes_str if categoria_comprovantes_str != '-' else categoria_extrato
            
            diferenca = abs(valor_extrato) - abs(valor_total_conciliado)
            observacao = ''
            
            if diferenca > 0:  # Sem tolerância - diferença deve ser zero
                observacao = f'⚠️ Diferença de R$ {diferenca:.2f}'
            elif qtd_comprovantes > 1:
                observacao = f'Conciliado com {qtd_comprovantes} comprovantes'
            else:
                observacao = 'Perfeitamente conciliado'
        else:
            # Está pendente
            status = '❌ Pendente'
            qtd_comprovantes = 0
            valor_total_conciliado = 0
            ids_comprovantes = '-'
            ref_sigra_str = '-'
            cliente_str = '-'
            # Para pendentes, mostra categoria do extrato se existir
            categoria_str = categoria_extrato
            diferenca = abs(valor_extrato)
            observacao = 'Não encontrou comprovantes correspondentes'
        
        status_lista.append({
            'ID Extrato': id_extrato,
            'Data': data_extrato,
            'Valor Extrato': abs(valor_extrato) if valor_extrato else 0,
            'Favorecido/Descrição': str(favorecido)[:60] if favorecido else '',
            'Status': status,
            'Qtd Comprovantes': qtd_comprovantes,
            'Valor Total Conciliado': valor_total_conciliado if valor_total_conciliado > 0 else 0,
            'Diferença': diferenca if diferenca > 0 else 0,
            'Ref. Sigra': ref_sigra_str,
            'Categoria': categoria_str,
            'Cliente': cliente_str,
            'IDs Comprovantes': ids_comprovantes,
            'Observação': observacao
        })
    
    df_status = pd.DataFrame(status_lista)
    
    # Debug: verifica se as colunas foram criadas
    print(f"\n[DEBUG] ===== VERIFICAÇÃO STATUS EXTRATO =====")
    print(f"[DEBUG] Total de linhas no DataFrame: {len(df_status)}")
    print(f"[DEBUG] Colunas criadas no DataFrame Status Extrato:")
    for idx, col in enumerate(df_status.columns, 1):
        print(f"  [{idx}] '{col}'")
    
    # Verifica especificamente Ref. Sigra e Categoria
    if 'Ref. Sigra' in df_status.columns:
        refs_preenchidas = (df_status['Ref. Sigra'] != '-').sum()
        refs_com_valor = df_status[df_status['Ref. Sigra'] != '-']['Ref. Sigra'].head(5).tolist()
        print(f"[DEBUG] ✓ Coluna 'Ref. Sigra' encontrada: {refs_preenchidas}/{len(df_status)} linhas com valores")
        if refs_preenchidas > 0:
            print(f"[DEBUG]   Primeiros valores: {refs_com_valor}")
    else:
        print(f"[DEBUG] ✗ ERRO: Coluna 'Ref. Sigra' NÃO encontrada no DataFrame!")
    
    if 'Categoria' in df_status.columns:
        cats_preenchidas = (df_status['Categoria'] != '-').sum()
        cats_com_valor = df_status[df_status['Categoria'] != '-']['Categoria'].head(5).tolist()
        print(f"[DEBUG] ✓ Coluna 'Categoria' encontrada: {cats_preenchidas}/{len(df_status)} linhas com valores")
        if cats_preenchidas > 0:
            print(f"[DEBUG]   Primeiros valores: {cats_com_valor}")
    else:
        print(f"[DEBUG] ✗ ERRO: Coluna 'Categoria' NÃO encontrada no DataFrame!")
    
    print(f"[DEBUG] =======================================\n")
    
    return df_status


def formatar_planilha_excel(workbook, df_extrato, df_comprovantes, df_conciliacao,
                            df_extratos_pendentes, df_comprovantes_pendentes):
    """
    Formata a planilha Excel com cores, bordas e estilos para facilitar visualização.
    
    Args:
        workbook: Workbook do openpyxl
        df_extrato: DataFrame do extrato
        df_comprovantes: DataFrame dos comprovantes
        df_conciliacao: DataFrame da conciliação
        df_extratos_pendentes: DataFrame de extratos pendentes
        df_comprovantes_pendentes: DataFrame de comprovantes pendentes
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    
    # Cores padronizadas
    COR_CABECALHO = "4472C4"  # Azul
    COR_POSITIVO = "C6EFCE"   # Verde claro
    COR_NEGATIVO = "FFC7CE"   # Vermelho claro
    COR_PENDENTE = "FFEB9C"   # Amarelo claro
    COR_ALTERNADA = "F2F2F2"  # Cinza claro
    
    # Estilos
    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    preenchimento_cabecalho = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    borda_fina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alinhamento_centro = Alignment(horizontal='center', vertical='center')
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    
    # Função auxiliar para encontrar coluna de valor
    def encontrar_coluna_valor(sheet):
        """Encontra a coluna de valor baseado no cabeçalho."""
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).lower()
                if 'valor' in header and ('$' in header or 'r$' in header or 'rs' in header):
                    return col_idx
        # Se não encontrou, procura apenas por 'valor'
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).lower()
                if 'valor' in header:
                    return col_idx
        return None
    
    # Função para formatar uma aba
    def formatar_aba(sheet_name, df, col_valor=None, tipo='normal'):
        if sheet_name not in workbook.sheetnames:
            return
        
        sheet = workbook[sheet_name]
        
        # Se col_valor não foi fornecido, tenta encontrar
        if col_valor is None:
            col_valor = encontrar_coluna_valor(sheet)
        
        # Formata cabeçalho
        for cell in sheet[1]:
            cell.font = fonte_cabecalho
            cell.fill = preenchimento_cabecalho
            cell.border = borda_fina
            cell.alignment = alinhamento_centro
        
        # Congela primeira linha
        sheet.freeze_panes = 'A2'
        
        # Ajusta largura das colunas
        for idx, column in enumerate(sheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(idx)
            
            # Verifica cabeçalho
            header_cell = sheet.cell(row=1, column=idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))
            
            # Verifica células de dados
            for cell in column[1:]:  # Pula cabeçalho
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            # Define largura mínima e máxima
            adjusted_width = min(max(max_length + 2, 12), 60)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Formata células de dados
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = borda_fina
                
                # Formata valores monetários
                if col_valor and col_idx == col_valor:
                    try:
                        if cell.value is not None and cell.value != '':
                            valor = float(cell.value) if cell.value else 0
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = alinhamento_direita
                            
                            # Cores baseadas no valor
                            if tipo == 'pendente' or tipo == 'extrato':
                                if valor < 0:
                                    cell.fill = PatternFill(start_color=COR_NEGATIVO, end_color=COR_NEGATIVO, fill_type="solid")
                                elif valor > 0:
                                    cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                            elif tipo == 'comprovantes':
                                cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                    except:
                        pass
                
                # Formata datas
                header_cell = sheet.cell(row=1, column=col_idx)
                if header_cell and ('data' in str(header_cell.value).lower() or 'date' in str(header_cell.value).lower()):
                    cell.alignment = alinhamento_centro
                
                # Linhas alternadas (zebra striping) - apenas se não tiver cor especial
                if row_idx % 2 == 0 and tipo != 'pendente':
                    # Aplica cor alternada apenas se não for célula de valor monetário
                    if not col_valor or col_idx != col_valor:
                        try:
                            # Verifica se já tem cor especial (positivo/negativo)
                            if cell.fill and cell.fill.start_color:
                                rgb = str(cell.fill.start_color.rgb).upper()
                                # Se não for cor especial, aplica alternada
                                if rgb not in [COR_POSITIVO.replace('#', 'FF'), COR_NEGATIVO.replace('#', 'FF'), COR_PENDENTE.replace('#', 'FF')]:
                                    cell.fill = PatternFill(start_color=COR_ALTERNADA, end_color=COR_ALTERNADA, fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color=COR_ALTERNADA, end_color=COR_ALTERNADA, fill_type="solid")
                        except:
                            pass
        
        # Formata pendências com cor amarela de fundo
        if tipo == 'pendente':
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    try:
                        # Não sobrescreve cores de valores monetários (já formatadas acima)
                        if col_valor and col_idx == col_valor:
                            continue
                        
                        # Aplica cor de pendente
                        cell.fill = PatternFill(start_color=COR_PENDENTE, end_color=COR_PENDENTE, fill_type="solid")
                    except:
                        pass
    
    # Formata cada aba
    print("  Formatando aba 'extrato'...")
    col_valor = None
    if 'valor' in df_extrato.columns:
        col_valor = list(df_extrato.columns).index('valor') + 1
    formatar_aba('extrato', df_extrato, col_valor, 'extrato')
    
    print("  Formatando aba 'comprovantes'...")
    col_valor = None
    if 'valor' in df_comprovantes.columns:
        col_valor = list(df_comprovantes.columns).index('valor') + 1
    formatar_aba('comprovantes', df_comprovantes, col_valor, 'comprovantes')
    
    print("  Formatando aba 'conciliacao'...")
    if len(df_conciliacao) > 0:
        # Encontra colunas de valor na aba de conciliação
        if 'conciliacao' in workbook.sheetnames:
            sheet = workbook['conciliacao']
            col_valor_extrato = None
            col_valor_comprovante = None
            
            # Procura pelos cabeçalhos
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:
                    header = str(cell.value).lower()
                    if 'valor_extrato' in header or 'valor extrato' in header:
                        col_valor_extrato = col_idx
                    elif 'valor_comprovante' in header or 'valor comprovante' in header:
                        col_valor_comprovante = col_idx
            
            # Se não encontrou, tenta encontrar automaticamente
            if col_valor_extrato is None:
                col_valor_extrato = encontrar_coluna_valor(sheet)
            
            formatar_aba('conciliacao', df_conciliacao, col_valor_extrato, 'conciliacao')
            
            # Formata valor do comprovante também
            if col_valor_comprovante:
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    cell = row[col_valor_comprovante - 1]
                    try:
                        if cell.value is not None and cell.value != '':
                            valor = float(cell.value) if cell.value else 0
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = alinhamento_direita
                            if valor > 0:
                                cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                    except:
                        pass
    
    print("  Formatando aba 'pendencias_extratos'...")
    if len(df_extratos_pendentes) > 0:
        col_valor = None
        if 'valor' in df_extratos_pendentes.columns:
            col_valor = list(df_extratos_pendentes.columns).index('valor') + 1
        formatar_aba('pendencias_extratos', df_extratos_pendentes, col_valor, 'pendente')
    
    print("  Formatando aba 'pendencias_comprovantes'...")
    if len(df_comprovantes_pendentes) > 0:
        col_valor = None
        if 'valor' in df_comprovantes_pendentes.columns:
            col_valor = list(df_comprovantes_pendentes.columns).index('valor') + 1
        formatar_aba('pendencias_comprovantes', df_comprovantes_pendentes, col_valor, 'pendente')
    
    # Formata aba de Conciliação Bancária especialmente
    print("  Formatando aba 'Conciliação Bancária'...")
    if 'Conciliação Bancária' in workbook.sheetnames:
        sheet = workbook['Conciliação Bancária']
        
        # Formata cabeçalho
        for cell in sheet[1]:
            cell.font = fonte_cabecalho
            cell.fill = preenchimento_cabecalho
            cell.border = borda_fina
            cell.alignment = alinhamento_centro
        
        sheet.freeze_panes = 'A2'
        
        # Ajusta larguras
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15
        
        # Formata células
        from openpyxl.styles import Font as OpenpyxlFont
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = borda_fina
                
                categoria_cell = row[0] if len(row) > 0 else None
                if categoria_cell and categoria_cell.value:
                    categoria = str(categoria_cell.value).strip()
                    
                    # Formatação especial para títulos principais
                    if categoria in ['CONCILIAÇÃO BANCÁRIA', 'CRÉDITOS (ENTRADAS)', 'DÉBITOS (SAÍDAS)', 'RESUMO']:
                        for c in row:
                            c.font = OpenpyxlFont(bold=True, size=12)
                            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    
                    # Formata valores monetários
                    if col_idx == 3 and cell.value is not None:  # Coluna Valor
                        try:
                            if pd.notna(cell.value) and cell.value != '':
                                valor = float(cell.value) if cell.value else 0
                                cell.number_format = 'R$ #,##0.00'
                                cell.alignment = alinhamento_direita
                                
                                # Cores baseadas na categoria
                                if 'Créditos' in categoria or 'CRÉDITOS' in categoria:
                                    cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                                elif 'Débitos' in categoria or 'DÉBITOS' in categoria or 'DEBITOS' in categoria:
                                    cell.fill = PatternFill(start_color=COR_NEGATIVO, end_color=COR_NEGATIVO, fill_type="solid")
                                elif 'Saldo' in categoria:
                                    if valor >= 0:
                                        cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                                    else:
                                        cell.fill = PatternFill(start_color=COR_NEGATIVO, end_color=COR_NEGATIVO, fill_type="solid")
                                elif 'Conciliado' in categoria:
                                    cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                                elif 'Pendente' in categoria:
                                    cell.fill = PatternFill(start_color=COR_PENDENTE, end_color=COR_PENDENTE, fill_type="solid")
                        except:
                            pass
                    
                    # Formata coluna de Status
                    if col_idx == 4:  # Coluna Status
                        cell.alignment = alinhamento_centro
                        if cell.value:
                            status_text = str(cell.value)
                            if '✅' in status_text:
                                cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                            elif '❌' in status_text:
                                cell.fill = PatternFill(start_color=COR_NEGATIVO, end_color=COR_NEGATIVO, fill_type="solid")
                            elif '⚠️' in status_text:
                                cell.fill = PatternFill(start_color=COR_PENDENTE, end_color=COR_PENDENTE, fill_type="solid")
                            elif 'ℹ️' in status_text:
                                cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                    
                    # Indentação visual
                    if categoria.startswith('  '):
                        categoria_cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                        categoria_cell.font = OpenpyxlFont(size=10)
                    elif categoria.startswith('    -'):
                        categoria_cell.alignment = Alignment(horizontal='left', vertical='center', indent=4)
                        categoria_cell.font = OpenpyxlFont(size=10)
    
    # Formata aba de Status Extrato especialmente
    print("  Formatando aba 'Status Extrato'...")
    if 'Status Extrato' in workbook.sheetnames:
        sheet = workbook['Status Extrato']
        col_valor_extrato = None
        col_valor_conciliado = None
        col_diferenca = None
        col_status = None
        
        # Encontra colunas importantes
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).lower()
                if 'valor extrato' in header:
                    col_valor_extrato = col_idx
                elif 'valor total conciliado' in header:
                    col_valor_conciliado = col_idx
                elif 'diferença' in header:
                    col_diferenca = col_idx
                elif 'status' in header:
                    col_status = col_idx
        
        # Formata cabeçalho
        for cell in sheet[1]:
            cell.font = fonte_cabecalho
            cell.fill = preenchimento_cabecalho
            cell.border = borda_fina
            cell.alignment = alinhamento_centro
        
        sheet.freeze_panes = 'A2'
        
        # Ajusta larguras
        larguras = {
            'ID Extrato': 12,
            'Data': 12,
            'Valor Extrato': 18,
            'Favorecido/Descrição': 40,
            'Status': 18,
            'Qtd Comprovantes': 18,
            'Valor Total Conciliado': 22,
            'Diferença': 15,
            'Ref. Sigra': 20,
            'Categoria': 30,
            'Cliente': 35,  # Coluna F do PGTO MASTER
            'IDs Comprovantes': 30,
            'Observação': 40
        }
        
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value and str(cell.value) in larguras:
                column_letter = get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].width = larguras[str(cell.value)]
        
        # Formata células de dados
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = borda_fina
                
                # Formata valores monetários
                if col_valor_extrato and col_idx == col_valor_extrato:
                    try:
                        if cell.value is not None and cell.value != '':
                            valor = float(cell.value) if cell.value else 0
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = alinhamento_direita
                            cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                    except:
                        pass
                
                elif col_valor_conciliado and col_idx == col_valor_conciliado:
                    try:
                        if cell.value is not None and cell.value != '':
                            valor = float(cell.value) if cell.value else 0
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = alinhamento_direita
                            if valor > 0:
                                cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                    except:
                        pass
                
                elif col_diferenca and col_idx == col_diferenca:
                    try:
                        if cell.value is not None and cell.value != '':
                            valor = float(cell.value) if cell.value else 0
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = alinhamento_direita
                            if valor > 0.01:
                                cell.fill = PatternFill(start_color=COR_NEGATIVO, end_color=COR_NEGATIVO, fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                    except:
                        pass
                
                # Formata coluna de status
                elif col_status and col_idx == col_status:
                    cell.alignment = alinhamento_centro
                    if cell.value:
                        status_text = str(cell.value)
                        if '✅' in status_text or 'Conciliado' in status_text:
                            cell.fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type="solid")
                        elif '❌' in status_text or 'Pendente' in status_text:
                            cell.fill = PatternFill(start_color=COR_PENDENTE, end_color=COR_PENDENTE, fill_type="solid")
                
                # Linhas alternadas
                elif row_idx % 2 == 0:
                    try:
                        if not cell.fill or not cell.fill.start_color:
                            cell.fill = PatternFill(start_color=COR_ALTERNADA, end_color=COR_ALTERNADA, fill_type="solid")
                    except:
                        pass
                
                # Formata datas
                header_cell = sheet.cell(row=1, column=col_idx)
                if header_cell and header_cell.value and 'data' in str(header_cell.value).lower():
                    cell.alignment = alinhamento_centro


def _normalizar_colunas_ids_bb(df):
    """
    Normaliza variações de nomes de colunas de ID para evitar perda de vínculo
    ao ler planilhas de versões anteriores.
    """
    if df is None or len(df) == 0:
        return df

    def canon(nome):
        return ''.join(ch for ch in str(nome).strip().lower() if ch.isalnum())

    rename_map = {}
    for col in df.columns:
        chave = canon(col)
        if chave == 'idextrato':
            rename_map[col] = 'ID_extrato'
        elif chave == 'idcomprovante':
            rename_map[col] = 'ID_comprovante'
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def carregar_dados_existentes_bb(caminho_saida):
    """
    Carrega as abas do arquivo de conciliação BB existente (se existir).
    Usado para completar a planilha com novos dados em vez de substituir.
    """
    if not os.path.exists(caminho_saida):
        return None
    try:
        xl = pd.ExcelFile(caminho_saida, engine='openpyxl')
        result = {}
        for sheet in ['extrato', 'comprovantes', 'conciliacao', 'pendencias_extratos', 'pendencias_comprovantes', 'Status Extrato']:
            if sheet in xl.sheet_names:
                df = pd.read_excel(caminho_saida, sheet_name=sheet, engine='openpyxl')
                df = _normalizar_colunas_ids_bb(df)
                if sheet == 'Status Extrato':
                    result['status_extrato'] = df
                else:
                    result[sheet] = df
        if 'extrato' not in result or len(result.get('extrato', [])) == 0:
            return None
        return result
    except Exception as e:
        print(f"[AVISO] Não foi possível carregar planilha existente: {e}")
        return None


def mesclar_dados_bb(existentes, df_extrato, df_comprovantes, df_conciliacao,
                     df_extratos_pendentes, df_comprovantes_pendentes):
    """
    Mescla dados da execução atual com os já existentes na planilha BB.
    Renumera ID_extrato e ID_comprovante para não colidir.
    Retorna (df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes, df_comprovantes_pendentes).
    """
    df_ext = _normalizar_colunas_ids_bb(existentes.get('extrato'))
    df_comp = _normalizar_colunas_ids_bb(existentes.get('comprovantes'))
    df_conc = _normalizar_colunas_ids_bb(existentes.get('conciliacao'))
    df_pend_ext = _normalizar_colunas_ids_bb(existentes.get('pendencias_extratos'))
    df_pend_comp = _normalizar_colunas_ids_bb(existentes.get('pendencias_comprovantes'))
    df_extrato = _normalizar_colunas_ids_bb(df_extrato)
    df_comprovantes = _normalizar_colunas_ids_bb(df_comprovantes)
    df_conciliacao = _normalizar_colunas_ids_bb(df_conciliacao)
    df_extratos_pendentes = _normalizar_colunas_ids_bb(df_extratos_pendentes)
    df_comprovantes_pendentes = _normalizar_colunas_ids_bb(df_comprovantes_pendentes)
    max_id_extrato = int(df_ext['ID_extrato'].max()) if df_ext is not None and len(df_ext) > 0 and 'ID_extrato' in df_ext.columns else 0
    max_id_comprovante = int(df_comp['ID_comprovante'].max()) if df_comp is not None and len(df_comp) > 0 and 'ID_comprovante' in df_comp.columns else 0
    mapa_extrato = {}
    if len(df_extrato) > 0 and 'ID_extrato' in df_extrato.columns:
        for i, old_id in enumerate(df_extrato['ID_extrato'].values, start=1):
            mapa_extrato[old_id] = max_id_extrato + i
    mapa_comprovante = {}
    if len(df_comprovantes) > 0 and 'ID_comprovante' in df_comprovantes.columns:
        for i, old_id in enumerate(df_comprovantes['ID_comprovante'].values, start=1):
            mapa_comprovante[old_id] = max_id_comprovante + i
    df_extrato_novo = df_extrato.copy()
    df_comprovantes_novo = df_comprovantes.copy()
    if 'ID_extrato' in df_extrato_novo.columns:
        df_extrato_novo['ID_extrato'] = df_extrato_novo['ID_extrato'].map(mapa_extrato).fillna(df_extrato_novo['ID_extrato']).astype(int)
    if 'ID_comprovante' in df_comprovantes_novo.columns:
        df_comprovantes_novo['ID_comprovante'] = df_comprovantes_novo['ID_comprovante'].map(mapa_comprovante).fillna(df_comprovantes_novo['ID_comprovante']).astype(int)
    df_conc_novo = df_conciliacao.copy()
    if len(df_conc_novo) > 0:
        if 'ID_extrato' in df_conc_novo.columns:
            df_conc_novo['ID_extrato'] = df_conc_novo['ID_extrato'].map(mapa_extrato).fillna(df_conc_novo['ID_extrato'])
        if 'ID_comprovante' in df_conc_novo.columns:
            df_conc_novo['ID_comprovante'] = df_conc_novo['ID_comprovante'].map(mapa_comprovante).fillna(df_conc_novo['ID_comprovante'])
    df_pend_ext_novo = df_extratos_pendentes.copy()
    df_pend_comp_novo = df_comprovantes_pendentes.copy()
    if len(df_pend_ext_novo) > 0 and 'ID_extrato' in df_pend_ext_novo.columns:
        df_pend_ext_novo['ID_extrato'] = df_pend_ext_novo['ID_extrato'].map(mapa_extrato).fillna(df_pend_ext_novo['ID_extrato']).astype(int)
    if len(df_pend_comp_novo) > 0 and 'ID_comprovante' in df_pend_comp_novo.columns:
        df_pend_comp_novo['ID_comprovante'] = df_pend_comp_novo['ID_comprovante'].map(mapa_comprovante).fillna(df_pend_comp_novo['ID_comprovante']).astype(int)
    def concat_alinhado(df_antigo, df_novo):
        if df_antigo is None or len(df_antigo) == 0:
            return df_novo.copy() if df_novo is not None else df_novo
        if df_novo is None or len(df_novo) == 0:
            return df_antigo.copy()
        cols = list(df_antigo.columns) + [c for c in df_novo.columns if c not in df_antigo.columns]
        a, b = df_antigo.copy(), df_novo.copy()
        for c in cols:
            if c not in a.columns:
                a[c] = pd.NA
            if c not in b.columns:
                b[c] = pd.NA
        return pd.concat([a[cols], b[cols]], ignore_index=True)
    df_extrato = concat_alinhado(df_ext, df_extrato_novo)
    df_comprovantes = concat_alinhado(df_comp, df_comprovantes_novo)
    df_conciliacao = concat_alinhado(df_conc, df_conc_novo)
    df_extratos_pendentes = concat_alinhado(df_pend_ext, df_pend_ext_novo)
    df_comprovantes_pendentes = concat_alinhado(df_pend_comp, df_pend_comp_novo)
    return (df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes, df_comprovantes_pendentes)


def obter_data_inicio_status_extrato(caminho_saida):
    """
    Lê a aba 'Status Extrato' do arquivo existente e retorna a data mínima
    (dia seguinte ao último dia presente) para conciliar somente extratos novos.
    """
    if not os.path.exists(caminho_saida):
        return None

    try:
        df_status = pd.read_excel(caminho_saida, sheet_name='Status Extrato', engine='openpyxl')
    except Exception as e:
        print(f"[AVISO] Não foi possível ler 'Status Extrato' para calcular incremental: {e}")
        return None

    if df_status is None or len(df_status) == 0:
        return None

    # Procura coluna de data (o script grava com cabeçalho 'Data')
    colunas_possiveis = ['Data', 'data', 'Data Extrato', 'data_original']
    col_data = next((c for c in colunas_possiveis if c in df_status.columns), None)
    if col_data is None:
        return None

    datas = pd.to_datetime(df_status[col_data], errors='coerce', dayfirst=True)
    datas_validas = datas.dropna()
    if datas_validas.empty:
        return None

    ultimo_dia = datas_validas.max()
    # Dia seguinte ao último que já foi processado
    return pd.Timestamp(ultimo_dia).normalize() + pd.Timedelta(days=1)


def gerar_excel_final(df_extrato, df_comprovantes, df_conciliacao, 
                      df_extratos_pendentes, df_comprovantes_pendentes, caminho_saida,
                      data_inicio=None):
    """
    Gera arquivo Excel final com todas as abas para Power BI.
    Se o arquivo já existir, completa (acumula) os dados em vez de substituir.
    
    Args:
        df_extrato: DataFrame do extrato
        df_comprovantes: DataFrame dos comprovantes
        df_conciliacao: DataFrame da conciliação
        df_extratos_pendentes: DataFrame de extratos pendentes
        df_comprovantes_pendentes: DataFrame de comprovantes pendentes
        caminho_saida: Caminho do arquivo de saída
    """
    print("\n" + "="*80)
    print("GERANDO ARQUIVO EXCEL FINAL")
    print("="*80)
    
    # Tenta gerar o arquivo, se estiver aberto, usa nome alternativo
    caminho_final = caminho_saida
    tentativas = 0
    max_tentativas = 5
    
    while tentativas < max_tentativas:
        try:
            # Se o arquivo já existe, completa a planilha (acumula) em vez de substituir
            arquivo_existia = os.path.exists(caminho_final)
            dados_existentes = carregar_dados_existentes_bb(caminho_final)
            status_extrato_existente = None
            if dados_existentes is not None:
                print("[INFO] Planilha existente encontrada. Completando com dados desta execução (não substituindo).")
                status_extrato_existente = dados_existentes.get('status_extrato')
                (df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes,
                 df_comprovantes_pendentes) = mesclar_dados_bb(
                    dados_existentes, df_extrato, df_comprovantes, df_conciliacao,
                    df_extratos_pendentes, df_comprovantes_pendentes
                )
            elif arquivo_existia:
                # Segurança: nunca sobrescreve um arquivo existente se não conseguiu carregar os dados antigos.
                raise RuntimeError(
                    "Arquivo de saída já existe, mas não foi possível carregar os dados anteriores. "
                    "Processo interrompido para evitar perda de histórico de conciliação."
                )

            print(
                f"[INFO] Registros que serão salvos -> "
                f"Extrato: {len(df_extrato)} | Comprovantes: {len(df_comprovantes)} | "
                f"Conciliação: {len(df_conciliacao)} | Status Extrato: {len(df_extrato)}"
            )
            with pd.ExcelWriter(caminho_final, engine='openpyxl') as writer:
                # Aba 1: Extrato (dados originais + ID)
                df_extrato_export = df_extrato.copy()
                df_extrato_export.to_excel(writer, sheet_name='extrato', index=False)
                print("[OK] Aba 'extrato' criada")
                
                # Aba 2: Comprovantes (dados originais + ID)
                df_comprovantes_export = df_comprovantes.copy()
                df_comprovantes_export.to_excel(writer, sheet_name='comprovantes', index=False)
                print("[OK] Aba 'comprovantes' criada")
                
                # Aba 3: Conciliação (tabela ponte)
                if len(df_conciliacao) > 0:
                    df_conciliacao.to_excel(writer, sheet_name='conciliacao', index=False)
                    print("[OK] Aba 'conciliacao' criada")
                else:
                    # Cria aba vazia com estrutura
                    pd.DataFrame(columns=['ID_extrato', 'Data_extrato', 'Valor_extrato', 
                                         'ID_comprovante', 'Data_comprovante', 'Valor_comprovante']).to_excel(
                        writer, sheet_name='conciliacao', index=False)
                    print("[OK] Aba 'conciliacao' criada (vazia)")
                
                # Aba 4: Pendências - Extratos
                if len(df_extratos_pendentes) > 0:
                    df_extratos_pendentes.to_excel(writer, sheet_name='pendencias_extratos', index=False)
                    print("[OK] Aba 'pendencias_extratos' criada")
                else:
                    pd.DataFrame(columns=df_extrato.columns).to_excel(
                        writer, sheet_name='pendencias_extratos', index=False)
                    print("[OK] Aba 'pendencias_extratos' criada (vazia)")
                
                # Aba 5: Pendências - Comprovantes
                if len(df_comprovantes_pendentes) > 0:
                    df_comprovantes_pendentes.to_excel(writer, sheet_name='pendencias_comprovantes', index=False)
                    print("[OK] Aba 'pendencias_comprovantes' criada")
                else:
                    pd.DataFrame(columns=df_comprovantes.columns).to_excel(
                        writer, sheet_name='pendencias_comprovantes', index=False)
                    print("[OK] Aba 'pendencias_comprovantes' criada (vazia)")
                
                # Aba 6: Conciliação Bancária (formato tradicional)
                df_conciliacao_bancaria = criar_aba_conciliacao_bancaria(
                    df_extrato, df_comprovantes, df_conciliacao,
                    df_extratos_pendentes, df_comprovantes_pendentes
                )
                df_conciliacao_bancaria.to_excel(writer, sheet_name='Conciliação Bancária', index=False)
                print("[OK] Aba 'Conciliação Bancária' criada")
                
                # Aba 7: Status Extrato (visão consolidada do que está acontecendo com cada valor)
                df_status_extrato = criar_aba_status_extrato(df_extrato, df_conciliacao, df_extratos_pendentes, df_comprovantes)

                # Modo aditivo: preserva status existente e adiciona apenas linhas novas por data/conteúdo
                if status_extrato_existente is not None and len(status_extrato_existente) > 0:
                    status_antigo = status_extrato_existente.copy()
                    status_novo = df_status_extrato.copy()
                    # Se incremental estiver ativo, só considera candidatos a partir da data de corte
                    if data_inicio is not None and 'Data' in status_novo.columns:
                        data_inicio_ts = pd.Timestamp(data_inicio)
                        datas_novo = pd.to_datetime(status_novo['Data'], errors='coerce', dayfirst=True)
                        status_novo = status_novo[datas_novo >= data_inicio_ts].copy()

                    # Dedup por chave de negócio (independente de ID)
                    chave_cols = [c for c in ['Data', 'Valor Extrato', 'Favorecido/Descrição'] if c in status_antigo.columns and c in status_novo.columns]
                    if chave_cols:
                        antigo_chave = status_antigo[chave_cols].copy()
                        novo_chave = status_novo[chave_cols].copy()
                        for c in chave_cols:
                            if 'data' in c.lower():
                                antigo_chave[c] = pd.to_datetime(antigo_chave[c], errors='coerce', dayfirst=True).dt.strftime('%Y-%m-%d')
                                novo_chave[c] = pd.to_datetime(novo_chave[c], errors='coerce', dayfirst=True).dt.strftime('%Y-%m-%d')
                            elif 'valor' in c.lower():
                                antigo_chave[c] = pd.to_numeric(antigo_chave[c], errors='coerce').fillna(0).round(2).astype(str)
                                novo_chave[c] = pd.to_numeric(novo_chave[c], errors='coerce').fillna(0).round(2).astype(str)
                            else:
                                antigo_chave[c] = antigo_chave[c].fillna('').astype(str).str.strip().str.upper()
                                novo_chave[c] = novo_chave[c].fillna('').astype(str).str.strip().str.upper()

                        set_antigo = set(map(tuple, antigo_chave.values.tolist()))
                        mask_novos = ~novo_chave.apply(lambda r: tuple(r.values.tolist()) in set_antigo, axis=1)
                        apenas_novos = status_novo[mask_novos].copy()
                    else:
                        apenas_novos = status_novo.copy()

                    # Mantém sequência dos IDs no Status Extrato
                    if len(apenas_novos) > 0:
                        if 'ID Extrato' not in status_antigo.columns and 'ID_extrato' in status_antigo.columns:
                            status_antigo = status_antigo.rename(columns={'ID_extrato': 'ID Extrato'})
                        if 'ID Extrato' not in apenas_novos.columns and 'ID_extrato' in apenas_novos.columns:
                            apenas_novos = apenas_novos.rename(columns={'ID_extrato': 'ID Extrato'})

                        if 'ID Extrato' in status_antigo.columns and 'ID Extrato' in apenas_novos.columns:
                            ids_antigos = pd.to_numeric(status_antigo['ID Extrato'], errors='coerce')
                            max_id = int(ids_antigos.max()) if ids_antigos.notna().any() else 0
                            apenas_novos = apenas_novos.reset_index(drop=True)
                            apenas_novos['ID Extrato'] = range(max_id + 1, max_id + 1 + len(apenas_novos))

                    df_status_extrato = pd.concat([status_antigo, apenas_novos], ignore_index=True)
                    print(f"[INFO] Status Extrato em modo aditivo: {len(status_antigo)} existentes + {len(apenas_novos)} novos")

                df_status_extrato.to_excel(writer, sheet_name='Status Extrato', index=False)
                print("[OK] Aba 'Status Extrato' criada")
            
            # Formata o arquivo Excel
            print("\n[Aplicando formatação visual...]")
            from openpyxl import load_workbook
            from zipfile import BadZipFile
            import time

            formatado = False
            max_tentativas_formatacao = 5
            for tentativa_fmt in range(1, max_tentativas_formatacao + 1):
                workbook = None
                try:
                    workbook = load_workbook(caminho_final)
                    formatar_planilha_excel(
                        workbook, df_extrato, df_comprovantes, df_conciliacao,
                        df_extratos_pendentes, df_comprovantes_pendentes
                    )
                    workbook.save(caminho_final)
                    formatado = True
                    print("[OK] Formatação aplicada com sucesso!")
                    break
                except BadZipFile:
                    # Em drive de rede, o arquivo pode demorar a ficar íntegro após escrita.
                    if tentativa_fmt < max_tentativas_formatacao:
                        espera = tentativa_fmt * 1.5
                        print(f"[AVISO] Arquivo ainda não íntegro para formatação (tentativa {tentativa_fmt}/{max_tentativas_formatacao}). Aguardando {espera:.1f}s...")
                        time.sleep(espera)
                    else:
                        print("[AVISO] Não foi possível aplicar formatação visual agora (arquivo ainda em sincronização).")
                        print("[AVISO] O arquivo foi gerado e salvo; apenas a formatação foi ignorada nesta execução.")
                finally:
                    if workbook is not None:
                        try:
                            workbook.close()
                        except Exception:
                            pass
            
            # Se chegou aqui, conseguiu escrever
            break
            
        except PermissionError:
            tentativas += 1
            if tentativas < max_tentativas:
                # Tenta com nome alternativo (adiciona timestamp)
                from datetime import datetime
                base, ext = os.path.splitext(caminho_saida)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                caminho_final = f"{base}_{timestamp}{ext}"
                print(f"[AVISO] Arquivo está aberto. Tentando salvar como: {os.path.basename(caminho_final)}")
            else:
                print(f"\n[ERRO] Não foi possível gerar o arquivo após {max_tentativas} tentativas.")
                print("Por favor, feche o arquivo Excel se estiver aberto e tente novamente.")
                raise
    
    print(f"\n[OK] Arquivo Excel gerado: {caminho_final}")
    return caminho_final


# ============================================================================
# FUNÇÃO DE BUSCA DE ARQUIVOS
# ============================================================================

def buscar_arquivos_bb():
    """
    Busca automaticamente os arquivos do Banco do Brasil na pasta do dia de hoje.
    
    Returns:
        Tupla com (caminho_extrato, caminho_comprovantes)
    """
    # Obtém o diretório do script
    script_dir = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
    
    # Obtém a data de hoje no formato YYYY-MM-DD
    data_hoje = datetime.now().strftime('%Y-%m-%d')
    pasta_base = os.path.join(script_dir, 'downloads')
    pasta_downloads = os.path.join(pasta_base, data_hoje)
    
    # Se pasta do dia não existe, usa a mais recente disponível
    if not os.path.exists(pasta_downloads):
        if os.path.exists(pasta_base):
            pastas = []
            for item in os.listdir(pasta_base):
                caminho_item = os.path.join(pasta_base, item)
                if os.path.isdir(caminho_item) and len(item) == 10 and item.count('-') == 2:
                    try:
                        datetime.strptime(item, '%Y-%m-%d')
                        pastas.append((item, caminho_item))
                    except ValueError:
                        pass
            if pastas:
                pastas.sort(key=lambda x: x[0], reverse=True)
                pasta_downloads = pastas[0][1]
                print(f"[INFO] Pasta do dia não encontrada. Usando pasta mais recente: {pastas[0][0]}")
    
    print(f"\n[INFO] Buscando arquivos na pasta: {pasta_downloads}")
    
    if not os.path.exists(pasta_downloads):
        raise FileNotFoundError(f"Pasta não encontrada: {pasta_downloads}\n"
                               f"Certifique-se de que o script 'acessar_drive.py' foi executado e que há pastas no formato YYYY-MM-DD em downloads/.")
    
    # Lista todos os arquivos Excel na pasta para debug
    todos_arquivos = glob.glob(os.path.join(pasta_downloads, '*.xlsx'))
    todos_arquivos.extend(glob.glob(os.path.join(pasta_downloads, '*.xls')))
    print(f"[DEBUG] Total de arquivos Excel na pasta: {len(todos_arquivos)}")
    if len(todos_arquivos) > 0:
        print(f"[DEBUG] Arquivos encontrados:")
        for arquivo in todos_arquivos:
            print(f"  - {os.path.basename(arquivo)}")
    
    # Busca arquivo do extrato (contém "extrato bb" ou "bb" no nome, mas não "pgtos")
    # Busca com diferentes variações de maiúsculas/minúsculas e espaços
    arquivos_extrato = glob.glob(os.path.join(pasta_downloads, '*extrato*bb*.xlsx'))
    arquivos_extrato.extend(glob.glob(os.path.join(pasta_downloads, '*EXTRATO*BB*.xlsx')))
    arquivos_extrato.extend(glob.glob(os.path.join(pasta_downloads, '*Extrato*BB*.xlsx')))
    arquivos_extrato.extend(glob.glob(os.path.join(pasta_downloads, '*extrato*BB*.xlsx')))
    arquivos_extrato.extend(glob.glob(os.path.join(pasta_downloads, '*EXTRATO*bb*.xlsx')))
    
    # Também busca apenas "bb" mas exclui arquivos com "pgtos" ou "comprovantes"
    arquivos_bb = glob.glob(os.path.join(pasta_downloads, '*bb*.xlsx'))
    arquivos_bb.extend(glob.glob(os.path.join(pasta_downloads, '*BB*.xlsx')))
    # Remove arquivos que contêm "pgtos" ou "comprovantes" (esses são comprovantes)
    arquivos_bb = [f for f in arquivos_bb if 'pgtos' not in os.path.basename(f).lower() and 'comprovantes' not in os.path.basename(f).lower()]
    arquivos_extrato.extend(arquivos_bb)
    
    # Remove duplicatas
    arquivos_extrato = list(set(arquivos_extrato))
    
    # Busca arquivo de comprovantes (aceita "pgto" e "pgtos" no nome)
    arquivos_comprovantes = glob.glob(os.path.join(pasta_downloads, '*pgto*.xlsx'))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*PGTO*.xlsx')))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*Pgto*.xlsx')))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*pgtos*.xlsx')))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*PGTOS*.xlsx')))
    arquivos_comprovantes.extend(glob.glob(os.path.join(pasta_downloads, '*Pgtos*.xlsx')))
    
    # Remove duplicatas (case-insensitive)
    arquivos_extrato = list(set(arquivos_extrato))
    arquivos_comprovantes = list(set(arquivos_comprovantes))
    
    print(f"[DEBUG] Arquivos de extrato encontrados: {len(arquivos_extrato)}")
    for arquivo in arquivos_extrato:
        print(f"  - {os.path.basename(arquivo)}")
    
    print(f"[DEBUG] Arquivos de comprovantes encontrados: {len(arquivos_comprovantes)}")
    for arquivo in arquivos_comprovantes:
        print(f"  - {os.path.basename(arquivo)}")
    
    # Verifica se encontrou os arquivos
    if not arquivos_extrato:
        raise FileNotFoundError(f"Nenhum arquivo de extrato BB encontrado na pasta {pasta_downloads}\n"
                               f"Procurei por arquivos contendo 'extrato bb' ou 'bb' (sem 'pgto/pgtos') no nome.\n"
                               f"Arquivos disponíveis: {[os.path.basename(f) for f in todos_arquivos]}")
    
    if not arquivos_comprovantes:
        raise FileNotFoundError(f"Nenhum arquivo de comprovantes 'pgto/pgtos' encontrado na pasta {pasta_downloads}\n"
                               f"Procurei por arquivos contendo 'pgto' ou 'pgtos' no nome.\n"
                               f"Arquivos disponíveis: {[os.path.basename(f) for f in todos_arquivos]}")
    
    # Se encontrou múltiplos arquivos, usa o primeiro (ou pode implementar lógica mais sofisticada)
    caminho_extrato = arquivos_extrato[0]
    caminho_comprovantes = arquivos_comprovantes[0]
    
    if len(arquivos_extrato) > 1:
        print(f"[AVISO] Múltiplos arquivos de extrato encontrados. Usando: {os.path.basename(caminho_extrato)}")
    
    if len(arquivos_comprovantes) > 1:
        print(f"[AVISO] Múltiplos arquivos de comprovantes encontrados. Usando: {os.path.basename(caminho_comprovantes)}")
    
    print(f"[OK] Extrato encontrado: {os.path.basename(caminho_extrato)}")
    print(f"[OK] Comprovantes encontrados: {os.path.basename(caminho_comprovantes)}")
    
    return caminho_extrato, caminho_comprovantes


# ============================================================================
# FUNÇÃO PRINCIPAL
# ============================================================================

def conciliar_bb(caminho_extrato=None, caminho_comprovantes=None, data_inicio=None):
    """
    Executa a conciliação do Banco do Brasil e retorna os DataFrames.
    
    Args:
        caminho_extrato: Caminho do arquivo de extrato (se None, busca automaticamente)
        caminho_comprovantes: Caminho do arquivo de comprovantes (se None, busca automaticamente)
    
    Returns:
        Tupla com (df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes, df_comprovantes_pendentes)
    """
    # Se não foram fornecidos caminhos, busca automaticamente
    if caminho_extrato is None or caminho_comprovantes is None:
        caminho_extrato, caminho_comprovantes = buscar_arquivos_bb()
    
    # 1. Ler arquivos
    print("\n[BB] Lendo arquivos...")
    df_extrato_raw = ler_extrato(caminho_extrato)
    df_comprovantes_raw = ler_comprovantes(caminho_comprovantes)
    
    # 2. Preparar e normalizar dados
    print("\n[BB] Preparando e normalizando dados...")
    df_extrato = preparar_dados(df_extrato_raw, tipo='extrato')
    df_comprovantes = preparar_dados(df_comprovantes_raw, tipo='comprovantes')

    # Se existe histórico com último dia processado, filtra somente extratos novos
    if data_inicio is not None and 'data' in df_extrato.columns:
        df_extrato = df_extrato[df_extrato['data'] >= pd.Timestamp(data_inicio)].copy()
        print(f"[BB] Incremental ativo: filtrando extrato >= {pd.Timestamp(data_inicio).date()} -> {len(df_extrato)} linhas")
    
    # 3. Conciliação
    print("\n[BB] Realizando conciliação...")
    df_conciliacao, comprovantes_usados = conciliar_extrato_comprovantes(
        df_extrato, df_comprovantes
    )
    
    # 4. Identificar pendências
    print("\n[BB] Identificando pendências...")
    df_extratos_pendentes, df_comprovantes_pendentes = identificar_pendencias(
        df_extrato, df_comprovantes, df_conciliacao, comprovantes_usados
    )
    
    print(f"\n[BB] Extratos pendentes: {len(df_extratos_pendentes)}")
    print(f"[BB] Comprovantes pendentes: {len(df_comprovantes_pendentes)}")
    
    return df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes, df_comprovantes_pendentes


def main():
    """
    Função principal do script.
    """
    print("="*80)
    print("CONCILIAÇÃO DE EXTRATOS BANCÁRIOS - BANCO DO BRASIL")
    print("="*80)
    print(f"Data/Hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80)
    
    try:
        # Modo incremental: concilia somente a partir do dia seguinte ao último dia presente em 'Status Extrato'
        data_inicio = obter_data_inicio_status_extrato(CAMINHO_SAIDA)

        # Executa conciliação
        df_extrato, df_comprovantes, df_conciliacao, df_extratos_pendentes, df_comprovantes_pendentes = conciliar_bb(
            data_inicio=data_inicio
        )
        
        # 5. Gerar Excel final
        print("\n[5/5] Gerando arquivo Excel final...")
        caminho_excel_gerado = gerar_excel_final(
            df_extrato, df_comprovantes, df_conciliacao,
            df_extratos_pendentes, df_comprovantes_pendentes,
            CAMINHO_SAIDA,
            data_inicio=data_inicio
        )
        
        # Resumo final
        print("\n" + "="*80)
        print("RESUMO FINAL")
        print("="*80)
        print(f"Total de extratos: {len(df_extrato)}")
        print(f"Total de comprovantes: {len(df_comprovantes)}")
        extratos_conciliados = len(df_conciliacao['ID_extrato'].unique()) if len(df_conciliacao) > 0 else 0
        print(f"Conciliações encontradas: {len(df_conciliacao)} relacionamentos")
        print(f"Extratos conciliados: {extratos_conciliados}")
        print(f"Extratos pendentes: {len(df_extratos_pendentes)}")
        print(f"Comprovantes pendentes: {len(df_comprovantes_pendentes)}")
        if len(df_extrato) > 0:
            taxa = extratos_conciliados / len(df_extrato) * 100
            print(f"Taxa de conciliação: {taxa:.1f}%")
        else:
            print("Taxa de conciliação: N/A (nenhum extrato válido)")
        print("="*80)
        print("\n[OK] Processo concluído com sucesso!")
        print(f"Arquivo Excel gerado: {caminho_excel_gerado}")
        
    except Exception as e:
        print(f"\n[ERRO] Erro durante execução: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == '__main__':
    main()
