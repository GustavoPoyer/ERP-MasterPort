"""
Numerário Itaú: leitura do Excel e conciliação com extratos pendentes (após SIGRA).

Enriquecimento opcional:
- Coluna L "Mensagem (Email)": extrai REF cliente entre "para ao processo" e "no total de".
- Opcionalmente consulta o Sigra (Selenium) e preenche "ref sigra" (ID na URL).

Credenciais Sigra: SIGRA_EMAIL / SIGRA_PASSWORD no ambiente sobrepõem os valores padrão no código.
Não partilhe o repositório publicamente com senha em claro.
"""

from __future__ import annotations

import glob
import os
import re
import time
import unicodedata
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd

# Numerário (layout comum: aba "acompanhamento")
COLUNAS_NUMERARIO = {
    'data': 'Criação',
    'valor': 'Valor',
    'favorecido': 'Empresa',
    'descricao': 'Assunto (Email)',
    'mensagem': 'Mensagem (Email)',
    'id': 'ID',
}

COL_REF_CLIENTE = 'REF cliente'
COL_REF_SIGRA = 'ref sigra'

# Cabeçalhos do export de numerário (aba acompanhamento) — usados se o .xlsx vier truncado.
CABECALHOS_PLANILHA_NUMERARIO = [
    'ID',
    'Empresa',
    'CNPJ',
    'Moeda',
    'Valor',
    'Criação',
    'Vencimento',
    'Pagamento',
    'Conta Contábil',
    'Assunto (Email)',
    'Destinatários (Email)',
    'Mensagem (Email)',
]


def xlsx_openxml_e_valido(caminho: str) -> bool:
    """
    True se o ficheiro parece um .xlsx Office Open XML completo (ZIP com workbook + folha).
    """
    caminho = os.path.abspath(caminho)
    if not os.path.isfile(caminho):
        return False
    if os.path.getsize(caminho) < 500:
        return False
    try:
        with zipfile.ZipFile(caminho, 'r') as zf:
            nomes = set(zf.namelist())
    except zipfile.BadZipFile:
        return False
    tem_ct = '[Content_Types].xml' in nomes
    tem_wb = 'xl/workbook.xml' in nomes
    tem_folha = any(
        n.startswith('xl/worksheets/') and n.endswith('.xml')
        for n in nomes
    )
    return bool(tem_ct and tem_wb and tem_folha)


def garantir_xlsx_openxml_valido(caminho: str) -> None:
    """
    Um .xlsx válido é um ZIP com [Content_Types].xml e xl/workbook.xml.
    Arquivos truncados ou export errado geram KeyError no openpyxl sem mensagem clara.
    """
    caminho = os.path.abspath(caminho)
    if not os.path.isfile(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
    tamanho = os.path.getsize(caminho)
    if tamanho < 500:
        raise ValueError(
            f"Arquivo muito pequeno ({tamanho} bytes) para ser um .xlsx com dados: "
            f"{os.path.basename(caminho)}"
        )
    try:
        with zipfile.ZipFile(caminho, 'r') as zf:
            nomes = set(zf.namelist())
    except zipfile.BadZipFile as e:
        raise ValueError(
            f"O arquivo não é um ZIP válido (não é .xlsx íntegro): {os.path.basename(caminho)}"
        ) from e

    tem_ct = '[Content_Types].xml' in nomes
    tem_wb = 'xl/workbook.xml' in nomes
    tem_folha = any(
        n.startswith('xl/worksheets/') and n.endswith('.xml')
        for n in nomes
    )
    if not (tem_ct and tem_wb and tem_folha):
        raise ValueError(
            f"Excel incompleto ou corrompido: {os.path.basename(caminho)}. "
            f"Faltam partes obrigatórias do formato .xlsx (planilha/workbook). "
            f"Isso costuma acontecer com download interrompido ou exportação incorreta. "
            f"Solução: abra o arquivo no Excel (se abrir), use 'Salvar como' .xlsx, ou exporte/baixe de novo."
        )


def garantir_numerario_xlsx_usavel(caminho: str) -> None:
    """
    Garante que o ficheiro pode ser lido pelo openpyxl/pandas.

    Vários exports (ou gravação interrompida) geram ZIP com theme/styles mas sem
    xl/workbook.xml / folhas — o Excel até quebra ao abrir. Nesse caso fazemos
    backup do ficheiro e recriamos um .xlsx válido com cabeçalhos padrão.
    """
    caminho = os.path.abspath(caminho)
    if not os.path.isfile(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
    if xlsx_openxml_e_valido(caminho):
        return

    import shutil
    from openpyxl import Workbook

    bak = caminho + '.corrupto.bak'
    try:
        shutil.copy2(caminho, bak)
    except OSError:
        bak = None

    wb = Workbook()
    ws = wb.active
    ws.title = 'acompanhamento'
    ws.append(CABECALHOS_PLANILHA_NUMERARIO)
    wb.save(caminho)

    print(
        f"[AVISO] O ficheiro estava incompleto (ZIP sem workbook/planilhas - comum em export a meio). "
        f"Foi recriado um Excel válido com cabeçalhos na aba 'acompanhamento'."
    )
    if bak:
        print(f"        Cópia do ficheiro antigo: {os.path.basename(bak)}")
    print(
        "        Se tinha dados na origem, volte a colá-los neste ficheiro e guarde no Excel "
        "(Ficheiro -> Guardar) antes de correr de novo o enriquecimento."
    )

SIGRA_LOGIN_URL = 'https://app.sigraweb.com/#/login'
# XPath informado pelo usuário para o campo de busca (header)
XPATH_BUSCA_SIGRA = '/html/body/div[1]/header/div[1]/div[1]/form/input'

# Login Sigra (variáveis de ambiente SIGRA_EMAIL / SIGRA_PASSWORD)
SIGRA_EMAIL_PADRAO = ''
SIGRA_SENHA_PADRAO = ''


def _resolver_credenciais_sigra(
    email: Optional[str] = None,
    password: Optional[str] = None,
) -> tuple[str, str]:
    e = (email or os.environ.get('SIGRA_EMAIL', '').strip() or SIGRA_EMAIL_PADRAO).strip()
    p = (password or os.environ.get('SIGRA_PASSWORD', '').strip() or SIGRA_SENHA_PADRAO).strip()
    return e, p

# Textos reais variam: "para ao processo", "para o processo", "para o processo de", etc.
_PADROES_REF_CLIENTE_MENSAGEM = [
    re.compile(
        r'para\s+ao\s+processo(?:\s+de)?\s+(.+?)\s+no\s+total\s+de',
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        r'para\s+o\s+processo(?:\s+de)?\s+(.+?)\s+no\s+total\s+de',
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        r'pelo\s+processo(?:\s+de)?\s+(.+?)\s+no\s+total\s+de',
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        r'(?:para|pelo)\s+ao\s+processo(?:\s+de)?\s+(.+?)\s+no\s+valor\s+total',
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        r'(?:para|pelo)\s+o\s+processo(?:\s+de)?\s+(.+?)\s+no\s+valor\s+total',
        re.IGNORECASE | re.DOTALL,
    ),
]
# Último recurso: "processo XYZ ... no total de" (sem "para")
_RE_REF_PROCESSO_NO_TOTAL = re.compile(
    r'processo\s+(.+?)\s+no\s+total\s+de',
    re.IGNORECASE | re.DOTALL,
)


def _formatar_valor_br(valor):
    """Formata valor monetário no padrão brasileiro (ex.: 4.601,70)."""
    if pd.isna(valor) or valor == 0:
        return "0,00"
    valor_float = round(float(valor), 2)
    parte_inteira = int(abs(valor_float))
    parte_decimal = abs(valor_float) - parte_inteira
    centavos = int(round(parte_decimal * 100))
    parte_inteira_str = f"{parte_inteira:,}".replace(",", ".")
    parte_decimal_str = f"{centavos:02d}"
    return f"{parte_inteira_str},{parte_decimal_str}"


def _valores_equivalentes_centavos(valor_extrato_int, valor_comp_int):
    if valor_comp_int is None:
        return False
    return valor_comp_int == valor_extrato_int


def _normalizar_chave_coluna(nome: str) -> str:
    s = unicodedata.normalize('NFKD', str(nome).strip().lower())
    return ''.join(c for c in s if not unicodedata.combining(c))


def _achar_coluna_mensagem(columns) -> Optional[str]:
    """Nome real da coluna L / mensagem no DataFrame."""
    for c in columns:
        cl = _normalizar_chave_coluna(c)
        if 'mensagem' in cl and 'email' in cl:
            return c
    return None


def extrair_ref_cliente_mensagem(texto: Any) -> str:
    """
    Extrai o número do processo do texto da coluna Mensagem (Email).
    Trecho esperado: ... para ao processo MBS-018-2026 no total de ...
    """
    if pd.isna(texto) or texto is None:
        return ''
    s = str(texto).replace('\r\n', '\n').strip()
    if not s:
        return ''
    for rx in _PADROES_REF_CLIENTE_MENSAGEM:
        m = rx.search(s)
        if m:
            ref = m.group(1).strip()
            ref = re.sub(r'\s+', ' ', ref)
            ref = ref.strip(' ,.;:\n\t')
            if ref:
                return ref
    m2 = _RE_REF_PROCESSO_NO_TOTAL.search(s)
    if m2:
        ref = m2.group(1).strip()
        ref = re.sub(r'\s+', ' ', ref)
        ref = ref.strip(' ,.;:\n\t')
        if ref and len(ref) <= 200:
            return ref
    return ''


def enriquecer_ref_cliente_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Garante coluna REF cliente preenchida a partir de Mensagem (Email), se existir.
    Não sobrescreve REF cliente já preenchida (exceto se for só espaço).
    """
    out = df.copy()
    col_msg = _achar_coluna_mensagem(out.columns)
    if col_msg is None:
        print("[AVISO] Coluna 'Mensagem (Email)' não encontrada; REF cliente não derivada do texto.")
        if COL_REF_CLIENTE not in out.columns:
            out[COL_REF_CLIENTE] = ''
        return out

    if COL_REF_CLIENTE not in out.columns:
        out[COL_REF_CLIENTE] = ''

    preenchidos_antes = sum(
        1 for i in out.index if pd.notna(out.at[i, COL_REF_CLIENTE]) and str(out.at[i, COL_REF_CLIENTE]).strip()
    )
    for i in out.index:
        atual = out.at[i, COL_REF_CLIENTE]
        if pd.notna(atual) and str(atual).strip():
            continue
        ref = extrair_ref_cliente_mensagem(out.at[i, col_msg])
        out.at[i, COL_REF_CLIENTE] = ref

    preenchidos_depois = sum(
        1 for i in out.index if pd.notna(out.at[i, COL_REF_CLIENTE]) and str(out.at[i, COL_REF_CLIENTE]).strip()
    )
    if preenchidos_depois > preenchidos_antes:
        print(
            f"[INFO] REF cliente: {preenchidos_depois - preenchidos_antes} linha(s) preenchida(s) a partir da Mensagem "
            f"(total com REF: {preenchidos_depois}/{len(out)})."
        )

    return out


def extrair_processo_numerario(row) -> str:
    """
    Identificador do processo para conciliação:
    1) Mensagem (Email): padrão "para ao processo ... no total de"
    2) Assunto (Email): "processo XYZ"
    3) Texto inteiro do assunto como fallback
    """
    col_msg = COLUNAS_NUMERARIO.get('mensagem', 'Mensagem (Email)')
    msg = row.get(col_msg, row.get('Mensagem (Email)', ''))
    ref = extrair_ref_cliente_mensagem(msg)
    if ref:
        return ref

    col_desc = COLUNAS_NUMERARIO.get('descricao', 'Assunto (Email)')
    assunto = row.get(col_desc, row.get('Assunto (Email)', ''))
    if pd.isna(assunto) or not str(assunto).strip():
        return '-'
    s = str(assunto).strip()
    match = re.search(r'processo\s+([^\s,;]+)', s, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return s


def extrair_id_sigra_da_url(url: str) -> Optional[str]:
    """Ex.: https://app.sigraweb.com/#/importacao/1793905 -> 1793905"""
    if not url:
        return None
    m = re.search(r'#/(?:importacao|exportacao)/(\d+)', url, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r'/(?:importacao|exportacao)/(\d+)', url, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def _criar_driver_chrome_sigra(headless: bool = False):
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    opts = Options()
    if headless:
        opts.add_argument('--headless=new')
    opts.add_argument('--disable-gpu')
    opts.add_argument('--window-size=1400,900')
    opts.add_argument('--no-sandbox')
    opts.add_argument('--disable-dev-shm-usage')
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


def _driver_sessao_valida(driver) -> bool:
    """False se o browser foi fechado ou a sessao WebDriver expirou."""
    from selenium.common.exceptions import InvalidSessionIdException, WebDriverException

    try:
        _ = driver.session_id
        _ = driver.current_url
        return True
    except (InvalidSessionIdException, WebDriverException):
        return False


def _fechar_driver_seguro(driver) -> None:
    if driver is None:
        return
    try:
        driver.quit()
    except Exception:
        pass


def _e_erro_sessao_invalida(exc: BaseException) -> bool:
    from selenium.common.exceptions import InvalidSessionIdException

    if isinstance(exc, InvalidSessionIdException):
        return True
    return 'invalid session id' in str(exc).lower()


def _voltar_tela_busca_sigra(driver, timeout: int = 20) -> None:
    """Volta para uma rota onde o campo de busca do header existe (apos abrir um processo)."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    if not _driver_sessao_valida(driver):
        return
    wait = WebDriverWait(driver, timeout)
    for url in (
        'https://app.sigraweb.com/#/board',
        'https://app.sigraweb.com/#/',
        'https://app.sigraweb.com/',
    ):
        try:
            driver.get(url)
            time.sleep(0.8)
            wait.until(EC.presence_of_element_located((By.XPATH, XPATH_BUSCA_SIGRA)))
            return
        except Exception:
            continue


def _login_sigra_selenium(driver, email: str, password: str, timeout: int = 45) -> None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    wait = WebDriverWait(driver, timeout)
    driver.get(SIGRA_LOGIN_URL)
    time.sleep(1.5)

    email_el = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email']"))
    )
    pwd_el = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
    email_el.clear()
    email_el.send_keys(email)
    pwd_el.clear()
    pwd_el.send_keys(password)

    clicou = False
    for btn in driver.find_elements(By.TAG_NAME, 'button'):
        try:
            t = (btn.text or '').strip().lower()
            if 'entrar' in t:
                btn.click()
                clicou = True
                break
        except Exception:
            continue
    if not clicou:
        for el in driver.find_elements(By.XPATH, "//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'entrar agora')]"):
            try:
                el.click()
                clicou = True
                break
            except Exception:
                continue
    if not clicou:
        raise RuntimeError('Botão "Entrar" / "Entrar agora" não encontrado na página de login.')

    wait.until(
        EC.presence_of_element_located((By.XPATH, XPATH_BUSCA_SIGRA))
    )


def _pesquisar_e_abrir_processo_sigra(driver, ref_cliente: str, timeout: int = 25) -> Optional[str]:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    if not ref_cliente or not str(ref_cliente).strip():
        return None

    ref = str(ref_cliente).strip()
    wait = WebDriverWait(driver, timeout)
    inp = wait.until(EC.element_to_be_clickable((By.XPATH, XPATH_BUSCA_SIGRA)))
    inp.clear()
    time.sleep(0.2)
    inp.send_keys(ref)
    time.sleep(0.4)
    inp.send_keys(Keys.ENTER)

    time.sleep(1.0)

    links = driver.find_elements(
        By.CSS_SELECTOR,
        "a[href*='importacao/'], a[href*='exportacao/'], a[href*='#/importacao/'], a[href*='#/exportacao/']",
    )
    alvo = None
    for a in links:
        try:
            texto = (a.text or '').strip()
            href = a.get_attribute('href') or ''
            if ref in texto or ref in href:
                alvo = a
                break
        except Exception:
            continue
    if alvo is None and links:
        alvo = links[0]
    if alvo is None:
        try:
            alvo = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, ref)))
        except Exception:
            return None

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
        time.sleep(0.2)
        alvo.click()
    except Exception:
        driver.execute_script("arguments[0].click();", alvo)

    wait.until(lambda d: extrair_id_sigra_da_url(d.current_url) is not None)
    sid = extrair_id_sigra_da_url(driver.current_url)
    _voltar_tela_busca_sigra(driver)
    return sid


def enriquecer_numerario_com_sigra(
    df: pd.DataFrame,
    email: Optional[str] = None,
    password: Optional[str] = None,
    headless: bool = False,
    pausa_entre_linhas: float = 1.2,
) -> pd.DataFrame:
    """
    Preenche coluna ref sigra via busca no Sigra (uma sessão, reutiliza IDs por REF cliente).

    Requer: pip install selenium, Chrome instalado, webdriver-manager (recomendado).
    Credenciais: parâmetros email/password, ou SIGRA_* no ambiente, ou SIGRA_EMAIL_PADRAO / SIGRA_SENHA_PADRAO.
    """
    email, password = _resolver_credenciais_sigra(email=email, password=password)
    if not email or not password:
        raise ValueError(
            'Credenciais Sigra em falta: preencha SIGRA_EMAIL_PADRAO / SIGRA_SENHA_PADRAO no código '
            'ou defina SIGRA_EMAIL / SIGRA_PASSWORD no ambiente.'
        )

    out = enriquecer_ref_cliente_dataframe(df)
    if COL_REF_SIGRA not in out.columns:
        out[COL_REF_SIGRA] = ''

    driver = _criar_driver_chrome_sigra(headless=headless)
    cache: Dict[str, str] = {}

    def reiniciar_sessao(motivo: str) -> None:
        nonlocal driver
        print(f"  [AVISO] {motivo} A reiniciar Chrome e fazer login de novo...")
        _fechar_driver_seguro(driver)
        driver = _criar_driver_chrome_sigra(headless=headless)
        _login_sigra_selenium(driver, email, password)

    try:
        _login_sigra_selenium(driver, email, password)
        for i in out.index:
            ref = out.at[i, COL_REF_CLIENTE]
            if pd.isna(ref) or not str(ref).strip():
                continue
            ref_s = str(ref).strip()
            atual_sigra = out.at[i, COL_REF_SIGRA]
            if pd.notna(atual_sigra) and str(atual_sigra).strip():
                continue

            if ref_s in cache:
                out.at[i, COL_REF_SIGRA] = cache[ref_s]
                print(f"  [cache] {ref_s} -> {cache[ref_s]}")
                continue

            if not _driver_sessao_valida(driver):
                reiniciar_sessao('Sessao do browser invalida.')

            sid = None
            ultimo_erro: Optional[BaseException] = None
            for tentativa in range(2):
                try:
                    sid = _pesquisar_e_abrir_processo_sigra(driver, ref_s)
                    ultimo_erro = None
                    break
                except Exception as ex:
                    ultimo_erro = ex
                    if _e_erro_sessao_invalida(ex) and tentativa == 0:
                        reiniciar_sessao('Sessao perdida durante a busca.')
                        continue
                    break

            if ultimo_erro is not None and sid is None:
                print(f"  [ERRO] Sigra falhou para '{ref_s}': {ultimo_erro}")
            elif sid:
                out.at[i, COL_REF_SIGRA] = sid
                cache[ref_s] = sid
                print(f"  [OK] Sigra {ref_s} -> ref sigra {sid}")
            else:
                print(f"  [AVISO] Sem ID na URL para REF cliente: {ref_s}")

            time.sleep(pausa_entre_linhas)
    finally:
        _fechar_driver_seguro(driver)

    return out


def ler_numerario(caminho):
    """
    Lê o arquivo Excel do numerário (aba "acompanhamento").
    """
    try:
        garantir_numerario_xlsx_usavel(caminho)
        xl = pd.ExcelFile(caminho, engine='openpyxl')
        sheet_name = 'acompanhamento' if 'acompanhamento' in xl.sheet_names else 0
        df = pd.read_excel(caminho, sheet_name=sheet_name, engine='openpyxl')
        print(
            f"[INFO] Arquivo numerário lido: {len(df)} linhas"
            + (" (aba 'acompanhamento')" if sheet_name == 'acompanhamento' else "")
        )

        if any('Unnamed' in str(col) for col in df.columns):
            for skip_rows in range(0, 10):
                try:
                    df_teste = pd.read_excel(caminho, sheet_name=sheet_name, skiprows=skip_rows, engine='openpyxl')
                    if not any('Unnamed' in str(col) for col in df_teste.columns[:3]):
                        df = df_teste
                        print(f"[INFO] Cabeçalho encontrado na linha {skip_rows + 1}")
                        break
                except Exception:
                    continue

        df = df.dropna(how='all')
        df = enriquecer_ref_cliente_dataframe(df)
        print(f"[OK] Numerário carregado: {len(df)} registros (REF cliente derivada da Mensagem quando aplicável)")
        return df
    except Exception as e:
        print(f"[ERRO] Erro ao ler numerário: {e}")
        raise


def salvar_numerario_xlsx(df: pd.DataFrame, caminho: str, sheet_name: str = 'acompanhamento') -> None:
    """
    Grava o DataFrame na aba indicada (substitui só essa aba se o arquivo já existir).

    Não pode remover a aba e gravar o workbook a seguir: se era a única folha, o openpyxl
    fica sem folhas visíveis e falha com "At least one sheet must be visible".
    """
    caminho = os.path.abspath(caminho)
    d = os.path.dirname(caminho)
    if d:
        os.makedirs(d, exist_ok=True)

    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if not os.path.isfile(caminho):
        with pd.ExcelWriter(caminho, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    try:
        wb = load_workbook(caminho)
    except PermissionError as e:
        raise PermissionError(
            f"Não foi possível abrir '{os.path.basename(caminho)}'. "
            f"Feche o ficheiro no Excel (ou outro programa) e tente de novo."
        ) from e

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for s in wb.worksheets:
        s.sheet_state = 'visible'
    wb.active = wb[sheet_name]

    try:
        wb.save(caminho)
    except PermissionError as e:
        raise PermissionError(
            f"Não foi possível gravar '{os.path.basename(caminho)}'. "
            f"Feche o ficheiro no Excel e tente de novo."
        ) from e
    finally:
        wb.close()


def diretorio_numerario_itau() -> str:
    """Pasta onde está numerario_itau.py (Extratos-consolidados)."""
    return os.path.dirname(os.path.abspath(__file__))


def resolver_pasta_downloads_dia(data_referencia: Optional[str] = None) -> str:
    """
    Resolve .../Extratos-consolidados/downloads/AAAA-MM-DD.

    - Sem argumento: pasta do dia de hoje; se não existir, a data mais recente
      em downloads (mesma lógica de buscar_arquivos_itau_sigra).
    - Com data_referencia='2026-04-13': exige que essa pasta exista.
    """
    script_dir = diretorio_numerario_itau()
    pasta_base = os.path.join(script_dir, 'downloads')

    if data_referencia:
        pasta = os.path.join(pasta_base, data_referencia)
        if not os.path.isdir(pasta):
            raise FileNotFoundError(
                f"Pasta downloads do dia não encontrada: {pasta}"
            )
        return pasta

    data_hoje = datetime.now().strftime('%Y-%m-%d')
    pasta_hoje = os.path.join(pasta_base, data_hoje)
    if os.path.isdir(pasta_hoje):
        return pasta_hoje

    if not os.path.isdir(pasta_base):
        raise FileNotFoundError(f"Pasta base downloads não existe: {pasta_base}")

    pastas: List[str] = []
    for item in os.listdir(pasta_base):
        caminho_item = os.path.join(pasta_base, item)
        if os.path.isdir(caminho_item) and len(item) == 10 and item.count('-') == 2:
            try:
                datetime.strptime(item, '%Y-%m-%d')
                pastas.append(item)
            except ValueError:
                pass

    if pastas:
        pastas.sort()
        mais = pastas[-1]
        print(f"[AVISO] Pasta de hoje ({data_hoje}) não encontrada. Usando downloads/{mais}")
        return os.path.join(pasta_base, mais)

    raise FileNotFoundError(f"Nenhuma pasta AAAA-MM-DD em: {pasta_base}")


def listar_arquivos_numerario_xlsx(pasta: str) -> List[str]:
    """Lista .xlsx de numerário na pasta (mais recente primeiro)."""
    padroes = [
        os.path.join(pasta, '*numerario*.xlsx'),
        os.path.join(pasta, '*NUMERARIO*.xlsx'),
        os.path.join(pasta, '*numerário*.xlsx'),
        os.path.join(pasta, '*NUMERÁRIO*.xlsx'),
    ]
    found: List[str] = []
    for pat in padroes:
        found.extend(glob.glob(pat))
    unicos = {f for f in found if not os.path.basename(f).startswith('~$')}
    return sorted(unicos, key=lambda p: os.path.getmtime(p), reverse=True)


def localizar_arquivo_numerario_pasta_downloads(data_referencia: Optional[str] = None) -> str:
    """
    Escolhe .xlsx de numerário na pasta do dia (mais recente primeiro).
    Ignora ficheiros .xlsx incompletos/corrompidos e tenta o próximo candidato.
    """
    pasta = resolver_pasta_downloads_dia(data_referencia)
    print(f"[INFO] Pasta downloads: {pasta}")
    arqs = listar_arquivos_numerario_xlsx(pasta)
    if not arqs:
        raise FileNotFoundError(
            f"Nenhum .xlsx de numerário em {pasta} (ex.: *numerario*.xlsx)."
        )
    ignorados: List[str] = []
    for caminho in arqs:
        if xlsx_openxml_e_valido(caminho):
            if ignorados:
                print(
                    f"[AVISO] Ignorado(s) (Excel incompleto/corrompido): {', '.join(ignorados)}"
                )
            if caminho != arqs[0]:
                print(
                    f"[INFO] Usando primeiro ficheiro válido: {os.path.basename(caminho)} "
                    f"(o mais recente na pasta não era um .xlsx íntegro)"
                )
            else:
                print(f"[OK] Arquivo: {os.path.basename(caminho)}")
            return caminho
        ignorados.append(os.path.basename(caminho))

    # Nenhum íntegro: repara só o mais recente (evita estragar vários ficheiros)
    principal = arqs[0]
    print(
        f"[AVISO] Nenhum numerário .xlsx íntegro (tentados: {', '.join(ignorados)}). "
        f"A reparar: {os.path.basename(principal)}"
    )
    garantir_numerario_xlsx_usavel(principal)
    print(f"[OK] Arquivo: {os.path.basename(principal)}")
    return principal


def resolver_caminho_arquivo_numerario(
    arquivo: Optional[str],
    data_referencia: Optional[str] = None,
) -> str:
    """
    - arquivo None ou vazio: localiza automaticamente na pasta downloads do dia.
    - só nome (ex. Modelo.xlsx): busca dentro de downloads/AAAA-MM-DD.
    - caminho absoluto ou relativo válido: usa o arquivo informado.
    """
    if arquivo and os.path.isabs(arquivo):
        p = os.path.abspath(arquivo)
        if not os.path.isfile(p):
            raise FileNotFoundError(f"Arquivo não encontrado: {p}")
        return p

    pasta = resolver_pasta_downloads_dia(data_referencia)

    if not arquivo or not str(arquivo).strip():
        return localizar_arquivo_numerario_pasta_downloads(data_referencia)

    nome = str(arquivo).strip()
    if os.path.isfile(nome):
        return os.path.abspath(nome)

    candidato = os.path.join(pasta, os.path.basename(nome))
    if os.path.isfile(candidato):
        return os.path.abspath(candidato)

    raise FileNotFoundError(
        f"Arquivo não encontrado: {nome!r} (também não em {candidato!r})"
    )


def processar_arquivo_numerario_sigra(
    caminho_xlsx: str,
    com_sigra: Optional[bool] = None,
    salvar: bool = True,
    headless: bool = False,
) -> pd.DataFrame:
    """
    Lê o Excel, preenche REF cliente; em seguida consulta Sigra quando aplicável; grava.

    com_sigra:
      - None (padrão): abre o Chrome e preenche ref sigra se SIGRA_EMAIL e SIGRA_PASSWORD
        estiverem definidos no ambiente; caso contrário só REF cliente e aviso.
      - True: obriga Sigra (falha se faltar credenciais).
      - False: não acede ao Sigra.
    """
    garantir_numerario_xlsx_usavel(caminho_xlsx)
    df = pd.read_excel(caminho_xlsx, sheet_name='acompanhamento', engine='openpyxl')
    df = df.dropna(how='all')
    df = enriquecer_ref_cliente_dataframe(df)

    _em, _pw = _resolver_credenciais_sigra()
    credenciais_ok = bool(_em and _pw)

    if com_sigra is False:
        fazer_sigra = False
    elif com_sigra is True:
        fazer_sigra = True
    else:
        fazer_sigra = credenciais_ok
        if not fazer_sigra:
            print(
                "[INFO] Sigra nao executado apos REF cliente: preencha SIGRA_EMAIL_PADRAO / SIGRA_SENHA_PADRAO "
                "no codigo ou SIGRA_EMAIL / SIGRA_PASSWORD no ambiente, ou use --sigra."
            )

    if fazer_sigra:
        if COL_REF_SIGRA not in df.columns:
            df[COL_REF_SIGRA] = ''
        print("\n" + "=" * 60)
        print("PASSO SIGRA: login, busca por REF cliente, preenchimento de ref sigra")
        print("=" * 60)
        df = enriquecer_numerario_com_sigra(df, headless=headless)

    if salvar:
        salvar_numerario_xlsx(df, caminho_xlsx, sheet_name='acompanhamento')
        print(f"[OK] Arquivo atualizado: {caminho_xlsx}")

    return df


def encontrar_grupo_por_processo_numerario(extrato_row, numerario_disponivel, tolerancia_dias=0):
    """
    Concilia numerário pelo processo: agrupa registros do mesmo dia pelo processo
    (extraído de "Assunto (Email)"). Só soma registros do MESMO processo;
    não pode misturar mais de um processo por extrato.
    """
    valor_extrato = abs(extrato_row['valor'])
    data_extrato = extrato_row['data']
    if valor_extrato == 0 or pd.isna(data_extrato):
        return None
    valor_extrato_int = int(round(valor_extrato * 100, 0))
    data_extrato_norm = pd.Timestamp(data_extrato).normalize()

    def mesma_data(row):
        d = row['data']
        if pd.isna(d):
            return False
        return abs((pd.Timestamp(d).normalize() - data_extrato_norm).days) <= tolerancia_dias

    numerario_do_dia = numerario_disponivel[numerario_disponivel.apply(mesma_data, axis=1)]
    if len(numerario_do_dia) < 1:
        return None

    grupos = {}
    for idx, row in numerario_do_dia.iterrows():
        proc = extrair_processo_numerario(row)
        if proc not in grupos:
            grupos[proc] = []
        grupos[proc].append((idx, abs(float(row['valor']))))

    for proc, itens in grupos.items():
        if proc == '-':
            continue
        soma = sum(v for _, v in itens)
        soma_int = int(round(soma * 100, 0))
        if _valores_equivalentes_centavos(valor_extrato_int, soma_int):
            indices = [i for i, _ in itens]
            print(
                f"    [OK] Numerário grupo por processo ({proc}): {len(indices)} registro(s) = "
                f"R$ {_formatar_valor_br(valor_extrato)} | Data: {data_extrato_norm.strftime('%d/%m/%Y')}"
            )
            return indices
    return None


def conciliar_extrato_numerario(df_extrato_pendentes, df_numerario, conciliacoes_existentes):
    """
    Concilia extratos pendentes com numerário (passada após SIGRA).

    Depende de funções do módulo principal (match exato e formatação).
    """
    from conciliar_itau_sigra import (
        TOLERANCIA_DATA,
        encontrar_match_exato,
        formatar_valor_br,
    )

    print("\n" + "="*80)
    print("TERCEIRA PASSADA: CONCILIAÇÃO COM NUMERÁRIO")
    print("="*80)

    novas_conciliacoes = []
    numerario_usado = set()

    extratos_conciliados = set(c['ID_extrato'] for c in conciliacoes_existentes)

    df_extrato_pend = df_extrato_pendentes[
        (~df_extrato_pendentes['ID_extrato'].isin(extratos_conciliados))
    ].copy()

    if len(df_extrato_pend) == 0:
        print("Nenhum extrato pendente para conciliar com numerário.")
        return novas_conciliacoes

    df_extrato_ordenado = df_extrato_pend.sort_values('valor', key=abs, ascending=False).copy()

    total_extratos = len(df_extrato_ordenado)
    print(f"\nTotal de extratos pendentes: {total_extratos}")
    print(f"Total de registros no numerário: {len(df_numerario)}")

    for idx, extrato_row in df_extrato_ordenado.iterrows():
        valor_extrato_abs = abs(extrato_row['valor'])
        print(f"\nProcessando extrato {extrato_row['ID_extrato']}/{total_extratos} (R$ {formatar_valor_br(valor_extrato_abs)})")

        numerario_disponivel = df_numerario[~df_numerario.index.isin(numerario_usado)]

        if len(numerario_disponivel) == 0:
            print("  [INFO] Nenhum registro de numerário disponível")
            continue

        combinacao = encontrar_match_exato(extrato_row, numerario_disponivel)

        if not combinacao:
            combinacao = encontrar_grupo_por_processo_numerario(
                extrato_row, numerario_disponivel, tolerancia_dias=TOLERANCIA_DATA
            )

        if combinacao:
            numerario_usado.update(combinacao)

            for numerario_idx in combinacao:
                numerario_row = df_numerario.loc[numerario_idx]

                ref_processo = extrair_processo_numerario(numerario_row)
                if pd.isna(ref_processo) or not str(ref_processo).strip():
                    ref_processo = '-'
                else:
                    match_proc = re.search(r'processo\s+([^\s]+)', str(ref_processo), re.IGNORECASE)
                    if match_proc:
                        ref_processo = match_proc.group(1)

                id_sigra_cel = numerario_row.get(COL_REF_SIGRA, '')
                if pd.notna(id_sigra_cel) and str(id_sigra_cel).strip():
                    ref_processo = str(id_sigra_cel).strip()

                id_numerario = numerario_row.get('ID', '-')
                if pd.isna(id_numerario):
                    id_numerario = numerario_row.get('ID_comprovante', '-')

                novas_conciliacoes.append({
                    'ID_extrato': extrato_row['ID_extrato'],
                    'Data_extrato': extrato_row['data_original'],
                    'Valor_extrato': extrato_row['valor'],
                    'Favorecido_extrato': extrato_row['favorecido_original'],
                    'ID_comprovante': id_numerario,
                    'Data_comprovante': numerario_row['data_original'],
                    'Valor_comprovante': numerario_row['valor'],
                    'Favorecido_comprovante': numerario_row['favorecido_original'],
                    'Ref. Sigra': ref_processo,
                    'Categoria': '-',
                    'Cliente': '-',
                    'Origem': 'Numerário',
                })

            print(f"  [OK] Conciliação com numerário confirmada: {len(combinacao)} registro(s)")

    extratos_conciliados_numerario = set(c['ID_extrato'] for c in novas_conciliacoes)
    pendentes_apos_numerario = df_extrato_pend[~df_extrato_pend['ID_extrato'].isin(extratos_conciliados_numerario)]
    if len(pendentes_apos_numerario) > 0 and len(numerario_usado) < len(df_numerario):
        print("\n" + "-"*60)
        print("NUMERÁRIO - 2ª passada: mesmo dia e valor")
        print("-"*60)
        for _, extrato_row in pendentes_apos_numerario.iterrows():
            numerario_disp = df_numerario[~df_numerario.index.isin(numerario_usado)]
            if len(numerario_disp) == 0:
                break
            combinacao = encontrar_match_exato(extrato_row, numerario_disp)
            if combinacao:
                numerario_usado.update(combinacao)
                for numerario_idx in combinacao:
                    numerario_row = df_numerario.loc[numerario_idx]
                    ref_processo = extrair_processo_numerario(numerario_row)
                    id_sigra_cel = numerario_row.get(COL_REF_SIGRA, '')
                    if pd.notna(id_sigra_cel) and str(id_sigra_cel).strip():
                        ref_processo = str(id_sigra_cel).strip()
                    id_numerario = numerario_row.get('ID', '-')
                    if pd.isna(id_numerario):
                        id_numerario = numerario_row.get('ID_comprovante', '-')
                    novas_conciliacoes.append({
                        'ID_extrato': extrato_row['ID_extrato'],
                        'Data_extrato': extrato_row['data_original'],
                        'Valor_extrato': extrato_row['valor'],
                        'Favorecido_extrato': extrato_row['favorecido_original'],
                        'ID_comprovante': id_numerario,
                        'Data_comprovante': numerario_row['data_original'],
                        'Valor_comprovante': numerario_row['valor'],
                        'Favorecido_comprovante': numerario_row['favorecido_original'],
                        'Ref. Sigra': ref_processo,
                        'Categoria': '-',
                        'Cliente': '-',
                        'Origem': 'Numerário',
                    })
                print(f"  [OK] Conciliação numerário (2ª passada): extrato {extrato_row['ID_extrato']}")

    print(f"\n[OK] Conciliação com numerário concluída: {len(novas_conciliacoes)} novas conciliações")
    return novas_conciliacoes


if __name__ == '__main__':
    import argparse
    import sys

    p = argparse.ArgumentParser(
        description=(
            'Numerário: extrai REF cliente (coluna L), depois consulta Sigra se SIGRA_EMAIL/PASSWORD '
            'estiverem no ambiente. Sem argumento, usa downloads/AAAA-MM-DD (hoje ou pasta mais recente).'
        )
    )
    p.add_argument(
        'arquivo',
        nargs='?',
        default=None,
        help=(
            'Caminho completo do .xlsx, ou só o nome do arquivo dentro da pasta do dia, '
            'ou omita para localizar *numerario*.xlsx na pasta downloads do dia'
        ),
    )
    p.add_argument(
        '--data',
        default=None,
        metavar='AAAA-MM-DD',
        help='Usar esta data como pasta downloads (ex.: 2026-04-13) em vez do padrão (hoje)',
    )
    g = p.add_mutually_exclusive_group()
    g.add_argument(
        '--sigra',
        action='store_true',
        help='Obrigar passo Sigra (falha se faltar SIGRA_EMAIL / SIGRA_PASSWORD)',
    )
    g.add_argument(
        '--sem-sigra',
        action='store_true',
        help='Nao abrir o Sigra (so REF cliente), mesmo com credenciais no ambiente',
    )
    p.add_argument('--headless', action='store_true', help='Chrome em modo headless no passo Sigra')
    p.add_argument('--no-save', action='store_true', help='Não sobrescrever o arquivo')
    args = p.parse_args()
    if args.sigra:
        com_sigra = True
    elif args.sem_sigra:
        com_sigra = False
    else:
        com_sigra = None
    try:
        caminho = resolver_caminho_arquivo_numerario(args.arquivo, args.data)
        processar_arquivo_numerario_sigra(
            caminho,
            com_sigra=com_sigra,
            salvar=not args.no_save,
            headless=args.headless,
        )
    except (ValueError, FileNotFoundError, PermissionError) as e:
        print(f"\n[ERRO] {e}", file=sys.stderr)
        sys.exit(1)
