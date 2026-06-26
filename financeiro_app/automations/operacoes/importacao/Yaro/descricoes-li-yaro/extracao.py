# -*- coding: utf-8 -*-
"""
Extração de dados de Faturas Comerciais (Commercial Invoice) - formato YARO/Atlantic.
Lê PDFs e extrai: cabeçalho, importador, conta bancária, fabricante, itens e totais.
"""

import re
import json
import sys
import os
import queue
import shutil
import threading
from pathlib import Path
from typing import Optional

try:
    from pypdf import PdfReader
except ImportError:
    from PyPDF2 import PdfReader

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


def normalizar_texto(texto: str) -> str:
    """Substitui tabs por espaço e normaliza espaços múltiplos."""
    return re.sub(r'\s+', ' ', texto.replace('\t', ' ')).strip()


def _log_debug(msg: str) -> None:
    try:
        base = pasta_execucao()
        log_path = base / "ocr_debug.log"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"{msg}\n")
    except Exception:
        pass


def _ocr_pdf_para_texto(caminho_pdf: str) -> str:
    """
    OCR fallback para PDFs escaneados (sem texto selecionável).
    Retorna texto aproximado em ordem de leitura.
    """
    try:
        import numpy as np
        import pypdfium2 as pdfium
        from rapidocr_onnxruntime import RapidOCR
    except Exception as e:
        _log_debug(f"OCR import error: {e!r}")
        return ""

    def _linhas_ocr_em_ordem(resultado_ocr) -> list[str]:
        linhas_brutas: list[tuple[float, float, str]] = []
        for item in (resultado_ocr or []):
            try:
                box, txt, _score = item
            except Exception:
                continue
            txt_n = normalizar_texto(str(txt or ""))
            if not txt_n:
                continue
            xs = [float(p[0]) for p in box]
            ys = [float(p[1]) for p in box]
            x = sum(xs) / len(xs)
            y = sum(ys) / len(ys)
            linhas_brutas.append((y, x, txt_n))
        if not linhas_brutas:
            return []
        linhas_brutas.sort(key=lambda t: (t[0], t[1]))
        agrupadas: list[list[tuple[float, str]]] = []
        y_ref = None
        tol = 16.0
        for y, x, txt in linhas_brutas:
            if y_ref is None or abs(y - y_ref) <= tol:
                if not agrupadas:
                    agrupadas.append([])
                agrupadas[-1].append((x, txt))
                y_ref = y if y_ref is None else ((y_ref * 0.6) + (y * 0.4))
            else:
                agrupadas.append([(x, txt)])
                y_ref = y
        saida: list[str] = []
        for grupo in agrupadas:
            grupo.sort(key=lambda t: t[0])
            linha = " ".join(t for _, t in grupo).strip()
            if linha:
                saida.append(linha)
        return saida

    try:
        engine = RapidOCR()
        doc = pdfium.PdfDocument(caminho_pdf)
    except Exception as e:
        _log_debug(f"OCR init error: {e!r}")
        return ""

    paginas_txt: list[str] = []
    for i in range(len(doc)):
        try:
            page = doc[i]
            render = page.render(scale=2.0)
            pil_img = render.to_pil()
            img = np.array(pil_img)
            resultado, _ = engine(img)
            linhas = _linhas_ocr_em_ordem(resultado)
            if linhas:
                paginas_txt.append("\n".join(linhas))
        except Exception:
            _log_debug(f"OCR page {i+1} error")
            continue
    return "\n".join(paginas_txt)


def extrair_texto_pdf(caminho_pdf: str) -> str:
    """Extrai todo o texto de um PDF."""
    reader = PdfReader(caminho_pdf)
    partes = []
    for page in reader.pages:
        partes.append(page.extract_text() or "")
    texto = "\n".join(partes)
    if texto.strip():
        _log_debug(f"extract_text ok len={len(texto.strip())}")
        return texto
    # Fallback para PDFs escaneados.
    _log_debug("extract_text vazio, acionando OCR")
    texto_ocr = _ocr_pdf_para_texto(caminho_pdf)
    _log_debug(f"OCR texto len={len((texto_ocr or '').strip())}")
    return texto_ocr or texto


def extrair_cabecalho(texto: str) -> dict:
    """Extrai campos do cabeçalho da fatura (DATE, NUMBER, PAYMENT TERMS, etc.)."""
    texto_norm = normalizar_texto(texto)
    campos = {}

    padroes = [
        (r'DATE:\s*(\d{1,2}-[a-z]{3}-\d{2})', 'data'),
        (r'NUMBER:\s*([A-Z0-9\-]+)', 'numero'),
        (r'PAYMENT\s+TERMS:\s*([^.]+?)(?=\s+PLACE|\s+TERMS:|$)', 'condicoes_pagamento'),
        (r'PLACE\s+OF\s+LOADING:\s*([^T]+?)(?=\s+TERMS:|$)', 'local_carregamento'),
        (r'PORT\s+OF\s+DISCHARGE:\s*([^C]+?)(?=\s+COUNTRY|$)', 'porto_descarga'),
        (r'(?<!PAYMENT\s)TERMS:\s*(FOB|CIF|CFR|EXW|DAP|DDP|CIP)\b', 'termos'),  # Incoterm (não PAYMENT TERMS)
        (r'COUNTRY\s+OF\s+ORIGIN\s*/\s*MANUFACUTING:\s*(\w+)', 'pais_origem'),
    ]
    for regex, chave in padroes:
        m = re.search(regex, texto_norm, re.IGNORECASE | re.DOTALL)
        if m:
            campos[chave] = normalizar_texto(m.group(1))

    return campos


def extrair_importador(texto: str) -> dict:
    """Extrai dados do importador/comprador."""
    texto_norm = normalizar_texto(texto)
    # Nome: após IMPORTER / BUYER até LTDA/LTD (pode conter R em IMPORTACAO)
    m_nome = re.search(r'IMPORTER\s*/\s*BUYER:\s*(.+?(?:LTDA|LTD|S\.?A\.?|S\/A))', texto_norm, re.IGNORECASE)
    nome = normalizar_texto(m_nome.group(1)) if m_nome else ""

    m_cep = re.search(r'CEP\s*([\d\.\-]+)', texto_norm)
    cep = m_cep.group(1).strip() if m_cep else ""

    m_cnpj = re.search(r'CNPJ\s*([\d\.\/\-]+)', texto_norm)
    cnpj = m_cnpj.group(1).strip() if m_cnpj else ""

    # Endereço: entre fim do nome e CEP
    endereco = ""
    if m_nome and m_cep:
        trecho = texto_norm[m_nome.end() : texto_norm.find("CEP")]
        endereco = normalizar_texto(trecho).strip()

    return {
        "nome": nome,
        "endereco": endereco,
        "cep": cep,
        "cnpj": cnpj,
    }


def extrair_conta_bancaria(texto: str) -> dict:
    """Extrai informações da conta (banco, SWIFT, beneficiário)."""
    texto_norm = normalizar_texto(texto)
    campos = {}

    m = re.search(r'Beneficiary\s+Bank\s+Name:\s*(.+?)(?=Beneficiary\s+Bank\s+Address:)', texto_norm, re.IGNORECASE | re.DOTALL)
    campos["banco_beneficiario"] = normalizar_texto(m.group(1)) if m else ""

    m = re.search(r'Beneficiary\s+Bank\s+Address:\s*(.+?)(?=SWIFT)', texto_norm, re.IGNORECASE | re.DOTALL)
    campos["endereco_banco"] = normalizar_texto(m.group(1)) if m else ""

    m = re.search(r'SWIFT\s+Code:\s*(\w+)', texto_norm, re.IGNORECASE)
    campos["swift"] = m.group(1).strip() if m else ""

    m = re.search(r'Beneficiary\s+Name:\s*([^B]+?)(?=Beneficiary\s+Account:)', texto_norm, re.IGNORECASE)
    campos["beneficiario"] = normalizar_texto(m.group(1)) if m else ""

    m = re.search(r'Beneficiary\s+Account:\s*([\w\s]+?)(?=MANUFACTURER|$)', texto_norm, re.IGNORECASE)
    campos["conta_beneficiario"] = normalizar_texto(m.group(1)) if m else ""

    return campos


def extrair_fabricante(texto: str) -> dict:
    """Extrai nome e endereço do fabricante."""
    texto_norm = normalizar_texto(texto)
    m_nome = re.search(r'MANUFACTURER:\s*Name:\s*(.+?)(?=Address:)', texto_norm, re.IGNORECASE | re.DOTALL)
    m_end = re.search(r'MANUFACTURER:.*?Name:.*?Address:\s*(.+?)(?=DESCRIPTION|$)', texto_norm, re.IGNORECASE | re.DOTALL)
    return {
        "nome": normalizar_texto(m_nome.group(1)) if m_nome else "",
        "endereco": normalizar_texto(m_end.group(1)) if m_end else "",
    }


def extrair_itens(texto: str) -> list[dict]:
    """
    Extrai a tabela de itens (descrição, código YARO, padrão, NCM, quantidade, pesos, FOB).
    Linhas de produto contêm NCM no formato 4011.xx.xx.
    """
    linhas = texto.split('\n')
    itens = []
    # Padrão NCM tolerante: 4011.20.90 / 40112090 / 4011 20 90
    ncm_re = re.compile(r"\b(4011(?:[\.\s]?\d{2}){2})\b")

    def _normalizar_ncm(raw: str) -> str:
        digits = re.sub(r"\D", "", raw or "")
        if len(digits) == 8:
            return f"{digits[:4]}.{digits[4:6]}.{digits[6:8]}"
        return normalizar_texto(raw or "")

    def _is_num(tok: str) -> bool:
        t = (tok or "").replace("$", "").replace(",", "").strip()
        return bool(re.fullmatch(r"\d+(?:\.\d+)?", t))

    def _to_float(tok: str) -> float:
        return float((tok or "").replace("$", "").replace(",", "").strip())

    for i, linha in enumerate(linhas):
        linha_limpa = normalizar_texto(linha)
        if not linha_limpa or 'DESCRIPTION' in linha_limpa:
            continue
        # Cabeçalho da tabela; linhas de produto também podem conter "New Tires".
        if 'New Tires' in linha_limpa and not ncm_re.search(linha_limpa):
            continue
        # Linha de totais da tabela (ex: 6,710  60,272.00  $126,858.90)
        if re.match(r'^[\d,]+\.?\d*\s+[\d,]+\.?\d*\s+\$[\d,]+\.?\d*$', linha_limpa):
            break
        ncm_match = ncm_re.search(linha_limpa)
        if not ncm_match:
            continue

        # Quebrar por espaços preservando números com vírgula e $ 
        partes = re.split(r'\s+', linha_limpa)
        ncm = _normalizar_ncm(ncm_match.group(1))
        idx_ncm = None
        for j, p in enumerate(partes):
            if _normalizar_ncm(p) == ncm:
                idx_ncm = j
                break
        if idx_ncm is None:
            continue

        # Estrutura esperada: ... NCM QUANTITY [UNIT] NET_WEIGHT TOTAL_WEIGHT FOB TOTAL_FOB
        try:
            idx = idx_ncm + 1
            if idx >= len(partes) or not _is_num(partes[idx]):
                continue
            qtd = partes[idx].replace(",", "")
            idx += 1
            # Unidade opcional (ex.: PCS, PC)
            if idx < len(partes) and re.fullmatch(r"[A-Za-z]{1,4}", partes[idx]):
                idx += 1
            if idx + 3 >= len(partes):
                continue
            if not (_is_num(partes[idx]) and _is_num(partes[idx + 1]) and _is_num(partes[idx + 2]) and _is_num(partes[idx + 3])):
                continue
            peso_unit = _to_float(partes[idx])
            peso_total = _to_float(partes[idx + 1])
            fob_unit = _to_float(partes[idx + 2])
            fob_total = _to_float(partes[idx + 3])
        except (IndexError, ValueError):
            continue

        # Descrição: do início até antes do código YARO; padrão é CF + números
        desc_partes = partes[:idx_ncm]
        yaro_code = ""
        pattern = ""
        desc_fim = -1
        for k in range(len(desc_partes) - 1, -1, -1):
            if re.match(r'^CF\d+$', desc_partes[k], re.IGNORECASE):
                pattern = desc_partes[k]
                yaro_code = desc_partes[k - 1] if k > 0 else "0"
                desc_fim = k - 2  # descrição = tudo antes do yaro_code
                break
        descricao = " ".join(desc_partes[: desc_fim + 1]) if desc_fim >= 0 else " ".join(desc_partes[: len(desc_partes) - 2])

        itens.append({
            "descricao": descricao,
            "yaro_code": yaro_code,
            "pattern": pattern,
            "ncm": ncm,
            "quantidade": int(float(qtd)),
            "peso_unitario": float(peso_unit),
            "peso_total": float(peso_total),
            "fob_unitario_usd": float(fob_unit),
            "total_fob_usd": float(fob_total),
        })
    return itens


# --- Descrição via catálogo (catalogo-produtos-importacao) ---

CATALOGO_PADRAO = "catalogo-produtos-importacao (4).xlsx"


def carregar_catalogo(caminho_planilha: Optional[str]) -> list[dict]:
    """
    Carrega a planilha do catálogo de produtos (ex.: catalogo-produtos-importacao (4).xlsx).
    Colunas esperadas: Identificadores, Denominação, Descrição, NCM, Modalidade...
    Retorna lista de dicts com chaves: identificadores, denominacao, descricao, ncm, modalidade.
    """
    if not caminho_planilha or not Path(caminho_planilha).is_file():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wb = load_workbook(caminho_planilha, read_only=True, data_only=True)
    ws = wb.active
    # Forçar leitura de colunas e linhas (planilhas com tabela podem reportar max_row=1)
    rows = list(ws.iter_rows(min_row=1, max_row=10000, min_col=1, max_col=10, values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    header = [str(c).strip() if c else "" for c in rows[0]]
    # Mapear colunas por nome (case-insensitive, sem acento)
    def norm(s: str) -> str:
        s = (s or "").lower().replace(" ", "_").replace("ç", "c").replace("ã", "a").replace("õ", "o")
        return re.sub(r"[^a-z0-9_]", "", s)

    col_ident = col_denom = col_desc = None
    for i, h in enumerate(header):
        n = norm(h)
        if "identificador" in n:
            col_ident = i
        elif "denomina" in n:
            col_denom = i
        elif "descri" in n and "denomina" not in n:
            col_desc = i

    if col_denom is None or col_desc is None:
        return []

    catalogo = []
    for row in rows[1:]:
        if not row:
            continue
        denom = row[col_denom]
        desc = row[col_desc]
        if denom is None and desc is None:
            continue
        catalogo.append({
            "identificadores": str(row[col_ident]).strip() if col_ident is not None and col_ident < len(row) and row[col_ident] is not None else "",
            "denominacao": str(denom).strip() if denom is not None else "",
            "descricao": str(desc).strip() if desc is not None else "",
        })
    return catalogo


def buscar_descricao_no_catalogo(catalogo: list[dict], nome_item: str) -> str:
    """
    Busca no catálogo pela descrição do item (ex.: "205/75R16C 110/108R 8PR").
    Usa a medida base (ex.: 205/75R16) para alinhar com linhas do catálogo sem sufixo C/LT
    e escolhe a linha cujos tokens (índices, PR, etc.) melhor coincidem com o item da fatura.
    """
    def _norm_catalogo(s: str) -> str:
        # Comparação estável: um espaço entre palavras, maiúsculas, ()[],; viram espaço.
        # TL e demais tokens permanecem (tubeless entra na conta do match).
        s = normalizar_texto(s or "").upper()
        s = re.sub(r"[()\[\],;]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _tokens(s: str) -> set[str]:
        s = _norm_catalogo(s)
        return {t for t in s.split() if t}

    def _tokens_item_catalogo(s_item: str) -> set[str]:
        """Tokens do item com primeiro token na medida base (ex. 205/75R16C -> 205/75R16)."""
        partes = _norm_catalogo(s_item).split()
        if not partes:
            return set()
        base = extrair_medida_base_pneu(s_item)
        if base:
            partes = [base] + partes[1:]
        return {t for t in partes if t}

    nome = _norm_catalogo(nome_item)
    if not nome:
        return ""

    base_item = extrair_medida_base_pneu(nome_item)
    tokens_nome = _tokens_item_catalogo(nome_item)

    melhor_desc = ""
    melhor_score = -1

    for linha in catalogo:
        denom = linha.get("denominacao", "")
        if not denom:
            continue
        denom_norm = _norm_catalogo(denom)
        desc = linha.get("descricao", "")
        base_denom = extrair_medida_base_pneu(denom)
        toks_denom = _tokens(denom_norm)

        if nome == denom_norm:
            return desc
        if len(nome) >= 12 and denom_norm.startswith(nome):
            return desc

        score = -1

        if base_item and base_denom and _bases_medida_equivalentes(base_item, base_denom):
            if not toks_denom:
                continue
            inter = tokens_nome & toks_denom
            score = 5000 + 120 * len(inter) - 15 * max(0, len(toks_denom) - len(tokens_nome))
            score -= len(denom_norm) // 28

        elif tokens_nome and toks_denom and tokens_nome.issubset(toks_denom):
            score = 800 + 100 * len(tokens_nome) - (len(toks_denom) - len(tokens_nome))

        if score > melhor_score:
            melhor_score = score
            melhor_desc = desc

    return melhor_desc


def aplicar_descricoes_do_catalogo(
    dados: dict,
    caminho_catalogo: Optional[str] = None,
    caminho_certificados: Optional[str] = None,
) -> None:
    """
    Para cada item da fatura, monta a descrição pelo template + CERTIFICADOS INMETRO.
    O catálogo está desativado neste fluxo.
    Modifica dados in-place.
    """
    certificados = carregar_certificados_inmetro(caminho_certificados) if caminho_certificados else []

    def _eh_lt_ou_c(item: dict) -> bool:
        """True para medidas de Light Truck: prefixo LT, sufixo C ou token LT na descrição."""
        desc_item = item.get("descricao", "")
        parsed_item = parse_descricao_pneu(desc_item)
        medida_item = (parsed_item.get("medida") or "").upper()
        desc_up = normalizar_texto(desc_item).upper()
        return (
            medida_item.startswith("LT")
            or "LT" in medida_item
            or medida_item.endswith("C")
            or bool(re.search(r"\bLT\d", desc_up))
        )

    def _eh_carga_reforcado(item: dict) -> bool:
        """Carga reforçada: LT... (ex. LT175/70R14) ou sufixo ...C."""
        if item.get("ncm") != "4011.20.90":
            return False
        return _eh_lt_ou_c(item)

    def _eh_passeio_1_indice_aro_menor20(item: dict) -> bool:
        """
        Regra operacional solicitada:
        se houver 1 índice de carga (ex.: 109Q) e aro < 20,
        classifica como NCM 4011.10.00 (passageiro/veículos leves).
        """
        desc_item = item.get("descricao", "")
        parsed_item = parse_descricao_pneu(desc_item)
        carga = (parsed_item.get("indice_carga") or "").strip().upper()
        vel = (parsed_item.get("indice_velocidade") or "").strip().upper()
        if not carga or not vel:
            return False
        # Um único índice: "109Q" (não "117/114Q").
        if "/" in carga:
            return False
        if not re.fullmatch(r"\d+", carga):
            return False
        if not re.fullmatch(r"[A-Z]", vel):
            return False
        aro = extrair_aro_da_medida((parsed_item.get("medida") or desc_item))
        return 0 < aro < 20.0

    def _eh_tbr_onibus_caminhao(item: dict) -> bool:
        """TBR (ônibus/caminhões): NCM 4011.20.90 com aro >= 20."""
        if item.get("ncm") != "4011.20.90":
            return False
        # Prioridade operacional: LT/C é sempre Light Truck (judicial), não TBR.
        if _eh_lt_ou_c(item):
            return False
        desc_item = item.get("descricao", "")
        parsed_item = parse_descricao_pneu(desc_item)
        medida_item = parsed_item.get("medida") or desc_item
        aro = extrair_aro_da_medida(medida_item)
        return aro >= 20.0

    for item in dados.get("itens", []):
        # Regra operacional solicitada: toda medida LT/C é Light Truck judicial (NCM 4011.20.90).
        if _eh_lt_ou_c(item):
            item["ncm"] = "4011.20.90"
        # Para não LT/C, mantém a regra de 1 índice + aro < 20 => 4011.10.00.
        elif _eh_passeio_1_indice_aro_menor20(item):
            item["ncm"] = "4011.10.00"

        nome = item.get("descricao", "")
        # NCM 4011.10.00 (passeio, 1 índice de carga): descrição vem só do padrão INMETRO, não do catálogo (muitas linhas incorretas).
        if item.get("ncm") == "4011.10.00":
            cert = buscar_certificado(
                certificados,
                item.get("yaro_code", ""),
                item.get("pattern", ""),
                nome,
                str(item.get("marca", "")),
            )
            desc_std, compl_std = montar_descricao_detalhada(item, cert, certificados)
            item["descricao_detalhada"] = desc_std
            item["informacoes_complementares"] = compl_std
            item["descricao_origem"] = "Padrão NCM 4011.10.00 (INMETRO)" if cert else "Padrão NCM 4011.10.00 (sem certificado)"
            continue

        # NCM 4011.20.90 de carga reforçado (LT... / ...C): também forçar padrão INMETRO.
        if _eh_carga_reforcado(item):
            cert = buscar_certificado(
                certificados,
                item.get("yaro_code", ""),
                item.get("pattern", ""),
                nome,
                str(item.get("marca", "")),
            )
            desc_std, compl_std = montar_descricao_detalhada(item, cert, certificados)
            item["descricao_detalhada"] = desc_std
            item["informacoes_complementares"] = compl_std
            item["descricao_origem"] = "Padrão NCM 4011.20.90 carga reforçado (INMETRO)" if cert else "Padrão NCM 4011.20.90 carga reforçado (sem certificado)"
            continue

        # NCM 4011.20.90 TBR (ônibus/caminhões, 2 índices de carga): forçar padrão INMETRO.
        if _eh_tbr_onibus_caminhao(item):
            cert = buscar_certificado(
                certificados,
                item.get("yaro_code", ""),
                item.get("pattern", ""),
                nome,
                str(item.get("marca", "")),
            )
            desc_std, compl_std = montar_descricao_detalhada(item, cert, certificados)
            item["descricao_detalhada"] = desc_std
            item["informacoes_complementares"] = compl_std
            item["descricao_origem"] = "Padrão NCM 4011.20.90 TBR (INMETRO)" if cert else "Padrão NCM 4011.20.90 TBR (sem certificado)"
            continue

        # Montar descrição (template + CERTIFICADOS INMETRO)
        cert = buscar_certificado(
            certificados,
            item.get("yaro_code", ""),
            item.get("pattern", ""),
            nome,
            str(item.get("marca", "")),
        )
        desc_fallback, compl_fallback = montar_descricao_detalhada(item, cert, certificados)
        item["descricao_detalhada"] = desc_fallback
        item["informacoes_complementares"] = compl_fallback
        item["descricao_origem"] = "CERTIFICADOS INMETRO" if cert else "Não encontrada"


# --- (Legado: montagem de descrição por CERTIFICADOS INMETRO; mantido para referência) ---

TEXTO_PROCESSO_JUDICIAL = (
    "A classificação 4011.20.90 utilizada nesta adição está respaldada em decisão judicial "
    "proferida na Apelação Cível nº 5050748-16.2021.4.04.7000, da 1a. Turma do TRF da 4a. Região "
    "para todos os pneus novos destinados a veículos comerciais leves, C e LT, da categoria 3 do INMETRO."
)
TEXTO_ANALISAR_ANTI_DUMPING = "analisar anti dumping"


def parse_descricao_pneu(descricao: str) -> dict:
    """
    Extrai da descrição do pneu: medida, indice_carga, indice_velocidade, pr, tem_tl.
    Tolera ordem variada (ex.: "175/75R13 4PR 85T TL", "LT175/70R14 98/96S 10PR").
    """
    texto = normalizar_texto(descricao)
    partes = texto.split()
    if not partes:
        return {"medida": "", "indice_carga": "", "indice_velocidade": "", "pr": "", "tem_tl": False}

    partes_medida = partes
    if partes_medida and partes_medida[0].upper() == "PNEU":
        partes_medida = partes_medida[1:]

    texto_up = " ".join(partes_medida).upper()
    medida = ""
    indice_carga = ""
    indice_velocidade = ""
    pr = ""
    tem_tl = False

    # Medida com tolerância a espaço antes de R e aro decimal (ex.: 295/80 R22.5).
    for padrao in (
        r"\b(?:LT|P)?\d{2,4}/\d{2}Z?\s*R\d{2}(?:\.\d)?(?:C|LT)?\b",
        r"\b(?:LT)?\d+X[\d.]+\s*R\d{2}(?:LT)?\b",
    ):
        m_med = re.search(padrao, texto_up, re.IGNORECASE)
        if m_med:
            medida = re.sub(r"\s+", "", m_med.group(0).upper())
            break

    # Carga/velocidade tolerando token colado com PR (ex.: 154/149K16PR...).
    m_cv = re.search(r"(\d+(?:/\d+)?)([A-Z])(?=\d{1,2}PR|\b)", texto_up)
    if m_cv:
        indice_carga = m_cv.group(1)
        indice_velocidade = m_cv.group(2)

    # PR pode vir colado em outros textos.
    m_pr = re.search(r"(\d{1,2}PR)", texto_up)
    if m_pr:
        pr = m_pr.group(1)

    tem_tl = bool(re.search(r"\bTL\b", texto_up))

    return {
        "medida": medida,
        "indice_carga": indice_carga,
        "indice_velocidade": indice_velocidade,
        "pr": pr,
        "tem_tl": tem_tl,
    }


def extrair_medida_base_pneu(texto: str) -> str:
    """
    Normaliza o primeiro token da descrição para chave de catálogo.
    Ex.: 205/75R16C -> 205/75R16; P275/60R20 mantém P; LT175/70R14 mantém LT.
    """
    texto = normalizar_texto(texto or "")
    if not texto:
        return ""
    primeiro = texto.split()[0].upper()

    m_fl = re.match(
        r"^(?P<pre>LT)?(?P<core>\d+X[\d.]+R\d{2})(?P<suf>LT)?$",
        primeiro,
        re.IGNORECASE,
    )
    if m_fl:
        pre = (m_fl.group("pre") or "").upper()
        core = m_fl.group("core").upper()
        return f"{pre}{core}" if pre else core

    m_std = re.match(
        r"^(?P<pre>LT|P)?(?P<core>\d{2,4}/\d{2}Z?R\d{2})(?P<suf>C|LT)?$",
        primeiro,
        re.IGNORECASE,
    )
    if m_std:
        pre = (m_std.group("pre") or "").upper()
        core = m_std.group("core").upper()
        return f"{pre}{core}" if pre else core

    return primeiro


def extrair_aro_da_medida(texto: str) -> float:
    """
    Extrai o valor do aro a partir da medida (ou descrição).
    Ex.: 295/80 R22.5 -> 22.5; LT175/70R14 -> 14.
    Retorna 0.0 quando não identificado.
    """
    t = normalizar_texto(texto or "").upper()
    if not t:
        return 0.0
    m = re.search(r"R\s*([0-9]{2}(?:\.[0-9])?)", t)
    if not m:
        return 0.0
    try:
        return float(m.group(1))
    except ValueError:
        return 0.0


def _bases_medida_equivalentes(a: str, b: str) -> bool:
    """True se a mesma medida (ignora sufixo C/LT na borda e opcionalmente P/LT no prefixo)."""
    if not a or not b:
        return False
    A, B = a.upper(), b.upper()
    if A == B:
        return True

    def nucleo(s: str) -> str:
        return re.sub(r"^(LT|P)", "", s)

    return nucleo(A) == nucleo(B)


def carregar_certificados_inmetro(caminho_planilha: Optional[str]) -> list[dict]:
    """
    Carrega a planilha 'CERTIFICADOS INMETRO 2026' (ou outro Excel).
    Retorna lista de dicionários com colunas normalizadas (minúsculas, sem acento nas chaves).
    Busca por YARO CODE e PATTERN para cruzar com os itens da fatura.
    """
    if not caminho_planilha or not Path(caminho_planilha).is_file():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wb = load_workbook(caminho_planilha, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # Primeira linha = cabeçalho; normalizar nomes (YARO CODE -> yaro_code, etc.)
    header = [str(c).strip() if c else "" for c in rows[0]]

    def norm(s: str, idx: int) -> str:
        s = (s or "").lower().replace(" ", "_").replace("ç", "c").replace("ã", "a").replace("í", "i").replace("é", "e").replace("á", "a").replace("ó", "o").replace("ú", "u")
        s = re.sub(r"[^a-z0-9_]", "", s)
        if "yaro" in s and "cod" in s:
            return "yaro_code"
        if "pattern" in s or "padrao" in s:
            return "pattern"
        if "marca" in s:
            return "marca"
        if "modelo" in s:
            return "modelo"
        if "familia" in s:
            return "familia"
        if "certificado" in s and "inmetro" in s:
            return "certificado_inmetro"
        if "registro" in s and "in" in s:
            return "registro_in"
        if s == "rrc":
            return "rrc"
        if s == "g":
            return "g"
        if "db" in s or "ruido" in s:
            return "db"
        if "codigo" in s and "barras" in s:
            return "codigo_barras"
        if "descric" in s or "description" in s:
            return "descricao_cert"
        return s if s else f"col_{idx}"

    col_map = {norm(h, i): i for i, h in enumerate(header) if h is not None}
    certs = []
    for row in rows[1:]:
        if not row:
            continue
        item = {}
        for key, idx in col_map.items():
            if idx < len(row) and row[idx] is not None:
                val = row[idx]
                if isinstance(val, (int, float)) and "cod" in key and key != "rrc":
                    val = str(int(val)) if isinstance(val, float) and val == int(val) else str(val)
                item[key] = str(val).strip() if val != "" else ""
        # Extrair pattern do modelo quando não existir coluna Pattern (ex.: "CF1100 - 6939801700663" -> CF1100)
        if not item.get("pattern") and item.get("modelo"):
            modelo = str(item.get("modelo", "")).replace("\n", " ").strip()
            if " / " in modelo:
                item["pattern"] = modelo.split(" / ")[0].strip()
            elif " - " in modelo:
                item["pattern"] = modelo.split(" - ")[0].strip()
            elif "-" in modelo:
                item["pattern"] = modelo.split("-")[0].strip()
            else:
                item["pattern"] = modelo[:20].strip() if len(modelo) > 20 else modelo
        if item.get("yaro_code") or item.get("pattern") or item.get("descricao_cert"):
            certs.append(item)
    return certs


def _tokens_item_para_match(item_descricao: str) -> dict:
    """
    Extrai tokens relevantes do item para comparar com a descricao_cert do INMETRO, mesmo com ordem diferente.
    Ex.: "LT175/70R14 98/96S 10PR" -> medida="LT175/70R14", carga_vel="98/96S", pr="10PR"
    """
    texto = normalizar_texto(item_descricao or "")
    parsed = parse_descricao_pneu(texto)
    medida = parsed.get("medida", "") or (texto.split()[0] if texto.split() else "")
    carga = parsed.get("indice_carga", "")
    vel = parsed.get("indice_velocidade", "")
    carga_vel = f"{carga}{vel}" if carga and vel else ""
    pr = ""
    m_pr = re.search(r"\b(\d{1,2}PR)\b", texto.upper())
    if m_pr:
        pr = m_pr.group(1)
    return {"medida": medida, "carga_vel": carga_vel, "pr": pr}


def _score_match_cert(descricao_cert: str, item_descricao: str) -> int:
    """
    Pontua o quanto a descricao_cert corresponde ao item, tolerando reordenação:
    - medida precisa bater (forte)
    - carga/velocidade e PR (10PR, 8PR...) contam pontos, em qualquer ordem
    """
    if not descricao_cert or not item_descricao:
        return 0

    dc = normalizar_texto(descricao_cert).upper()
    it = normalizar_texto(item_descricao).upper()
    tok = _tokens_item_para_match(it)
    medida = (tok.get("medida") or "").upper()
    carga_vel = (tok.get("carga_vel") or "").upper()
    pr = (tok.get("pr") or "").upper()

    if not medida:
        return 0

    score = 0
    # Medida no início do certificado (mais confiável)
    if dc.startswith(medida):
        score += 80
    elif medida in dc[: max(40, len(medida) + 10)]:
        score += 60
    else:
        return 0  # sem medida, não é o item

    # Tokens complementares podem vir em outra ordem (ex.: 10PR antes de 98/96S)
    if carga_vel and carga_vel in dc:
        score += 25
    if pr and pr in dc:
        score += 20
    # Bônus se toda a descrição do item (sem pontuação) aparece logo no início
    if it and it in dc[:200]:
        score += 10
    return score


def _pattern_efetivo_certificado(c: dict) -> str:
    """Pattern da planilha ou extraído do modelo (ex.: CF710 / AB603... -> CF710)."""
    p = str(c.get("pattern", "")).strip().upper()
    if p:
        return p
    modelo = str(c.get("modelo", "")).replace("\n", " ").strip()
    if not modelo:
        return ""
    modelo_u = modelo.upper()
    if " / " in modelo_u:
        return modelo.split(" / ")[0].strip().upper()
    if " - " in modelo:
        return modelo.split(" - ")[0].strip().upper()
    if "-" in modelo:
        return modelo.split("-")[0].strip().upper()
    return modelo_u[:20].strip()


def _certificado_bate_no_pattern(c: dict, pat_u: str) -> bool:
    """True se a linha do certificado corresponde ao pattern da fatura (CF1100, CF710, etc.)."""
    if not pat_u:
        return False
    if _pattern_efetivo_certificado(c) == pat_u:
        return True
    modelo_u = normalizar_texto(str(c.get("modelo", ""))).upper()
    if not modelo_u:
        return False
    # Ex.: "CF710 / AB603255502", "CF710 - 6938601718838"
    for prefix in (pat_u + " /", pat_u + " -", pat_u + "/", pat_u + "-"):
        if modelo_u.startswith(prefix):
            return True
    return bool(re.search(rf"(?<![A-Z0-9]){re.escape(pat_u)}(?![A-Z0-9])", modelo_u))


def _buscar_certificado_pattern_depois_medida(
    certificados: list[dict],
    pat: str,
    item_descricao: Optional[str],
) -> Optional[dict]:
    """
    Fallback: todas as linhas com o mesmo pattern; escolhe pela medida em descricao_cert.
    Usado quando yaro_code/pattern exato não fecha candidato com a medida correta.
    """
    pat_u = (pat or "").strip().upper()
    if not pat_u or not item_descricao:
        return None
    parsed = parse_descricao_pneu(item_descricao)
    medida_item = (parsed.get("medida") or "").upper().replace(" ", "")
    if not medida_item:
        return None
    base_item = extrair_medida_base_pneu(item_descricao)
    candidatos_pat = [c for c in certificados if _certificado_bate_no_pattern(c, pat_u)]
    if not candidatos_pat:
        return None
    melhor = None
    melhor_score = -1
    for c in candidatos_pat:
        desc_cert = c.get("descricao_cert") or ""
        sc = _score_match_cert(desc_cert, item_descricao)
        if sc < 55 and desc_cert:
            dc = normalizar_texto(desc_cert).upper().replace(" ", "")
            mi = medida_item.replace(" ", "")
            if mi and (mi in dc or (base_item and base_item.upper().replace(" ", "") in dc)):
                sc = max(sc, 55)
        if sc > melhor_score:
            melhor_score = sc
            melhor = c
    if melhor and melhor_score >= 55:
        return melhor
    return None


def _buscar_certificado_por_descricao_global(
    certificados: list[dict],
    item_descricao: Optional[str],
) -> Optional[dict]:
    """
    Fallback amplo por descrição do item (medida/carga/PR), útil quando não há
    pattern confiável no item (ex.: alguns layouts Omni).
    """
    if not item_descricao:
        return None
    melhor = None
    melhor_score = -1
    for c in certificados:
        desc_cert = c.get("descricao_cert") or ""
        sc = _score_match_cert(desc_cert, item_descricao)
        if sc > melhor_score:
            melhor_score = sc
            melhor = c
    if melhor and melhor_score >= 60:
        return melhor
    return None


def _filtrar_certificados_por_marca(certificados: list[dict], marca_preferida: str) -> list[dict]:
    """Filtra certificados por marca (coluna Marca), case-insensitive."""
    marca_u = normalizar_texto(marca_preferida or "").upper()
    if not marca_u:
        return certificados
    filtrados = []
    for c in certificados:
        marca_c = normalizar_texto(str(c.get("marca", ""))).upper()
        if marca_c == marca_u:
            filtrados.append(c)
    return filtrados


def _buscar_certificado_impl(
    certificados: list[dict],
    yaro_code: str,
    pattern: str,
    item_descricao: Optional[str] = None,
) -> Optional[dict]:
    """
    Busca o certificado que corresponde ao item.
    Prioriza match por yaro_code + pattern + medida/descrição do item (descricao_cert no Excel),
    para não misturar dados de outro pneu (ex.: 31x10.50R15LT não pode usar cert de 205/75R16C).
    """
    yaro = str(yaro_code).strip() if yaro_code else ""
    pat = str(pattern).strip() if pattern else ""
    medida_item = ""
    if item_descricao:
        parsed = parse_descricao_pneu(item_descricao)
        medida_item = parsed.get("medida", "") or item_descricao

    candidatos = []
    yaro_vazio = not yaro or yaro == "0"
    cy_vazio = lambda cy: not cy or cy == "0"
    pat_u = pat.upper() if pat else ""
    for c in certificados:
        cy = str(c.get("yaro_code", "")).strip()
        cp = _pattern_efetivo_certificado(c)
        match_yaro_pat = False
        if yaro and pat and cy == yaro and cp == pat_u:
            match_yaro_pat = True
        elif yaro and cy == yaro and (not pat or cp == pat_u):
            match_yaro_pat = True
        elif pat_u and cp == pat_u and (not yaro or cy == yaro or (yaro_vazio and cy_vazio(cy))):
            match_yaro_pat = True
        # Alguns certificados não vêm com Código YARO preenchido; se houver descrição do item,
        # aceitar candidatos por pattern e filtrar pela medida no score (evita misturar itens).
        elif item_descricao and pat_u and cp == pat_u and cy_vazio(cy):
            match_yaro_pat = True
        elif item_descricao and pat_u and cy_vazio(cy) and _certificado_bate_no_pattern(c, pat_u):
            match_yaro_pat = True
        if not match_yaro_pat:
            continue
        candidatos.append(c)

    if not candidatos:
        # Quando não há pattern no item (comum em Omni), tentar casar por descrição completa.
        if item_descricao and not pat:
            return _buscar_certificado_por_descricao_global(certificados, item_descricao)
        return None
    # Se temos descrição do item e os certs têm descricao_cert, escolher o cert desta medida (tolerante a ordem)
    if medida_item or item_descricao:
        melhor = None
        melhor_score = 0
        for c in candidatos:
            desc_cert = c.get("descricao_cert") or ""
            sc = _score_match_cert(desc_cert, item_descricao or medida_item)
            if sc > melhor_score:
                melhor_score = sc
                melhor = c
        if melhor and melhor_score >= 60:
            return melhor
        # Há candidatos por yaro/pattern mas nenhum é desta medida — não usar cert de outro item
        if any(c.get("descricao_cert") for c in candidatos):
            fb = _buscar_certificado_pattern_depois_medida(certificados, pat, item_descricao)
            if fb:
                return fb
            return None
    return candidatos[0]


def buscar_certificado(
    certificados: list[dict],
    yaro_code: str,
    pattern: str,
    item_descricao: Optional[str] = None,
    marca_preferida: str = "",
) -> Optional[dict]:
    """Busca certificado: fluxo yaro/pattern + medida; se vazio, fallback pattern → medida em toda a base."""
    base = _filtrar_certificados_por_marca(certificados, marca_preferida)
    if marca_preferida and not base:
        return None
    r = _buscar_certificado_impl(base, yaro_code, pattern, item_descricao)
    if r is not None:
        return r
    pat = str(pattern or "").strip()
    if pat and item_descricao:
        return _buscar_certificado_pattern_depois_medida(base, pat, item_descricao)
    return None


def _valor_cert(cert: Optional[dict], key: str, default: str = "") -> str:
    if not cert:
        return default
    v = cert.get(key)
    if v is None:
        return default
    return str(v).strip() or default


def _codigo_barras_valido(cb: str) -> bool:
    """
    True apenas quando há código de barras utilizável (numérico típico de EAN/UPC).
    Textos como "Não existente", "N/A" ou sequências curtas contam como ausência.
    """
    s = (cb or "").strip()
    if not s:
        return False
    s_lower = s.lower()
    if any(
        frag in s_lower
        for frag in (
            "não exist",
            "nao exist",
            "inexistente",
            "não cadastr",
            "nao cadastr",
            "sem código",
            "sem codigo",
            "não se aplica",
            "nao se aplica",
            "n/a",
            "s/n",
            "n.i.",
            "não inform",
            "nao inform",
        )
    ):
        return False
    digits = re.sub(r"\D", "", s)
    return len(digits) >= 8


def _fragmento_modelo_e_codigo_barras(modelo: str, codigo_barras: str) -> str:
    """Trecho 'MODELO ...' com ou sem sufixo de código de barras."""
    modelo_txt = (modelo or "").strip()
    cb = (codigo_barras or "").strip()
    if _codigo_barras_valido(cb):
        # Evita duplicar quando o código já faz parte do texto do modelo.
        if cb and cb in modelo_txt:
            return f"MODELO {modelo_txt}"
        return f"MODELO {modelo_txt} - {cb} (codigo de barras)"
    return f"MODELO {modelo_txt}"


def _montar_prefixo_descricao_40111000(
    medida: str,
    indice_carga: str,
    indice_velocidade: str,
    pr: str,
    tem_tl: bool,
) -> str:
    """Prefixo antes de // : ex. 175/75R13 4PR 85T TL, //"""
    partes = [medida]
    if pr:
        partes.append(pr)
    if indice_carga and indice_velocidade:
        partes.append(f"{indice_carga}{indice_velocidade}")
    elif indice_carga:
        partes.append(indice_carga)
    if tem_tl:
        partes.append("TL")
    return " ".join(partes) + ", //"


def _resumo_medida_pneu(
    medida: str,
    indice_carga: str,
    indice_velocidade: str,
    pr: str,
    tem_tl: bool,
) -> str:
    """Resumo curto da medida para prefixo do cabeçalho."""
    partes = [medida]
    medida_up = normalizar_texto(medida).upper()
    if pr:
        if pr.upper() in medida_up:
            pr = ""
    if pr:
        partes.append(pr)
    if indice_carga and indice_velocidade:
        carga_vel = f"{indice_carga}{indice_velocidade}".upper()
        if carga_vel not in medida_up:
            partes.append(carga_vel)
    elif indice_carga:
        if indice_carga.upper() not in medida_up:
            partes.append(indice_carga)
    if tem_tl:
        if "TL" not in medida_up:
            partes.append("TL")
    return " ".join([p for p in partes if p]).strip()


def _resumo_medida_do_certificado(cert: Optional[dict]) -> str:
    """
    Extrai o prefixo de medida/índices diretamente da descricao_cert do INMETRO.
    Ex.: "225/65R16C 8PR ... (RRC):(X);..." -> "225/65R16C 8PR ..."
    """
    bruto = _valor_cert(cert, "descricao_cert", "")
    if not bruto:
        return ""
    texto = normalizar_texto(bruto)
    texto = re.split(r"\(\s*RRC\s*\)\s*:", texto, maxsplit=1, flags=re.IGNORECASE)[0]
    return texto.strip(" ,;")


def _cabecalho_do_certificado(cert: Optional[dict]) -> str:
    """
    Retorna o cabeçalho completo vindo da descricao_cert (até antes do //),
    preservando o formato do certificado.
    Ex.: "33X12.50R20LT 114Q TL (RRC); (G):A; n/adB"
    """
    bruto = _valor_cert(cert, "descricao_cert", "")
    if not bruto:
        return ""
    texto = normalizar_texto(bruto)
    texto = texto.split("//", 1)[0]
    return texto.strip(" ,.;")


def _indices_do_certificado(cert: Optional[dict]) -> tuple[str, str, str]:
    """
    Extrai RRC, G e dB da descricao_cert quando não vierem em colunas separadas.
    Aceita formatos como:
    - "(RRC):(E);(G):C;72dB"
    - "(RRC):E; (G):C; 72dB"
    """
    bruto = _valor_cert(cert, "descricao_cert", "")
    if not bruto:
        return ("", "", "")
    t = normalizar_texto(bruto)
    m = re.search(
        r"\(\s*RRC\s*\)\s*:\s*\(?\s*([A-Z])\s*\)?\s*;\s*\(\s*G\s*\)\s*:\s*\(?\s*([A-Z])\s*\)?\s*;\s*([0-9]{2,3})\s*dB",
        t,
        re.IGNORECASE,
    )
    if not m:
        return ("", "", "")
    return (m.group(1).upper(), m.group(2).upper(), m.group(3))


def _cabecalho_certificado_compativel_com_item(item_desc: str, cabecalho_cert: str) -> bool:
    """
    True quando o cabeçalho do certificado representa o mesmo item da fatura.
    Compara medida e, quando presentes em ambos, carga/velocidade/PR/TL.
    """
    if not item_desc or not cabecalho_cert:
        return False
    p_item = parse_descricao_pneu(item_desc)
    p_cert = parse_descricao_pneu(cabecalho_cert)

    medida_item = p_item.get("medida", "") or item_desc
    medida_cert = p_cert.get("medida", "") or cabecalho_cert
    base_item = extrair_medida_base_pneu(medida_item)
    base_cert = extrair_medida_base_pneu(medida_cert)
    if not (base_item and base_cert and _bases_medida_equivalentes(base_item, base_cert)):
        return False

    carga_item = (p_item.get("indice_carga", "") or "").upper()
    carga_cert = (p_cert.get("indice_carga", "") or "").upper()
    if carga_item and carga_cert and carga_item != carga_cert:
        return False

    vel_item = (p_item.get("indice_velocidade", "") or "").upper()
    vel_cert = (p_cert.get("indice_velocidade", "") or "").upper()
    if vel_item and vel_cert and vel_item != vel_cert:
        return False

    pr_item = (p_item.get("pr", "") or "").upper()
    pr_cert = (p_cert.get("pr", "") or "").upper()
    if pr_item and pr_cert and pr_item != pr_cert:
        return False

    tl_item = bool(p_item.get("tem_tl", False))
    tl_cert = bool(p_cert.get("tem_tl", False))
    if tl_item and not tl_cert:
        return False

    return True


def montar_descricao_detalhada(
    item: dict,
    cert: Optional[dict],
    certificados: list[dict],
) -> tuple[str, str]:
    """
    Monta a descrição detalhada do item e texto de informações complementares (processo judicial).
    Retorna (descricao_detalhada, informacoes_complementares).
    """
    desc_raw = item.get("descricao", "")
    ncm = item.get("ncm", "")
    parsed = parse_descricao_pneu(desc_raw)
    medida = parsed["medida"] or desc_raw
    indice_carga = parsed["indice_carga"]
    indice_velocidade = parsed["indice_velocidade"]
    pr_item = parsed.get("pr", "")
    tem_tl = parsed.get("tem_tl", False)

    if cert is None:
        cert = buscar_certificado(
            certificados,
            item.get("yaro_code", ""),
            item.get("pattern", ""),
            item.get("descricao", ""),
            str(item.get("marca", "")),
        )

    marca = _valor_cert(cert, "marca", str(item.get("marca", "")).strip() or "YARO")
    modelo = _valor_cert(cert, "modelo", "")
    familia = _valor_cert(cert, "familia", "")
    cert_inmetro = _valor_cert(cert, "certificado_inmetro", "") or _valor_cert(cert, "certificado", "") or "N/I"
    registro_in = _valor_cert(cert, "registro_in", "") or "N/I"
    rrc = _valor_cert(cert, "rrc", "X")
    g = _valor_cert(cert, "g", "X")
    db = _valor_cert(cert, "db", "XX")
    codigo_barras = _valor_cert(cert, "codigo_barras", "")
    cabecalho_cert = _cabecalho_do_certificado(cert)
    resumo_cert = _resumo_medida_do_certificado(cert)
    rrc_cert, g_cert, db_cert = _indices_do_certificado(cert)
    if (not rrc or rrc.upper() == "X") and rrc_cert:
        rrc = rrc_cert
    if (not g or g.upper() == "X") and g_cert:
        g = g_cert
    if (not db or db.upper() == "XX") and db_cert:
        db = db_cert
    if resumo_cert:
        parsed_cert = parse_descricao_pneu(resumo_cert)
        medida_cert = parsed_cert.get("medida", "")
        cert_compativel_com_item = _cabecalho_certificado_compativel_com_item(desc_raw, resumo_cert)
        # Evita "deslocar" a coluna J em relação à A quando o certificado veio de outra medida.
        if not cert_compativel_com_item:
            resumo_cert = ""
            cabecalho_cert = ""
        if medida_cert and cert_compativel_com_item:
            medida = medida_cert
        if not indice_carga and cert_compativel_com_item:
            indice_carga = parsed_cert.get("indice_carga", "")
        if not indice_velocidade and cert_compativel_com_item:
            indice_velocidade = parsed_cert.get("indice_velocidade", "")
        if not pr_item and cert_compativel_com_item:
            pr_item = parsed_cert.get("pr", "")
        if not tem_tl and cert_compativel_com_item:
            tem_tl = parsed_cert.get("tem_tl", False)

    # Pneus de passeio NCM 4011.10.00 (um índice de carga): usa layout base de passageiro.
    if ncm == "4011.10.00":
        resumo_medida = resumo_cert or _resumo_medida_pneu(medida, indice_carga, indice_velocidade, pr_item, tem_tl)
        cabecalho = cabecalho_cert or f"{resumo_medida or medida} (RRC):{rrc}; (G):{g}; {db}dB"
        texto_4011 = (
            f"{cabecalho} // PNEU NOVO DE BORRACHA PARA USO EM CARROS DE PASSEIO E VEÍCULOS LEVES, "
            f"MEDIDA {medida}, CONSTRUÇÃO RADIAL COM INDICE DE CARGA {indice_carga} E INDICE DE VELOCIDADE {indice_velocidade}, "
            f"MARCA {marca}, MODELO {modelo}, FAMILIA {familia}, DO CERTIFICADO INMETRO {cert_inmetro}, REGISTRO IN: {registro_in}"
        )
        return (texto_4011, "")

    # Tipo por regra operacional:
    # - LT/...C é sempre carga reforçada (Light Truck judicial), inclusive aro >= 20.
    # - Demais 4011.20.90 com aro >= 20 => TBR; restante => carga reforçada.
    # - NCM diferente de 4011.20.90 sem LT/C => passeio.
    desc_up = normalizar_texto(desc_raw).upper()
    medida_up = medida.upper()
    eh_lt_ou_c = (
        medida_up.startswith("LT")
        or "LT" in medida_up
        or medida_up.endswith("C")
        or bool(re.search(r"\bLT\d", desc_up))
    )
    aro = extrair_aro_da_medida(medida or desc_raw)
    eh_aro_grande = aro >= 20.0
    eh_4011_20_90 = ncm == "4011.20.90"
    if eh_lt_ou_c:
        tipo = "carga_lt_c"
    elif eh_4011_20_90:
        tipo = "tbr" if eh_aro_grande else "carga_lt_c"
    else:
        tipo = "passeio"

    # Cabeçalho comum: MEDIDA (RRC):X; (G):X; XXdB (outros NCM / não 4011.10.00)
    resumo_medida = resumo_cert or _resumo_medida_pneu(medida, indice_carga, indice_velocidade, pr_item, tem_tl)
    cabecalho = cabecalho_cert or f"{resumo_medida or medida} (RRC):{rrc}; (G):{g}; {db}dB"

    if tipo == "passeio":
        texto = (
            f"{cabecalho} // PNEU NOVO DE BORRACHA PARA USO EM CARROS DE PASSEIO E VEÍCULOS LEVES, "
            f"MEDIDA {medida}, CONSTRUÇÃO RADIAL COM INDICE DE CARGA {indice_carga} E INDICE DE VELOCIDADE {indice_velocidade}, "
            f"MARCA {marca}, MODELO {modelo}, FAMILIA {familia}, DO CERTIFICADO INMETRO {cert_inmetro}, REGISTRO IN: {registro_in}"
        )
        return (texto, "")

    if tipo == "carga_lt_c":
        frag_modelo = _fragmento_modelo_e_codigo_barras(modelo, codigo_barras)
        texto = (
            f"{cabecalho} // PNEU NOVO DE BORRACHA, *PNEU DE CARGA REFORCADO, COM ESTRUTURA E CAPACIDADE DE CARGA,* "
            f"PARA USO EM VEÍCULOS COMERCIAIS, LEVES E REBOCADOS, MEDIDA {medida}, CONSTRUÇÃO RADIAL COM INDICE DE CARGA {indice_carga} E INDICE DE VELOCIDADE {indice_velocidade}, "
            f"MARCA {marca}, {frag_modelo}, FAMILIA {familia}, DO CERTIFICADO INMETRO {cert_inmetro}, REGISTRO IN: {registro_in}"
        )
        return (texto, TEXTO_PROCESSO_JUDICIAL)

    # TBR (ônibus/caminhões)
    frag_modelo = _fragmento_modelo_e_codigo_barras(modelo, codigo_barras)
    texto = (
        f"{cabecalho} // PNEU NOVO DE BORRACHA DO TIPO UTILIZADO EM ÔNIBUS (AUTOCARROS) OU CAMINHÕES, "
        f"MEDIDA {medida}, CONSTRUÇÃO RADIAL COM INDICE DE CARGA {indice_carga} E INDICE DE VELOCIDADE {indice_velocidade}, "
        f"MARCA {marca}, {frag_modelo}, FAMILIA {familia}, DO CERTIFICADO INMETRO {cert_inmetro}, REGISTRO IN: {registro_in}"
    )
    return (texto, TEXTO_ANALISAR_ANTI_DUMPING)


def aplicar_descricoes_detalhadas(dados: dict, caminho_certificados: Optional[str] = None) -> None:
    """
    Para cada item em dados['itens'], busca certificado e preenche
    descricao_detalhada e informacoes_complementares (quando houver processo judicial).
    Modifica dados in-place.
    """
    certs = carregar_certificados_inmetro(caminho_certificados)
    for item in dados.get("itens", []):
        cert = buscar_certificado(
            certs,
            item.get("yaro_code", ""),
            item.get("pattern", ""),
            item.get("descricao", ""),
            str(item.get("marca", "")),
        )
        desc, compl = montar_descricao_detalhada(item, cert, certs)
        item["descricao_detalhada"] = desc
        item["informacoes_complementares"] = compl


def extrair_totais(texto: str) -> dict:
    """Extrai BRANDS, Total Amount, Total Unit, Total Containers, Total weight."""
    texto_norm = normalizar_texto(texto)
    totais = {}

    m = re.search(r'BRANDS:\s*\.?\s*([^.]+?)(?=\.|Total)', texto_norm, re.IGNORECASE)
    totais["marcas"] = normalizar_texto(m.group(1)) if m else ""

    m = re.search(r'Total\s+Amount\s*:\s*(\$?[\d,]+\.?\d*)', texto_norm, re.IGNORECASE)
    totais["valor_total"] = (m.group(1).replace('$', '').replace(',', '').strip() if m else "")

    m = re.search(r'Total\s+Unit\s*:\s*([\d,]+)', texto_norm, re.IGNORECASE)
    totais["total_unidades"] = (m.group(1).replace(',', '').strip() if m else "")

    m = re.search(r'Total\s+Containers\s*:\s*([^\s.]+)', texto_norm, re.IGNORECASE)
    totais["total_conteineres"] = m.group(1).strip() if m else ""

    m = re.search(r'Total\s+Gross/Net\s+weight\s*:\s*([\d,\.]+)', texto_norm, re.IGNORECASE)
    totais["peso_total_geral"] = (m.group(1).replace(',', '').strip() if m else "")

    return totais


def _eh_layout_latitude(texto: str) -> bool:
    t = (texto or "").upper()
    return "LATITUDE(SINGAPORE) INDUSTRIAL PTE.LTD." in t and "PRODUCT CODE" in t and "INVOICE NO." in t


def _eh_layout_omni(texto: str) -> bool:
    t = (texto or "").upper()
    return "OMNI UNITED (S) PTE LTD" in t and "COMMERCIAL INVOICE" in t and "ITEM CODE" in t


def extrair_cabecalho_latitude(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    campos = {}
    m = re.search(r"INVOICE\s+NO\.\s*([A-Z0-9\-\/]+)", texto_norm, re.IGNORECASE)
    if m:
        campos["numero"] = m.group(1).strip()
    m = re.search(r"DATE:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{4})", texto_norm, re.IGNORECASE)
    if m:
        campos["data"] = normalizar_texto(m.group(1))
    m = re.search(r"FROM:\s*([A-Z\s]+?)\s+TO:\s*([A-Z\s]+?)(?=\s+PRODUCT\s+CODE|\s+GRAND\s+TOTAL|$)", texto_norm, re.IGNORECASE)
    if m:
        campos["local_carregamento"] = normalizar_texto(m.group(1))
        campos["porto_descarga"] = normalizar_texto(m.group(2))
    m = re.search(r"PAYMENT\s+TERMS:\s*(.+?)\s*$", texto_norm, re.IGNORECASE)
    if m:
        campos["condicoes_pagamento"] = normalizar_texto(m.group(1))
    m = re.search(r"Country\s+of\s+ORIGINAL\s+([A-Z]+)", texto_norm, re.IGNORECASE)
    if m:
        campos["pais_origem"] = m.group(1).strip().upper()
    return campos


def extrair_importador_latitude(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    nome = ""
    endereco = ""
    cep = ""
    cnpj = ""
    m = re.search(r"THE\s+BUYER:\s*(.+?)\s+INVOICE\s+NO\.", texto_norm, re.IGNORECASE)
    if m:
        nome = normalizar_texto(m.group(1))
    m = re.search(r"ADDRESS:\s*(.+?)\s+CNPJ:", texto_norm, re.IGNORECASE)
    if m:
        endereco = normalizar_texto(m.group(1))
        m_cep = re.search(r"ZIP\s+CODE:\s*([0-9\.\-]+)", endereco, re.IGNORECASE)
        if m_cep:
            cep = m_cep.group(1).strip()
    m = re.search(r"CNPJ:\s*([0-9\.\-\/]+)", texto_norm, re.IGNORECASE)
    if m:
        cnpj = m.group(1).strip()
    return {"nome": nome, "endereco": endereco, "cep": cep, "cnpj": cnpj}


def extrair_fabricante_latitude(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    nome = ""
    endereco = ""
    m = re.search(r"The\s+Manufacturer:\s*(.+?)\s+ADDRESS:", texto_norm, re.IGNORECASE)
    if m:
        nome = normalizar_texto(m.group(1))
    m = re.search(r"The\s+Manufacturer:.*?ADDRESS:\s*(.+?)\s+FROM:", texto_norm, re.IGNORECASE)
    if m:
        endereco = normalizar_texto(m.group(1))
    return {"nome": nome, "endereco": endereco}


def extrair_itens_latitude(texto: str) -> list[dict]:
    itens = []
    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    padrao_linha = re.compile(
        r"^(?P<product_code>\d+)\s+"
        r"(?P<ean>\d+)\s+"
        r"(?P<brand>[A-Z0-9\-]+)\s+"
        r"(?P<pattern>[A-Z0-9\-]+)\s+"
        r"(?P<size>[0-9A-Z\/\.\-]+)\s+"
        r"(?P<li>\d+/\d+)\s+"
        r"(?P<si>[A-Z])\s+"
        r"(?P<pr>\d+PR)\s+"
        r"(?P<family>[A-Z0-9]+)\s+"
        r"(?P<cert_no>[A-Z0-9\-\/]+)\s+"
        r"(?P<registro>\d+/\d{4})\s+"
        r"(?P<qty>[\d,]+)\s+"
        r"(?P<container>[\d\.]+)\s+\$"
        r"(?P<fob_unit>[\d,]+\.\d{2})\s+\$"
        r"(?P<fob_total>[\d,]+\.\d{2})$",
        re.IGNORECASE,
    )
    for linha in linhas:
        m = padrao_linha.match(normalizar_texto(linha))
        if not m:
            continue
        size = m.group("size").upper()
        li = m.group("li")
        si = m.group("si").upper()
        pr = m.group("pr").upper()
        descricao = f"{size} {li}{si} {pr}"
        qty = int(m.group("qty").replace(",", ""))
        fob_unit = float(m.group("fob_unit").replace(",", ""))
        total = float(m.group("fob_total").replace(",", ""))
        itens.append(
            {
                "descricao": descricao,
                "yaro_code": m.group("product_code"),
                "pattern": m.group("pattern").upper(),
                "ncm": "4011.20.90",
                "quantidade": qty,
                "peso_unitario": 0.0,
                "peso_total": 0.0,
                "fob_unitario_usd": fob_unit,
                "total_fob_usd": total,
            }
        )
    return itens


def extrair_totais_latitude(texto: str, itens: list[dict]) -> dict:
    texto_norm = normalizar_texto(texto)
    totais = {
        "marcas": "",
        "valor_total": "",
        "total_unidades": "",
        "total_conteineres": "",
        "peso_total_geral": "",
    }
    marcas = sorted({i.get("pattern", "") for i in itens if i.get("pattern")})
    totais["marcas"] = ", ".join(marcas)
    m = re.search(r"GRAND\s+TOTAL:\s*([\d,]+)\s+([\d\.]+)\s+\$([\d,]+\.\d{2})", texto_norm, re.IGNORECASE)
    if m:
        totais["total_unidades"] = m.group(1).replace(",", "")
        totais["total_conteineres"] = m.group(2)
        totais["valor_total"] = m.group(3).replace(",", "")
    m = re.search(r"Total\s+Gross\s+Weight\s*([0-9\.,]+)", texto_norm, re.IGNORECASE)
    if m:
        totais["peso_total_geral"] = m.group(1).replace(",", "")
    return totais


def extrair_fatura_latitude(caminho_pdf: str, texto: str) -> dict:
    itens = extrair_itens_latitude(texto)
    return {
        "cabecalho": extrair_cabecalho_latitude(texto),
        "importador": extrair_importador_latitude(texto),
        "conta_bancaria": {},
        "fabricante": extrair_fabricante_latitude(texto),
        "itens": itens,
        "totais": extrair_totais_latitude(texto, itens),
        "arquivo": str(Path(caminho_pdf).name),
    }


def extrair_cabecalho_omni(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    campos = {}
    m = re.search(r"CNPJ:\s*[0-9\.\-\/]+\s+([0-9\/]+)", texto_norm, re.IGNORECASE)
    if m:
        campos["numero"] = m.group(1).strip()
    m = re.search(r"\b([0-9]{1,2}-[A-Za-z]{3}-[0-9]{2})\b", texto_norm, re.IGNORECASE)
    if m:
        campos["data"] = m.group(1)
    m = re.search(r"PORT\s+OF\s+LOADING\s+(.+?)(?=\s+[0-9]{1,2}-[A-Za-z]{3}-[0-9]{2}\b)", texto_norm, re.IGNORECASE)
    if m:
        campos["local_carregamento"] = normalizar_texto(m.group(1))
    m = re.search(r"INCOTERM\s+([A-Z]{3})", texto_norm, re.IGNORECASE)
    if m:
        campos["termos"] = m.group(1).upper()
    m = re.search(r"PAYMENT\s+TERMS\s*:\s*(.+?)(?=\s+DUE\s+DATE:|\s+MANUFACTURER:|$)", texto_norm, re.IGNORECASE)
    if m:
        campos["condicoes_pagamento"] = normalizar_texto(m.group(1))
    m = re.search(r"COUNTRY\s+OF\s+ORIGIN:\s*([A-Z]+)", texto_norm, re.IGNORECASE)
    if m:
        campos["pais_origem"] = m.group(1).upper()
    return campos


def extrair_importador_omni(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    nome = ""
    endereco = ""
    cep = ""
    cnpj = ""

    m = re.search(r"Bill\s+To:.*?((?:[A-Z0-9&\.\-]+\s+){1,12}(?:LTDA|LTD|S\.?A\.?|S\/A))", texto_norm, re.IGNORECASE)
    if m:
        nome = normalizar_texto(m.group(1))
    m = re.search(r"((?:Rodovia|Rua|Av\.?|Avenida).+?Zip\s+Code:\s*[0-9\.\-]+)", texto_norm, re.IGNORECASE)
    if m:
        endereco = normalizar_texto(m.group(1))
        m_cep = re.search(r"Zip\s+Code:\s*([0-9\.\-]+)", endereco, re.IGNORECASE)
        if m_cep:
            cep = m_cep.group(1).strip()
    m = re.search(r"CNPJ:\s*([0-9\.\-\/]+)", texto_norm, re.IGNORECASE)
    if m:
        cnpj = m.group(1).strip()

    return {"nome": nome, "endereco": endereco, "cep": cep, "cnpj": cnpj}


def extrair_conta_bancaria_omni(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    campos = {}
    m = re.search(r"Bank\s+Name:\s*(.+?)\s+Swift\s+Code:", texto_norm, re.IGNORECASE)
    campos["banco_beneficiario"] = normalizar_texto(m.group(1)) if m else ""
    m = re.search(r"Swift\s+Code:\s*([A-Z0-9]+)", texto_norm, re.IGNORECASE)
    campos["swift"] = m.group(1).strip() if m else ""
    m = re.search(r"Beneficiary:\s*(.+?)\s+Account\s+No\.:", texto_norm, re.IGNORECASE)
    campos["beneficiario"] = normalizar_texto(m.group(1)) if m else ""
    m = re.search(r"Account\s+No\.\s*:\s*([A-Z0-9\-]+)", texto_norm, re.IGNORECASE)
    campos["conta_beneficiario"] = m.group(1).strip() if m else ""
    campos["endereco_banco"] = ""
    return campos


def extrair_fabricante_omni(texto: str) -> dict:
    texto_norm = normalizar_texto(texto)
    nome = ""
    endereco = ""
    m = re.search(r"MANUFACTURER:\s*(.+?LTD\.)\s*(.+?)(?=\s+FOR\s+OMNI\s+UNITED)", texto_norm, re.IGNORECASE)
    if m:
        nome = normalizar_texto(m.group(1))
        endereco = normalizar_texto(m.group(2))
    else:
        m = re.search(r"MANUFACTURER:\s*(.+?)(?=\s+FOR\s+OMNI\s+UNITED|$)", texto_norm, re.IGNORECASE)
        if m:
            nome = normalizar_texto(m.group(1))
    return {"nome": nome, "endereco": endereco}


def _extrair_sizes_omni(texto: str) -> list[str]:
    """Extrai valores da coluna Size para usar como part number."""
    sizes = []
    for linha_raw in texto.splitlines():
        linha = normalizar_texto(linha_raw)
        if not linha:
            continue
        if "RENEG.RT+" not in linha.upper():
            continue
        m = re.match(
            r"^(?P<medida>[A-Z0-9xX\/\.\-]+)\s+(?P<carga_vel>\d+(?:/\d+)?[A-Z])(?:\s+(?P<faixa>[A-Z]))?",
            linha,
            re.IGNORECASE,
        )
        if not m:
            continue
        medida = m.group("medida").upper().replace("X", "x")
        carga_vel = (m.group("carga_vel") or "").upper()
        faixa = (m.group("faixa") or "").upper()
        if re.search(r"[0-9]", medida) and "R" in medida:
            size = f"{medida} {carga_vel}".strip()
            if faixa:
                size = f"{size} {faixa}"
            sizes.append(size)
    return sizes


def _normalizar_size_linha_omni(size_raw: str) -> str:
    """
    Normaliza o texto da coluna Size vindo da própria linha da tabela Omni.
    Ex.: "LT265/65R18 117/114Q D RENEG.RT+" -> "LT265/65R18 117/114Q D"
    """
    s = normalizar_texto(size_raw or "")
    if not s:
        return ""
    s_up = s.upper()
    # Remove sufixo comercial/família após o size principal.
    s_up = re.sub(r"\s+RENEG\.RT\+.*$", "", s_up, flags=re.IGNORECASE).strip()
    m = re.search(
        r"(?P<medida>(?:LT|P)?\d{2,4}/\d{2}Z?\s*R\d{2}(?:\.\d)?(?:C|LT)?|(?:LT)?\d+X[\d.]+\s*R\d{2}(?:LT)?)\s+"
        r"(?P<carga_vel>\d+(?:/\d+)?[A-Z])(?:\s+(?P<faixa>[A-Z]))?",
        s_up,
        re.IGNORECASE,
    )
    if not m:
        # Importante: não retornar texto cru.
        # Se não casar com padrão de medida de pneu, não é Size válido.
        return ""
    medida = re.sub(r"\s+", "", m.group("medida").upper()).replace("X", "x")
    carga_vel = (m.group("carga_vel") or "").upper()
    faixa = (m.group("faixa") or "").upper()
    out = f"{medida} {carga_vel}".strip()
    if faixa:
        out = f"{out} {faixa}"
    return out


def _extrair_sizes_por_item_code_omni(texto: str) -> dict[str, list[str]]:
    """
    Extrai Size por Item Code (mantendo ordem de aparição), para alinhar a coluna A
    com a ordem já correta de B..I quando o PDF não traz o Size na mesma linha numérica.
    """
    mapa: dict[str, list[str]] = {}
    padrao_inicio_item = re.compile(
        r"(?:^|\b)(?:\d+\s+)?(?P<code>RANCCN\d{4}|RND\d{4})/?\b(?P<resto>.*)$",
        re.IGNORECASE,
    )
    for linha_raw in texto.splitlines():
        linha = normalizar_texto(linha_raw)
        if not linha:
            continue
        m = padrao_inicio_item.search(linha)
        if not m:
            continue
        code = re.sub(r"[^A-Z0-9]", "", (m.group("code") or "").upper())
        resto = m.group("resto") or ""
        size_norm = _normalizar_size_linha_omni(resto)
        if (
            not code
            or not size_norm
            or "R" not in size_norm.upper()
            or not re.search(r"\d", size_norm)
        ):
            continue
        mapa.setdefault(code, []).append(size_norm)
    return mapa


def _extrair_itens_omni_por_pdfplumber(caminho_pdf: str) -> list[dict]:
    """
    Extração preferencial para Omni via leitura tabular.
    Mantém a ordem exata da fatura (Container + S/N), evitando embaralhamento de part number.
    """
    if pdfplumber is None:
        return []

    padrao_code = re.compile(r"^(RANCCN\d{4}|RND\d{4})$", re.IGNORECASE)

    def _to_float(v: str) -> float:
        s = normalizar_texto(v or "")
        s = s.replace("USD", "").replace("$", "").replace(",", "").strip()
        return float(s) if s else 0.0

    def _to_int(v: str) -> int:
        s = normalizar_texto(v or "").replace(",", "").strip()
        return int(float(s)) if s else 0

    itens: list[dict] = []
    container_seq = 0

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    for row in table:
                        if not row:
                            continue
                        c0 = normalizar_texto(((row[0] if len(row) > 0 else "") or "").replace("\n", " "))
                        c1 = normalizar_texto(((row[1] if len(row) > 1 else "") or "").replace("\n", " "))
                        c2 = normalizar_texto(((row[2] if len(row) > 2 else "") or "").replace("\n", " "))
                        c3 = normalizar_texto(((row[3] if len(row) > 3 else "") or "").replace("\n", " "))
                        c5 = normalizar_texto(((row[5] if len(row) > 5 else "") or "").replace("\n", " "))
                        c6 = normalizar_texto(((row[6] if len(row) > 6 else "") or "").replace("\n", " "))
                        c7 = normalizar_texto(((row[7] if len(row) > 7 else "") or "").replace("\n", " "))
                        c8 = normalizar_texto(((row[8] if len(row) > 8 else "") or "").replace("\n", " "))

                        if c0.upper().startswith("CONTAINER#:"):
                            container_seq += 1
                            continue
                        if not c0.isdigit():
                            continue
                        if not c1:
                            continue

                        code = re.sub(r"[^A-Z0-9]", "", c1.upper())
                        if not padrao_code.match(code):
                            continue

                        part_number = _normalizar_size_linha_omni(c2) or code
                        itens.append(
                            {
                                "descricao": part_number,
                                "yaro_code": code,
                                "marca": "RADAR",
                                "pattern": "",
                                "ncm": "4011.20.90",
                                "quantidade": _to_int(c3),
                                "peso_unitario": _to_float(c7),
                                "peso_total": _to_float(c8),
                                "fob_unitario_usd": _to_float(c5),
                                "total_fob_usd": _to_float(c6),
                                "part_number": part_number,
                                "item_code_omni": code,
                                "_container_seq": container_seq if container_seq > 0 else 10_000,
                                "_sn": int(c0),
                            }
                        )
    except Exception:
        return []

    if not itens:
        return []

    itens.sort(key=lambda x: (int(x.get("_container_seq", 10_000)), int(x.get("_sn", 10_000))))
    for it in itens:
        it.pop("_container_seq", None)
        it.pop("_sn", None)
    return itens


def extrair_itens_omni(texto: str, caminho_pdf: Optional[str] = None) -> list[dict]:
    # Caminho principal: tabela estruturada via pdfplumber (ordem fiel da fatura).
    if caminho_pdf:
        itens_tabela = _extrair_itens_omni_por_pdfplumber(caminho_pdf)
        if itens_tabela:
            return itens_tabela

    # Fallback legado: parser textual.
    itens = []
    sizes_por_code = _extrair_sizes_por_item_code_omni(texto)
    idx_por_code: dict[str, int] = {}
    linhas = [normalizar_texto(l) for l in texto.splitlines() if normalizar_texto(l)]
    pendentes_sem_size: list[int] = []
    container_seq = 0
    container_atual = ""
    em_tabela_itens = False
    padrao_container = re.compile(r"^Container#:\s*(?P<container>[A-Z0-9]+)", re.IGNORECASE)
    padrao_fim_tabela = re.compile(r"^(Less\s+Advance|Total\s+Amount\b|TOTAL\b)", re.IGNORECASE)
    padrao_linha_com_size = re.compile(
        r"^(?P<sn>\d+)\s+(?P<item_code>[A-Z0-9/]+)\s+"
        r"(?P<size>.+?)\s+"
        r"(?P<qty>[\d,]+)\s+PC\s+"
        r"(?P<fob_unit>[\d,]+\.\d{2})\s+USD\s+"
        r"(?P<fob_total>[\d,]+\.\d{2})\s+"
        r"(?P<peso_unit>[\d,]+\.\d{2})\s+"
        r"(?P<peso_total>[\d,]+\.\d{2})$",
        re.IGNORECASE,
    )
    padrao_linha = re.compile(
        r"^(?P<sn>\d+)\s+(?P<item_code>[A-Z0-9/]+)\s+"
        r"(?P<qty>[\d,]+)\s+PC\s+"
        r"(?P<fob_unit>[\d,]+\.\d{2})\s+USD\s+"
        r"(?P<fob_total>[\d,]+\.\d{2})\s+"
        r"(?P<peso_unit>[\d,]+\.\d{2})\s+"
        r"(?P<peso_total>[\d,]+\.\d{2})$",
        re.IGNORECASE,
    )
    for linha in linhas:
        m_container = padrao_container.match(linha)
        if m_container:
            container_seq += 1
            container_atual = m_container.group("container").upper()
            em_tabela_itens = True
            continue
        if em_tabela_itens and padrao_fim_tabela.match(linha):
            em_tabela_itens = False
            continue
        m_size = padrao_linha_com_size.match(linha)
        part_number = ""
        if m_size:
            m = m_size
            part_number = _normalizar_size_linha_omni(m.group("size"))
        else:
            m = padrao_linha.match(linha)
        if not m:
            # Alguns PDFs quebram a coluna Size em linha separada.
            # Vincula esse Size ao próximo item pendente para manter a ordem B..I.
            if em_tabela_itens and pendentes_sem_size:
                size_linha_solta = _normalizar_size_linha_omni(linha)
                if size_linha_solta and "R" in size_linha_solta.upper() and re.search(r"\d", size_linha_solta):
                    idx_item = pendentes_sem_size.pop(0)
                    itens[idx_item]["descricao"] = size_linha_solta
                    itens[idx_item]["part_number"] = size_linha_solta
            continue
        qty = int(m.group("qty").replace(",", ""))
        fob_unit = float(m.group("fob_unit").replace(",", ""))
        fob_total = float(m.group("fob_total").replace(",", ""))
        peso_unit = float(m.group("peso_unit").replace(",", ""))
        peso_total = float(m.group("peso_total").replace(",", ""))
        sn = int(m.group("sn"))
        code = re.sub(r"[^A-Z0-9]", "", (m.group("item_code") or "").upper())
        itens.append(
            {
                "descricao": part_number,
                # Omni: part number vem da coluna Size, mas o vínculo com certificados
                # costuma usar o item code (RANCCN..., RND...).
                "yaro_code": code,
                "marca": "RADAR",
                "pattern": "",
                "ncm": "4011.20.90",
                "quantidade": qty,
                "peso_unitario": peso_unit,
                "peso_total": peso_total,
                "fob_unitario_usd": fob_unit,
                "total_fob_usd": fob_total,
                "part_number": part_number,
                "item_code_omni": code,
                "_container_seq": container_seq if container_seq > 0 else 10_000,
                "_sn": sn,
                "_container": container_atual,
            }
        )
        if not part_number:
            pendentes_sem_size.append(len(itens) - 1)
    itens.sort(key=lambda x: (int(x.get("_container_seq", 10_000)), int(x.get("_sn", 10_000))))
    # Fallback por Item Code apenas para itens ainda sem Size.
    # Não usa fallback global por índice para não embaralhar a ordem.
    for it in itens:
        if (it.get("part_number") or "").strip():
            continue
        code = it.get("item_code_omni", "")
        lst = sizes_por_code.get(code, [])
        idx = idx_por_code.get(code, 0)
        if idx < len(lst):
            it["descricao"] = lst[idx]
            it["part_number"] = lst[idx]
            idx_por_code[code] = idx + 1
        else:
            it["descricao"] = code
            it["part_number"] = code
    for it in itens:
        it.pop("_container_seq", None)
        it.pop("_sn", None)
        it.pop("_container", None)
    return itens


def extrair_totais_omni(texto: str, itens: list[dict]) -> dict:
    texto_norm = normalizar_texto(texto)
    totais = {
        "marcas": "",
        "valor_total": "",
        "total_unidades": "",
        "total_conteineres": "",
        "peso_total_geral": "",
    }
    m = re.search(r"DESCRIPTION\s+OF\s+GOODS:\s*(.+?)\s+COUNTRY\s+OF\s+ORIGIN:", texto_norm, re.IGNORECASE)
    if m:
        totais["marcas"] = normalizar_texto(m.group(1))

    containers = re.findall(r"Container#:\s*([A-Z0-9]+)", texto, re.IGNORECASE)
    if containers:
        totais["total_conteineres"] = str(len(containers))

    for linha in [normalizar_texto(l) for l in texto.splitlines() if normalizar_texto(l)]:
        m = re.match(r"^([0-9,]+)\s+\$?([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})$", linha)
        if m:
            totais["total_unidades"] = m.group(1).replace(",", "")
            totais["valor_total"] = m.group(2).replace(",", "")
            totais["peso_total_geral"] = m.group(3).replace(",", "")

    if not totais["total_unidades"] and itens:
        totais["total_unidades"] = str(sum(i.get("quantidade", 0) for i in itens))
    if not totais["valor_total"] and itens:
        totais["valor_total"] = f"{sum(i.get('total_fob_usd', 0.0) for i in itens):.2f}"
    if not totais["peso_total_geral"] and itens:
        totais["peso_total_geral"] = f"{sum(i.get('peso_total', 0.0) for i in itens):.2f}"
    return totais


def extrair_fatura_omni(caminho_pdf: str, texto: str) -> dict:
    itens = extrair_itens_omni(texto, caminho_pdf=caminho_pdf)
    return {
        "cabecalho": extrair_cabecalho_omni(texto),
        "importador": extrair_importador_omni(texto),
        "conta_bancaria": extrair_conta_bancaria_omni(texto),
        "fabricante": extrair_fabricante_omni(texto),
        "itens": itens,
        "totais": extrair_totais_omni(texto, itens),
        "arquivo": str(Path(caminho_pdf).name),
    }


def extrair_fatura(caminho_pdf: str, fornecedor: str = "auto") -> dict:
    """
    Extrai todas as informações de uma fatura comercial (PDF).
    fornecedor: "auto" (detecta pelo texto), "atlantic", "latitude" ou "omni".
    Retorna um dicionário com: cabecalho, importador, conta_bancaria, fabricante, itens, totais.
    """
    texto = extrair_texto_pdf(caminho_pdf)
    modo = (fornecedor or "auto").strip().lower()
    if modo not in ("auto", "atlantic", "latitude", "omni"):
        modo = "auto"
    if modo == "omni":
        return extrair_fatura_omni(caminho_pdf, texto)
    if modo == "latitude":
        return extrair_fatura_latitude(caminho_pdf, texto)
    if modo == "atlantic":
        return {
            "cabecalho": extrair_cabecalho(texto),
            "importador": extrair_importador(texto),
            "conta_bancaria": extrair_conta_bancaria(texto),
            "fabricante": extrair_fabricante(texto),
            "itens": extrair_itens(texto),
            "totais": extrair_totais(texto),
            "arquivo": str(Path(caminho_pdf).name),
        }
    if _eh_layout_omni(texto):
        return extrair_fatura_omni(caminho_pdf, texto)
    if _eh_layout_latitude(texto):
        return extrair_fatura_latitude(caminho_pdf, texto)
    return {
        "cabecalho": extrair_cabecalho(texto),
        "importador": extrair_importador(texto),
        "conta_bancaria": extrair_conta_bancaria(texto),
        "fabricante": extrair_fabricante(texto),
        "itens": extrair_itens(texto),
        "totais": extrair_totais(texto),
        "arquivo": str(Path(caminho_pdf).name),
    }


def salvar_json(dados: dict, caminho_saida: str) -> None:
    """Salva os dados extraídos em JSON (UTF-8)."""
    with open(caminho_saida, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def salvar_excel(dados: dict, caminho_saida: str) -> None:
    """Gera planilha Excel profissional com Resumo e Itens."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo da Fatura"

    # Estilos mais modernos
    titulo_font = Font(bold=True, size=16, color="1F2933")  # cinza escuro
    secao_font = Font(bold=True, size=12, color="1F2933")
    label_font = Font(bold=True, size=10, color="102A43")
    valor_font = Font(size=10, color="102A43")
    cabecalho_fill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")  # azul petróleo
    cabecalho_font = Font(bold=True, color="FFFFFF", size=11)
    secao_fill = PatternFill(start_color="E4E7EB", end_color="E4E7EB", fill_type="solid")  # cinza claro
    linha_par_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    linha_impar_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    total_fill = PatternFill(start_color="D9E2EC", end_color="D9E2EC", fill_type="solid")
    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def escrever_secao(ws, linha: int, titulo: str, itens: list[tuple]) -> int:
        ws.cell(linha, 1, titulo).font = secao_font
        ws.cell(linha, 1).fill = secao_fill
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
        linha += 1
        for label, valor in itens:
            if label and valor is not None and str(valor).strip():
                ws.cell(linha, 1, label).font = label_font
                ws.cell(linha, 2, valor).font = valor_font
                ws.cell(linha, 1).border = borda_fina
                ws.cell(linha, 2).border = borda_fina
                ws.cell(linha, 1).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(linha, 2).alignment = Alignment(wrap_text=True, vertical="top")
                linha += 1
        return linha + 1

    linha = 1
    # Título
    titulo_cell = ws_resumo.cell(linha, 1, "FATURA COMERCIAL — DADOS EXTRAÍDOS")
    titulo_cell.font = titulo_font
    titulo_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_resumo.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    ws_resumo.row_dimensions[linha].height = 26
    linha += 2

    cab = dados.get("cabecalho", {})
    linha = escrever_secao(ws_resumo, linha, "Cabeçalho", [
        ("Número da fatura", cab.get("numero")),
        ("Data", cab.get("data")),
        ("Condições de pagamento", cab.get("condicoes_pagamento")),
        ("Local de carregamento", cab.get("local_carregamento")),
        ("Termos (Incoterm)", cab.get("termos")),
        ("Porto de descarga", cab.get("porto_descarga")),
        ("País de origem", cab.get("pais_origem")),
    ])

    imp = dados.get("importador", {})
    linha = escrever_secao(ws_resumo, linha, "Importador / Comprador", [
        ("Nome / Razão social", imp.get("nome")),
        ("Endereço", imp.get("endereco")),
        ("CEP", imp.get("cep")),
        ("CNPJ", imp.get("cnpj")),
    ])

    conta = dados.get("conta_bancaria", {})
    linha = escrever_secao(ws_resumo, linha, "Dados bancários", [
        ("Banco beneficiário", conta.get("banco_beneficiario")),
        ("Endereço do banco", conta.get("endereco_banco")),
        ("Código SWIFT", conta.get("swift")),
        ("Beneficiário", conta.get("beneficiario")),
        ("Conta do beneficiário", conta.get("conta_beneficiario")),
    ])

    fab = dados.get("fabricante", {})
    linha = escrever_secao(ws_resumo, linha, "Fabricante", [
        ("Nome", fab.get("nome")),
        ("Endereço", fab.get("endereco")),
    ])

    totais = dados.get("totais", {})
    linha = escrever_secao(ws_resumo, linha, "Totais gerais", [
        ("Marcas", totais.get("marcas")),
        ("Valor total (US$)", totais.get("valor_total")),
        ("Total de unidades", totais.get("total_unidades")),
        ("Total de contêineres", totais.get("total_conteineres")),
        ("Peso total (kg)", totais.get("peso_total_geral")),
    ])

    ws_resumo.column_dimensions["A"].width = 30
    ws_resumo.column_dimensions["B"].width = 70
    ws_resumo.column_dimensions["C"].width = 5

    # Aba Itens
    ws_itens = wb.create_sheet("Itens da fatura", 1)
    colunas = [
        ("Descrição", "descricao"),
        ("Código YARO", "yaro_code"),
        ("Pattern", "pattern"),
        ("NCM", "ncm"),
        ("Quantidade", "quantidade"),
        ("Peso unit. (kg)", "peso_unitario"),
        ("Peso total (kg)", "peso_total"),
        ("FOB unit. (US$)", "fob_unitario_usd"),
        ("Total FOB (US$)", "total_fob_usd"),
        ("Descrição detalhada", "descricao_detalhada"),
        ("Descrição (origem)", "descricao_origem"),
        ("Informações complementares", "informacoes_complementares"),
    ]
    for col, (nome, _) in enumerate(colunas, 1):
        cell = ws_itens.cell(1, col, nome)
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina
    ws_itens.row_dimensions[1].height = 30
    # Congela o cabeçalho e aplica filtros
    ws_itens.freeze_panes = "A2"
    ws_itens.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"
    for row_idx, item in enumerate(dados.get("itens", []), 2):
        row_fill = linha_par_fill if row_idx % 2 == 0 else linha_impar_fill
        ws_itens.row_dimensions[row_idx].height = None
        for col_idx, (_, key) in enumerate(colunas, 1):
            val = item.get(key)
            cell = ws_itens.cell(row_idx, col_idx, val)
            cell.border = borda_fina
            cell.fill = row_fill
            if key in ("fob_unitario_usd", "total_fob_usd") and isinstance(val, (int, float)):
                cell.number_format = '"US$"#,##0.00'
            elif key in ("peso_unitario", "peso_total") and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            elif key == "quantidade" and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            if key in ("descricao", "descricao_detalhada", "informacoes_complementares"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif key in ("quantidade", "peso_unitario", "peso_total", "fob_unitario_usd", "total_fob_usd"):
                cell.alignment = Alignment(horizontal="right", wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(colunas) + 1):
        largura_padrao = 18
        if col == 1:
            largura_padrao = 30
        if col in (10, 12):  # Descrição detalhada, Informações complementares
            largura_padrao = 55
        if col == 11:  # Descrição (origem): Catálogo / Não encontrada
            largura_padrao = 16
        ws_itens.column_dimensions[get_column_letter(col)].width = largura_padrao

    # Linha de totais na aba Itens
    total_linha = len(dados.get("itens", [])) + 2
    ws_itens.row_dimensions[total_linha].height = 24
    ws_itens.cell(total_linha, 1, "TOTAL").font = secao_font
    ws_itens.cell(total_linha, 5, sum(i.get("quantidade", 0) for i in dados.get("itens", [])))
    ws_itens.cell(total_linha, 6, "").number_format = "#,##0.00"
    ws_itens.cell(total_linha, 7, sum(i.get("peso_total", 0) for i in dados.get("itens", []))).number_format = "#,##0.00"
    ws_itens.cell(total_linha, 8, "")
    ws_itens.cell(total_linha, 9, sum(i.get("total_fob_usd", 0) for i in dados.get("itens", []))).number_format = '"US$"#,##0.00'
    for c in range(1, len(colunas) + 1):
        ws_itens.cell(total_linha, c).font = secao_font
        ws_itens.cell(total_linha, c).border = borda_fina
        ws_itens.cell(total_linha, c).fill = total_fill

    wb.save(caminho_saida)


def pasta_execucao() -> Path:
    """Retorna pasta do script (python) ou do executável (PyInstaller)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resolver_recurso_app(nome_arquivo: str) -> Optional[Path]:
    """
    Localiza arquivos de recurso (logo, etc.) tanto no modo script quanto no executável.
    """
    candidatos: list[Path] = []
    base_exec = pasta_execucao()
    candidatos.append(base_exec / nome_arquivo)
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidatos.append(Path(meipass) / nome_arquivo)
    else:
        candidatos.append(Path(__file__).resolve().parent / nome_arquivo)
    for p in candidatos:
        if p.is_file():
            return p
    return None


def pasta_saida_arquivos(pasta_base: Path) -> Path:
    """
    Resolve a pasta padrão de saída.
    Prioridade:
    1) <pasta_base>/Saida
    2) <pasta_base>/../Saida
    3) <pasta_base>/YARO_APP/Saida
    4) <pasta_base>/../YARO_APP/Saida
    5) fallback: cria <pasta_base>/Saida
    """
    candidatos = [
        pasta_base / "Saida",
        pasta_base.parent / "Saida",
        pasta_base / "YARO_APP" / "Saida",
        pasta_base.parent / "YARO_APP" / "Saida",
    ]
    for p in candidatos:
        if p.is_dir():
            return p
    destino = pasta_base / "Saida"
    destino.mkdir(parents=True, exist_ok=True)
    return destino


def _pastas_busca_arquivos_apoio(pasta_app: Path, pasta_pdf: Path) -> list[Path]:
    """Pastas candidatas para localizar catálogo e certificados entre máquinas."""
    candidatos = [
        pasta_app,
        pasta_app.parent,
        pasta_app.parent.parent,
        pasta_app / "YARO_APP",
        pasta_app.parent / "YARO_APP",
        Path.cwd(),
        Path.cwd().parent,
        pasta_pdf,
        pasta_pdf.parent,
    ]
    saida: list[Path] = []
    vistos: set[str] = set()
    for p in candidatos:
        try:
            rp = p.resolve()
        except Exception:
            rp = p
        key = str(rp).lower()
        if key in vistos:
            continue
        vistos.add(key)
        if rp.is_dir():
            saida.append(rp)
    return saida


def _resolver_arquivo_apoio(
    pastas_busca: list[Path],
    nome_exato: str,
    padroes_fallback: list[str],
) -> Optional[Path]:
    for pasta in pastas_busca:
        p = pasta / nome_exato
        if p.is_file():
            return p
    for pasta in pastas_busca:
        for padrao in padroes_fallback:
            for p in pasta.glob(padrao):
                if p.is_file():
                    return p
    return None


def processar_pdf(
    pdf_path: Path,
    pasta_base: Optional[Path] = None,
    fornecedor: str = "auto",
) -> tuple[Path, Path, dict]:
    """Processa um PDF e gera os arquivos de saída (JSON/Excel). fornecedor: auto, atlantic, latitude ou omni."""
    pasta = pasta_base or pasta_execucao()
    if not pdf_path.is_file():
        raise FileNotFoundError(f"Arquivo não encontrado: {pdf_path}")

    dados = extrair_fatura(str(pdf_path), fornecedor=fornecedor)

    # Catálogo/certificados: busca robusta em múltiplas pastas (evita diferença entre PCs).
    pastas_busca = _pastas_busca_arquivos_apoio(pasta, pdf_path.parent)
    catalogo_path = _resolver_arquivo_apoio(
        pastas_busca,
        CATALOGO_PADRAO,
        ["*catalogo*produtos*importacao*.xlsx"],
    )
    cert_path = _resolver_arquivo_apoio(
        pastas_busca,
        "CERTIFICADOS INMETRO 2026.xlsx",
        ["*CERTIFICADOS*INMETRO*.xlsx"],
    )
    aplicar_descricoes_do_catalogo(
        dados,
        str(catalogo_path) if catalogo_path else None,
        str(cert_path) if cert_path else None,
    )

    pasta_saida = pasta_saida_arquivos(pasta)
    saida_json = pasta_saida / f"{pdf_path.stem} - Dados.json"
    saida_excel = pasta_saida / f"{pdf_path.stem} - Fatura.xlsx"
    salvar_json(dados, str(saida_json))
    salvar_excel(dados, str(saida_excel))
    return saida_json, saida_excel, dados


def abrir_interface() -> None:
    """Interface visual clean para selecionar a fatura e executar a extração."""
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    # Tema clean moderno escuro (degradê em tons grafite/índigo)
    C_NAVY = "#0b1020"
    C_NAVY_CARD = "#1a2340"
    C_BG = "#0f172a"
    C_SURFACE = "#111827"
    C_BORDER = "#1f2937"
    C_BORDER_STRONG = "#334155"
    C_TEXT = "#e5e7eb"
    C_MUTED = "#93a3b8"
    C_MUTED_LIGHT = "#cbd5e1"
    C_ACCENT = "#6366f1"
    C_ACCENT_HOVER = "#4f46e5"
    C_ACCENT_SOFT = "#1e293b"
    C_BTN_SECONDARY_BG = "#1f2937"
    C_BTN_SECONDARY_HOVER = "#273449"
    C_DROP_INNER = "#0b1220"
    C_HEADING_ON_NAVY = "#f8fafc"
    C_ACCENT_LINE = "#7c3aed"
    C_BADGE_FG = "#bfdbfe"

    root = tk.Tk()
    width, height = 920, 720
    if sys.platform == "win32":
        try:
            import ctypes

            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("com.masterport.yaro.extrator")
        except Exception:
            pass
    app_icon_path = resolver_recurso_app("yaro.ico")
    if app_icon_path is not None:
        try:
            root.iconbitmap(default=str(app_icon_path))
        except Exception:
            pass
    root.title("MasterPort Comex | Pneus — Extrator YARO")
    root.resizable(False, False)
    root.configure(bg=C_BG)
    root.update_idletasks()
    pos_x = (root.winfo_screenwidth() - width) // 2
    pos_y = (root.winfo_screenheight() - height) // 3
    root.geometry(f"{width}x{height}+{max(0, pos_x)}+{max(0, pos_y)}")

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure(
        "Modern.TCombobox",
        fieldbackground=C_SURFACE,
        background=C_SURFACE,
        foreground=C_TEXT,
        arrowcolor=C_MUTED,
        bordercolor=C_BORDER_STRONG,
        padding=(10, 8),
    )
    style.map(
        "Modern.TCombobox",
        fieldbackground=[("readonly", C_SURFACE)],
        selectbackground=[("readonly", C_ACCENT_SOFT)],
        selectforeground=[("readonly", C_TEXT)],
    )
    style.configure(
        "App.Primary.TButton",
        background=C_ACCENT,
        foreground="#ffffff",
        borderwidth=0,
        focuscolor="none",
        padding=(20, 11),
        font=("Segoe UI", 10, "bold"),
    )
    style.map(
        "App.Primary.TButton",
        background=[("active", C_ACCENT_HOVER), ("disabled", "#475569")],
        foreground=[("disabled", "#cbd5e1")],
    )
    style.configure(
        "App.Secondary.TButton",
        background=C_BTN_SECONDARY_BG,
        foreground=C_TEXT,
        borderwidth=0,
        focuscolor="none",
        padding=(18, 11),
        font=("Segoe UI", 10),
    )
    style.map(
        "App.Secondary.TButton",
        background=[("active", C_BTN_SECONDARY_HOVER)],
        foreground=[("active", C_TEXT)],
    )
    style.configure(
        "App.Ghost.TButton",
        background=C_SURFACE,
        foreground=C_MUTED,
        borderwidth=0,
        focuscolor="none",
        padding=(16, 11),
        font=("Segoe UI", 10),
    )
    style.map(
        "App.Ghost.TButton",
        background=[("active", C_DROP_INNER)],
        foreground=[("active", C_TEXT)],
    )

    caminho_var = tk.StringVar()
    status_var = tk.StringVar(value="Aguardando a Commercial Invoice (PDF).")
    pasta_saida_var = tk.StringVar(value="")
    nome_arquivo_var = tk.StringVar(value="")
    opcoes_fornecedor_ui = (
        "Detectar automaticamente",
        "Atlantic",
        "Latitude",
        "Omni",
    )
    mapa_fornecedor = {
        "Detectar automaticamente": "auto",
        "Atlantic": "atlantic",
        "Latitude": "latitude",
        "Omni": "omni",
    }
    fornecedor_var = tk.StringVar(value=opcoes_fornecedor_ui[0])
    _loading_ctl: dict = {"start": lambda: None, "stop": lambda: None}
    _extracao_q: queue.Queue = queue.Queue()
    _extracao_poll_after: list[Optional[str]] = [None]

    def _caminho_de_drop(item) -> str:
        if isinstance(item, str):
            return item.strip().strip("\x00")
        for enc in ("utf-8", sys.getfilesystemencoding() or "mbcs", "latin-1"):
            try:
                return item.decode(enc).strip().strip("\x00")
            except (UnicodeDecodeError, AttributeError):
                continue
        return item.decode("utf-8", errors="replace").strip().strip("\x00")

    def aplicar_arquivo_escolhido(caminho: str) -> None:
        caminho_var.set(caminho)
        nome_arquivo_var.set(Path(caminho).name)
        status_var.set("PDF selecionado. Clique em Extrair dados.")

    def selecionar_arquivo() -> None:
        arquivo = filedialog.askopenfilename(
            title="Selecione a Commercial Invoice (PDF)",
            filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")],
        )
        if arquivo:
            aplicar_arquivo_escolhido(arquivo)

    def abrir_pasta_saida() -> None:
        pasta = pasta_saida_var.get().strip()
        if not pasta:
            messagebox.showinfo("Saída", "Execute uma extração para gerar os arquivos primeiro.")
            return
        if not Path(pasta).is_dir():
            messagebox.showerror("Saída", f"Pasta não encontrada: {pasta}")
            return
        os.startfile(pasta)

    logo_img_tk = None
    icon_img_tk = None
    logo_path = resolver_recurso_app("yaro.jpg")
    if logo_path:
        try:
            from PIL import Image, ImageTk

            img = Image.open(logo_path)
            icon_img = img.copy()
            if hasattr(Image, "Resampling"):
                icon_img.thumbnail((32, 32), Image.Resampling.LANCZOS)
            else:
                icon_img.thumbnail((32, 32), Image.LANCZOS)
            icon_img_tk = ImageTk.PhotoImage(icon_img)
            root.iconphoto(True, icon_img_tk)
            if hasattr(Image, "Resampling"):
                img.thumbnail((180, 64), Image.Resampling.LANCZOS)
            else:
                img.thumbnail((180, 64), Image.LANCZOS)
            logo_img_tk = ImageTk.PhotoImage(img)
            # Mantém referência viva para o Tk não descartar a imagem.
            root._logo_img_tk = logo_img_tk
            root._icon_img_tk = icon_img_tk
        except Exception:
            logo_img_tk = None

    # pady no Frame não aceita tupla em algumas versões do Tk (ex.: Python 3.14).
    cabecalho = tk.Frame(root, bg=C_NAVY, padx=32, pady=20)
    cabecalho.pack(fill="x")
    cabecalho_topo = tk.Frame(cabecalho, bg=C_NAVY)
    cabecalho_topo.pack(fill="x")
    cabecalho_left = tk.Frame(cabecalho_topo, bg=C_NAVY)
    cabecalho_left.pack(side="left", fill="x", expand=True)
    faixa = tk.Frame(cabecalho_left, bg=C_NAVY_CARD, padx=14, pady=6)
    faixa.pack(anchor="w")
    tk.Label(
        faixa,
        text="COMÉRCIO EXTERIOR  ·  IMPORTAÇÃO DE PNEUS",
        bg=C_NAVY_CARD,
        fg=C_BADGE_FG,
        font=("Segoe UI", 8, "bold"),
    ).pack(anchor="w")
    tk.Label(
        cabecalho_left,
        text="MasterPort para YARO",
        bg=C_NAVY,
        fg=C_HEADING_ON_NAVY,
        font=("Segoe UI", 11),
    ).pack(anchor="w", pady=(12, 0))
    tk.Label(
        cabecalho_left,
        text="Extrator de Commercial Invoice",
        bg=C_NAVY,
        fg=C_HEADING_ON_NAVY,
        font=("Segoe UI", 20, "bold"),
    ).pack(anchor="w", pady=(4, 0))
    if logo_img_tk is not None:
        tk.Label(
            cabecalho_topo,
            image=logo_img_tk,
            bg=C_NAVY,
        ).pack(side="right", anchor="ne", padx=(12, 0), pady=(4, 0))
    tk.Frame(root, bg=C_ACCENT_LINE, height=3).pack(fill="x")

    shell = tk.Frame(root, bg=C_BG, padx=32, pady=24)
    shell.pack(fill="both", expand=True)

    corpo = tk.Frame(shell, bg=C_BG)
    corpo.pack(fill="both", expand=True, pady=(22, 0))

    # Cartão só na altura do conteúdo (sem expand vertical); botões não somem no rodapé.
    card_rim = tk.Frame(corpo, bg=C_BORDER, padx=1, pady=1)
    card_rim.pack(fill="x", anchor="nw")
    card = tk.Frame(card_rim, bg=C_SURFACE, padx=26, pady=24)
    card.pack(fill="x", anchor="nw")

    tk.Label(
        card,
        text="Documento de embarque",
        bg=C_SURFACE,
        fg=C_TEXT,
        font=("Segoe UI", 12, "bold"),
    ).pack(anchor="w")
    tk.Label(
        card,
        text="PDF da invoice do exportador (ex.: Atlantic, Latitude, Omni). Mesmo arquivo usado no processo comex.",
        bg=C_SURFACE,
        fg=C_MUTED_LIGHT,
        font=("Segoe UI", 10),
    ).pack(anchor="w", pady=(2, 14))

    linha_forn = tk.Frame(card, bg=C_SURFACE)
    linha_forn.pack(fill="x", pady=(0, 14))
    tk.Label(
        linha_forn,
        text="Fornecedor",
        bg=C_SURFACE,
        fg=C_TEXT,
        font=("Segoe UI", 10, "bold"),
    ).pack(side="left")
    combo_fornecedor = ttk.Combobox(
        linha_forn,
        textvariable=fornecedor_var,
        values=opcoes_fornecedor_ui,
        state="readonly",
        width=34,
        font=("Segoe UI", 10),
        style="Modern.TCombobox",
    )
    combo_fornecedor.pack(side="right")

    drop_outer = tk.Frame(card, bg=C_BORDER, padx=1, pady=1)
    drop_outer.pack(fill="x")
    drop = tk.Frame(drop_outer, bg=C_DROP_INNER, padx=22, pady=22)
    drop.pack(fill="x")
    icone_wrap = tk.Frame(drop, bg=C_ACCENT, width=52, height=52)
    icone_wrap.pack(pady=(2, 12))
    icone_wrap.pack_propagate(False)
    icone_seta = tk.Label(
        icone_wrap, text="↑", bg=C_ACCENT, fg="#ffffff", font=("Segoe UI", 20), cursor="hand2"
    )
    icone_seta.place(relx=0.5, rely=0.5, anchor="center")
    texto_drop = tk.Label(
        drop,
        text="Arraste a Commercial Invoice ou clique para selecionar",
        bg=C_DROP_INNER,
        fg=C_TEXT,
        cursor="hand2",
        font=("Segoe UI", 13),
    )
    texto_drop.pack()
    drop_sub = tk.Label(
        drop,
        text="Gera planilha e JSON para cruzar com catálogo, certificado INMETRO e NCM 4011 (pneus).",
        bg=C_DROP_INNER,
        fg=C_MUTED,
        font=("Segoe UI", 9),
        cursor="hand2",
    )
    drop_sub.pack(pady=(6, 0))
    nome_label = tk.Label(
        drop,
        textvariable=nome_arquivo_var,
        bg=C_DROP_INNER,
        fg=C_ACCENT,
        font=("Segoe UI", 10, "bold"),
    )
    nome_label.pack(pady=(12, 0))
    for w in (drop, texto_drop, nome_label, icone_wrap, icone_seta, drop_sub):
        w.bind("<Button-1>", lambda _e: selecionar_arquivo())

    # Arrastar e soltar (Windows): tk não oferece DnD nativo; usamos windnd nos HWNDs da zona.
    if sys.platform == "win32":
        try:
            import windnd
        except ImportError:
            windnd = None  # type: ignore
        if windnd is not None:
            _drop_after_id: list[Optional[str]] = [None]

            def _ao_soltar_arquivos(file_list) -> None:
                if not file_list:
                    return

                def processar() -> None:
                    _drop_after_id[0] = None
                    candidato = None
                    for raw in file_list:
                        p = _caminho_de_drop(raw)
                        if p.lower().endswith(".pdf"):
                            candidato = p
                            break
                        candidato = candidato or p
                    if not candidato:
                        return
                    if not candidato.lower().endswith(".pdf"):
                        messagebox.showwarning("Formato", "Solte apenas arquivos PDF (.pdf).")
                        return
                    if not Path(candidato).is_file():
                        messagebox.showerror("Arquivo", f"Não foi possível abrir:\n{candidato}")
                        return
                    aplicar_arquivo_escolhido(candidato)

                if _drop_after_id[0] is not None:
                    try:
                        root.after_cancel(_drop_after_id[0])
                    except Exception:
                        pass
                _drop_after_id[0] = root.after(20, processar)

            alvos_drop = (drop_outer, drop, icone_wrap, icone_seta, texto_drop, drop_sub, nome_label)
            for alvo in alvos_drop:
                try:
                    windnd.hook_dropfiles(alvo, func=_ao_soltar_arquivos, force_unicode=True)
                except Exception:
                    pass

    loading_box = tk.Frame(card, bg=C_SURFACE)
    loading_canvas = tk.Canvas(
        loading_box,
        height=72,
        bg="#0b1220",
        highlightthickness=1,
        highlightbackground=C_BORDER,
        bd=0,
    )
    loading_canvas.pack(fill="x")
    tk.Label(
        loading_box,
        text="Carregando · o navio indica que a extração está em andamento",
        bg=C_SURFACE,
        fg=C_TEXT,
        font=("Segoe UI", 9, "italic"),
    ).pack(anchor="w", pady=(4, 0))

    _navio_anim: dict = {
        "after_id": None,
        "ativo": False,
        "ship_x": 28.0,
        "dv": 1.0,
        "dv_mag": 1.0,
        "fundo_w": None,
    }
    _MS_NAVIO = 72
    _MS_POLL_EXTRACAO = 120
    _NAVIO_TRAVESSIA_MS = 2000  # um percurso até a borda (~2 s)

    def _garantir_fundo_mar(c: tk.Canvas, w: int, h: int) -> None:
        if _navio_anim.get("fundo_w") == w and c.find_withtag("fundo"):
            return
        _navio_anim["fundo_w"] = w
        c.delete("fundo")
        c.create_rectangle(0, 0, w, h, fill="#111827", width=0, tags=("fundo",))
        y_onda = h * 0.52
        c.create_rectangle(0, y_onda, w, h, fill="#1e3a8a", width=0, tags=("fundo",))
        c.create_line(0, y_onda, w, y_onda, fill="#3b82f6", width=2, tags=("fundo",))

    def _pintar_apenas_navio(c: tk.Canvas, hx: float, h: int) -> None:
        c.delete("navio")
        c.create_polygon(
            hx,
            h * 0.48,
            hx + 88,
            h * 0.48,
            hx + 80,
            h * 0.63,
            hx + 8,
            h * 0.63,
            fill=C_NAVY,
            outline=C_NAVY_CARD,
            width=1,
            tags=("navio",),
        )
        c.create_rectangle(
            hx + 40, h * 0.36, hx + 68, h * 0.48, fill=C_NAVY_CARD, outline="", tags=("navio",)
        )
        c.create_line(
            hx + 54, h * 0.36, hx + 54, h * 0.26, fill=C_ACCENT_LINE, width=2, tags=("navio",)
        )
        c.create_rectangle(
            hx + 18, h * 0.42, hx + 34, h * 0.48, fill="#fbbf24", outline="", tags=("navio",)
        )
        c.create_rectangle(
            hx + 70,
            h * 0.42,
            hx + 86,
            h * 0.48,
            fill="#f1f5f9",
            outline=C_BORDER_STRONG,
            width=1,
            tags=("navio",),
        )

    def _tick_navio_carregamento() -> None:
        if not _navio_anim["ativo"]:
            return
        c = loading_canvas
        w = max(int(c.winfo_width()), 320)
        h = 72
        _garantir_fundo_mar(c, w, h)
        dv_mag = float(_navio_anim["dv_mag"])
        _navio_anim["ship_x"] += _navio_anim["dv"]
        if _navio_anim["ship_x"] > w - 100:
            _navio_anim["ship_x"] = float(w - 100)
            _navio_anim["dv"] = -dv_mag
        elif _navio_anim["ship_x"] < 18:
            _navio_anim["ship_x"] = 18.0
            _navio_anim["dv"] = dv_mag
        _pintar_apenas_navio(c, _navio_anim["ship_x"], h)
        _navio_anim["after_id"] = root.after(_MS_NAVIO, _tick_navio_carregamento)

    def _loading_start() -> None:
        if _navio_anim["after_id"] is not None:
            try:
                root.after_cancel(_navio_anim["after_id"])
            except Exception:
                pass
            _navio_anim["after_id"] = None
        _navio_anim["fundo_w"] = None
        _navio_anim["ativo"] = True
        _navio_anim["ship_x"] = 28.0
        loading_box.pack(fill="x", pady=(20, 10), after=drop_outer)
        loading_canvas.update_idletasks()
        w0 = max(int(loading_canvas.winfo_width()), 320)
        # Percurso ship_x 28 → limite direito w-100, em ~2 s.
        perna_px = max(float(w0 - 100) - 28.0, 80.0)
        _navio_anim["dv_mag"] = perna_px * _MS_NAVIO / float(_NAVIO_TRAVESSIA_MS)
        _navio_anim["dv"] = _navio_anim["dv_mag"]
        _garantir_fundo_mar(loading_canvas, w0, 72)
        _pintar_apenas_navio(loading_canvas, _navio_anim["ship_x"], 72)
        _navio_anim["after_id"] = root.after(_MS_NAVIO, _tick_navio_carregamento)

    def _loading_stop() -> None:
        _navio_anim["ativo"] = False
        _navio_anim["fundo_w"] = None
        if _navio_anim["after_id"] is not None:
            try:
                root.after_cancel(_navio_anim["after_id"])
            except Exception:
                pass
            _navio_anim["after_id"] = None
        try:
            loading_canvas.delete("all")
        except tk.TclError:
            pass
        loading_box.pack_forget()

    _loading_ctl["start"] = _loading_start
    _loading_ctl["stop"] = _loading_stop

    def _mostrar_janela_conclusao(saida_json: Path, saida_excel: Path, dados: dict) -> None:
        win = tk.Toplevel(root)
        win.title("Embarque concluído · MasterPort")
        win.transient(root)
        win.resizable(False, False)
        win.configure(bg=C_SURFACE)
        win.grab_set()

        W, H = 540, 460
        head = tk.Frame(win, bg=C_NAVY, padx=22, pady=18)
        head.pack(fill="x")
        tk.Label(
            head,
            text="✓  Operação concluída",
            fg=C_ACCENT_LINE,
            bg=C_NAVY,
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")
        tk.Label(
            head,
            text="Documentação gerada com sucesso",
            fg=C_HEADING_ON_NAVY,
            bg=C_NAVY,
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w", pady=(8, 0))
        tk.Label(
            head,
            text="Arquivos prontos para conferência comex e cadastro",
            fg=C_MUTED_LIGHT,
            bg=C_NAVY,
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(4, 0))

        tk.Frame(win, bg=C_ACCENT_LINE, height=3).pack(fill="x")

        body = tk.Frame(win, bg=C_SURFACE, padx=22, pady=20)
        body.pack(fill="both", expand=True)

        ok_canvas = tk.Canvas(body, width=72, height=72, bg=C_SURFACE, highlightthickness=0)
        ok_canvas.pack(pady=(4, 14))
        ok_canvas.create_oval(4, 4, 68, 68, fill=C_ACCENT_SOFT, outline=C_ACCENT, width=2)
        ok_canvas.create_text(36, 36, text="✓", fill=C_ACCENT_HOVER, font=("Segoe UI", 32, "bold"))

        n_itens = len(dados.get("itens", []))
        tk.Label(
            body,
            text=f"{n_itens} itens extraídos da invoice",
            bg=C_SURFACE,
            fg=C_TEXT,
            font=("Segoe UI", 12, "bold"),
        ).pack(anchor="w")

        def linha_arquivo(titulo: str, caminho: Path) -> None:
            fr = tk.Frame(body, bg=C_DROP_INNER, highlightthickness=1, highlightbackground=C_BORDER)
            fr.pack(fill="x", pady=(12, 0))
            tk.Label(
                fr,
                text=titulo,
                bg=C_DROP_INNER,
                fg=C_TEXT,
                font=("Segoe UI", 9, "bold"),
            ).pack(anchor="w", padx=12, pady=(10, 0))
            tk.Label(
                fr,
                text=str(caminho),
                bg=C_DROP_INNER,
                fg=C_MUTED,
                font=("Segoe UI", 9),
                wraplength=480,
                justify="left",
            ).pack(anchor="w", padx=12, pady=(2, 10))

        linha_arquivo("JSON — dados estruturados", saida_json)
        linha_arquivo("Excel — fatura formatada", saida_excel)

        bf = tk.Frame(body, bg=C_SURFACE)
        bf.pack(fill="x", pady=(24, 0))

        def abrir_pasta_saida_dialogo() -> None:
            p = saida_excel.parent
            if p.is_dir():
                os.startfile(str(p))

        def fechar_conclusao() -> None:
            try:
                win.grab_release()
            except Exception:
                pass
            win.destroy()

        ttk.Button(
            bf,
            text="Abrir pasta de saída",
            command=abrir_pasta_saida_dialogo,
            style="App.Secondary.TButton",
        ).pack(side="left")
        ttk.Button(
            bf,
            text="OK",
            command=fechar_conclusao,
            style="App.Primary.TButton",
        ).pack(side="right", padx=(12, 0))

        win.protocol("WM_DELETE_WINDOW", fechar_conclusao)

        win.update_idletasks()
        rx, ry = root.winfo_rootx(), root.winfo_rooty()
        rw, rh = root.winfo_width(), root.winfo_height()
        win.geometry(f"{W}x{H}+{rx + max(0, (rw - W) // 2)}+{ry + max(0, (rh - H) // 2)}")
        win.focus_set()

    def _finalizar_extracao_na_ui(ok: bool, payload) -> None:
        _loading_ctl["stop"]()
        botao_extrair.config(state="normal")
        botao_buscar.config(state="normal")
        botao_fechar.config(state="normal")
        combo_fornecedor.config(state="readonly")
        if ok:
            saida_json, saida_excel, dados = payload
            pasta_saida_var.set(str(Path(saida_excel).parent))
            status_var.set("Extração concluída com sucesso.")
            _mostrar_janela_conclusao(Path(saida_json), Path(saida_excel), dados)
        else:
            status_var.set("Falha na extração.")
            messagebox.showerror("Erro na extração", str(payload))

    def _poll_fila_extracao() -> None:
        try:
            tag, payload = _extracao_q.get_nowait()
        except queue.Empty:
            _extracao_poll_after[0] = root.after(_MS_POLL_EXTRACAO, _poll_fila_extracao)
            return
        _extracao_poll_after[0] = None
        _finalizar_extracao_na_ui(tag == "ok", payload)

    def executar_extracao() -> None:
        caminho = caminho_var.get().strip()
        if not caminho:
            messagebox.showwarning("Aviso", "Selecione um PDF antes de extrair.")
            return
        while True:
            try:
                _extracao_q.get_nowait()
            except queue.Empty:
                break
        if _extracao_poll_after[0] is not None:
            try:
                root.after_cancel(_extracao_poll_after[0])
            except Exception:
                pass
            _extracao_poll_after[0] = None

        status_var.set("Processando embarque da documentação…")
        botao_extrair.config(state="disabled")
        botao_buscar.config(state="disabled")
        botao_fechar.config(state="disabled")
        combo_fornecedor.config(state="disabled")
        _loading_ctl["start"]()
        root.update_idletasks()

        codigo_forn = mapa_fornecedor.get(fornecedor_var.get().strip(), "auto")
        pdf_p = Path(caminho)

        def worker() -> None:
            try:
                res = processar_pdf(pdf_p, fornecedor=codigo_forn)
                _extracao_q.put(("ok", res))
            except Exception as e:
                _extracao_q.put(("err", e))

        threading.Thread(target=worker, daemon=True).start()
        _poll_fila_extracao()

    botoes = tk.Frame(card, bg=C_SURFACE)
    botoes.pack(fill="x", pady=(8, 0))
    botao_extrair = ttk.Button(
        botoes,
        text="Extrair invoice",
        command=executar_extracao,
        style="App.Primary.TButton",
    )
    botao_extrair.pack(side="left")

    botao_buscar = ttk.Button(
        botoes,
        text="Abrir pasta de saída",
        command=abrir_pasta_saida,
        style="App.Secondary.TButton",
    )
    botao_buscar.pack(side="left", padx=(12, 0))

    botao_fechar = ttk.Button(
        botoes,
        text="Fechar",
        command=root.destroy,
        style="App.Ghost.TButton",
    )
    botao_fechar.pack(side="right")

    tk.Label(
        card,
        textvariable=status_var,
        anchor="w",
        bg=C_SURFACE,
        fg=C_MUTED,
        font=("Segoe UI", 10),
    ).pack(fill="x", pady=(16, 0))
    tk.Label(
        shell,
        text="Saída: pasta “Saida” — use os arquivos na conferência aduaneira e no cadastro dos pneus.",
        bg=C_BG,
        fg=C_MUTED_LIGHT,
        font=("Segoe UI", 9),
        justify="left",
        wraplength=width - 64,
    ).pack(anchor="w", pady=(16, 0))

    root.mainloop()


def _modo_plataforma() -> bool:
    """True quando executado pelo card de automação do KIVO (upload via site)."""
    return bool(os.environ.get("OPERACOES_INPUT_FOLDER", "").strip())


def _listar_pdfs_entrada(pasta: Path) -> list[Path]:
    encontrados: list[Path] = []
    for item in pasta.iterdir():
        if item.is_file() and item.suffix.lower() == ".pdf":
            encontrados.append(item)
    return sorted(encontrados, key=lambda p: p.name.lower())


def _imprimir_resumo(dados: dict) -> None:
    print("\n--- Resumo ---")
    print("Número:", dados.get("cabecalho", {}).get("numero"))
    print("Data:", dados.get("cabecalho", {}).get("data"))
    print("Importador:", dados.get("importador", {}).get("nome"))
    print("Total itens:", len(dados.get("itens", [])))
    print("Valor total:", dados.get("totais", {}).get("valor_total"))


def _import_runtime_paths():
    ops_dir = Path(__file__).resolve().parents[3]
    ops_path = str(ops_dir)
    if ops_path not in sys.path:
        sys.path.insert(0, ops_path)
    from runtime_paths import get_form_value, get_slot_files, resolve_input_folder

    return get_form_value, get_slot_files, resolve_input_folder


def _executar_plataforma() -> None:
    """Processa PDF(s) enviados pelo card Operações e grava saída para download no site."""
    get_form_value, get_slot_files, resolve_input_folder = _import_runtime_paths()

    input_folder_raw = resolve_input_folder()
    if not input_folder_raw:
        raise FileNotFoundError("Pasta de entrada inválida (OPERACOES_INPUT_FOLDER).")
    input_folder = Path(input_folder_raw)
    output_path = os.environ.get("OPERACOES_OUTPUT_PATH", "").strip()
    fornecedor = (
        get_form_value("fornecedor")
        or os.environ.get("OPERACOES_FORNECEDOR", "auto").strip()
        or "auto"
    )

    pdfs: list[Path] = []
    for slot in ("fatura", "pdf", "arquivo", "anexo"):
        pdfs = [Path(path) for path in get_slot_files(slot) if str(path).lower().endswith(".pdf")]
        if pdfs:
            break
    if not pdfs:
        pdfs = _listar_pdfs_entrada(input_folder)
    if not pdfs:
        raise FileNotFoundError("Anexe pelo menos um PDF de fatura comercial (.pdf).")

    pasta = pasta_execucao()
    ultimo_json: Path | None = None
    ultimo_excel: Path | None = None
    ultimos_dados: dict | None = None

    for pdf_path in pdfs:
        print(f"Processando: {pdf_path.name}")
        saida_json, saida_excel, dados = processar_pdf(pdf_path, pasta, fornecedor=fornecedor)
        ultimo_json, ultimo_excel, ultimos_dados = saida_json, saida_excel, dados
        print(f"JSON: {saida_json}")
        print(f"Excel: {saida_excel}")

    if ultimo_excel is None or ultimos_dados is None:
        raise RuntimeError("Nenhum arquivo foi gerado.")

    destino_excel = ultimo_excel
    if output_path:
        destino = Path(output_path)
        destino.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(ultimo_excel, destino)
        destino_excel = destino
        if ultimo_json and ultimo_json.is_file():
            destino_json = destino.parent / ultimo_json.name
            shutil.copy2(ultimo_json, destino_json)
            print(f"JSON de apoio: {destino_json}")

    print(f"Arquivo Excel gerado: {destino_excel}")
    _imprimir_resumo(ultimos_dados)


def _resolver_pdf_cli(argv: list[str]) -> Optional[Path]:
    """
    Resolve caminho do PDF recebido via linha de comando.
    Tolera cenários onde o caminho chega quebrado em múltiplos argumentos.
    """
    if not argv:
        return None
    candidatos = []
    if len(argv) > 1:
        candidatos.append(" ".join(argv))
    candidatos.append(argv[0])
    for raw in candidatos:
        p = Path(str(raw).strip().strip('"').strip("'"))
        if p.is_file():
            return p
    return None


def main():
    if _modo_plataforma():
        _executar_plataforma()
        return

    pasta = pasta_execucao()
    if len(sys.argv) > 1:
        pdf_path = _resolver_pdf_cli(sys.argv[1:])
        if pdf_path is None:
            raise FileNotFoundError(f"Arquivo não encontrado: {' '.join(sys.argv[1:]).strip()}")
        print(f"Processando: {pdf_path.name}")
        saida_json, saida_excel, dados = processar_pdf(pdf_path, pasta)
    else:
        abrir_interface()
        return

    print(f"JSON: {saida_json}")
    print(f"Excel: {saida_excel}")
    print(f"Arquivo Excel gerado: {saida_excel}")
    _imprimir_resumo(dados)


if __name__ == "__main__":
    main()
